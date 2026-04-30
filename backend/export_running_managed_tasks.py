import json
import re
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "story_creator.db"
EXPORT_DIR = BASE_DIR / "exports"

STATUS_FILLS = {
    "pending": "FFF4CC",
    "processing": "D9EAF7",
    "completed": "D9EAD3",
    "failed": "F4CCCC",
}


def _fetch_running_managed_rows():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    rows = cur.execute(
        """
        SELECT
            u.id AS user_id,
            u.username AS username,
            s.id AS script_id,
            e.id AS episode_id,
            ms.id AS session_id,
            ms.status AS session_status,
            ms.provider AS provider,
            ms.variant_count AS variant_count,
            ms.total_shots AS total_shots,
            ms.completed_shots AS completed_shots,
            ms.created_at AS session_created_at,
            ms.completed_at AS session_completed_at,
            mt.id AS managed_task_id,
            mt.status AS task_status,
            mt.shot_id AS result_shot_id,
            mt.shot_stable_id AS shot_stable_id,
            mt.video_path AS video_path,
            mt.error_message AS error_message,
            mt.task_id AS upstream_task_id,
            mt.created_at AS task_created_at,
            mt.completed_at AS task_completed_at,
            rs.shot_number AS result_shot_number,
            rs.variant_index AS result_variant_index,
            orig.id AS original_shot_id,
            orig.shot_number AS original_shot_number
        FROM managed_tasks mt
        JOIN managed_sessions ms ON mt.session_id = ms.id
        JOIN episodes e ON ms.episode_id = e.id
        JOIN scripts s ON e.script_id = s.id
        JOIN users u ON s.user_id = u.id
        LEFT JOIN storyboard_shots rs ON mt.shot_id = rs.id
        LEFT JOIN storyboard_shots orig
            ON orig.stable_id = mt.shot_stable_id
           AND orig.variant_index = 0
        WHERE ms.status = 'running'
        ORDER BY u.username ASC, ms.id ASC, mt.id ASC
        """
    ).fetchall()
    conn.close()
    return rows


def _group_rows(rows):
    export = {
        "exported_at": datetime.now().isoformat(),
        "database_path": str(DB_PATH),
        "session_status_filter": "running",
        "user_count": 0,
        "session_count": 0,
        "task_count": 0,
        "status_counts": {
            "pending": 0,
            "processing": 0,
            "completed": 0,
            "failed": 0,
        },
        "users": [],
    }

    users = {}
    session_seen = set()

    for row in rows:
        user_key = str(row["user_id"])
        username = row["username"] or f"user_{row['user_id']}"
        if user_key not in users:
            users[user_key] = {
                "user_id": row["user_id"],
                "username": username,
                "script_ids": set(),
                "episode_ids": set(),
                "session_count": 0,
                "task_count": 0,
                "status_counts": {
                    "pending": 0,
                    "processing": 0,
                    "completed": 0,
                    "failed": 0,
                },
                "sessions": {},
            }

        user_bucket = users[user_key]
        user_bucket["script_ids"].add(row["script_id"])
        user_bucket["episode_ids"].add(row["episode_id"])
        user_bucket["task_count"] += 1

        task_status = str(row["task_status"] or "").strip().lower()
        if task_status in user_bucket["status_counts"]:
            user_bucket["status_counts"][task_status] += 1
        if task_status in export["status_counts"]:
            export["status_counts"][task_status] += 1

        session_key = str(row["session_id"])
        if session_key not in user_bucket["sessions"]:
            user_bucket["sessions"][session_key] = {
                "session_id": row["session_id"],
                "script_id": row["script_id"],
                "episode_id": row["episode_id"],
                "status": row["session_status"],
                "provider": row["provider"],
                "variant_count": row["variant_count"],
                "total_shots": row["total_shots"],
                "completed_shots": row["completed_shots"],
                "created_at": row["session_created_at"],
                "completed_at": row["session_completed_at"],
                "task_count": 0,
                "status_counts": {
                    "pending": 0,
                    "processing": 0,
                    "completed": 0,
                    "failed": 0,
                },
                "tasks": [],
            }
            user_bucket["session_count"] += 1
            session_seen.add(row["session_id"])

        session_bucket = user_bucket["sessions"][session_key]
        session_bucket["task_count"] += 1
        if task_status in session_bucket["status_counts"]:
            session_bucket["status_counts"][task_status] += 1

        session_bucket["tasks"].append({
            "managed_task_id": row["managed_task_id"],
            "status": row["task_status"],
            "result_shot_id": row["result_shot_id"],
            "result_shot_number": row["result_shot_number"],
            "result_variant_index": row["result_variant_index"],
            "original_shot_id": row["original_shot_id"],
            "original_shot_number": row["original_shot_number"],
            "shot_stable_id": row["shot_stable_id"],
            "upstream_task_id": row["upstream_task_id"],
            "video_path": row["video_path"],
            "error_message": row["error_message"],
            "created_at": row["task_created_at"],
            "completed_at": row["task_completed_at"],
        })

    for user in users.values():
        user["script_ids"] = sorted(user["script_ids"])
        user["episode_ids"] = sorted(user["episode_ids"])
        user["sessions"] = list(user["sessions"].values())
        export["users"].append(user)

    export["users"].sort(key=lambda item: str(item["username"]).lower())
    export["user_count"] = len(export["users"])
    export["session_count"] = len(session_seen)
    export["task_count"] = len(rows)
    return export


def _safe_sheet_name(name, used_names):
    base = re.sub(r"[:\\\\/*?\\[\\]]", "_", str(name or "Sheet")).strip() or "Sheet"
    base = base[:31]
    candidate = base
    index = 2
    while candidate in used_names:
        suffix = f"_{index}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        index += 1
    used_names.add(candidate)
    return candidate


def _style_sheet(ws, header_fill="1F4E78", freeze="A2"):
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            value_lines = value.splitlines() or [value]
            max_line = max(len(line) for line in value_lines)
            length = max(length, max_line)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 48)


def _task_result_label(task):
    shot_number = task.get("result_shot_number")
    variant_index = task.get("result_variant_index")
    if shot_number is None:
        return "待生成"
    try:
        shot_number = int(shot_number)
        variant_index = int(variant_index or 0)
    except (TypeError, ValueError):
        return str(shot_number)
    if variant_index > 0:
        return f"{shot_number}_{variant_index}"
    return str(shot_number)


def _append_summary_sheet(wb, export):
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ["导出时间", export["exported_at"]],
        ["数据库", export["database_path"]],
        ["托管状态过滤", export["session_status_filter"]],
        ["用户数", export["user_count"]],
        ["运行中会话数", export["session_count"]],
        ["运行中会话任务数", export["task_count"]],
        ["Pending", export["status_counts"]["pending"]],
        ["Processing", export["status_counts"]["processing"]],
        ["Completed", export["status_counts"]["completed"]],
        ["Failed", export["status_counts"]["failed"]],
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80


def _append_sessions_sheet(wb, export):
    ws = wb.create_sheet("Sessions")
    ws.append([
        "用户名", "用户ID", "剧本ID", "片段ID", "会话ID", "状态", "服务商",
        "variant_count", "total_shots", "completed_shots",
        "任务总数", "pending", "processing", "completed", "failed",
        "创建时间", "完成时间"
    ])

    for user in export["users"]:
        for session in user["sessions"]:
            ws.append([
                user["username"],
                user["user_id"],
                session["script_id"],
                session["episode_id"],
                session["session_id"],
                session["status"],
                session["provider"],
                session["variant_count"],
                session["total_shots"],
                session["completed_shots"],
                session["task_count"],
                session["status_counts"]["pending"],
                session["status_counts"]["processing"],
                session["status_counts"]["completed"],
                session["status_counts"]["failed"],
                session["created_at"],
                session["completed_at"] or "",
            ])

    _style_sheet(ws)


def _append_tasks_sheet(wb, title, tasks_rows):
    ws = wb.create_sheet(title)
    ws.append([
        "用户名", "用户ID", "会话ID", "片段ID", "服务商",
        "托管任务ID", "任务状态", "原始镜头", "结果镜头",
        "结果镜头ID", "Stable ID", "上游任务ID",
        "创建时间", "完成时间", "错误信息", "视频地址"
    ])

    for row in tasks_rows:
        ws.append(row)

    _style_sheet(ws)

    status_col = 7
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status = str(row[status_col - 1].value or "").strip().lower()
        fill_color = STATUS_FILLS.get(status)
        if fill_color:
            row[status_col - 1].fill = PatternFill("solid", fgColor=fill_color)


def _build_task_rows(export):
    rows = []
    rows_by_user = defaultdict(list)
    for user in export["users"]:
        for session in user["sessions"]:
            for task in session["tasks"]:
                row = [
                    user["username"],
                    user["user_id"],
                    session["session_id"],
                    session["episode_id"],
                    session["provider"],
                    task["managed_task_id"],
                    task["status"],
                    task["original_shot_number"] if task["original_shot_number"] is not None else "",
                    _task_result_label(task),
                    task["result_shot_id"] if task["result_shot_id"] is not None else "",
                    task["shot_stable_id"] or "",
                    task["upstream_task_id"] or "",
                    task["created_at"] or "",
                    task["completed_at"] or "",
                    task["error_message"] or "",
                    task["video_path"] or "",
                ]
                rows.append(row)
                rows_by_user[user["username"]].append(row)
    return rows, rows_by_user


def _write_xlsx(export, output_path):
    wb = Workbook()
    _append_summary_sheet(wb, export)
    _append_sessions_sheet(wb, export)

    all_task_rows, rows_by_user = _build_task_rows(export)
    _append_tasks_sheet(wb, "AllTasks", all_task_rows)

    used_names = {ws.title for ws in wb.worksheets}
    for username, task_rows in sorted(rows_by_user.items(), key=lambda item: str(item[0]).lower()):
        sheet_name = _safe_sheet_name(f"user_{username}", used_names)
        _append_tasks_sheet(wb, sheet_name, task_rows)

    wb.save(output_path)


def main():
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = EXPORT_DIR / f"running_managed_tasks_by_user_{timestamp}.json"
    xlsx_path = EXPORT_DIR / f"running_managed_tasks_by_user_{timestamp}.xlsx"

    rows = _fetch_running_managed_rows()
    export = _group_rows(rows)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(export, f, ensure_ascii=False, indent=2)

    _write_xlsx(export, xlsx_path)

    print(json.dumps({
        "json_path": str(json_path),
        "xlsx_path": str(xlsx_path),
        "user_count": export["user_count"],
        "session_count": export["session_count"],
        "task_count": export["task_count"],
        "status_counts": export["status_counts"],
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
