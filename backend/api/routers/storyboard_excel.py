import json
import os
import re
import uuid
from io import BytesIO
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

import models
from api.routers.episodes import _delete_episode_storyboard_shots
from auth import get_current_user
from database import get_db


router = APIRouter()
@router.post("/api/episodes/{episode_id}/import-storyboard")

async def import_storyboard(
    episode_id: int,
    file: UploadFile = File(...),
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """导入分镜表（xls）并生成镜头"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in [".xls", ".xlsx"]:
        raise HTTPException(status_code=400, detail="仅支持.xls或.xlsx格式的分镜表")

    try:
        from openpyxl import load_workbook
        from io import BytesIO
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少依赖openpyxl，请先安装")

    content = await file.read()
    try:
        # 去掉data_only=True，确保读取最新保存的值
        wb = load_workbook(filename=BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析失败: {str(e)}")

    if ws.max_row < 2:
        raise HTTPException(status_code=400, detail="表格为空")

    def normalize_header(value: str) -> str:
        text = str(value).strip() if value else ""
        text = text.replace("\n", "").replace("\r", "").replace(" ", "")
        text = text.replace("（", "(").replace("）", ")")
        return text

    def find_column(header_map, keywords):
        for key, idx in header_map.items():
            for kw in keywords:
                if kw in key:
                    return idx
        return None

    def cell_to_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value)
        return str(value).strip()

    def parse_shot_number(value, fallback):
        if value is None or value == "":
            return fallback
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value) if value.is_integer() else fallback
        text = str(value).strip()
        if not text:
            return fallback
        try:
            return int(text)
        except Exception:
            match = re.search(r"\d+", text)
            return int(match.group(0)) if match else fallback

    def parse_subjects_with_type(value) -> List[dict]:
        """解析主体列表，支持格式：萧景珩(角色)、书房(场景)

        Returns:
            List[dict]: [{"name": "萧景珩", "type": "角色"}, ...]
        """
        text = cell_to_text(value)
        if not text:
            return []

        # 支持多种分隔符：顿号、逗号、分号等
        normalized = re.sub(r"[，、/;；|]+", "、", text)
        parts = [part.strip() for part in normalized.split("、") if part.strip()]

        subjects = []
        allowed_types = {"角色", "场景"}
        for part in parts:
            # 尝试匹配 "名称(类型)" 格式
            match = re.match(r"^(.+?)\(([^)]+)\)$", part)
            if match:
                name = match.group(1).strip()
                subject_type = match.group(2).strip()
            else:
                # 没有括号，默认为角色
                name = part
                subject_type = "角色"

            if name and subject_type in allowed_types:
                subjects.append({"name": name, "type": subject_type})

        return subjects

    # 读取表头（openpyxl的行列从1开始）
    headers = [normalize_header(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    header_map = {headers[i]: i + 1 for i in range(len(headers)) if headers[i]}  # 存储column号（1-based）

    shot_idx = find_column(header_map, ["镜号"])
    subjects_idx = find_column(header_map, ["角色/场景", "角色场景", "主体"])
    excerpt_idx = find_column(header_map, ["对应的原剧本段落", "原剧本段落"])
    dialogue_idx = find_column(header_map, ["对白", "台词"])
    storyboard_prompt_idx = find_column(header_map, ["分镜提示词"])
    duration_idx = find_column(header_map, ["时长"])

    # 至少需要有"角色/场景"或"原剧本段落"或"对白"或"分镜提示词"之一
    if subjects_idx is None and excerpt_idx is None and dialogue_idx is None and storyboard_prompt_idx is None:
        raise HTTPException(status_code=400, detail="未找到必要列：至少需要有【角色/场景】或【原剧本段落】或【对白】或【分镜提示词】之一")

    rows_data = []
    for r in range(2, ws.max_row + 1):  # 从第2行开始读取数据
        raw_shot = ws.cell(row=r, column=shot_idx).value if shot_idx is not None else None
        shot_number = parse_shot_number(raw_shot, r - 1)  # r-1 作为fallback（因为第2行对应镜号1）

        # 解析新的6列结构
        subjects = parse_subjects_with_type(ws.cell(row=r, column=subjects_idx).value) if subjects_idx is not None else []
        script_excerpt = cell_to_text(ws.cell(row=r, column=excerpt_idx).value) if excerpt_idx is not None else ""
        dialogue = cell_to_text(ws.cell(row=r, column=dialogue_idx).value) if dialogue_idx is not None else ""
        storyboard_prompt = cell_to_text(ws.cell(row=r, column=storyboard_prompt_idx).value) if storyboard_prompt_idx is not None else ""

        # 调试日志：打印第一行数据
        if r == 2:
            print(f"  镜号={shot_number}, 主体={subjects}")

        # 跳过空行
        if not subjects and not script_excerpt and not dialogue and not storyboard_prompt:
            continue

        duration = 15
        if duration_idx is not None:
            raw_duration = ws.cell(row=r, column=duration_idx).value
            parsed_duration = parse_shot_number(raw_duration, None)
            if parsed_duration in (10, 15):
                duration = parsed_duration

        rows_data.append({
            "shot_number": shot_number,
            "subjects": subjects,
            "script_excerpt": script_excerpt,
            "dialogue": dialogue,
            "storyboard_prompt": storyboard_prompt,
            "duration": duration
        })

    if not rows_data:
        raise HTTPException(status_code=400, detail="表格无有效数据")

    # 获取剧集的主体库
    library = db.query(models.StoryLibrary).filter(
        models.StoryLibrary.episode_id == episode.id
    ).first()

    if not library:
        raise HTTPException(status_code=500, detail="主体库不存在")

    # 现有主体卡片映射
    existing_cards = db.query(models.SubjectCard).filter(
        models.SubjectCard.library_id == library.id
    ).all()
    name_to_id = {card.name: card.id for card in existing_cards}

    # 根据Excel中的主体自动补齐主体卡片
    created_subjects = []
    for row in rows_data:
        for subject in row.get("subjects", []):
            name = subject["name"]
            subject_type = subject["type"]
            if subject_type not in ("角色", "场景"):
                continue
            if name not in name_to_id:
                new_card = models.SubjectCard(
                    library_id=library.id,
                    name=name,
                    card_type=subject_type
                )
                db.add(new_card)
                db.flush()
                name_to_id[name] = new_card.id
                created_subjects.append(f"{name}({subject_type})")

    # ⚠️ 替换模式：删除该episode的所有旧镜头
    deleted_count = _delete_episode_storyboard_shots(episode_id, db)
    db.commit()

    # 创建新导入的镜头
    for idx, row in enumerate(rows_data):
        # 获取主体ID列表
        selected_ids = []
        for subject in row.get("subjects", []):
            name = subject["name"]
            if name in name_to_id:
                selected_ids.append(name_to_id[name])

        # 打印每个镜头的对白字段（用于调试）
        dialogue_preview = row['dialogue'][:50] + '...' if len(row['dialogue']) > 50 else row['dialogue']
        storyboard_prompt_preview = row['storyboard_prompt'][:50] + '...' if len(row['storyboard_prompt']) > 50 else row['storyboard_prompt']

        for _ in [None]:
            new_shot = models.StoryboardShot(
                episode_id=episode_id,
                shot_number=int(row["shot_number"]),
                variant_index=0,
                prompt_template="",
                script_excerpt=row["script_excerpt"],  # 原剧本段落
                storyboard_dialogue=row["dialogue"],  # 对白
                sora_prompt=row["storyboard_prompt"],  # ✅ 保存Excel中的分镜提示词
                selected_card_ids=json.dumps(selected_ids),
                selected_sound_card_ids=None,
                aspect_ratio="16:9",
                duration=row["duration"],
                storyboard_video_model="",
                storyboard_video_model_override_enabled=False,
                duration_override_enabled=True,
            )
            db.add(new_shot)

    db.commit()

    # 验证：查询刚保存的数据
    saved_shots = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id
    ).order_by(models.StoryboardShot.shot_number).limit(3).all()

    for shot in saved_shots:
        dialogue_preview = (shot.storyboard_dialogue or '')[:50] + '...' if len(shot.storyboard_dialogue or '') > 50 else (shot.storyboard_dialogue or '')
        sora_prompt_preview = (shot.sora_prompt or '')[:50] + '...' if len(shot.sora_prompt or '') > 50 else (shot.sora_prompt or '')


    return {
        "message": "导入成功（已替换所有镜头）",
        "imported_shots": len(rows_data),
        "deleted_shots": deleted_count,
        "created_subjects": len(created_subjects),
        "created_subject_names": created_subjects[:10]  # 显示前10个
    }

@router.get("/api/episodes/{episode_id}/export-storyboard")

async def export_storyboard(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """导出分镜表为Excel文件（.xlsx）"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # 获取所有镜头
    shots = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id
    ).order_by(
        models.StoryboardShot.shot_number.asc(),
        models.StoryboardShot.variant_index.asc()
    ).all()

    if not shots:
        raise HTTPException(status_code=400, detail="没有镜头数据")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少依赖openpyxl，请先安装")

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "分镜表"

    # 表头
    headers = ["镜号", "角色/场景", "原剧本段落", "对白", "分镜提示词", "时长"]
    ws.append(headers)

    # 设置表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 设置列宽
    column_widths = [8, 30, 40, 30, 50, 8]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # 获取主体库以便查询主体详情
    library = db.query(models.StoryLibrary).filter(
        models.StoryLibrary.episode_id == episode.id
    ).first()

    card_map = {}
    if library:
        cards = db.query(models.SubjectCard).filter(
            models.SubjectCard.library_id == library.id
        ).all()
        card_map = {card.id: card for card in cards}

    # 填充数据
    for shot in shots:
        # 解析主体列表
        try:
            selected_ids = json.loads(shot.selected_card_ids or "[]")
        except:
            selected_ids = []

        subjects_text_parts = []
        for card_id in selected_ids:
            if card_id in card_map:
                card = card_map[card_id]
                subjects_text_parts.append(f"{card.name}({card.card_type})")

        subjects_text = "、".join(subjects_text_parts)

        # 构建行数据
        row_data = [
            shot.shot_number,
            subjects_text,
            shot.script_excerpt or "",
            shot.storyboard_dialogue or "",
            shot.sora_prompt or "",  # 分镜提示词使用sora_prompt
            shot.duration
        ]
        ws.append(row_data)

        # 设置单元格自动换行
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 生成文件名
    safe_episode_name = re.sub(r'[\\/*?:"<>|]', '_', episode.name or f"片段{episode_id}")
    filename = f"{safe_episode_name}_分镜表.xlsx"
    output_path = os.path.join("uploads", f"export_{uuid.uuid4().hex[:8]}_{filename}")

    # 保存文件
    wb.save(output_path)

    return FileResponse(
        path=output_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



