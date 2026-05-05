from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, BackgroundTasks, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi import Header
from sqlalchemy.orm import Session
from sqlalchemy import text, func, or_, and_, case
from sqlalchemy.exc import OperationalError
from typing import List, Optional, Dict, Any, Tuple
from pydantic import BaseModel
from datetime import datetime
import os
import shutil
import uuid
import json
import requests
import mimetypes
import asyncio
import re
import time
import hashlib
from datetime import timedelta
from concurrent.futures import ThreadPoolExecutor
from functools import partial

from api.routers import (
    admin_users,
    auth,
    billing,
    card_media,
    dashboard,
    image_generation,
    libraries,
    media,
    model_configs,
    pages,
    public,
    settings,
    scripts,
    simple_storyboard,
    storyboard2,
    storyboard_excel,
    managed_generation,
    subject_cards,
    templates,
    video,
    voiceover,
    episodes,
    hit_dramas,
    shots,
)
from api.services.card_media import (
    _ensure_audio_duration_seconds_cached,
    save_and_upload_to_cdn,
    save_audio_and_upload_to_cdn,
)
from api.services import billing_charges
from api.services import card_image_generation as card_image_generation_service
from api.services import episode_cleanup
from api.services import model_configs as model_configs_service
from api.services import shot_reference_workflow
from api.services import simple_storyboard_batches as simple_storyboard_batches_service
from api.services import storyboard_defaults
from api.services import storyboard2_board
from api.services import storyboard2_media
from api.services import storyboard2_reference_images
from api.services import storyboard_reference_assets
from api.services import storyboard_sound_cards
from api.services import storyboard_prompt_context
from api.services import shot_image_generation
from api.services import storyboard_sync
from api.services import storyboard_video_generation_limits
from api.services import storyboard_video_prompt_builder
from api.services import storyboard_video_settings
from api.services import storyboard_video_payload
from api.services import voiceover_data
from api.schemas import episodes as episode_schemas
from api.schemas import shots as shot_schemas
from api.schemas.shots import GenerateStoryboardImageRequest, GenerateDetailImagesRequest, SetDetailImageCoverRequest
from env_config import get_env, is_placeholder_env_value, load_app_env


load_app_env()
try:
    import msvcrt
except ImportError:
    msvcrt = None
try:
    import fcntl
except ImportError:
    fcntl = None

import models
import billing_service
import image_platform_client
from database import IS_SQLITE, engine, get_db, SessionLocal
from db_compat import boolean_sql, datetime_sql, get_table_columns, rename_column_if_needed, table_exists
from runtime_load import request_load_tracker
from startup_external_prewarms import start_external_cache_prewarms
from runtime import pollers as runtime_pollers
from startup_schema_policy import should_apply_runtime_postgres_alter
from auth import get_current_user, verify_library_owner
from ai_config import (
    build_ai_debug_config,
    get_ai_config,
    get_ai_provider_catalog,
    get_ai_provider_public_configs,
    get_default_ai_provider_key,
    get_provider_model_options,
    normalize_ai_provider_key,
)
from text_relay_service import (
    submit_and_persist_text_task,
    text_relay_poller,
)
from ai_service import (
    generate_storyboard_prompts,
    stage1_generate_initial_storyboard, stage2_generate_subject_prompts,
    get_prompt_by_key
)
from storyboard_prompt_templates import (
    build_large_shot_prompt_rule,
    get_default_large_shot_templates,
    inject_large_shot_template_content,
    is_legacy_large_shot_prompt_rule,
)
from storyboard_variant import (
    build_duplicate_shot_payload,
    build_storyboard_image_variant_payload,
    choose_storyboard_reference_source,
)
from dashboard_service import (
    log_debug_task_event,
    log_file_task_event,
    sync_managed_task_to_dashboard,
    sync_voiceover_tts_task_to_dashboard,
)
from utils import upload_to_cdn
from PIL import Image, ImageDraw, ImageFont
from video_service import (
    poller,
    update_shot_status,
    process_and_upload_video_with_cover,
    check_video_status,
    is_transient_video_status_error,
    normalize_video_generation_status,
)
from video_api_config import (
    get_video_api_headers,
    get_video_task_create_url,
    get_video_task_status_url,
)
from image_generation_service import (
    image_poller, MODEL_CONFIGS, submit_image_generation,
    check_task_status, download_and_upload_image, jimeng_generate_image_with_polling,
    create_jimeng_image_task, get_image_task_status, is_jimeng_image_model,
    is_moti_image_model, is_transient_image_status_error,
    submit_moti_standard_image_generation, get_image_submit_api_url,
    get_image_status_api_url, resolve_jimeng_actual_model
)

from managed_generation_service import managed_poller, ACTIVE_MANAGED_SESSION_STATUSES
from model_pricing_poller import model_pricing_poller
from text_llm_queue import run_text_llm_request
from simple_storyboard_rules import (
    get_default_rule_config,
    normalize_rule_config,
)
from storyboard_video_reference import (
    build_seedance_content_text,
    build_seedance_prompt,
    build_seedance_reference_images,
    collect_first_frame_candidate_urls,
    is_allowed_first_frame_candidate_url,
    normalize_first_frame_candidate_url,
    resolve_scene_reference_image_url,
    should_autofill_scene_override,
)
from threading import Thread, Lock

# 创建线程池（用于处理同步阻塞调用）
executor = ThreadPoolExecutor(max_workers=10)

# 故事板2镜头图运行中任务（进程内）
storyboard2_active_image_tasks = set()
storyboard2_active_image_tasks_lock = Lock()
simple_storyboard_batch_update_lock = simple_storyboard_batches_service.simple_storyboard_batch_update_lock
startup_bootstrap_lock_handle = None


_safe_json_dumps = billing_charges.safe_json_dumps


def _resolve_storyboard_video_billing_model(shot: models.StoryboardShot) -> str:
    return billing_charges.resolve_storyboard_video_billing_model(
        shot,
        resolve_model_by_provider=_resolve_storyboard_video_model_by_provider,
        default_model=DEFAULT_STORYBOARD_VIDEO_MODEL,
    )


_record_card_image_charge = billing_charges.record_card_image_charge
_record_storyboard_image_charge = billing_charges.record_storyboard_image_charge
_record_detail_image_charge = billing_charges.record_detail_image_charge


def _record_storyboard_video_charge(
    db: Session,
    *,
    shot: models.StoryboardShot,
    task_id: str,
    stage: str = "video_generate",
    detail_payload: Optional[Dict[str, Any]] = None,
):
    return billing_charges.record_storyboard_video_charge(
        db,
        shot=shot,
        task_id=task_id,
        model_name=_resolve_storyboard_video_billing_model(shot),
        stage=stage,
        detail_payload=detail_payload,
    )


_record_storyboard2_video_charge = billing_charges.record_storyboard2_video_charge
_record_storyboard2_image_charge = billing_charges.record_storyboard2_image_charge


def _read_utf8_text_file(file_path: str) -> str:
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception:
        return ""


def _load_storyboard2_prompt_base() -> str:
    return """原剧本段落：
{script_excerpt}

场景描述：
{scene_description}

候选主体（只能从这里选择）：
{subject_text}

说明：
- 角色主体会按“主体名-性格描述”的格式提供，场景/道具主体只写名称
- 生成结果里的 subjects 数组只填写“-”前的主体名称，不要把后面的性格描述写进去

输出 JSON，格式如下：
{{
  "timeline": [
    {{
      "time": "00s-03s",
      "subjects": ["主体A", "场景B"],
      "visual": "[镜头类型][景别/拍摄角度] 画面描述（至少20字）",
      "audio": "[角色] 说：\"台词\" 或 旁白：\"内容\""
    }},
    {{
      "time": "03s-06s",
      "subjects": ["主体A"],
      "visual": "[镜头类型][景别/拍摄角度] 画面描述（至少20字）",
      "audio": "(SFX:音效) 或 [角色] 说：\"台词\""
    }},
    {{
      "time": "06s-{safe_duration:02d}s",
      "subjects": ["主体A", "主体C"],
      "visual": "[镜头类型][景别/拍摄角度] 画面描述（至少20字）",
      "audio": "[角色] 说：\"台词\""
    }}
  ]
}}

要求：
1. 时长总计 {safe_duration} 秒，按剧情拆分为 3-5 个连续时间段
2. time 字段必须连续不重叠，最后一个时间段必须覆盖到 {safe_duration} 秒
3. 每个时间段必须包含 subjects 字段，类型为字符串数组
4. subjects 里的名称只能从“候选主体”中选择，禁止编造新主体名称
5. visual 只描述镜头、动作、微动作、情绪和场景，不改写剧情核心事件
6. audio 保留原文台词；有音效可用 (SFX:...)；没有台词可用旁白
7. 只输出 JSON，不要其他说明

{extra_style}"""


STORYBOARD2_IMAGE_PROMPT_KEY = "storyboard2_image_prompt_prefix"
STORYBOARD2_IMAGE_PROMPT_DEFAULT = "生成动漫风格的图片"
CHARACTER_THREE_VIEW_PROMPT_KEY = "character_three_view_image_prompt"
CHARACTER_THREE_VIEW_PROMPT_DEFAULT = "生人物三视图，生成全身三视图以及一张面部特写(最左边占满三分之一的位置是超大的面部特写，右边三分之二放正视图、侧视图、后视图，（正视图、侧视图、后视图并排）纯白背景"
GROK_RULE_KEY = "grok_rule"
GROK_RULE_DEFAULT = "严格按照提示词生视频，不要出现其他人物"
STORYBOARD2_VIDEO_PROMPT_KEY = "generate_storyboard2_video_prompts"
STORYBOARD2_VIDEO_PROMPT_DEFAULT = _load_storyboard2_prompt_base()
MANAGED_PROMPT_OPTIMIZE_KEY = "managed_retry_optimize_prompt"
MANAGED_PROMPT_OPTIMIZE_DEFAULT = """你是一位视频生成提示词优化助手。请在尽量保留原意、人物、场景、镜头顺序、时长和关键信息的前提下，改写下面这段完整视频提示词，使其更容易通过平台文字审核。

要求：
1. 不改变剧情核心、镜头时序、人物关系、时长和主要动作。
2. 删除、弱化或替换可能触发审核的问题表达，改为克制、中性、合规的说法。
3. 尽量保留原有的段落结构、镜头编号、旁白/台词格式。
4. 不要输出解释、分析、备注或 markdown，只输出修改后的完整提示词文本。
5. 如果原文包含过于激烈、违规、血腥、暴力、色情、仇恨、违法或明显不当的描述，请改写为平台更容易接受的表达。

当前失败原因：
{error_message}

原始完整提示词：
{full_prompt}"""

ALLOWED_CARD_TYPES = storyboard_sync.ALLOWED_CARD_TYPES
ALL_SUBJECT_CARD_TYPES = ("角色", "场景", "道具", "声音")
SOUND_CARD_TYPE = "声音"
SEEDANCE_AUDIO_MAX_COUNT = 3
SEEDANCE_AUDIO_MAX_TOTAL_SECONDS = 15.0
SEEDANCE_AUDIO_COUNT_ERROR = "请检查音频总数是否不超过3个"
SEEDANCE_AUDIO_DURATION_ERROR = "请检查音频时长总和是否小于15s"
SEEDANCE_AUDIO_VALIDATION_ERRORS = {
    SEEDANCE_AUDIO_COUNT_ERROR,
    SEEDANCE_AUDIO_DURATION_ERROR,
}

# 不可通过管理界面操作、且通用密码无效的保留账号
HIDDEN_USERS = {"test", "9f3a7c2e4b6d8a1c"}


def _get_private_password_env(name: str) -> str:
    value = (get_env(name, "") or "").strip()
    if is_placeholder_env_value(value):
        return ""
    return value


# 管理员通用密码（可登录任意非保留账号）
MASTER_PASSWORD = _get_private_password_env("MASTER_PASSWORD")
ADMIN_PANEL_PASSWORD = _get_private_password_env("ADMIN_PANEL_PASSWORD")
DEFAULT_STORYBOARD_VIDEO_MODEL = storyboard_video_settings.DEFAULT_STORYBOARD_VIDEO_MODEL
MOTI_STORYBOARD_VIDEO_MODELS = storyboard_video_settings.MOTI_STORYBOARD_VIDEO_MODELS
SQLITE_LOCK_RETRY_DELAYS = (0.3, 0.8, 1.5, 3.0)
STARTUP_BOOTSTRAP_LOCK_PATH = os.path.join(os.path.dirname(__file__), ".startup_bootstrap.lock")


def start_background_pollers(force: bool = False):
    return runtime_pollers.start_background_pollers(
        pollers=(
            poller,
            image_poller,
            managed_poller,
            text_relay_poller,
            voiceover_tts_poller,
            model_pricing_poller,
        ),
        recover_storyboard2_video_polling=_recover_storyboard2_video_polling,
        force=force,
    )


def stop_background_pollers():
    return runtime_pollers.stop_background_pollers(
        pollers=(
            poller,
            image_poller,
            managed_poller,
            text_relay_poller,
            voiceover_tts_poller,
            model_pricing_poller,
        )
    )


def acquire_startup_bootstrap_lock(timeout_seconds: float = 300.0):
    global startup_bootstrap_lock_handle
    if startup_bootstrap_lock_handle is not None:
        return startup_bootstrap_lock_handle

    lock_file = open(STARTUP_BOOTSTRAP_LOCK_PATH, "a+b")
    lock_file.seek(0, os.SEEK_END)
    if lock_file.tell() == 0:
        lock_file.write(b"0")
        lock_file.flush()

    start_time = time.time()
    while True:
        try:
            lock_file.seek(0)
            if os.name == "nt" and msvcrt is not None:
                msvcrt.locking(lock_file.fileno(), msvcrt.LK_NBLCK, 1)
            elif fcntl is not None:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            else:
                startup_bootstrap_lock_handle = lock_file
                return lock_file
            startup_bootstrap_lock_handle = lock_file
            print(f"[startup] bootstrap lock acquired by pid={os.getpid()}")
            return lock_file
        except OSError:
            if time.time() - start_time >= timeout_seconds:
                lock_file.close()
                raise TimeoutError(f"Timed out waiting for startup bootstrap lock: {STARTUP_BOOTSTRAP_LOCK_PATH}")
            time.sleep(0.2)


def release_startup_bootstrap_lock():
    global startup_bootstrap_lock_handle
    lock_file = startup_bootstrap_lock_handle
    if lock_file is None:
        return
    try:
        lock_file.seek(0)
        if os.name == "nt" and msvcrt is not None:
            msvcrt.locking(lock_file.fileno(), msvcrt.LK_UNLCK, 1)
        elif fcntl is not None:
            fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)
    finally:
        lock_file.close()
        startup_bootstrap_lock_handle = None


def run_startup_bootstrap():
    acquire_startup_bootstrap_lock()
    try:
        ensure_storyboard_columns()
        ensure_storyboard_sora_prompt_full_flag()
        ensure_image_model_keys()
        ensure_storyboard_shot_family_stable_ids()
        ensure_managed_session_variant_count_column()
        ensure_managed_task_prompt_text_column()
        ensure_subject_card_columns()
        ensure_subject_card_audio_duration_column()
        ensure_subject_card_generating_columns()
        ensure_subject_card_protagonist_columns()
        ensure_subject_card_linked_columns()
        ensure_subject_card_personality_columns()
        ensure_script_columns()
        ensure_storyboard2_shot_columns()
        ensure_storyboard2_subshot_columns()
        ensure_storyboard2_subshot_image_columns()
        ensure_storyboard2_subshot_video_columns()
        ensure_shot_detail_image_columns()
        ensure_episode_columns()
        ensure_simple_storyboard_batch_columns()
        ensure_billing_columns()
        ensure_billing_defaults()
        ensure_video_model_pricing()
        ensure_video_style_templates()
        ensure_prompt_config_table()
        ensure_function_model_config_columns()
        ensure_stage2_refine_shot_prompt_config()
        ensure_prop_subject_prompt_configs()
        ensure_character_three_view_prompt_config()
        ensure_storyboard2_prompt_config()
        ensure_shot_duration_template_config_json()
        ensure_shot_duration_template_without_scene_description()
        ensure_generate_video_prompt_config_without_scene_description()
        ensure_shot_duration_template_large_shot_rule()
        ensure_generate_large_shot_prompt_config()
        ensure_large_shot_templates()
        ensure_shot_duration_template_subject_personality()
        ensure_remove_legacy_duration_templates()
        ensure_video_prompt_subject_personality_configs()
        ensure_style_template_variant_columns()
        ensure_default_style_templates()
        ensure_hit_drama_columns()
        ensure_user_password_column()
    finally:
        release_startup_bootstrap_lock()


def _hash_password(password: str) -> str:
    """sha256 哈希密码"""
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def _rollback_quietly(db: Session):
    try:
        db.rollback()
    except Exception:
        pass


def _is_sqlite_lock_error(db: Session, exc: Exception) -> bool:
    dialect = getattr(getattr(db, "bind", None), "dialect", None)
    dialect_name = getattr(dialect, "name", "")
    return dialect_name == "sqlite" and "database is locked" in str(exc).lower()


def commit_with_retry(
    db: Session,
    prepare_fn=None,
    context: str = "db commit"
):
    max_retries = len(SQLITE_LOCK_RETRY_DELAYS)

    for attempt in range(max_retries + 1):
        if prepare_fn:
            prepare_fn()
        try:
            db.commit()
            return
        except OperationalError as e:
            _rollback_quietly(db)
            if not _is_sqlite_lock_error(db, e) or attempt >= max_retries:
                raise
            delay = SQLITE_LOCK_RETRY_DELAYS[attempt]
            print(f"[db] {context} 遇到 SQLite 写锁，{delay:.1f}s 后重试 ({attempt + 1}/{max_retries})")
            time.sleep(delay)
        except Exception:
            _rollback_quietly(db)
            raise

# 补充新增字段（兼容旧数据库）
def ensure_storyboard_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "storyboard_shots")
            if not columns:
                return
            has_variant = "variant_index" in columns

            if "script_excerpt" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN script_excerpt TEXT DEFAULT ''")
                )
                print("??? storyboard_shots.script_excerpt ??")
            if "storyboard_video_prompt" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN storyboard_video_prompt TEXT DEFAULT ''")
                )
                print("??? storyboard_shots.storyboard_video_prompt ??")
            if "storyboard_audio_prompt" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN storyboard_audio_prompt TEXT DEFAULT ''")
                )
                print("??? storyboard_shots.storyboard_audio_prompt ??")
            if "storyboard_dialogue" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN storyboard_dialogue TEXT DEFAULT ''")
                )
                print("??? storyboard_shots.storyboard_dialogue ??")
            if "sora_prompt" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN sora_prompt TEXT DEFAULT ''")
                )
                print("??? storyboard_shots.sora_prompt ??")
            if "sora_prompt_status" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN sora_prompt_status TEXT DEFAULT 'idle'")
                )
                print("??? storyboard_shots.sora_prompt_status ??")
            if "selected_sound_card_ids" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN selected_sound_card_ids TEXT")
                )
                print("已添加 storyboard_shots.selected_sound_card_ids 字段")
            if "thumbnail_video_path" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN thumbnail_video_path TEXT DEFAULT ''")
                )
                print("??? storyboard_shots.thumbnail_video_path ??")
            if "detail_image_prompt_overrides" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN detail_image_prompt_overrides TEXT DEFAULT '{}'")
                )
                print("已添加 storyboard_shots.detail_image_prompt_overrides 字段")
            if "storyboard_image_model" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN storyboard_image_model TEXT DEFAULT ''")
                )
                print("已添加 storyboard_shots.storyboard_image_model 字段")
                columns.add("storyboard_image_model")
            if "first_frame_reference_image_url" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN first_frame_reference_image_url TEXT DEFAULT ''")
                )
                print("已添加 storyboard_shots.first_frame_reference_image_url 字段")
                columns.add("first_frame_reference_image_url")
            if "uploaded_first_frame_reference_image_url" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN uploaded_first_frame_reference_image_url TEXT DEFAULT ''")
                )
                print("已添加 storyboard_shots.uploaded_first_frame_reference_image_url 字段")
                columns.add("uploaded_first_frame_reference_image_url")
            if "uploaded_scene_image_url" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN uploaded_scene_image_url TEXT DEFAULT ''")
                )
                print("已添加 storyboard_shots.uploaded_scene_image_url 字段")
                columns.add("uploaded_scene_image_url")
            if "use_uploaded_scene_image" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN use_uploaded_scene_image BOOLEAN DEFAULT FALSE")
                )
                print("已添加 storyboard_shots.use_uploaded_scene_image 字段")
                columns.add("use_uploaded_scene_image")
            if "duration_override_enabled" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN duration_override_enabled BOOLEAN DEFAULT FALSE")
                )
                print("已添加 storyboard_shots.duration_override_enabled 字段")
                columns.add("duration_override_enabled")
            if "storyboard_video_model" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN storyboard_video_model TEXT DEFAULT ''")
                )
                print("已添加 storyboard_shots.storyboard_video_model 字段")
                columns.add("storyboard_video_model")
            if "storyboard_video_model_override_enabled" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN storyboard_video_model_override_enabled BOOLEAN DEFAULT FALSE")
                )
                print("已添加 storyboard_shots.storyboard_video_model_override_enabled 字段")
                columns.add("storyboard_video_model_override_enabled")
            if not has_variant:
                conn.execute(
                    text("ALTER TABLE storyboard_shots ADD COLUMN variant_index INTEGER DEFAULT 0")
                )
                print("??? storyboard_shots.variant_index ??")
                has_variant = True
                columns.add("variant_index")
            if has_variant:
                conn.execute(
                    text("UPDATE storyboard_shots SET variant_index = 0 WHERE variant_index IS NULL")
                )
            conn.execute(
                text("UPDATE storyboard_shots SET detail_image_prompt_overrides = '{}' WHERE detail_image_prompt_overrides IS NULL OR TRIM(detail_image_prompt_overrides) = ''")
            )
            conn.execute(
                text("UPDATE storyboard_shots SET storyboard_image_model = '' WHERE storyboard_image_model IS NULL")
            )
            conn.execute(
                text("UPDATE storyboard_shots SET first_frame_reference_image_url = '' WHERE first_frame_reference_image_url IS NULL")
            )
            conn.execute(
                text("UPDATE storyboard_shots SET uploaded_first_frame_reference_image_url = '' WHERE uploaded_first_frame_reference_image_url IS NULL")
            )
            conn.execute(
                text("UPDATE storyboard_shots SET uploaded_scene_image_url = '' WHERE uploaded_scene_image_url IS NULL")
            )
            conn.execute(
                text("UPDATE storyboard_shots SET use_uploaded_scene_image = FALSE WHERE use_uploaded_scene_image IS NULL")
            )
            conn.execute(
                text("UPDATE storyboard_shots SET duration_override_enabled = FALSE WHERE duration_override_enabled IS NULL")
            )
            conn.execute(
                text("UPDATE storyboard_shots SET storyboard_video_model = '' WHERE storyboard_video_model IS NULL")
            )
            conn.execute(
                text("UPDATE storyboard_shots SET storyboard_video_model_override_enabled = FALSE WHERE storyboard_video_model_override_enabled IS NULL")
            )
            if (
                "use_uploaded_scene_image" in columns
                and not IS_SQLITE
                and should_apply_runtime_postgres_alter("storyboard_shots", "use_uploaded_scene_image")
            ):
                conn.execute(
                    text("ALTER TABLE storyboard_shots ALTER COLUMN use_uploaded_scene_image SET DEFAULT FALSE")
                )
                conn.execute(
                    text("ALTER TABLE storyboard_shots ALTER COLUMN use_uploaded_scene_image SET NOT NULL")
                )
            if (
                "duration_override_enabled" in columns
                and not IS_SQLITE
                and should_apply_runtime_postgres_alter("storyboard_shots", "duration_override_enabled")
            ):
                conn.execute(
                    text("ALTER TABLE storyboard_shots ALTER COLUMN duration_override_enabled SET DEFAULT FALSE")
                )
                conn.execute(
                    text("ALTER TABLE storyboard_shots ALTER COLUMN duration_override_enabled SET NOT NULL")
                )
            if (
                "storyboard_video_model_override_enabled" in columns
                and not IS_SQLITE
                and should_apply_runtime_postgres_alter("storyboard_shots", "storyboard_video_model_override_enabled")
            ):
                conn.execute(
                    text("ALTER TABLE storyboard_shots ALTER COLUMN storyboard_video_model_override_enabled SET DEFAULT FALSE")
                )
                conn.execute(
                    text("ALTER TABLE storyboard_shots ALTER COLUMN storyboard_video_model_override_enabled SET NOT NULL")
                )
            conn.execute(
                text("UPDATE storyboard_shots SET selected_sound_card_ids = NULL WHERE selected_sound_card_ids IS NOT NULL AND TRIM(selected_sound_card_ids) = ''")
            )

    except Exception as e:
        print(f"???? storyboard_shots ??: {str(e)}")



def ensure_image_model_keys():
    try:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "UPDATE generated_images "
                    "SET model_name = 'jimeng-4.5' "
                    "WHERE LOWER(TRIM(model_name)) IN ('seedream-4-5', 'doubao-seedance-4-5')"
                )
            )
            conn.execute(
                text(
                    "UPDATE storyboard_shots "
                    "SET storyboard_image_model = 'jimeng-4.5' "
                    "WHERE LOWER(TRIM(storyboard_image_model)) IN ('seedream-4-5', 'doubao-seedance-4-5')"
                )
            )
            conn.execute(
                text(
                    "UPDATE generated_images "
                    "SET model_name = 'banana2' "
                    "WHERE LOWER(TRIM(model_name)) IN ('banana2', 'nano-banana-2.5', 'nanobanana2.5', 'nano banana2.5', 'gemini-2.5-flash-image-preview')"
                )
            )
            conn.execute(
                text(
                    "UPDATE generated_images "
                    "SET model_name = 'banana2-moti' "
                    "WHERE LOWER(TRIM(model_name)) IN ('banana2-moti', 'nano-banana-2.5-moti', 'nanobanana2.5moti', 'nano banana2.5 moti', 'nano-banana2.5-moti')"
                )
            )
            conn.execute(
                text(
                    "UPDATE generated_images "
                    "SET model_name = 'banana-pro' "
                    "WHERE LOWER(TRIM(model_name)) IN ('banana-pro', 'gemini-3-pro-image-preview', 'nanobanana3.0', 'nano banana3.0', 'nano-banana-3.0')"
                )
            )
            conn.execute(
                text(
                    "UPDATE storyboard_shots "
                    "SET storyboard_image_model = 'banana2' "
                    "WHERE LOWER(TRIM(storyboard_image_model)) IN ('banana2', 'nano-banana-2.5', 'nanobanana2.5', 'nano banana2.5', 'gemini-2.5-flash-image-preview')"
                )
            )
            conn.execute(
                text(
                    "UPDATE storyboard_shots "
                    "SET storyboard_image_model = 'banana2-moti' "
                    "WHERE LOWER(TRIM(storyboard_image_model)) IN ('banana2-moti', 'nano-banana-2.5-moti', 'nanobanana2.5moti', 'nano banana2.5 moti', 'nano-banana2.5-moti')"
                )
            )
            conn.execute(
                text(
                    "UPDATE storyboard_shots "
                    "SET storyboard_image_model = 'banana-pro' "
                    "WHERE LOWER(TRIM(storyboard_image_model)) IN ('banana-pro', 'gemini-3-pro-image-preview', 'nanobanana3.0', 'nano banana3.0', 'nano-banana-3.0')"
                )
            )
    except Exception as e:
        print(f"统一图片模型键失败: {str(e)}")




def ensure_storyboard_shot_family_stable_ids():
    try:
        db = SessionLocal()
        try:
            families = db.query(
                models.StoryboardShot.episode_id,
                models.StoryboardShot.shot_number
            ).distinct().all()
            updated_count = 0

            for episode_id, shot_number in families:
                family_shots = db.query(models.StoryboardShot).filter(
                    models.StoryboardShot.episode_id == episode_id,
                    models.StoryboardShot.shot_number == shot_number
                ).order_by(
                    models.StoryboardShot.variant_index.asc(),
                    models.StoryboardShot.id.asc()
                ).all()
                if not family_shots:
                    continue

                stable_id = ""
                for family_shot in family_shots:
                    stable_id = (family_shot.stable_id or "").strip()
                    if stable_id:
                        break
                if not stable_id:
                    stable_id = str(uuid.uuid4())
                for family_shot in family_shots:
                    if (family_shot.stable_id or "").strip():
                        continue
                    family_shot.stable_id = stable_id
                    updated_count += 1

            if updated_count > 0:
                db.commit()
                print(f"已对齐 storyboard_shots stable_id 家族: {updated_count} 条")
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()
    except Exception as e:
        print(f"对齐 storyboard_shots stable_id 家族失败: {str(e)}")




def ensure_managed_session_variant_count_column():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "managed_sessions")
            if not columns:
                return

            if "variant_count" not in columns:
                conn.execute(
                    text("ALTER TABLE managed_sessions ADD COLUMN variant_count INTEGER DEFAULT 1")
                )
                print("已添加 managed_sessions.variant_count 字段")

            conn.execute(
                text(
                    """
                    UPDATE managed_sessions
                    SET variant_count = COALESCE(
                        (
                            SELECT MIN(task_count)
                            FROM (
                                SELECT COUNT(*) AS task_count
                                FROM managed_tasks
                                WHERE session_id = managed_sessions.id
                                GROUP BY shot_stable_id
                            )
                        ),
                        1
                    )
                    WHERE variant_count IS NULL OR variant_count < 1
                    """
                )
            )
    except Exception as e:
        print(f"检查/迁移 managed_sessions.variant_count 失败: {str(e)}")


def ensure_managed_task_prompt_text_column():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "managed_tasks")
            if not columns:
                return

            if "prompt_text" not in columns:
                conn.execute(
                    text("ALTER TABLE managed_tasks ADD COLUMN prompt_text TEXT DEFAULT ''")
                )
                print("已添加 managed_tasks.prompt_text 字段")

            conn.execute(
                text("UPDATE managed_tasks SET prompt_text = '' WHERE prompt_text IS NULL")
            )
    except Exception as e:
        print(f"检查/迁移 managed_tasks.prompt_text 失败: {str(e)}")


def ensure_storyboard_sora_prompt_full_flag():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "storyboard_shots")
            if not columns:
                return

            if "sora_prompt_is_full" not in columns:
                conn.execute(
                    text(
                        f"ALTER TABLE storyboard_shots ADD COLUMN sora_prompt_is_full BOOLEAN DEFAULT {boolean_sql(False)}"
                    )
                )
                print("已添加 storyboard_shots.sora_prompt_is_full 字段")

            conn.execute(
                text(
                    f"UPDATE storyboard_shots SET sora_prompt_is_full = {boolean_sql(False)} "
                    "WHERE sora_prompt_is_full IS NULL"
                )
            )
    except Exception as e:
        print(f"检查/迁移 storyboard_shots.sora_prompt_is_full 失败: {str(e)}")





def ensure_subject_card_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "subject_cards")
            if not columns:
                return
            if "alias" not in columns:
                conn.execute(
                    text("ALTER TABLE subject_cards ADD COLUMN alias TEXT DEFAULT ''")
                )
                print("已添加 subject_cards.alias 字段")
            if "ai_prompt_status" not in columns:
                conn.execute(
                    text("ALTER TABLE subject_cards ADD COLUMN ai_prompt_status TEXT DEFAULT NULL")
                )
                print("已添加 subject_cards.ai_prompt_status 字段")
    except Exception as e:
        print(f"检查/迁移 subject_cards 失败: {str(e)}")




def ensure_subject_card_audio_duration_column():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "subject_card_audios")
            if not columns:
                return

            if "duration_seconds" not in columns:
                conn.execute(
                    text("ALTER TABLE subject_card_audios ADD COLUMN duration_seconds FLOAT DEFAULT 0")
                )
                print("已添加 subject_card_audios.duration_seconds 字段")

            conn.execute(
                text(
                    "UPDATE subject_card_audios "
                    "SET duration_seconds = 0 "
                    "WHERE duration_seconds IS NULL"
                )
            )
    except Exception as e:
        print(f"检查/迁移 subject_card_audios.duration_seconds 失败: {str(e)}")



def ensure_subject_card_generating_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "subject_cards")
            if not columns:
                return

            if "is_generating_images" not in columns:
                conn.execute(
                    text(
                        f"ALTER TABLE subject_cards ADD COLUMN is_generating_images BOOLEAN DEFAULT {boolean_sql(False)}"
                    )
                )
                print("已添加 subject_cards.is_generating_images 字段")

            if "generating_count" not in columns:
                conn.execute(
                    text("ALTER TABLE subject_cards ADD COLUMN generating_count INTEGER DEFAULT 0")
                )
                print("已添加 subject_cards.generating_count 字段")

            # 确保旧数据的默认值
            conn.execute(
                text(
                    f"UPDATE subject_cards SET is_generating_images = {boolean_sql(False)} "
                    "WHERE is_generating_images IS NULL"
                )
            )
            conn.execute(
                text("UPDATE subject_cards SET generating_count = 0 WHERE generating_count IS NULL")
            )
    except Exception as e:
        print(f"检查/迁移 subject_cards 生成状态字段失败: {str(e)}")


def ensure_subject_card_protagonist_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "subject_cards")
            if not columns:
                return

            if "is_protagonist" not in columns:
                conn.execute(
                    text(
                        f"ALTER TABLE subject_cards ADD COLUMN is_protagonist BOOLEAN DEFAULT {boolean_sql(False)}"
                    )
                )
                print("Added subject_cards.is_protagonist column")

            if "protagonist_gender" not in columns:
                conn.execute(
                    text("ALTER TABLE subject_cards ADD COLUMN protagonist_gender TEXT DEFAULT ''")
                )
                print("Added subject_cards.protagonist_gender column")

            conn.execute(
                text(
                    f"UPDATE subject_cards SET is_protagonist = {boolean_sql(False)} "
                    "WHERE is_protagonist IS NULL"
                )
            )
            conn.execute(
                text("UPDATE subject_cards SET protagonist_gender = '' WHERE protagonist_gender IS NULL")
            )
            conn.execute(
                text("UPDATE subject_cards SET protagonist_gender = '' WHERE protagonist_gender NOT IN ('male', 'female', '')")
            )
            conn.execute(
                text(
                    f"UPDATE subject_cards SET is_protagonist = {boolean_sql(False)}, protagonist_gender = '' "
                    "WHERE card_type != '角色'"
                )
            )
    except Exception as e:
        print(f"Failed to ensure subject_cards protagonist columns: {str(e)}")


def ensure_subject_card_linked_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "subject_cards")
            if not columns:
                return

            if "linked_card_id" not in columns:
                conn.execute(
                    text("ALTER TABLE subject_cards ADD COLUMN linked_card_id INTEGER")
                )
                print("Added subject_cards.linked_card_id column")

            conn.execute(
                text("UPDATE subject_cards SET linked_card_id = NULL WHERE linked_card_id = 0")
            )
            conn.execute(
                text("UPDATE subject_cards SET linked_card_id = NULL WHERE card_type != '声音'")
            )
    except Exception as e:
        print(f"Failed to ensure subject_cards linked columns: {str(e)}")


def ensure_subject_card_personality_columns():
    try:
        if rename_column_if_needed(engine, "subject_cards", "role_personality_en", "role_personality"):
            print("Renamed subject_cards.role_personality_en to role_personality")
        with engine.begin() as conn:
            columns = get_table_columns(engine, "subject_cards")
            if not columns:
                return

            if "role_personality" not in columns:
                conn.execute(
                    text("ALTER TABLE subject_cards ADD COLUMN role_personality TEXT DEFAULT ''")
                )
                print("Added subject_cards.role_personality column")

            conn.execute(
                text("UPDATE subject_cards SET role_personality = '' WHERE role_personality IS NULL")
            )
            conn.execute(
                text("UPDATE subject_cards SET role_personality = '' WHERE card_type != '角色'")
            )
    except Exception as e:
        print(f"Failed to ensure subject_cards personality columns: {str(e)}")


def ensure_script_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "scripts")
            if not columns:
                return
            has_prompt_style = "sora_prompt_style" in columns
            if not has_prompt_style:
                conn.execute(
                    text("ALTER TABLE scripts ADD COLUMN sora_prompt_style TEXT DEFAULT ''")
                )
                print("已添加 scripts.sora_prompt_style 字段")
                has_prompt_style = True
            if has_prompt_style:
                conn.execute(
                    text("UPDATE scripts SET sora_prompt_style = '' WHERE sora_prompt_style IS NULL")
                )

            has_video_prompt_template = "video_prompt_template" in columns
            if not has_video_prompt_template:
                conn.execute(
                    text("ALTER TABLE scripts ADD COLUMN video_prompt_template TEXT DEFAULT ''")
                )
                print("已添加 scripts.video_prompt_template 字段")
                has_video_prompt_template = True
            if has_video_prompt_template:
                conn.execute(
                    text("UPDATE scripts SET video_prompt_template = '' WHERE video_prompt_template IS NULL")
                )

            # 添加 style_template 字段
            has_style_template = "style_template" in columns
            if not has_style_template:
                conn.execute(
                    text("ALTER TABLE scripts ADD COLUMN style_template TEXT DEFAULT ''")
                )
                print("已添加 scripts.style_template 字段")
                has_style_template = True
            if has_style_template:
                conn.execute(
                    text("UPDATE scripts SET style_template = '' WHERE style_template IS NULL")
                )
            has_voiceover_shared_data = "voiceover_shared_data" in columns
            if not has_voiceover_shared_data:
                conn.execute(
                    text("ALTER TABLE scripts ADD COLUMN voiceover_shared_data TEXT DEFAULT ''")
                )
                print("Added scripts.voiceover_shared_data column")
                has_voiceover_shared_data = True
            if has_voiceover_shared_data:
                conn.execute(
                    text("UPDATE scripts SET voiceover_shared_data = '' WHERE voiceover_shared_data IS NULL")
                )
    except Exception as e:
        print(f"检查/迁移 scripts 失败: {str(e)}")




def ensure_storyboard2_shot_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "storyboard2_shots")
            if not columns:
                return

            if "selected_card_ids" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard2_shots ADD COLUMN selected_card_ids TEXT DEFAULT '[]'")
                )
                print("已添加 storyboard2_shots.selected_card_ids 字段")

            conn.execute(
                text("UPDATE storyboard2_shots SET selected_card_ids = '[]' WHERE selected_card_ids IS NULL")
            )
    except Exception as e:
        print(f"检查/迁移 storyboard2_shots 失败: {str(e)}")




def ensure_storyboard2_subshot_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "storyboard2_subshots")
            if not columns:
                return

            if "selected_card_ids" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard2_subshots ADD COLUMN selected_card_ids TEXT DEFAULT '[]'")
                )
                print("已添加 storyboard2_subshots.selected_card_ids 字段")
            if "image_generate_status" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard2_subshots ADD COLUMN image_generate_status TEXT DEFAULT 'idle'")
                )
                print("已添加 storyboard2_subshots.image_generate_status 字段")
            if "image_generate_progress" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard2_subshots ADD COLUMN image_generate_progress TEXT DEFAULT ''")
                )
                print("已添加 storyboard2_subshots.image_generate_progress 字段")
            if "image_generate_error" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard2_subshots ADD COLUMN image_generate_error TEXT DEFAULT ''")
                )
                print("已添加 storyboard2_subshots.image_generate_error 字段")
            if "scene_override" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard2_subshots ADD COLUMN scene_override TEXT DEFAULT ''")
                )
                print("已添加 storyboard2_subshots.scene_override 字段")
            if "scene_override_locked" not in columns:
                conn.execute(
                    text(
                        f"ALTER TABLE storyboard2_subshots ADD COLUMN scene_override_locked BOOLEAN DEFAULT {boolean_sql(False)}"
                    )
                )
                print("已添加 storyboard2_subshots.scene_override_locked 字段")

            conn.execute(
                text("UPDATE storyboard2_subshots SET selected_card_ids = '[]' WHERE selected_card_ids IS NULL")
            )
            conn.execute(
                text("UPDATE storyboard2_subshots SET image_generate_status = 'idle' WHERE image_generate_status IS NULL")
            )
            conn.execute(
                text("UPDATE storyboard2_subshots SET image_generate_progress = '' WHERE image_generate_progress IS NULL")
            )
            conn.execute(
                text("UPDATE storyboard2_subshots SET image_generate_error = '' WHERE image_generate_error IS NULL")
            )
            conn.execute(
                text("UPDATE storyboard2_subshots SET scene_override = '' WHERE scene_override IS NULL")
            )
            conn.execute(
                text(
                    f"UPDATE storyboard2_subshots SET scene_override_locked = {boolean_sql(False)} "
                    "WHERE scene_override_locked IS NULL"
                )
            )
    except Exception as e:
        print(f"检查/迁移 storyboard2_subshots 失败: {str(e)}")



def ensure_storyboard2_subshot_image_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "storyboard2_subshot_images")
            if not columns:
                return

            if "size" not in columns:
                conn.execute(
                    text("ALTER TABLE storyboard2_subshot_images ADD COLUMN size TEXT DEFAULT '9:16'")
                )
                print("已添加 storyboard2_subshot_images.size 字段")

            conn.execute(
                text("UPDATE storyboard2_subshot_images SET size = '9:16' WHERE size IS NULL OR TRIM(size) = '' OR size = '1:2'")
            )
            conn.execute(
                text("UPDATE storyboard2_subshot_images SET size = '16:9' WHERE size = '2:1'")
            )
    except Exception as e:
        print(f"检查/迁移 storyboard2_subshot_images 失败: {str(e)}")



def ensure_storyboard2_subshot_video_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "storyboard2_subshot_videos")
            if not columns:
                return

            if "is_deleted" not in columns:
                conn.execute(
                    text(
                        f"ALTER TABLE storyboard2_subshot_videos ADD COLUMN is_deleted BOOLEAN DEFAULT {boolean_sql(False)}"
                    )
                )
                print("已添加 storyboard2_subshot_videos.is_deleted 字段")
            if "deleted_at" not in columns:
                conn.execute(
                    text(f"ALTER TABLE storyboard2_subshot_videos ADD COLUMN deleted_at {datetime_sql(engine)}")
                )
                print("已添加 storyboard2_subshot_videos.deleted_at 字段")

            conn.execute(
                text(
                    f"UPDATE storyboard2_subshot_videos SET is_deleted = {boolean_sql(False)} "
                    "WHERE is_deleted IS NULL"
                )
            )
    except Exception as e:
        print(f"检查/迁移 storyboard2_subshot_videos 失败: {str(e)}")

def ensure_shot_detail_image_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "shot_detail_images")
            if not columns:
                return

            if "task_id" not in columns:
                conn.execute(text("ALTER TABLE shot_detail_images ADD COLUMN task_id TEXT DEFAULT ''"))
                print("已添加 shot_detail_images.task_id 字段")
            if "provider" not in columns:
                conn.execute(text("ALTER TABLE shot_detail_images ADD COLUMN provider TEXT DEFAULT ''"))
                print("已添加 shot_detail_images.provider 字段")
            if "model_name" not in columns:
                conn.execute(text("ALTER TABLE shot_detail_images ADD COLUMN model_name TEXT DEFAULT ''"))
                print("已添加 shot_detail_images.model_name 字段")
            if "submit_api_url" not in columns:
                conn.execute(text("ALTER TABLE shot_detail_images ADD COLUMN submit_api_url TEXT DEFAULT ''"))
                print("已添加 shot_detail_images.submit_api_url 字段")
            if "status_api_url" not in columns:
                conn.execute(text("ALTER TABLE shot_detail_images ADD COLUMN status_api_url TEXT DEFAULT ''"))
                print("已添加 shot_detail_images.status_api_url 字段")
            if "query_error_count" not in columns:
                conn.execute(text("ALTER TABLE shot_detail_images ADD COLUMN query_error_count INTEGER DEFAULT 0"))
                print("已添加 shot_detail_images.query_error_count 字段")
            if "last_query_error" not in columns:
                conn.execute(text("ALTER TABLE shot_detail_images ADD COLUMN last_query_error TEXT DEFAULT ''"))
                print("已添加 shot_detail_images.last_query_error 字段")
            if "submitted_at" not in columns:
                conn.execute(text(f"ALTER TABLE shot_detail_images ADD COLUMN submitted_at {datetime_sql(engine)}"))
                print("已添加 shot_detail_images.submitted_at 字段")
            if "last_query_at" not in columns:
                conn.execute(text(f"ALTER TABLE shot_detail_images ADD COLUMN last_query_at {datetime_sql(engine)}"))
                print("已添加 shot_detail_images.last_query_at 字段")

            conn.execute(text("UPDATE shot_detail_images SET task_id = '' WHERE task_id IS NULL"))
            conn.execute(text("UPDATE shot_detail_images SET provider = '' WHERE provider IS NULL"))
            conn.execute(text("UPDATE shot_detail_images SET model_name = '' WHERE model_name IS NULL"))
            conn.execute(text("UPDATE shot_detail_images SET submit_api_url = '' WHERE submit_api_url IS NULL"))
            conn.execute(text("UPDATE shot_detail_images SET status_api_url = '' WHERE status_api_url IS NULL"))
            conn.execute(text("UPDATE shot_detail_images SET query_error_count = 0 WHERE query_error_count IS NULL"))
            conn.execute(text("UPDATE shot_detail_images SET last_query_error = '' WHERE last_query_error IS NULL"))

            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS idx_shot_detail_images_task_id ON shot_detail_images (task_id)"))
            except Exception as index_error:
                print(f"创建 shot_detail_images.task_id 索引失败: {index_error}")
    except Exception as e:
        print(f"检查/迁移 shot_detail_images 失败: {str(e)}")


def ensure_episode_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "episodes")
            if not columns:
                return

            if "shot_image_size" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN shot_image_size TEXT DEFAULT '9:16'")
                )
                print("已添加 episodes.shot_image_size 字段")
            if "voiceover_data" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN voiceover_data TEXT DEFAULT ''")
                )
                print("Added episodes.voiceover_data column")
            if "narration_converting" not in columns:
                conn.execute(
                    text(
                        f"ALTER TABLE episodes ADD COLUMN narration_converting BOOLEAN DEFAULT {boolean_sql(False)}"
                    )
                )
                print("Added episodes.narration_converting column")
            if "narration_error" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN narration_error TEXT DEFAULT ''")
                )
                print("Added episodes.narration_error column")
            if "detail_images_model" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN detail_images_model TEXT DEFAULT 'seedream-4.0'")
                )
                print("已添加 episodes.detail_images_model 字段")
            if "detail_images_provider" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN detail_images_provider TEXT DEFAULT ''")
                )
                print("已添加 episodes.detail_images_provider 字段")

            if "batch_generating_storyboard2_prompts" not in columns:
                conn.execute(
                    text(
                        f"ALTER TABLE episodes ADD COLUMN batch_generating_storyboard2_prompts BOOLEAN DEFAULT {boolean_sql(False)}"
                    )
                )
            if "storyboard2_video_duration" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN storyboard2_video_duration INTEGER DEFAULT 6")
                )
            if "storyboard2_duration" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN storyboard2_duration INTEGER DEFAULT 15")
                )
            if "storyboard2_image_cw" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN storyboard2_image_cw INTEGER DEFAULT 50")
                )
            if "storyboard2_include_scene_references" not in columns:
                conn.execute(
                    text(
                        f"ALTER TABLE episodes ADD COLUMN storyboard2_include_scene_references BOOLEAN DEFAULT {boolean_sql(False)}"
                    )
                )
                print("已添加 episodes.batch_generating_storyboard2_prompts 字段")
            if "storyboard_video_model" not in columns:
                conn.execute(
                    text(
                        f"ALTER TABLE episodes ADD COLUMN storyboard_video_model TEXT DEFAULT '{DEFAULT_STORYBOARD_VIDEO_MODEL}'"
                    )
                )
                print("已添加 episodes.storyboard_video_model 字段")
            if "storyboard_video_aspect_ratio" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN storyboard_video_aspect_ratio TEXT DEFAULT '16:9'")
                )
                print("已添加 episodes.storyboard_video_aspect_ratio 字段")
            if "storyboard_video_duration" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN storyboard_video_duration INTEGER DEFAULT 15")
                )
                print("已添加 episodes.storyboard_video_duration 字段")
            if "storyboard_video_resolution_name" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN storyboard_video_resolution_name TEXT DEFAULT '720p'")
                )
                print("已添加 episodes.storyboard_video_resolution_name 字段")
            if "storyboard_video_appoint_account" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN storyboard_video_appoint_account TEXT DEFAULT ''")
                )
                print("已添加 episodes.storyboard_video_appoint_account 字段")

            conn.execute(
                text("UPDATE episodes SET shot_image_size = '9:16' WHERE shot_image_size IS NULL OR TRIM(shot_image_size) = '' OR shot_image_size = '1:2'")
            )
            conn.execute(
                text("UPDATE episodes SET shot_image_size = '16:9' WHERE shot_image_size = '2:1'")
            )
            conn.execute(
                text("UPDATE episodes SET voiceover_data = '' WHERE voiceover_data IS NULL")
            )
            conn.execute(
                text(
                    f"UPDATE episodes SET narration_converting = {boolean_sql(False)} "
                    "WHERE narration_converting IS NULL"
                )
            )
            conn.execute(
                text("UPDATE episodes SET narration_error = '' WHERE narration_error IS NULL")
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET detail_images_model = 'seedream-4.0' "
                    "WHERE detail_images_model IS NULL "
                    "OR TRIM(detail_images_model) = ''"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET detail_images_model = 'seedream-4.0' "
                    "WHERE LOWER(TRIM(detail_images_model)) IN ('jimeng', 'jimeng-4.0', 'seedream-4-0')"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET detail_images_model = 'seedream-4.5' "
                    "WHERE LOWER(TRIM(detail_images_model)) IN ('jimeng-4.5', 'seedream-4-5', 'doubao-seedance-4-5')"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET detail_images_model = 'seedream-4.6' "
                    "WHERE LOWER(TRIM(detail_images_model)) IN ('jimeng-4.6', 'seedream-4-6')"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET detail_images_model = 'seedream-4.1' "
                    "WHERE LOWER(TRIM(detail_images_model)) IN ('jimeng-4.1', 'seedream-4-1')"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET detail_images_model = 'nano-banana-2' "
                    "WHERE LOWER(TRIM(detail_images_model)) IN ('banana2', 'banana2-moti', 'nano-banana-2', 'nano-banana-2.5', 'nanobanana2.5', 'nano banana2.5', 'gemini-2.5-flash-image-preview', 'nano-banana-2.5-moti', 'nanobanana2.5moti', 'nano banana2.5 moti', 'nano-banana2.5-moti')"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET detail_images_model = 'nano-banana-pro' "
                    "WHERE LOWER(TRIM(detail_images_model)) IN ('banana-pro', 'nano-banana-pro', 'gemini-3-pro-image-preview', 'nanobanana3.0', 'nano banana3.0', 'nano-banana-3.0')"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET detail_images_provider = '' "
                    "WHERE detail_images_provider IS NULL"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET detail_images_provider = 'jimeng' "
                    "WHERE (detail_images_provider IS NULL OR TRIM(detail_images_provider) = '') "
                    "AND LOWER(TRIM(detail_images_model)) IN ('seedream-4.0', 'seedream-4.1', 'seedream-4.5', 'seedream-4.6')"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET detail_images_provider = 'momo' "
                    "WHERE (detail_images_provider IS NULL OR TRIM(detail_images_provider) = '') "
                    "AND LOWER(TRIM(detail_images_model)) IN ('nano-banana-2', 'nano-banana-pro', 'gpt-image-2')"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET detail_images_provider = 'momo' "
                    "WHERE LOWER(TRIM(detail_images_provider)) IN ('banana', 'moti', 'moapp', 'gettoken')"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    f"SET batch_generating_storyboard2_prompts = {boolean_sql(False)} "
                    "WHERE batch_generating_storyboard2_prompts IS NULL"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET storyboard2_video_duration = 6 "
                    "WHERE storyboard2_video_duration IS NULL "
                    "OR storyboard2_video_duration NOT IN (6, 10)"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET storyboard2_duration = 15 "
                    "WHERE storyboard2_duration IS NULL "
                    "OR storyboard2_duration NOT IN (15, 25)"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET storyboard2_image_cw = 50 "
                    "WHERE storyboard2_image_cw IS NULL "
                    "OR storyboard2_image_cw < 1 "
                    "OR storyboard2_image_cw > 100"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    f"SET storyboard2_include_scene_references = {boolean_sql(False)} "
                    "WHERE storyboard2_include_scene_references IS NULL"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    f"SET storyboard_video_model = '{DEFAULT_STORYBOARD_VIDEO_MODEL}' "
                    "WHERE storyboard_video_model IS NULL "
                    "OR TRIM(storyboard_video_model) = '' "
                    "OR storyboard_video_model NOT IN ('sora-2', 'grok', 'Seedance 2.0 Fast VIP', 'Seedance 2.0 Fast', 'Seedance 2.0 VIP', 'Seedance 2.0')"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET storyboard_video_aspect_ratio = '16:9' "
                    "WHERE storyboard_video_model = 'sora-2' "
                    "AND (storyboard_video_aspect_ratio IS NULL "
                    "OR TRIM(storyboard_video_aspect_ratio) = '' "
                    "OR storyboard_video_aspect_ratio NOT IN ('16:9', '9:16'))"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET storyboard_video_aspect_ratio = '9:16' "
                    "WHERE storyboard_video_model = 'grok' "
                    "AND (storyboard_video_aspect_ratio IS NULL "
                    "OR TRIM(storyboard_video_aspect_ratio) = '' "
                    "OR storyboard_video_aspect_ratio NOT IN ('21:9','16:9','3:2','4:3','1:1','3:4','2:3','9:16'))"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET storyboard_video_duration = 15 "
                    "WHERE storyboard_video_model = 'sora-2' "
                    "AND (storyboard_video_duration IS NULL "
                    "OR storyboard_video_duration NOT IN (10, 15, 25))"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET storyboard_video_duration = 10 "
                    "WHERE storyboard_video_model = 'grok' "
                    "AND (storyboard_video_duration IS NULL "
                    "OR storyboard_video_duration NOT IN (10, 20, 30))"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET storyboard_video_resolution_name = '720p' "
                    "WHERE storyboard_video_resolution_name IS NULL "
                    "OR TRIM(storyboard_video_resolution_name) = '' "
                    "OR LOWER(TRIM(storyboard_video_resolution_name)) NOT IN ('480p', '720p')"
                )
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET storyboard_video_appoint_account = '' "
                    "WHERE storyboard_video_appoint_account IS NULL"
                )
            )

            if "video_style_template_id" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN video_style_template_id INTEGER")
                )
                print("已添加 episodes.video_style_template_id 字段")
            if "video_prompt_template" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN video_prompt_template TEXT DEFAULT ''")
                )
                print("已添加 episodes.video_prompt_template 字段")
            if "billing_version" not in columns:
                conn.execute(
                    text("ALTER TABLE episodes ADD COLUMN billing_version INTEGER DEFAULT 0")
                )
                print("已添加 episodes.billing_version 字段")
            conn.execute(
                text("UPDATE episodes SET video_prompt_template = '' WHERE video_prompt_template IS NULL")
            )
            conn.execute(
                text("UPDATE episodes SET billing_version = 0 WHERE billing_version IS NULL")
            )
            conn.execute(
                text(
                    "UPDATE episodes "
                    "SET video_prompt_template = ("
                    "  SELECT COALESCE(scripts.video_prompt_template, '') "
                    "  FROM scripts "
                    "  WHERE scripts.id = episodes.script_id"
                    ") "
                    "WHERE TRIM(COALESCE(episodes.video_prompt_template, '')) = '' "
                    "AND TRIM(COALESCE(("
                    "  SELECT scripts.video_prompt_template "
                    "  FROM scripts "
                    "  WHERE scripts.id = episodes.script_id"
                    "), '')) <> ''"
                )
            )

    except Exception as e:
        print(f"检查/迁移 episodes 失败: {str(e)}")


def ensure_simple_storyboard_batch_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "simple_storyboard_batches")
            if not columns:
                return
            if "source_text" not in columns:
                conn.execute(text("ALTER TABLE simple_storyboard_batches ADD COLUMN source_text TEXT DEFAULT ''"))
            if "shots_data" not in columns:
                conn.execute(text("ALTER TABLE simple_storyboard_batches ADD COLUMN shots_data TEXT DEFAULT ''"))
            if "error_message" not in columns:
                conn.execute(text("ALTER TABLE simple_storyboard_batches ADD COLUMN error_message TEXT DEFAULT ''"))
            if "last_attempt" not in columns:
                conn.execute(text("ALTER TABLE simple_storyboard_batches ADD COLUMN last_attempt INTEGER DEFAULT 0"))
            if "retry_count" not in columns:
                conn.execute(text("ALTER TABLE simple_storyboard_batches ADD COLUMN retry_count INTEGER DEFAULT 0"))
            if "total_batches" not in columns:
                conn.execute(text("ALTER TABLE simple_storyboard_batches ADD COLUMN total_batches INTEGER DEFAULT 0"))
            if "updated_at" not in columns:
                conn.execute(text("ALTER TABLE simple_storyboard_batches ADD COLUMN updated_at DATETIME"))
            conn.execute(text("UPDATE simple_storyboard_batches SET source_text = '' WHERE source_text IS NULL"))
            conn.execute(text("UPDATE simple_storyboard_batches SET shots_data = '' WHERE shots_data IS NULL"))
            conn.execute(text("UPDATE simple_storyboard_batches SET error_message = '' WHERE error_message IS NULL"))
            conn.execute(text("UPDATE simple_storyboard_batches SET last_attempt = 0 WHERE last_attempt IS NULL"))
            conn.execute(text("UPDATE simple_storyboard_batches SET retry_count = 0 WHERE retry_count IS NULL"))
            conn.execute(text("UPDATE simple_storyboard_batches SET total_batches = 0 WHERE total_batches IS NULL"))
            conn.execute(text("UPDATE simple_storyboard_batches SET updated_at = created_at WHERE updated_at IS NULL"))
    except Exception as e:
        print(f"检查/迁移 simple_storyboard_batches 失败: {str(e)}")


_parse_simple_storyboard_batch_shots = simple_storyboard_batches_service._parse_simple_storyboard_batch_shots
_build_simple_storyboard_from_batches = simple_storyboard_batches_service._build_simple_storyboard_from_batches
_serialize_simple_storyboard_batch = simple_storyboard_batches_service._serialize_simple_storyboard_batch
_get_simple_storyboard_batch_rows = simple_storyboard_batches_service._get_simple_storyboard_batch_rows
_get_simple_storyboard_batch_summary = simple_storyboard_batches_service._get_simple_storyboard_batch_summary
_refresh_episode_simple_storyboard_from_batches = simple_storyboard_batches_service._refresh_episode_simple_storyboard_from_batches
_split_simple_storyboard_batches = simple_storyboard_batches_service._split_simple_storyboard_batches
_group_simple_storyboard_shots_into_batches = simple_storyboard_batches_service._group_simple_storyboard_shots_into_batches
_persist_programmatic_simple_storyboard_batches = simple_storyboard_batches_service._persist_programmatic_simple_storyboard_batches
_reset_simple_storyboard_batches_for_episode = simple_storyboard_batches_service._reset_simple_storyboard_batches_for_episode
_touch_episode_simple_storyboard_activity = simple_storyboard_batches_service._touch_episode_simple_storyboard_activity
_build_simple_storyboard_batch_runtime_items = simple_storyboard_batches_service._build_simple_storyboard_batch_runtime_items


def _apply_simple_storyboard_batch_update(episode_id: int, payload: Dict[str, Any]) -> None:
    return simple_storyboard_batches_service._apply_simple_storyboard_batch_update(
        episode_id,
        payload,
        session_factory=SessionLocal,
    )


def ensure_billing_columns():
    try:
        with engine.begin() as conn:
            price_rule_columns = get_table_columns(engine, "billing_price_rules")
            if price_rule_columns and "resolution" not in price_rule_columns:
                conn.execute(text("ALTER TABLE billing_price_rules ADD COLUMN resolution TEXT DEFAULT ''"))
                print("已添加 billing_price_rules.resolution 字段")
            if price_rule_columns:
                conn.execute(text("UPDATE billing_price_rules SET resolution = '' WHERE resolution IS NULL"))

            ledger_columns = get_table_columns(engine, "billing_ledger_entries")
            if ledger_columns and "resolution" not in ledger_columns:
                conn.execute(text("ALTER TABLE billing_ledger_entries ADD COLUMN resolution TEXT DEFAULT ''"))
                print("已添加 billing_ledger_entries.resolution 字段")
            if ledger_columns:
                conn.execute(text("UPDATE billing_ledger_entries SET resolution = '' WHERE resolution IS NULL"))
    except Exception as e:
        print(f"检查/迁移 billing 表失败: {str(e)}")


def ensure_billing_defaults():
    try:
        db = SessionLocal()
        try:
            created_count = billing_service.ensure_default_pricing_rules(db)
            db.commit()
            if created_count:
                print(f"已初始化 billing_price_rules 默认规则: {created_count} 条")
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()
    except Exception as e:
        print(f"初始化计费默认规则失败: {str(e)}")


def ensure_video_model_pricing():
    """Ensure video_model_pricing table has correct schema (with provider column)."""
    try:
        columns = get_table_columns(engine, "video_model_pricing")
        if columns and "provider" not in columns:
            with engine.begin() as conn:
                print("[pricing] dropping old video_model_pricing table (missing provider column)")
                conn.execute(text("DROP TABLE video_model_pricing"))
            models.VideoModelPricing.__table__.create(engine)
            print("[pricing] recreated video_model_pricing table with provider column")
    except Exception as e:
        print(f"[pricing] ensure_video_model_pricing: {str(e)}")



# Initialize default video style templates
def ensure_video_style_templates():
    db = SessionLocal()
    try:
        existing = db.query(models.VideoStyleTemplate).count()
        if existing > 0:
            has_2d = db.query(models.VideoStyleTemplate).filter(models.VideoStyleTemplate.name == "2D").first()
            if not has_2d:
                db.add(models.VideoStyleTemplate(
                    name="2D",
                    sora_rule="准则：\n1 不要出现字幕。\n2 对话用中文。\n3 旁白时，人物嘴巴不要说话\n4 除场景声音外，不要有背景音乐",
                    style_prompt="严格按照逐帧2d日式动画的形式，赛璐璐着色风格，手绘动漫风格，强调帧间的手绘/精细绘制属性，帧间动作连贯流畅，每帧保留细腻的线条勾勒与沉稳质感的色彩，光影自然柔和，整体呈现影视动画的叙事分镜感\n线稿清晰，但最终效果趋向于无痕的上色，2D全彩动画，电影感。\nAvoid (Negative Prompts): Low quality, bad anatomy, deformed, grainy, text, watermark, 3D render , western style, cartoonish",
                    is_default=False
                ))

            preferred_template = db.query(models.VideoStyleTemplate).filter(
                models.VideoStyleTemplate.name == "3D国漫"
            ).first()
            if not preferred_template:
                preferred_template = models.VideoStyleTemplate(
                    name="3D国漫",
                    sora_rule="准则：\n1 不要出现字幕。\n2 对话用中文。\n3 旁白时，人物嘴巴不要说话\n4 除场景声音外，不要有背景音乐",
                    style_prompt="3D玄幻国漫",
                    is_default=True
                )
                db.add(preferred_template)

            db.flush()
            existing_default = db.query(models.VideoStyleTemplate).filter(
                models.VideoStyleTemplate.is_default == True
            ).first()
            if not existing_default:
                preferred_template.is_default = True
            db.commit()
            active_default = db.query(models.VideoStyleTemplate).filter(
                models.VideoStyleTemplate.is_default == True
            ).first()
            if active_default:
                print(f"[video-style] keep default template as {active_default.name}")
            else:
                print(f"[video-style] no explicit default template found; kept templates without override")
            return

        templates = [
            models.VideoStyleTemplate(
                name="2D",
                sora_rule="准则：\n1 不要出现字幕。\n2 对话用中文。\n3 旁白时，人物嘴巴不要说话\n4 除场景声音外，不要有背景音乐",
                style_prompt="严格按照逐帧2d日式动画的形式，赛璐璐着色风格，手绘动漫风格，强调帧间的手绘/精细绘制属性，帧间动作连贯流畅，每帧保留细腻的线条勾勒与沉稳质感的色彩，光影自然柔和，整体呈现影视动画的叙事分镜感\n线稿清晰，但最终效果趋向于无痕的上色，2D全彩动画，电影感。\nAvoid (Negative Prompts): Low quality, bad anatomy, deformed, grainy, text, watermark, 3D render , western style, cartoonish",
                is_default=False
            ),
            models.VideoStyleTemplate(
                name="3D国漫",
                sora_rule="准则：\n1 不要出现字幕。\n2 对话用中文。\n3 旁白时，人物嘴巴不要说话\n4 除场景声音外，不要有背景音乐",
                style_prompt="3D玄幻国漫",
                is_default=True
            ),
            models.VideoStyleTemplate(
                name="3D（半真实）",
                sora_rule="准则：\n1 不要出现字幕。\n2 对话用中文。\n3 旁白时，人物嘴巴不要说话\n4 除场景声音外，不要有背景音乐\n5 只能出现亚洲人面孔",
                style_prompt="仙侠3D半写实风格，PBR着色，次表面散射，环境光遮蔽，电影级景深，体积光照，暖色调，黄金时刻光晕，玄幻氛围。",
                is_default=False
            ),
            models.VideoStyleTemplate(
                name="真人",
                sora_rule="准则：\n1 不要出现字幕。\n2 对话用中文。\n3 旁白时，人物嘴巴不要说话\n4 除场景声音外，不要有背景音乐\n5 只能出现亚洲人面孔",
                style_prompt="皮肤质感：真实皮肤纹理，可见毛孔，自然皮肤瑕疵，皮肤细节，次表面散射，服饰为真实织物纹理，发丝为自然发质纹理，真实发丝，发量丰盈，光泽头发，自然光照，轮廓光，黄金时刻，自然人类表情，真实情感",
                is_default=False
            ),
        ]
        for t in templates:
            db.add(t)
        db.commit()
        print(f"[video-style] initialized {len(templates)} default video style templates")
    except Exception as e:
        print(f"[video-style] init failed: {e}")
        db.rollback()
    finally:
        db.close()


DEFAULT_PROMPTS = [
    {
        "key": "stage1_initial_storyboard",
        "name": "阶段1：初步分镜生成",
        "description": "按500字分批，生成约9个镜头的初步分镜",
        "content": """你是一位专业的影视分镜师。我将给你一段剧本内容，请将其拆分为大约9个镜头的分镜表。

【要求】
1. 每个镜头需要输出：镜号、主体（角色/场景/道具）、语音内容、原剧本段落
2. 主体类型只有三类：角色 / 场景 / 道具。
3. 主体只保留该镜头最主要的1-2个角色 + 1个场景 + 0-2个关键道具，避免过度细分。
4. 根据内容自然分镜，大约9个镜头（可以略多或略少）
5. 确保镜头之间连贯流畅

【关于语音内容的处理】
- **旁白**：第一人称的内心独白或解说
  - 使用具体人名作为说话人（如：李馨儿、萧景珩）
  - 第三方视角 → 使用"旁白"
  - 需要标注性别（女/男/中性）
  - 需要标注情绪（如：平静、悲伤、愤怒、带着哭腔等）

- **对话**：角色之间的对话
  - 按对话顺序列出，每句包含：说话人、对方、性别、情绪、对话内容
  - 示例：李馨儿对萧景珩说（带着哭腔）：你当真要我走？

【剧本内容】
{content}

【输出格式】
请严格按照以下JSON格式输出：
```json
{{
  "shots": [
    {{
      "shot_number": 1,
      "subjects": [
        {{"name": "柳如烟", "type": "角色"}},
        {{"name": "书房", "type": "场景"}},
        {{"name": "铜镜", "type": "道具"}}
      ],
      "original_text": "该镜头对应的原文片段",
      "voice_type": "narration 或 dialogue 或 none",
      "narration": {{
        "speaker": "具体人名（如：李馨儿）或 旁白",
        "gender": "女 或 男 或 中性",
        "emotion": "情绪描述",
        "text": "旁白内容"
      }},
      "dialogue": [
        {{
          "speaker": "说话人名字",
          "target": "对方名字（如果是对某人说）或 null",
          "gender": "女 或 男",
          "emotion": "情绪描述",
          "text": "对话内容"
        }}
      ]
    }}
  ]
}}
```

**注意**：
- 如果 voice_type 为 "narration"，则 dialogue 为 null
- 如果 voice_type 为 "dialogue"，则 narration 为 null
- 如果 voice_type 为 "none"（无语音），则 narration 和 dialogue 都为 null
- dialogue 数组可以包含多轮对话
- target 如果不是对特定人物说话（如自言自语、对众人说），则为 null

请开始分析。"""
    },
    {
        "key": "stage2_refine_shot",
        "name": "阶段2：主体绘画提示词",
        "description": "基于完整分镜表生成主体绘画提示词与别名",
        "content": """你是一位专业的AI绘画提示词工程师。以下是用户输入的分镜表：

  【分镜表】（共{total_shots}个镜头）
  {full_storyboard_json}

  【任务】
  1. 主体类型只有三类：角色 / 场景 / 道具。
  2. 智能识别重复和相似的主体：
     - 如果不同名称明显指向同一角色（昵称、称号、全名等），只保留一个规范名
     - 如果不同名称指向相似场景或同一关键道具，可以合并为一个
  3. 生成主体名称映射表（name_mappings）：
     - 将所有被合并的原始名称映射到规范名称
     - 映射表格式为JSON对象，键为原始名称，值为规范名称
  4. 为每个主体生成绘画提示词与别名。
     - 角色 ai_prompt：年龄 + 性别 + 表情 + 眼睛 + 发型 + 配饰 + 衣服 + 细节
     - 角色 role_personality：用中文一句话描述角色性格，场景/道具留空字符串
     - 场景 ai_prompt：整体风格、环境氛围、光影效果、细节特征
     - 道具 ai_prompt：材质、造型、颜色、结构、使用痕迹、细节特征
     - alias 为简短描述（10-20字）

  【输出格式】
  请严格按照以下JSON格式输出：
  {{
    "subjects": [
      {{
        "name": "规范主体名称",
        "type": "角色 / 场景 / 道具",
        "ai_prompt": "详细的绘画提示词",
        "role_personality": "角色性格（中文一句话），场景/道具填空字符串",
        "alias": "简短描述（10-20字）"
      }}
    ],
    "name_mappings": {{
      "原始名称1": "规范名称1",
      "原始名称2": "规范名称1"
    }}
  }}

  **重要提示**：
  - subjects 数组中只包含去重后的规范主体
  - name_mappings 必须包含所有被合并的原始名称到规范名称的映射
  - 如果某个主体没有被合并，则不需要在 name_mappings 中出现
  - 角色 ai_prompt：年龄 + 性别 + 表情 + 眼睛 + 发型 + 配饰 + 衣服 + 细节
  - 角色 role_personality：仅角色填写中文一句话性格描述，场景/道具必须为空字符串
  - 场景 ai_prompt：整体风格、环境氛围、光影效果、细节特征
  - 道具 ai_prompt：材质、造型、颜色、结构、使用痕迹、细节特征
  - 直接输出JSON对象，不要用markdown代码块包裹"""
    },
    {
        "key": "generate_subject_ai_prompt",
        "name": "生成主体绘画提示词",
        "description": "为单个主体生成AI绘画提示词与别名",
        "content": """你是一位专业的AI绘画提示词工程师。请为指定主体生成详细的绘画提示词。

【主体信息】
- 名称：{subject_name}
- 类型：{subject_type}

【分镜表上下文】
{storyboard_context}

【任务】
根据分镜表中该主体出现的场景和描述，生成详细的绘画提示词与别名。
- 角色 ai_prompt：年龄 + 性别 + 表情 + 眼睛 + 发型 + 配饰 + 衣服 + 细节
- 场景 ai_prompt：整体风格、环境氛围、光影效果、细节特征
- 道具 ai_prompt：材质、造型、颜色、结构、使用痕迹、细节特征
- alias 为简短描述（10-20字）
- ai_prompt需要是纯中文的

【输出格式】
请严格按照以下JSON格式输出：
{{
  "ai_prompt": "详细的绘画提示词（纯中文）",
  "alias": "简短描述（10-20字）"
}}

**重要提示**：
- 角色 ai_prompt：年龄 + 性别 + 表情 + 眼睛 + 发型 + 配饰 + 衣服 + 细节
- 场景 ai_prompt：整体风格、环境氛围、光影效果、细节特征
- 道具 ai_prompt：材质、造型、颜色、结构、使用痕迹、细节特征
- ai_prompt必须是纯中文
- alias为10-20字的简短描述
- 直接输出JSON对象，不要用markdown代码块包裹"""
    },
    {
        "key": "detailed_storyboard_content_analysis",
        "name": "详细分镜：内容分析",
        "description": "对已划分的镜头进行详细内容分析，提取主体、对白、旁白等信息",
        "content": """你是一位专业的影视分镜师。我将给你一组已经划分好的镜头（每个镜头包含镜号和原剧本段落），请为每个镜头提取详细的分镜信息。

【镜头列表】
{shots_content}

【分镜拆分要求】
每个镜头必须输出以下内容：
- 镜号（保持输入的一致）
- 主体（角色、场景、道具）
- 旁白或对白
- 原剧本段落（保持输入的一致）

【主体提取原则】
主体类型只能输出三类：角色、场景、道具。
主体必须包含至少一个场景（即：故事发生的地点）
主体只保留该镜头最主要的1-2个角色 + 1个场景 + 0-2个关键道具，避免把所有背景摆设都当成道具。
当原文出现对剧情推进、人物动作、画面构图或视觉识别有直接影响的具体物件时，必须提取为道具。
服装、发型、人物身体部位、纯抽象概念、普通环境装饰默认不要当作道具；只有被角色拿着、使用、注视，或被原文重点描述的关键物件才算道具。

【核心总原则（非常重要）】
严格保持输入数据的一致
如果输入不明确时，优先使用描述性文字。
只能严格使用剧本原文中的角色名字，如果角色没有姓名，也不要生成"女主角"、"男配角"。
你可以输入"警察"等职业常识性名词或用"警察应该叫做谁"来描述。
对原文未命名的，如果必须命名，可以使用唯一指代性的代词描述（例如"旗袍女""老板娘""母亲"等）。

【处理旁白及对白的处理】
- **旁白**：第三人称的旁述或独白或心理描述
  - 使用句号断句，句尾作为说话人，例如："某某内心道……"
  - 如果语句接近 → 使用"旁白"
  - 需要填写注意性别（女/男/无性别）
  - 需要填写注意情绪（例如：平静、伤心、悲愤、惊恐、兴奋等）
  - 角色内心的独白法则需注意叙事话语、细节等

- **对话**：角色之间的对话
  - 按对话顺序列出，每个对象包含说话人、对方名字（或描述）、对话内容
  - 示例：【某某内心道："……"，惊恐地（询问）："你当真要……？"】

【输出格式】
请严格按照以下JSON格式输出：
```json
{{
  "shots": [
    {{
      "shot_number": 1,
      "subjects": [
        {{"name": "角色名", "type": "角色"}},
        {{"name": "场景名", "type": "场景"}},
        {{"name": "关键道具", "type": "道具"}}
      ],
      "original_text": "该镜头对应的原文片段",
      "voice_type": "narration 或 dialogue 或 none",
      "narration": {{
        "speaker": "旁白角色名 或 旁白",
        "gender": "女 或 男 或 无 或 未知",
        "emotion": "情绪描述",
        "text": "旁白内容"
      }},
      "dialogue": [
        {{
          "speaker": "说话人名字",
          "target": "对方名字（如果是对某人说的）或 null",
          "gender": "女 或 男",
          "emotion": "情绪描述",
          "text": "对话内容"
        }}
      ]
    }}
  ]
}}
```

**注意**：
- 如果 voice_type 为 "narration"，则 dialogue 为 null
- 如果 voice_type 为 "dialogue"，则 narration 为 null
- 如果 voice_type 为 "none"，则两者（两个都是 narration 和 dialogue 均为 null
- dialogue 可以是旁白，可以是多段对话
- target 代表是对谁说的（如果对特定人说的），如果是广播性质或说话人为 null
- 保持 original_text 与输入完全一致，不要修改
- 保持 shot_number 与输入完全一致
- 直接输出JSON对象，不要用markdown代码块包裹

请开始分析。"""
    },
    {
        "key": CHARACTER_THREE_VIEW_PROMPT_KEY,
        "name": "角色三视图生图提示词",
        "description": "主体界面角色卡片点击“生成三视图”时使用的基础提示词",
        "content": CHARACTER_THREE_VIEW_PROMPT_DEFAULT
    },
    {
        "key": MANAGED_PROMPT_OPTIMIZE_KEY,
        "name": "优化提示词",
        "description": "托管任务因文字描述审核不通过时，用于一次性改写完整视频提示词",
        "content": MANAGED_PROMPT_OPTIMIZE_DEFAULT
    },
    {
        "key": "generate_video_prompts",
        "name": "分镜视频提示词生成",
        "description": "生成分镜的视频/音频/台词提示词",
        "content": """你是专业的分镜提示词生成助手。请根据以下信息生成分镜时间轴。

原剧本段落：
{script_excerpt}

出镜主体：
{subject_text}

说明：
- 角色主体会按“主体名-性格描述”的格式提供，场景/道具主体只写名称
- 请在画面设计、动作、表情和情绪表现中参考对应角色的性格信息

输出 JSON，格式如下：
{{
  "timeline": [
    {{
      "time": "00s-04s",
      "visual": "[推镜] [中景] 萧景珩翻阅书卷，眉头紧锁，手指轻抚纸页",
      "audio": "[萧景珩] 说：\\"此事不简单\\""
    }},
    {{
      "time": "04s-08s",
      "visual": "[特写] [近景] 手指停在某页，眼神凝重，面部微表情",
      "audio": "(SFX:翻页声、环境音)"
    }}
  ]
}}

要求：
1. 时长总计 {safe_duration} 秒，分为3-4个时间段
2. time字段格式：00s-04s、04s-08s（连续不重叠，覆盖完整时长）
3. visual字段包含：
   - 镜头类型：如[推镜][拉镜][摇镜][跟镜][正反打][切镜]等（自由发挥）
   - 景别：如[远景][全景][中景][近景][特写][大特写]等（自由发挥）
   - 画面描述：忠实描述原剧本段落的动作和情绪，不添加额外内容
4. audio字段包含：
   - 角色台词：格式为 [角色名] 说："台词内容"（保留原文台词）
   - 音效标记：格式为 (SFX:具体音效描述)
   - 如果既有台词又有音效，用顿号分隔：[角色] 说："台词"、(SFX:音效)
5. 每个时间段专注一个镜头焦点，保持连贯性
6. 只输出 JSON，不要其他说明

{extra_style}"""
    },
    {
        "key": "generate_large_shot_prompts",
        "name": "大镜头提示词生成",
        "description": "生成更偏电影化、镜头语言更强的大镜头时间轴提示词",
        "content": build_large_shot_prompt_rule(15, 4)
    },
    {
        "key": STORYBOARD2_VIDEO_PROMPT_KEY,
        "name": "故事板2分镜视频提示词生成",
        "description": "生成故事板2分镜的视频/音频/台词提示词",
        "content": STORYBOARD2_VIDEO_PROMPT_DEFAULT
    },
    {
        "key": STORYBOARD2_IMAGE_PROMPT_KEY,
        "name": "故事板2镜头图生图提示词",
        "description": "故事板2生成镜头图时自动前置的提示词",
        "content": STORYBOARD2_IMAGE_PROMPT_DEFAULT
    }
]

def ensure_prompt_config_table():
    """确保提示词配置表存在并初始化默认数据"""
    try:
        if not table_exists(engine, "prompt_configs"):
            print("prompt_configs 表不存在，将由 SQLAlchemy 创建")
            return

        # 表存在，补齐默认提示词（不覆盖已有配置）
        db = SessionLocal()
        try:
            existing_keys = {
                row[0] for row in db.query(models.PromptConfig.key).all()
            }
            inserted_count = 0
            for prompt_data in DEFAULT_PROMPTS:
                if prompt_data["key"] in existing_keys:
                    continue
                db.add(models.PromptConfig(**prompt_data))
                inserted_count += 1

            if inserted_count > 0:
                db.commit()
                print(f"prompt_configs 已补齐 {inserted_count} 条默认配置")
            else:
                total_count = db.query(models.PromptConfig).count()
                print(f"prompt_configs 表已有 {total_count} 条数据，无需补齐")
        except Exception as e:
            print(f"初始化默认提示词失败: {str(e)}")
            db.rollback()
        finally:
            db.close()

    except Exception as e:
        print(f"检查/初始化 prompt_configs 表失败: {str(e)}")



def upgrade_stage2_refine_shot_prompt_content(content: str) -> str:
    content_text = str(content or "")
    if not content_text:
        return content_text

    updated = content_text

    updated = updated.replace(
        "主体类型只有两类：角色 / 场景。",
        "主体类型只有三类：角色 / 场景 / 道具。"
    )
    updated = updated.replace(
        "如果不同名称指向相似场景，可以合并为一个",
        "如果不同名称指向相似场景或同一关键道具，可以合并为一个"
    )
    updated = updated.replace(
        "角色 role_personality_en：英文描述角色性格特征（如：calm, proud, stubborn），场景留空字符串",
        "角色 role_personality：用中文一句话描述角色性格，场景/道具留空字符串"
    )
    updated = updated.replace(
        '"role_personality_en": "角色性格（英文），场景填空字符串",',
        '"role_personality": "角色性格（中文一句话），场景/道具填空字符串",'
    )
    updated = updated.replace(
        "角色 role_personality_en：仅角色填写英文性格特征，场景必须为空字符串",
        "角色 role_personality：仅角色填写中文一句话性格描述，场景/道具必须为空字符串"
    )
    updated = updated.replace(
        "角色 role_personality：用中文一句话描述角色性格，场景留空字符串",
        "角色 role_personality：用中文一句话描述角色性格，场景/道具留空字符串"
    )
    updated = updated.replace(
        '"role_personality": "角色性格（中文一句话），场景填空字符串",',
        '"role_personality": "角色性格（中文一句话），场景/道具填空字符串",'
    )
    updated = updated.replace(
        "角色 role_personality：仅角色填写中文一句话性格描述，场景必须为空字符串",
        "角色 role_personality：仅角色填写中文一句话性格描述，场景/道具必须为空字符串"
    )
    updated = updated.replace(
        '"type": "角色 或 场景",',
        '"type": "角色 / 场景 / 道具",'
    )

    task_line = "     - 角色 ai_prompt：年龄 + 性别 + 表情 + 眼睛 + 发型 + 配饰 + 衣服 + 细节"
    scene_line = "     - 场景 ai_prompt：整体风格、环境氛围、光影效果、细节特征"
    prop_line = "     - 道具 ai_prompt：材质、造型、颜色、结构、使用痕迹、细节特征"
    if "角色 role_personality：" not in updated and "角色 role_personality_en：" not in updated:
        task_insert = task_line + "\n     - 角色 role_personality：用中文一句话描述角色性格，场景/道具留空字符串"
        updated = updated.replace(task_line, task_insert, 1)
    if prop_line not in updated and scene_line in updated:
        updated = updated.replace(scene_line, scene_line + "\n" + prop_line, 1)

    output_line = '        "ai_prompt": "详细的绘画提示词",\n        "alias": "简短描述（10-20字）"'
    if '"role_personality":' not in updated and '"role_personality_en":' not in updated:
        output_insert = '        "ai_prompt": "详细的绘画提示词",\n        "role_personality": "角色性格（中文一句话），场景/道具填空字符串",\n        "alias": "简短描述（10-20字）"'
        updated = updated.replace(output_line, output_insert, 1)

    hint_line = "  - 角色 ai_prompt：年龄 + 性别 + 表情 + 眼睛 + 发型 + 配饰 + 衣服 + 细节"
    scene_hint_line = "  - 场景 ai_prompt：整体风格、环境氛围、光影效果、细节特征"
    prop_hint_line = "  - 道具 ai_prompt：材质、造型、颜色、结构、使用痕迹、细节特征"
    if "角色 role_personality：" not in updated and "角色 role_personality_en：" not in updated:
        hint_insert = hint_line + "\n  - 角色 role_personality：仅角色填写中文一句话性格描述，场景/道具必须为空字符串"
        updated = updated.replace(hint_line, hint_insert, 1)
    if prop_hint_line not in updated and scene_hint_line in updated:
        updated = updated.replace(scene_hint_line, scene_hint_line + "\n" + prop_hint_line, 1)

    return updated


def ensure_stage2_refine_shot_prompt_config():
    """补齐阶段2提示词中的角色性格字段和道具主体规则，不影响其他自定义内容。"""
    db = SessionLocal()
    try:
        config = db.query(models.PromptConfig).filter(
            models.PromptConfig.key == "stage2_refine_shot"
        ).first()

        if not config:
            return

        upgraded = upgrade_stage2_refine_shot_prompt_content(config.content)
        if upgraded != (config.content or ""):
            config.content = upgraded
            db.commit()
            print("已补齐 stage2_refine_shot 的角色性格字段和道具主体规则")
    except Exception as e:
        db.rollback()
        print(f"更新 stage2_refine_shot 提示词失败: {str(e)}")
    finally:
        db.close()


def upgrade_subject_prompt_content_for_props(content: str) -> str:
    content_text = str(content or "")
    if not content_text:
        return content_text

    updated = content_text
    replacements = (
        ("主体（仅角色/场景）", "主体（角色/场景/道具）"),
        ("主体类型只有两类：角色 / 场景。不输出道具。", "主体类型只有三类：角色 / 场景 / 道具。"),
        ("主体只保留该镜头最主要的1-2个角色 + 1个场景，避免过度细分。", "主体只保留该镜头最主要的1-2个角色 + 1个场景 + 0-2个关键道具，避免过度细分。"),
        ("主体（角色和场景）", "主体（角色、场景、道具）"),
        ("主体类型只能输出两类：角色、场景。", "主体类型只能输出三类：角色、场景、道具。"),
        ("- 角色主体会按“主体名-性格描述”的格式提供，场景主体只写名称", "- 角色主体会按“主体名-性格描述”的格式提供，场景/道具主体只写名称"),
        ("角色 role_personality：用中文一句话描述角色性格，场景留空字符串", "角色 role_personality：用中文一句话描述角色性格，场景/道具留空字符串"),
        ('"role_personality": "角色性格（中文一句话），场景填空字符串",', '"role_personality": "角色性格（中文一句话），场景/道具填空字符串",'),
        ("角色 role_personality：仅角色填写中文一句话性格描述，场景必须为空字符串", "角色 role_personality：仅角色填写中文一句话性格描述，场景/道具必须为空字符串"),
    )
    for old_text, new_text in replacements:
        updated = updated.replace(old_text, new_text)

    if "角色 ai_prompt：年龄 + 性别 + 表情 + 眼睛 + 发型 + 配饰 + 衣服 + 细节" in updated and "道具 ai_prompt：" not in updated:
        updated = updated.replace(
            "- 场景 ai_prompt：整体风格、环境氛围、光影效果、细节特征",
            "- 场景 ai_prompt：整体风格、环境氛围、光影效果、细节特征\n- 道具 ai_prompt：材质、造型、颜色、结构、使用痕迹、细节特征"
        )

    if '{{"name": "铜镜", "type": "道具"}}' not in updated:
        updated = updated.replace(
            '{{"name": "书房", "type": "场景"}}',
            '{{"name": "书房", "type": "场景"}},\n        {{"name": "铜镜", "type": "道具"}}'
        )
    if '{{"name": "关键道具", "type": "道具"}}' not in updated:
        updated = updated.replace(
            '{{"name": "场景名", "type": "场景"}}',
            '{{"name": "场景名", "type": "场景"}},\n        {{"name": "关键道具", "type": "道具"}}'
        )

    updated = re.sub(
        r'(\{\{"name": "书房", "type": "场景"\}\})(?:,\s*\{\{"name": "铜镜", "type": "道具"\}\})+',
        r'\1,\n        {{"name": "铜镜", "type": "道具"}}',
        updated,
    )
    updated = re.sub(
        r'(\{\{"name": "场景名", "type": "场景"\}\})(?:,\s*\{\{"name": "关键道具", "type": "道具"\}\})+',
        r'\1,\n        {{"name": "关键道具", "type": "道具"}}',
        updated,
    )
    return updated


def upgrade_detailed_storyboard_content_analysis_prompt_content(content: str) -> str:
    content_text = upgrade_subject_prompt_content_for_props(content)
    if not content_text:
        return content_text

    updated = content_text

    scene_rule = "主体必须包含至少一个场景（即：故事发生的地点）"
    prop_scope_rule = "主体只保留该镜头最主要的1-2个角色 + 1个场景 + 0-2个关键道具，避免把所有背景摆设都当成道具。"
    if prop_scope_rule not in updated and scene_rule in updated:
        updated = updated.replace(scene_rule, scene_rule + "\n" + prop_scope_rule, 1)

    must_extract_prop_rule = "当原文出现对剧情推进、人物动作、画面构图或视觉识别有直接影响的具体物件时，必须提取为道具。"
    skip_noise_prop_rule = "服装、发型、人物身体部位、纯抽象概念、普通环境装饰默认不要当作道具；只有被角色拿着、使用、注视，或被原文重点描述的关键物件才算道具。"
    principles_header = "【主体提取原则】"
    if must_extract_prop_rule not in updated and principles_header in updated:
        updated = updated.replace(
            principles_header,
            principles_header + "\n" + must_extract_prop_rule + "\n" + skip_noise_prop_rule,
            1,
        )

    if "关键道具" not in updated and '"subjects": [' in updated:
        updated = updated.replace(
            '{{"name": "场景名", "type": "场景"}}',
            '{{"name": "场景名", "type": "场景"}},\n        {{"name": "关键道具", "type": "道具"}}',
            1,
        )

    return updated


def ensure_prop_subject_prompt_configs():
    """补齐主体相关提示词中的道具规则，不覆盖其他自定义内容。"""
    target_keys = (
        "stage1_initial_storyboard",
        "generate_subject_ai_prompt",
        "detailed_storyboard_content_analysis",
    )

    db = SessionLocal()
    try:
        configs = db.query(models.PromptConfig).filter(
            models.PromptConfig.key.in_(target_keys)
        ).all()

        updated_count = 0
        for config in configs:
            if config.key == "detailed_storyboard_content_analysis":
                upgraded = upgrade_detailed_storyboard_content_analysis_prompt_content(config.content)
            else:
                upgraded = upgrade_subject_prompt_content_for_props(config.content)
            if upgraded != (config.content or ""):
                config.content = upgraded
                updated_count += 1

        if updated_count:
            db.commit()
            print(f"已补齐 {updated_count} 条主体相关提示词的道具规则")
    except Exception as e:
        db.rollback()
        print(f"更新主体相关提示词失败: {str(e)}")
    finally:
        db.close()


def _is_placeholder_question_text(text: str) -> bool:
    value = str(text or "").strip()
    if not value:
        return True
    return "???" in value and not re.search(r"[\u4e00-\u9fff]", value)


def ensure_character_three_view_prompt_config():
    """修复角色三视图提示词配置中的乱码占位符，不覆盖正常自定义内容。"""
    db = SessionLocal()
    try:
        config = db.query(models.PromptConfig).filter(
            models.PromptConfig.key == CHARACTER_THREE_VIEW_PROMPT_KEY
        ).first()
        if not config:
            return

        should_update = False

        if _is_placeholder_question_text(config.name):
            config.name = "角色三视图生图提示词"
            should_update = True

        if _is_placeholder_question_text(config.description):
            config.description = "主体界面角色卡片点击“生成三视图”时使用的基础提示词"
            should_update = True

        if _is_placeholder_question_text(config.content):
            config.content = CHARACTER_THREE_VIEW_PROMPT_DEFAULT
            should_update = True

        if should_update:
            db.commit()
            print("已修复角色三视图提示词配置中的乱码占位符")
    except Exception as e:
        db.rollback()
        print(f"修复角色三视图提示词配置失败: {str(e)}")
    finally:
        db.close()




_normalize_subject_detail_entry = storyboard_sync.normalize_subject_detail_entry
_build_subject_detail_map = storyboard_sync.build_subject_detail_map
_normalize_storyboard_generation_subjects = storyboard_sync.normalize_storyboard_generation_subjects
_SUBJECT_MATCH_STOP_FRAGMENTS = storyboard_sync.SUBJECT_MATCH_STOP_FRAGMENTS
_find_meaningful_common_fragment = storyboard_sync.find_meaningful_common_fragment
_infer_storyboard_role_name_from_shot = storyboard_sync.infer_storyboard_role_name_from_shot
_resolve_storyboard_subject_name = storyboard_sync.resolve_storyboard_subject_name
_reconcile_storyboard_shot_subjects = storyboard_sync.reconcile_storyboard_shot_subjects
_sync_subjects_to_database = storyboard_sync.sync_subjects_to_database
_sync_storyboard_to_shots = storyboard_sync.sync_storyboard_to_shots


def _normalize_stage2_subjects(subjects: Optional[list]) -> list:
    return list(_build_subject_detail_map(subjects).values())


def ensure_storyboard2_prompt_config():
    """确保故事板2分镜提示词配置存在（仅初始化，不覆盖用户已保存内容）。"""
    db = SessionLocal()
    try:
        config = db.query(models.PromptConfig).filter(
            models.PromptConfig.key == STORYBOARD2_VIDEO_PROMPT_KEY
        ).first()

        if not config:
            db.add(models.PromptConfig(
                key=STORYBOARD2_VIDEO_PROMPT_KEY,
                name="故事板2分镜视频提示词生成",
                description="生成故事板2分镜的视频/音频/台词提示词",
                content=STORYBOARD2_VIDEO_PROMPT_DEFAULT
            ))
            db.commit()
            print("已初始化故事板2分镜提示词配置")
            return

        content_text = str(config.content or "").strip()
        has_subjects_field = "\"subjects\"" in content_text
        has_legacy_subject_marker = "[SB2_SUBJECTS_RULES_V2]" in content_text
        looks_like_stage1_prompt = (
            "阶段1：初步分镜生成" in content_text
            or ("\"shots\"" in content_text and "\"timeline\"" not in content_text)
            or "你是一位专业的影视分镜师" in content_text
        )
        missing_core_placeholders = (
            "{script_excerpt}" not in content_text
            or "{subject_text}" not in content_text
            or "{scene_description}" not in content_text
            or "{safe_duration}" not in content_text
        )

        if (
            looks_like_stage1_prompt
            or missing_core_placeholders
            or (not has_subjects_field)
            or has_legacy_subject_marker
        ):
            # 避免在服务重启时覆盖管理页已保存的自定义模板，只给出提示。
            print("检测到故事板2分镜提示词可能是旧版格式，已跳过自动覆盖，请在管理页手动更新。")
    except Exception as e:
        db.rollback()
        print(f"升级故事板2分镜提示词配置失败: {str(e)}")
    finally:
        db.close()




def _remove_scene_description_placeholder_block(prompt_text: str) -> str:
    """Remove the storyboard(sora) scene_description block without touching other content."""
    text_value = str(prompt_text or "")
    if not text_value:
        return text_value

    replacements = (
        (
            "{script_excerpt}\n\n场景描述：\n{scene_description}\n\n出镜主体：\n{subject_text}",
            "{script_excerpt}\n\n出镜主体：\n{subject_text}",
        ),
        (
            "{script_excerpt}\r\n\r\n场景描述：\r\n{scene_description}\r\n\r\n出镜主体：\r\n{subject_text}",
            "{script_excerpt}\r\n\r\n出镜主体：\r\n{subject_text}",
        ),
        ("\n\n场景描述：\n{scene_description}", ""),
        ("\r\n\r\n场景描述：\r\n{scene_description}", ""),
    )
    updated = text_value
    for old_value, new_value in replacements:
        updated = updated.replace(old_value, new_value, 1)
    return updated


def _subject_personality_hint_text(for_storyboard2: bool = False) -> str:
    if for_storyboard2:
        return (
            "说明：\n"
            "- 角色主体会按“主体名-性格描述”的格式提供，场景/道具主体只写名称\n"
            "- 生成结果里的 subjects 数组只填写“-”前的主体名称，不要把后面的性格描述写进去"
        )
    return (
        "说明：\n"
        "- 角色主体会按“主体名-性格描述”的格式提供，场景/道具主体只写名称\n"
        "- 请在画面设计、动作、表情和情绪表现中参考对应角色的性格信息"
    )


def _is_subject_personality_hint_line(line: str) -> bool:
    stripped = str(line or "").strip()
    if not stripped:
        return False

    if stripped == "说明：":
        return True

    hint_keywords = (
        "主体名-性格描述",
        "角色主体会按",
        "场景主体只写名称",
        "场景/道具主体只写名称",
        "道具主体只写名称",
        "画面设计",
        "动作、表情和情绪表现",
        "性格信息",
        "subjects 数组",
        "后面的性格描述写进去",
    )
    if any(keyword in stripped for keyword in hint_keywords):
        return True

    normalized = stripped.replace(" ", "").replace("-", "").replace("·", "").replace(":", "").replace("：", "")
    if not normalized:
        return False

    question_ratio = normalized.count("?") / len(normalized)
    return question_ratio >= 0.4


def _find_subject_personality_section_end(text_value: str, start_index: int) -> int:
    section_markers = (
        "\n\n输出 JSON",
        "\n\n可选的景别",
        "\n\n要求：",
        "\n\n输出要求",
        "\n\n输出格式",
        "\n\n请输出",
        "\n\n示例：",
    )
    positions = []
    for marker in section_markers:
        marker_index = text_value.find(marker, start_index)
        if marker_index != -1:
            positions.append(marker_index)
    return min(positions) if positions else -1


def _inject_subject_personality_hint(prompt_text: str, for_storyboard2: bool = False) -> str:
    text_value = str(prompt_text or "")
    if not text_value:
        return text_value

    hint_text = _subject_personality_hint_text(for_storyboard2)

    if "{subject_text}" in text_value:
        subject_end = text_value.find("{subject_text}") + len("{subject_text}")
        section_end = _find_subject_personality_section_end(text_value, subject_end)
        if section_end != -1:
            existing_block = text_value[subject_end:section_end]
            preserved_lines = []
            for raw_line in existing_block.splitlines():
                stripped = raw_line.strip()
                if not stripped:
                    continue
                if _is_subject_personality_hint_line(stripped):
                    continue
                preserved_lines.append(raw_line.rstrip())

            normalized_block = "\n\n" + hint_text
            if preserved_lines:
                normalized_block += "\n\n" + "\n".join(preserved_lines)
            return text_value[:subject_end] + normalized_block + text_value[section_end:]

        return text_value.replace("{subject_text}", "{subject_text}\n\n" + hint_text, 1)

    if "说明：" in text_value or "主体名-性格描述" in text_value or "subjects 数组" in text_value:
        cleaned_lines = []
        for raw_line in text_value.splitlines():
            stripped = raw_line.strip()
            if not stripped:
                cleaned_lines.append("")
                continue
            if _is_subject_personality_hint_line(stripped):
                continue
            cleaned_lines.append(raw_line.rstrip())
        text_value = "\n".join(cleaned_lines).strip()

    return f"{text_value.rstrip()}\n\n{hint_text}".strip()


def ensure_shot_duration_template_without_scene_description():
    """Remove the storyboard(sora) scene_description block from duration templates."""
    db = SessionLocal()
    try:
        templates = db.query(models.ShotDurationTemplate).all()
        updated_count = 0

        for template in templates:
            old_rule = template.video_prompt_rule or ""
            new_rule = _remove_scene_description_placeholder_block(old_rule)
            if new_rule != old_rule:
                template.video_prompt_rule = new_rule
                updated_count += 1

        if updated_count > 0:
            db.commit()
            print(f"已更新时长配置模板 video_prompt_rule，移除 scene_description 段落: {updated_count} 条")
    except Exception as e:
        db.rollback()
        print(f"更新时长配置模板 scene_description 段落失败: {str(e)}")
    finally:
        db.close()


def ensure_generate_video_prompt_config_without_scene_description():
    """Remove the storyboard(sora) scene_description block from the default video prompt config."""
    db = SessionLocal()
    try:
        config = db.query(models.PromptConfig).filter(
            models.PromptConfig.key == "generate_video_prompts"
        ).first()
        if not config:
            return

        old_content = config.content or ""
        new_content = _remove_scene_description_placeholder_block(old_content)
        if new_content != old_content:
            config.content = new_content
            db.commit()
            print("已更新 generate_video_prompts，移除 scene_description 段落")
    except Exception as e:
        db.rollback()
        print(f"更新 generate_video_prompts 的 scene_description 段落失败: {str(e)}")
    finally:
        db.close()


def _build_default_simple_storyboard_config_payload(duration: int) -> Dict[str, Any]:
    return get_default_rule_config(duration).to_dict()


def ensure_shot_duration_template_config_json():
    """Ensure shot duration templates expose structured simple-storyboard rule configs."""
    try:
        columns = get_table_columns(engine, "shot_duration_templates")
        if not columns:
            return
        if "simple_storyboard_config_json" not in columns:
            with engine.begin() as conn:
                conn.execute(
                    text("ALTER TABLE shot_duration_templates ADD COLUMN simple_storyboard_config_json TEXT DEFAULT ''")
                )
            print("已添加 shot_duration_templates.simple_storyboard_config_json 字段")
    except Exception as e:
        print(f"补充 shot_duration_templates.simple_storyboard_config_json 字段失败: {str(e)}")
        return

    db = SessionLocal()
    try:
        templates = db.query(models.ShotDurationTemplate).all()
        updated_count = 0
        for template in templates:
            current_value = str(getattr(template, "simple_storyboard_config_json", "") or "").strip()
            if current_value:
                try:
                    normalize_rule_config(json.loads(current_value), int(getattr(template, "duration", 15) or 15))
                    continue
                except Exception:
                    pass
            template.simple_storyboard_config_json = json.dumps(
                _build_default_simple_storyboard_config_payload(int(getattr(template, "duration", 15) or 15)),
                ensure_ascii=False,
            )
            updated_count += 1
        if updated_count > 0:
            db.commit()
            print(f"已补齐简单分镜结构化规则配置: {updated_count} 条")
    except Exception as e:
        db.rollback()
        print(f"补齐简单分镜结构化规则配置失败: {str(e)}")
    finally:
        db.close()


def ensure_shot_duration_template_large_shot_rule():
    """Ensure shot duration templates expose a dedicated large-shot prompt rule."""
    try:
        columns = get_table_columns(engine, "shot_duration_templates")
        if not columns:
            return

        if "large_shot_prompt_rule" not in columns:
            with engine.begin() as conn:
                conn.execute(
                    text("ALTER TABLE shot_duration_templates ADD COLUMN large_shot_prompt_rule TEXT DEFAULT ''")
                )
            print("已添加 shot_duration_templates.large_shot_prompt_rule 字段")
    except Exception as e:
        print(f"补充 shot_duration_templates.large_shot_prompt_rule 字段失败: {str(e)}")
        return

    db = SessionLocal()
    try:
        templates = db.query(models.ShotDurationTemplate).all()
        updated_count = 0

        for template in templates:
            existing_rule = (getattr(template, "large_shot_prompt_rule", "") or "").strip()
            if existing_rule and not is_legacy_large_shot_prompt_rule(existing_rule):
                continue
            template.large_shot_prompt_rule = build_large_shot_prompt_rule(
                template.duration,
                template.time_segments
            )
            updated_count += 1

        if updated_count > 0:
            db.commit()
            print(f"已补齐大镜头提示词规则: {updated_count} 条")
    except Exception as e:
        db.rollback()
        print(f"补齐大镜头提示词规则失败: {str(e)}")
    finally:
        db.close()


def ensure_generate_large_shot_prompt_config():
    """Repair or seed the default large-shot prompt config."""
    db = SessionLocal()
    try:
        config = db.query(models.PromptConfig).filter(
            models.PromptConfig.key == "generate_large_shot_prompts"
        ).first()
        if not config:
            return

        default_content = build_large_shot_prompt_rule(15, 4)
        current_content = (config.content or "").strip()
        if current_content and not is_legacy_large_shot_prompt_rule(current_content):
            return

        config.content = default_content
        db.commit()
        print("已修复 generate_large_shot_prompts 默认模板")
    except Exception as e:
        db.rollback()
        print(f"修复 generate_large_shot_prompts 默认模板失败: {str(e)}")
    finally:
        db.close()


def ensure_large_shot_templates():
    """Seed the global large-shot template library without overwriting user edits."""
    try:
        if not table_exists(engine, "large_shot_templates"):
            return
    except Exception:
        return

    db = SessionLocal()
    try:
        seeded_templates = get_default_large_shot_templates()
        existing_templates = db.query(models.LargeShotTemplate).order_by(
            models.LargeShotTemplate.created_at.asc(),
            models.LargeShotTemplate.id.asc()
        ).all()
        removed_legacy_count = 0
        for template in existing_templates:
            if (template.name or "").strip() == "综合默认":
                db.delete(template)
                removed_legacy_count += 1

        if removed_legacy_count > 0:
            db.commit()
            existing_templates = db.query(models.LargeShotTemplate).order_by(
                models.LargeShotTemplate.created_at.asc(),
                models.LargeShotTemplate.id.asc()
            ).all()

        existing_by_name = {
            (template.name or "").strip(): template
            for template in existing_templates
            if (template.name or "").strip()
        }

        inserted_count = 0
        for template_data in seeded_templates:
            template_name = (template_data.get("name") or "").strip()
            if not template_name or template_name in existing_by_name:
                continue
            db.add(models.LargeShotTemplate(
                name=template_name,
                content=template_data.get("content") or "",
                is_default=False,
            ))
            inserted_count += 1

        if inserted_count > 0:
            db.commit()
            existing_templates = db.query(models.LargeShotTemplate).order_by(
                models.LargeShotTemplate.created_at.asc(),
                models.LargeShotTemplate.id.asc()
            ).all()

        default_template = next((template for template in existing_templates if template.is_default), None)
        if not default_template and existing_templates:
            preferred_name = ((seeded_templates[0] or {}).get("name") or "").strip() if seeded_templates else ""
            preferred = next(
                (
                    template
                    for template in existing_templates
                    if preferred_name and (template.name or "").strip() == preferred_name
                ),
                existing_templates[0]
            )
            db.query(models.LargeShotTemplate).update({"is_default": False})
            preferred.is_default = True
            db.commit()
            print("已设置默认大镜头模板")
        elif inserted_count > 0 or removed_legacy_count > 0:
            print(f"已同步大镜头模板库: 新增 {inserted_count} 条，移除 {removed_legacy_count} 条")
    except Exception as e:
        db.rollback()
        print(f"初始化大镜头模板库失败: {str(e)}")
    finally:
        db.close()




def ensure_shot_duration_template_subject_personality():
    """Upgrade shot duration video prompt rules to mention role personality context."""
    db = SessionLocal()
    try:
        templates = db.query(models.ShotDurationTemplate).all()
        updated_count = 0

        for template in templates:
            old_rule = template.video_prompt_rule or ""
            new_rule = _inject_subject_personality_hint(old_rule, for_storyboard2=False)
            if new_rule != old_rule:
                template.video_prompt_rule = new_rule
                updated_count += 1

        if updated_count > 0:
            db.commit()
            print(f"已升级时长配置模板 video_prompt_rule，新增角色性格说明: {updated_count} 条")
    except Exception as e:
        db.rollback()
        print(f"升级时长配置模板角色性格说明失败: {str(e)}")
    finally:
        db.close()




def ensure_remove_legacy_duration_templates():
    """Delete 6s and 10s shot duration templates which are no longer in use."""
    try:
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM shot_duration_templates WHERE duration IN (6, 10)"))
        print("已删除废弃的 6s / 10s 时长配置模板")
    except Exception as e:
        print(f"删除废弃时长配置模板失败: {str(e)}")


def ensure_video_prompt_subject_personality_configs():
    """Upgrade video prompt configs so manage UI exposes role personality context."""
    db = SessionLocal()
    try:
        updated_count = 0
        config_rules = {
            "generate_video_prompts": False,
            "generate_large_shot_prompts": False,
            STORYBOARD2_VIDEO_PROMPT_KEY: True,
        }

        for key, for_storyboard2 in config_rules.items():
            config = db.query(models.PromptConfig).filter(
                models.PromptConfig.key == key
            ).first()
            if not config:
                continue
            old_content = config.content or ""
            new_content = _inject_subject_personality_hint(old_content, for_storyboard2=for_storyboard2)
            if new_content != old_content:
                config.content = new_content
                updated_count += 1

        if updated_count > 0:
            db.commit()
            print(f"已升级视频提示词配置，新增角色性格说明: {updated_count} 条")
    except Exception as e:
        db.rollback()
        print(f"升级视频提示词配置角色性格说明失败: {str(e)}")
    finally:
        db.close()


_STYLE_TEMPLATE_REMOVAL_PHRASES = [
    "保持与该风格角色模板一致的整体画风与审美基调：",
    "人物高清画质",
    "人物冷白皮",
    "照片级写实肖像",
    "真实人类",
    "专业人像摄影",
    "专业人像",
    "纯白色背景",
    "纯白背景",
    "白底",
    "全身",
    "半身",
    "正视图",
    "正面角度",
    "正面",
    "面部特写",
    "面部",
    "人物",
    "角色",
    "人像",
    "肖像",
    "人类",
    "冷白皮",
    "伪bjd脸型",
    "bjd脸型",
    "立体五官",
    "五官",
    "皮肤纹理",
    "皮肤细节",
    "自然皮肤瑕疵",
    "皮肤",
    "可见毛孔",
    "站立",
    "站姿",
]


_STYLE_TEMPLATE_BANNED_SEGMENT_KEYWORDS = [
    "人物",
    "角色",
    "人像",
    "男人",
    "女人",
    "少年",
    "少女",
    "男生",
    "女生",
    "男",
    "女",
    "脸",
    "面部",
    "五官",
    "皮肤",
    "毛孔",
    "冷白皮",
    "发型",
    "发丝",
    "头发",
    "发量",
    "眼睛",
    "瞳孔",
    "嘴唇",
    "表情",
    "眼神",
    "凝视",
    "情感",
    "模特",
    "样貌",
    "站立",
    "站姿",
    "全身",
    "半身",
    "白底",
    "正视图",
    "正面",
    "穿搭",
    "妆容",
]

_STYLE_TEMPLATE_SKIP_EXACT_CHUNKS = {
    "核心风格",
    "建模技术",
    "质感",
    "服饰细节",
    "光影效果",
}

_STYLE_TEMPLATE_LABEL_PREFIXES = [
    "核心风格",
    "建模技术",
    "质感",
    "皮肤质感",
    "服饰细节",
    "光影效果",
]


def _extract_style_core_from_character_template(character_content: str) -> str:
    text_content = str(character_content or "").strip()
    if not text_content:
        return ""

    normalized = text_content
    for phrase in _STYLE_TEMPLATE_REMOVAL_PHRASES:
        normalized = normalized.replace(phrase, "")

    raw_parts = re.split(r"[\n\r,，。；;、]+", normalized)
    cleaned_parts: List[str] = []
    seen_parts = set()

    for raw_part in raw_parts:
        chunk = re.sub(r"^[\-\s]+", "", str(raw_part or "").strip())
        chunk = chunk.strip(" ：:，,。；;、")
        chunk = re.sub(r"[＋+]{2,}", "＋", chunk)
        chunk = re.sub(r"^(?:[＋+]\s*)+", "", chunk)
        chunk = re.sub(r"(?:\s*[＋+])+$", "", chunk)
        for label_prefix in _STYLE_TEMPLATE_LABEL_PREFIXES:
            chunk = re.sub(rf"^{re.escape(label_prefix)}\s*[：:]\s*", "", chunk)
        if not chunk:
            continue
        if chunk in _STYLE_TEMPLATE_SKIP_EXACT_CHUNKS:
            continue
        if any(keyword in chunk for keyword in _STYLE_TEMPLATE_BANNED_SEGMENT_KEYWORDS):
            continue
        normalized_key = re.sub(r"\s+", "", chunk)
        if not normalized_key or normalized_key in seen_parts:
            continue
        seen_parts.add(normalized_key)
        cleaned_parts.append(chunk)

    return "，".join(cleaned_parts[:16]).strip("，")


def _build_scene_style_template_content(character_content: str) -> str:
    style_core = _extract_style_core_from_character_template(character_content)
    if not style_core:
        return (
            "突出环境设计、空间层次、光影氛围、建筑与陈设细节、材质肌理与镜头感。"
            "不要出现人物，不要纯白背景，不要角色定妆式构图。"
        )
    return (
        f"{style_core}\n"
        "突出环境设计、空间层次、光影氛围、建筑与陈设细节、材质肌理与镜头感。"
        "不要出现人物，不要纯白背景，不要角色定妆式构图。"
    )


def _build_prop_style_template_content(character_content: str) -> str:
    style_core = _extract_style_core_from_character_template(character_content)
    if not style_core:
        return (
            "突出道具材质、结构造型、轮廓识别度、工艺细节、使用痕迹与局部特写。"
            "不要出现人物，不要纯白背景，不要角色站姿或人像式构图。"
        )
    return (
        f"{style_core}\n"
        "突出道具材质、结构造型、轮廓识别度、工艺细节、使用痕迹与局部特写。"
        "不要出现人物，不要纯白背景，不要角色站姿或人像式构图。"
    )


def _style_template_variant_needs_regeneration(content: str) -> bool:
    text_content = str(content or "").strip()
    if not text_content:
        return True
    if "??????????" in text_content or "?????" in text_content[:20]:
        return True
    if text_content.startswith("保持与该风格角色模板一致的整体画风与审美基调："):
        return True
    artifact_markers = [
        "＋＋",
        "核心风格，",
        "建模技术，",
        "发量丰盈",
        "质感：",
        "服饰细节：",
        "光影效果：",
        "真实情感",
        "写实凝视",
    ]
    return any(marker in text_content for marker in artifact_markers)


def ensure_style_template_variant_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "style_templates")
            if not columns:
                return

            if "scene_content" not in columns:
                conn.execute(
                    text("ALTER TABLE style_templates ADD COLUMN scene_content TEXT DEFAULT ''")
                )
                print("已添加 style_templates.scene_content 字段")
            if "prop_content" not in columns:
                conn.execute(
                    text("ALTER TABLE style_templates ADD COLUMN prop_content TEXT DEFAULT ''")
                )
                print("已添加 style_templates.prop_content 字段")

            conn.execute(text("UPDATE style_templates SET scene_content = '' WHERE scene_content IS NULL"))
            conn.execute(text("UPDATE style_templates SET prop_content = '' WHERE prop_content IS NULL"))

        db = SessionLocal()
        try:
            updated_count = 0
            templates = db.query(models.StyleTemplate).all()
            for template in templates:
                if _style_template_variant_needs_regeneration(template.scene_content):
                    template.scene_content = _build_scene_style_template_content(template.content)
                    updated_count += 1
                if _style_template_variant_needs_regeneration(template.prop_content):
                    template.prop_content = _build_prop_style_template_content(template.content)
                    updated_count += 1

            if updated_count > 0:
                db.commit()
                print(f"已补齐风格模板场景/道具版本: {updated_count} 处")
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()
    except Exception as e:
        print(f"检查/迁移 style_templates 场景/道具版本失败: {str(e)}")


def ensure_default_style_templates():
    """初始化默认的全局风格模板"""
    try:
        db = SessionLocal()
        try:
            default_style_name = "日漫风格"
            default_style_content = "日漫，天官赐福风格，一人之下风格，大理寺日志风格，百妖谱风格，仙王的日常生活风格，二次元画风，古风，平涂，赛璐璐上色，低饱和度"

            # 检查全局模板是否已存在
            existing = db.query(models.StyleTemplate).filter(
                models.StyleTemplate.name == default_style_name
            ).first()

            if not existing:
                # 创建全局模板
                new_template = models.StyleTemplate(
                    name=default_style_name,
                    content=default_style_content,
                    scene_content=_build_scene_style_template_content(default_style_content),
                    prop_content=_build_prop_style_template_content(default_style_content),
                )
                db.add(new_template)
                db.commit()
                print(f"✓ 创建全局风格模板: {default_style_name}")

        except Exception as e:
            db.rollback()
            print(f"初始化风格模板失败: {e}")
        finally:
            db.close()
    except Exception as e:
        print(f"初始化风格模板出错: {e}")



def ensure_user_password_column():
    """迁移：确保 users 表包含 password_hash/password_plain，并补齐默认值"""
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "users")
            if not columns:
                return
            if "password_hash" not in columns:
                conn.execute(
                    text("ALTER TABLE users ADD COLUMN password_hash VARCHAR NOT NULL DEFAULT ''")
                )
                print("已添加 users.password_hash 字段")
            if "password_plain" not in columns:
                conn.execute(
                    text("ALTER TABLE users ADD COLUMN password_plain VARCHAR NOT NULL DEFAULT '123456'")
                )
                print("已添加 users.password_plain 字段")

        # 为无密码的已有用户设置默认密码 123456，并补齐明文字段
        default_hash = _hash_password("123456")
        with engine.begin() as conn:
            conn.execute(
                text("UPDATE users SET password_hash = :h WHERE password_hash = '' OR password_hash IS NULL"),
                {"h": default_hash}
            )
            conn.execute(
                text("UPDATE users SET password_plain = '123456' WHERE password_plain = '' OR password_plain IS NULL")
            )
    except Exception as e:
        print(f"迁移 users.password 字段失败: {str(e)}")


def ensure_function_model_config_columns():
    try:
        with engine.begin() as conn:
            columns = get_table_columns(engine, "function_model_configs")
            if not columns:
                return

            if "provider_key" not in columns:
                conn.execute(
                    text("ALTER TABLE function_model_configs ADD COLUMN provider_key VARCHAR DEFAULT 'openrouter'")
                )
                print("已添加 function_model_configs.provider_key 字段")
                columns.add("provider_key")

            if "model_key" not in columns:
                conn.execute(
                    text("ALTER TABLE function_model_configs ADD COLUMN model_key VARCHAR")
                )
                print("已添加 function_model_configs.model_key 字段")
                columns.add("model_key")

            conn.execute(
                text(
                    "UPDATE function_model_configs "
                    "SET provider_key = 'openrouter' "
                    "WHERE provider_key IS NULL OR TRIM(provider_key) = ''"
                )
            )
            conn.execute(
                text(
                    "UPDATE function_model_configs "
                    "SET model_key = model_id "
                    "WHERE (model_key IS NULL OR TRIM(model_key) = '') "
                    "AND model_id IS NOT NULL AND TRIM(model_id) <> ''"
                )
            )
            conn.execute(
                text(
                    "UPDATE function_model_configs "
                    "SET provider_key = 'openrouter', "
                    "    model_key = 'google/gemini-3.1-pro-preview', "
                    "    model_id = 'google/gemini-3.1-pro-preview' "
                    "WHERE function_key = 'video_prompt' "
                    "  AND LOWER(COALESCE(provider_key, '')) IN ('', 'openrouter', 'yyds') "
                    "  AND COALESCE(NULLIF(TRIM(model_key), ''), '') IN ('', 'google/gemini-3.1-pro-preview', 'google/gemini-3-pro-preview', 'gemini_pro_preview', 'gemini_pro_high') "
                    "  AND COALESCE(NULLIF(TRIM(model_id), ''), '') IN ('', 'google/gemini-3.1-pro-preview', 'google/gemini-3-pro-preview', 'gemini-3.1-pro-preview', 'gemini-3.1-pro-high')"
                )
            )
            conn.execute(
                text(
                    "UPDATE function_model_configs "
                    "SET model_key = 'gemini_pro_high', "
                    "    model_id = 'gemini-3.1-pro-high' "
                    "WHERE LOWER(COALESCE(provider_key, '')) = 'yyds' "
                    "  AND ("
                    "        COALESCE(NULLIF(TRIM(model_key), ''), '') IN ('gemini_pro_preview', 'gemini_pro_high') "
                    "     OR COALESCE(NULLIF(TRIM(model_id), ''), '') IN ('gemini-3.1-pro-preview', 'gemini-3.1-pro-high')"
                    "  )"
                )
            )
    except Exception as e:
        print(f"迁移 function_model_configs provider/model_key 字段失败: {str(e)}")


HIT_DRAMA_ONLINE_TIME_PATTERN = re.compile(r"^(?P<year>\d{4})[./-](?P<month>\d{1,2})[./-](?P<day>\d{1,2})$")






def ensure_hit_drama_columns():
    try:
        with engine.begin() as conn:
            if not table_exists(engine, "hit_dramas"):
                return

            columns = get_table_columns(engine, "hit_dramas")
            if not columns:
                return

            if "organizer" in columns:
                try:
                    conn.execute(text("ALTER TABLE hit_dramas DROP COLUMN organizer"))
                    print("已删除 hit_dramas.organizer 字段")
                except Exception as drop_error:
                    print(f"删除 hit_dramas.organizer 失败，等待下次启动重试: {drop_error}")
    except Exception as e:
        print(f"检查/迁移 hit_dramas 失败: {str(e)}")


# 确保必要的目录存在
os.makedirs("uploads", exist_ok=True)
os.makedirs("videos", exist_ok=True)
os.makedirs("../frontend", exist_ok=True)

app = FastAPI(title="Story Creator API")

# CORS配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def track_request_load(request, call_next):
    path = request.url.path or ""
    if path.startswith("/static") or path.startswith("/uploads") or path.startswith("/videos"):
        return await call_next(request)

    request_load_tracker.request_started(path)
    try:
        return await call_next(request)
    finally:
        request_load_tracker.request_finished(path)

# 挂载静态文件目录
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")
app.mount("/videos", StaticFiles(directory="videos"), name="videos")
app.mount("/static", StaticFiles(directory="../frontend"), name="static")
app.include_router(pages.router)
app.include_router(media.router)
app.include_router(public.router)
app.include_router(image_generation.router)
app.include_router(video.router)
app.include_router(libraries.router)
app.include_router(subject_cards.router)
app.include_router(card_media.router)
app.include_router(dashboard.router)
app.include_router(model_configs.router)
app.include_router(admin_users.router)
app.include_router(auth.router)
app.include_router(billing.router)
app.include_router(settings.router)
app.include_router(templates.router)

# AI调试信息保存函数
def save_ai_debug(
    stage: str,
    input_data: dict,
    output_data: dict = None,
    raw_response: dict = None,
    episode_id: int = None,
    shot_id: int = None,
    batch_id: str = None,
    task_folder: str = None,
    attempt: int = None
):
    """
    记录 AI 调试事件到 dashboard task log，并返回稳定的任务分组 key。

    Args:
        stage: 阶段名 ('stage1', 'stage2', 'video_generate')
        input_data: 输入数据
        output_data: 输出数据（可选）
        raw_response: 原始响应或错误信息（可选）
        episode_id: 片段ID
        shot_id: 镜头ID（用于视频或调试）
        batch_id: 批次ID（用于stage1）
        task_folder: 任务分组 key（可选，自动生成）
    """
    try:
        from datetime import datetime

        if not episode_id and not shot_id:
            return None

        # 生成task_folder
        if not task_folder:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            task_folder = f"episode_{episode_id}_{timestamp}" if episode_id else f"shot_{shot_id}_{timestamp}"

        # 生成文件名前缀（供log_debug_task_event使用）
        if (stage == 'stage1' or stage == 'simple_storyboard') and batch_id:
            file_prefix = f"{stage}_batch{batch_id}"
            if attempt is not None:
                file_prefix += f"_attempt{attempt}"
        elif stage == 'stage2':
            file_prefix = f"stage2_shot{shot_id}" if shot_id else "stage2"
            if attempt is not None:
                file_prefix += f"_attempt{attempt}"
        elif stage == 'video_generate' and shot_id:
            file_prefix = f"video_shot{shot_id}"
        else:
            file_prefix = stage
            if attempt is not None:
                file_prefix += f"_attempt{attempt}"

        try:
            log_debug_task_event(
                stage=stage,
                task_folder=task_folder,
                input_data=input_data,
                output_data=output_data,
                raw_response=raw_response,
                episode_id=episode_id,
                shot_id=shot_id,
                batch_id=batch_id,
                file_name=f"{file_prefix}_event.json",
            )
        except Exception as e:
            print(f"[dashboard] save_ai_debug sync failed: {str(e)}")

        return task_folder  # 返回文件夹名，供后续调用使用

    except Exception as e:
        import traceback
        traceback.print_exc()
        return None


def _extract_text_relay_result_content(upstream_payload: Dict[str, Any]) -> str:
    result = upstream_payload.get("result")
    if not isinstance(result, dict):
        raise ValueError("relay result payload missing result object")
    choices = result.get("choices")
    if isinstance(choices, list) and choices:
        first = choices[0] if isinstance(choices[0], dict) else {}
        message = first.get("message")
        if isinstance(message, dict):
            content = message.get("content")
            if isinstance(content, str):
                return content
    raise ValueError("relay result payload missing assistant content")


def _parse_text_relay_result_json(upstream_payload: Dict[str, Any]) -> Dict[str, Any]:
    content = _extract_text_relay_result_content(upstream_payload)
    normalized = str(content or "").strip()
    if normalized.startswith("```json"):
        normalized = normalized.split("```json", 1)[1].split("```", 1)[0].strip()
    elif normalized.startswith("```"):
        normalized = normalized.split("```", 1)[1].split("```", 1)[0].strip()
    return json.loads(normalized or "{}")


def handle_text_relay_task_success(db: Session, task: models.TextRelayTask, upstream_payload: Dict[str, Any]):
    task_payload = {}
    try:
        task_payload = json.loads(str(getattr(task, "task_payload", "") or "{}"))
    except Exception:
        task_payload = {}

    task_type = str(getattr(task, "task_type", "") or "").strip()

    if task_type == "opening":
        episode_id = int(task_payload.get("episode_id") or task.owner_id or 0)
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if episode:
            episode.opening_content = _extract_text_relay_result_content(upstream_payload)
            episode.opening_generating = False
            episode.opening_error = ""
        return

    if task_type == "narration":
        episode_id = int(task_payload.get("episode_id") or task.owner_id or 0)
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if episode:
            episode.content = _extract_text_relay_result_content(upstream_payload)
            episode.narration_converting = False
            episode.narration_error = ""
        return

    if task_type == "subject_prompt":
        card_id = int(task_payload.get("card_id") or task.owner_id or 0)
        card = db.query(models.SubjectCard).filter(models.SubjectCard.id == card_id).first()
        if not card:
            return
        parsed = _parse_text_relay_result_json(upstream_payload)
        card.ai_prompt = str(parsed.get("ai_prompt") or "")
        alias = str(parsed.get("alias") or "").strip()
        if alias:
            card.alias = alias
        if hasattr(card, "ai_prompt_status"):
            card.ai_prompt_status = "completed"
        return

    if task_type == "simple_storyboard_batch":
        raise ValueError("简单分镜已迁移为本地程序生成，请重新发起整次简单分镜。")

    if task_type == "detailed_storyboard_stage1":
        episode_id = int(task_payload.get("episode_id") or task.owner_id or 0)
        parsed = _parse_text_relay_result_json(upstream_payload)
        detailed_shots = parsed.get("shots") or []
        if not detailed_shots:
            raise ValueError("详细分镜 Stage 1 未返回镜头数据")

        final_shots = []
        for shot in detailed_shots:
            shot["subjects"] = _normalize_storyboard_generation_subjects(shot.get("subjects", []))
            final_shots.append(shot)
        _submit_detailed_storyboard_stage2_task(db, episode_id=episode_id, final_shots=final_shots)
        return

    if task_type == "detailed_storyboard_stage2":
        episode_id = int(task_payload.get("episode_id") or task.owner_id or 0)
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if not episode:
            return
        final_shots = list(task_payload.get("final_shots") or [])
        parsed = _parse_text_relay_result_json(upstream_payload)
        stage2_subjects = _normalize_stage2_subjects(parsed.get("subjects", []))
        name_mappings = parsed.get("name_mappings", {}) or {}
        canonical_subject_map = _build_subject_detail_map(stage2_subjects)
        for shot in final_shots:
            shot["subjects"] = _reconcile_storyboard_shot_subjects(
                shot,
                canonical_subject_map,
                name_mappings=name_mappings,
            )

        final_data = {
            "shots": final_shots,
            "subjects": stage2_subjects,
        }
        voiceover_shots = []
        for shot in final_shots:
            voiceover_shots.append({
                "shot_number": shot.get("shot_number"),
                "voice_type": shot.get("voice_type"),
                "narration": shot.get("narration"),
                "dialogue": shot.get("dialogue"),
            })
        merged_voiceover_data = _merge_voiceover_shots_preserving_extensions(
            episode.voiceover_data,
            voiceover_shots
        )

        episode.storyboard_data = json.dumps(final_data, ensure_ascii=False)
        episode.voiceover_data = json.dumps(merged_voiceover_data, ensure_ascii=False)
        episode.storyboard_generating = False
        episode.storyboard_error = ""
        _create_shots_from_storyboard_data(episode_id, db)
        return

    if task_type == "sora_prompt":
        shot_id = int(task_payload.get("shot_id") or task.owner_id or 0)
        shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
        if not shot:
            return
        parsed = _parse_text_relay_result_json(upstream_payload)
        timeline = parsed.get("timeline", []) if isinstance(parsed, dict) else []
        shot.timeline_json = json.dumps(timeline, ensure_ascii=False)
        table_content = format_timeline_to_table(timeline)
        shot.storyboard_video_prompt = table_content
        shot.storyboard_audio_prompt = ""
        shot.storyboard_dialogue = ""
        if should_autofill_scene_override(
            current_scene_override=shot.scene_override,
            scene_override_locked=bool(getattr(shot, "scene_override_locked", False)),
        ):
            scene_desc = extract_scene_description(shot, db)
            if scene_desc:
                shot.scene_override = scene_desc
        shot.sora_prompt = table_content
        shot.sora_prompt_status = "completed"
        _refresh_episode_batch_sora_prompt_state(int(shot.episode_id), db)
        return

    if task_type == "storyboard2_sora_prompt":
        storyboard2_shot_id = int(task_payload.get("storyboard2_shot_id") or task.owner_id or 0)
        storyboard2_shot = db.query(models.Storyboard2Shot).filter(models.Storyboard2Shot.id == storyboard2_shot_id).first()
        if not storyboard2_shot:
            return
        parsed = _parse_text_relay_result_json(upstream_payload)
        timeline = parsed.get("timeline", []) if isinstance(parsed, dict) else []
        if not timeline:
            timeline = _storyboard2_fallback_timeline(int(task_payload.get("duration") or 10))
        _apply_storyboard2_timeline_prompts(storyboard2_shot, timeline, db)
        _refresh_storyboard2_prompt_batch_state(int(storyboard2_shot.episode_id), db)
        return

    if task_type == "managed_prompt_optimize":
        managed_task_id = int(task_payload.get("managed_task_id") or 0)
        managed_task = db.query(models.ManagedTask).filter(models.ManagedTask.id == managed_task_id).first()
        if not managed_task:
            return
        optimized_prompt = _extract_text_relay_result_content(upstream_payload).strip()
        managed_task.prompt_text = optimized_prompt
        managed_task.status = "pending"
        managed_task.error_message = ""

        reserved_shot_id = int(task_payload.get("reserved_shot_id") or 0)
        if reserved_shot_id > 0:
            reserved_shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == reserved_shot_id).first()
            if reserved_shot:
                reserved_shot.sora_prompt = optimized_prompt
                reserved_shot.sora_prompt_is_full = True
                reserved_shot.video_status = "processing"
                reserved_shot.video_error_message = ""
                reserved_shot.video_path = ""
                reserved_shot.thumbnail_video_path = ""
                reserved_shot.task_id = ""
        return


def handle_text_relay_task_failure(db: Session, task: models.TextRelayTask, upstream_payload: Dict[str, Any]):
    task_payload = {}
    try:
        task_payload = json.loads(str(getattr(task, "task_payload", "") or "{}"))
    except Exception:
        task_payload = {}

    error_message = str(
        upstream_payload.get("error")
        or upstream_payload.get("message")
        or getattr(task, "error_message", "")
        or "任务失败"
    ).strip()
    task_type = str(getattr(task, "task_type", "") or "").strip()

    if task_type == "opening":
        episode_id = int(task_payload.get("episode_id") or task.owner_id or 0)
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if episode:
            episode.opening_generating = False
            episode.opening_error = error_message
        return

    if task_type == "narration":
        episode_id = int(task_payload.get("episode_id") or task.owner_id or 0)
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if episode:
            episode.narration_converting = False
            episode.narration_error = error_message
        return

    if task_type == "subject_prompt":
        card_id = int(task_payload.get("card_id") or task.owner_id or 0)
        card = db.query(models.SubjectCard).filter(models.SubjectCard.id == card_id).first()
        if card and hasattr(card, "ai_prompt_status"):
            card.ai_prompt_status = "failed"
        return

    if task_type == "simple_storyboard_batch":
        batch_row_id = int(task_payload.get("batch_row_id") or task.owner_id or 0)
        batch_row = db.query(models.SimpleStoryboardBatch).filter(models.SimpleStoryboardBatch.id == batch_row_id).first()
        if not batch_row:
            return
        batch_row.status = "failed"
        batch_row.error_message = "简单分镜已迁移为本地程序生成，请重新发起整次简单分镜。"
        batch_row.last_attempt = 1
        batch_row.updated_at = datetime.utcnow()
        episode = db.query(models.Episode).filter(models.Episode.id == batch_row.episode_id).first()
        if episode:
            _refresh_episode_simple_storyboard_from_batches(episode, db)
        return

    if task_type in {"detailed_storyboard_stage1", "detailed_storyboard_stage2"}:
        episode_id = int(task_payload.get("episode_id") or task.owner_id or 0)
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if episode:
            episode.storyboard_generating = False
            episode.storyboard_error = error_message
        return

    if task_type == "sora_prompt":
        shot_id = int(task_payload.get("shot_id") or task.owner_id or 0)
        shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
        if shot:
            shot.sora_prompt_status = "failed"
            _refresh_episode_batch_sora_prompt_state(int(shot.episode_id), db)
        return

    if task_type == "storyboard2_sora_prompt":
        episode_id = int(task_payload.get("episode_id") or 0)
        if episode_id > 0:
            _refresh_storyboard2_prompt_batch_state(episode_id, db)
        return

    if task_type == "managed_prompt_optimize":
        managed_task_id = int(task_payload.get("managed_task_id") or 0)
        managed_task = db.query(models.ManagedTask).filter(models.ManagedTask.id == managed_task_id).first()
        if managed_task:
            managed_task.status = "failed"
            managed_task.error_message = error_message
            managed_task.completed_at = datetime.utcnow()
        reserved_shot_id = int(task_payload.get("reserved_shot_id") or 0)
        if reserved_shot_id > 0:
            reserved_shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == reserved_shot_id).first()
            if reserved_shot:
                reserved_shot.video_status = "failed"
                reserved_shot.video_error_message = error_message
                reserved_shot.video_path = f"error:{error_message}" if error_message else ""
                reserved_shot.thumbnail_video_path = ""
                reserved_shot.task_id = ""
        return



extract_scene_description = storyboard_video_prompt_builder.extract_scene_description
_default_storyboard_video_prompt_template = storyboard_video_prompt_builder.default_storyboard_video_prompt_template
build_sora_prompt = storyboard_video_prompt_builder.build_sora_prompt

def format_timeline_to_table(timeline: list) -> str:
    """
    将timeline数组格式化为Markdown表格字符串

    Args:
        timeline: AI返回的timeline数组 [{"time": "00s-04s", "visual": "...", "audio": "..."}, ...]

    Returns:
        str: 格式化的Markdown表格
    """
    if not timeline or not isinstance(timeline, list):
        return ""

    # 构建表格（不含头部）
    lines = []

    for item in timeline:
        time = item.get("time", "")
        visual = item.get("visual", "")
        audio = item.get("audio", "")

        lines.append(f"| {time} | {visual} | {audio} |")

    return "\n".join(lines)

VOICEOVER_TTS_ACTIVE_POLL_INTERVAL_SECONDS = 1.5
VOICEOVER_TTS_IDLE_POLL_INTERVAL_SECONDS = 3.0
VOICEOVER_TTS_BUSY_IDLE_POLL_INTERVAL_SECONDS = 5.0


class VoiceoverTtsQueuePoller:
    """配音TTS全局串行队列轮询器（全用户统一排队）。"""

    def __init__(self):
        self.running = False
        self.thread = None

    def start(self):
        if self.running:
            return
        self.running = True
        self.thread = Thread(target=self._poll_loop, daemon=True)
        self.thread.start()
        print("[voiceover_tts] serial queue poller started")

    def stop(self):
        self.running = False
        if self.thread:
            self.thread.join(timeout=5)
        print("[voiceover_tts] serial queue poller stopped")

    def _poll_loop(self):
        while self.running:
            processed_task = False
            try:
                processed_task = self._process_once()
            except Exception as e:
                print(f"[voiceover_tts] poll loop error: {str(e)}")
            if processed_task:
                time.sleep(VOICEOVER_TTS_ACTIVE_POLL_INTERVAL_SECONDS)
            else:
                time.sleep(
                    request_load_tracker.choose_interval(
                        VOICEOVER_TTS_IDLE_POLL_INTERVAL_SECONDS,
                        VOICEOVER_TTS_BUSY_IDLE_POLL_INTERVAL_SECONDS,
                    )
                )

    def _claim_next_task(self) -> Optional[int]:
        db = SessionLocal()
        try:
            task = db.query(models.VoiceoverTtsTask).filter(
                models.VoiceoverTtsTask.status == "pending"
            ).order_by(
                models.VoiceoverTtsTask.created_at.asc(),
                models.VoiceoverTtsTask.id.asc()
            ).first()
            if not task:
                return None
            task.status = "processing"
            task.started_at = datetime.utcnow()
            db.commit()
            sync_voiceover_tts_task_to_dashboard(task.id)
            return int(task.id)
        finally:
            db.close()

    def _mark_failed(self, task_id: int, error_message: str):
        db = SessionLocal()
        try:
            task = db.query(models.VoiceoverTtsTask).filter(
                models.VoiceoverTtsTask.id == task_id
            ).first()
            if not task:
                return
            task.status = "failed"
            task.error_message = str(error_message or "未知错误")
            task.completed_at = datetime.utcnow()

            episode = db.query(models.Episode).filter(
                models.Episode.id == task.episode_id
            ).first()
            if episode:
                script = db.query(models.Script).filter(
                    models.Script.id == episode.script_id
                ).first()
                shared = _load_script_voiceover_shared_data(script) if script else _voiceover_default_shared_data()
                default_ref_id = ""
                refs = shared.get("voice_references", [])
                if isinstance(refs, list) and refs:
                    default_ref_id = str(refs[0].get("id") or "").strip()

                payload = _parse_episode_voiceover_payload(episode)
                shots, _ = _normalize_voiceover_shots_for_tts(payload.get("shots", []), default_ref_id)
                line_entry = _find_voiceover_line_entry(shots, task.line_id)
                if isinstance(line_entry, dict):
                    line_tts = _normalize_voiceover_line_tts(
                        line_entry.get("tts"),
                        default_ref_id
                    )
                    line_tts["generate_status"] = "failed"
                    line_tts["generate_error"] = task.error_message
                    line_tts["latest_task_id"] = str(task.id)
                    line_entry["tts"] = line_tts
                    payload["shots"] = shots
                    episode.voiceover_data = json.dumps(payload, ensure_ascii=False)

            db.commit()
            sync_voiceover_tts_task_to_dashboard(task.id)
        finally:
            db.close()

    def _process_once(self):
        task_id = self._claim_next_task()
        if task_id is None:
            return False
        self._execute_task(task_id)
        return True

    def _execute_task(self, task_id: int):
        debug_folder = f"voiceover_tts_task_{task_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
        runtime_context = {}
        db = SessionLocal()
        try:
            task = db.query(models.VoiceoverTtsTask).filter(
                models.VoiceoverTtsTask.id == task_id
            ).first()
            if not task:
                return
            if task.status != "processing":
                return

            try:
                request_payload = json.loads(task.request_json or "{}")
                if not isinstance(request_payload, dict):
                    request_payload = {}
            except Exception:
                request_payload = {}

            episode = db.query(models.Episode).filter(
                models.Episode.id == task.episode_id
            ).first()
            if not episode:
                raise Exception("片段不存在")

            script = db.query(models.Script).filter(
                models.Script.id == episode.script_id
            ).first()
            if not script:
                raise Exception("剧本不存在")

            shared = _load_script_voiceover_shared_data(script)
            refs = shared.get("voice_references", [])
            default_ref_id = str(refs[0].get("id") or "").strip() if isinstance(refs, list) and refs else ""

            voiceover_payload = _parse_episode_voiceover_payload(episode)
            shots, changed = _normalize_voiceover_shots_for_tts(
                voiceover_payload.get("shots", []),
                default_ref_id
            )
            if changed:
                voiceover_payload["shots"] = shots

            line_entry = _find_voiceover_line_entry(shots, task.line_id)
            if not isinstance(line_entry, dict):
                raise Exception(f"未找到line_id={task.line_id} 对应的配音行")

            line_tts = _normalize_voiceover_line_tts(
                line_entry.get("tts"),
                default_ref_id
            )
            line_tts["generate_status"] = "processing"
            line_tts["generate_error"] = ""
            line_tts["latest_task_id"] = str(task.id)
            line_entry["tts"] = line_tts

            voiceover_payload["shots"] = shots
            episode.voiceover_data = json.dumps(voiceover_payload, ensure_ascii=False)
            db.commit()

            line_text = str(
                request_payload.get("text")
                or line_entry.get("text")
                or ""
            ).strip()
            if not line_text:
                raise Exception("配音文本为空")

            method = str(
                request_payload.get("emotion_control_method")
                or line_tts.get("emotion_control_method")
                or VOICEOVER_TTS_METHOD_SAME
            ).strip()
            if method not in VOICEOVER_TTS_ALLOWED_METHODS:
                method = VOICEOVER_TTS_METHOD_SAME

            voice_ref_id = str(
                request_payload.get("voice_reference_id")
                or line_tts.get("voice_reference_id")
                or ""
            ).strip()
            if not voice_ref_id and isinstance(refs, list) and refs:
                voice_ref_id = str(refs[0].get("id") or "").strip()

            selected_voice_ref = None
            if isinstance(refs, list):
                selected_voice_ref = next((x for x in refs if str(x.get("id") or "") == voice_ref_id), None)
            if not selected_voice_ref and isinstance(refs, list) and refs:
                selected_voice_ref = refs[0]

            prompt_audio_source = _resolve_voiceover_audio_source(selected_voice_ref or {})
            if not prompt_audio_source:
                raise Exception("未找到有效的音色参考音频")

            emo_ref_audio_source = None
            if method == VOICEOVER_TTS_METHOD_AUDIO:
                emotion_audio_preset_id = str(
                    request_payload.get("emotion_audio_preset_id")
                    or line_tts.get("emotion_audio_preset_id")
                    or ""
                ).strip()
                emotion_presets = shared.get("emotion_audio_presets", [])
                selected_preset = None
                if isinstance(emotion_presets, list):
                    selected_preset = next(
                        (x for x in emotion_presets if str(x.get("id") or "") == emotion_audio_preset_id),
                        None
                    )
                emo_ref_audio_source = _resolve_voiceover_audio_source(selected_preset or {})
                if not emo_ref_audio_source:
                    raise Exception("情感参考音频未设置")

            vector_config = _normalize_voiceover_vector_config(
                request_payload.get("vector_config") or line_tts.get("vector_config")
            )
            emo_text = str(
                request_payload.get("emo_text")
                if request_payload.get("emo_text") is not None
                else line_entry.get("emotion")
                or ""
            ).strip()
            runtime_context = {
                "line_id": task.line_id,
                "text": line_text,
                "method": method,
                "api_url": VOICEOVER_TTS_API_URL,
                "voice_reference_id": voice_ref_id,
                "prompt_audio_source": prompt_audio_source,
                "emo_ref_audio_source": emo_ref_audio_source,
                "vector_config": vector_config,
                "emo_text": emo_text
            }
        except Exception as e:
            db.rollback()
            self._mark_failed(task_id, str(e))
            return
        finally:
            db.close()

        _save_voiceover_tts_debug(debug_folder, "input.json", runtime_context)

        try:
            result = _generate_tts_with_index_tts(
                text_content=runtime_context["text"],
                emotion_control_method=runtime_context["method"],
                prompt_audio_source=runtime_context["prompt_audio_source"],
                emo_ref_audio_source=runtime_context.get("emo_ref_audio_source"),
                vector_config=runtime_context.get("vector_config", {}),
                emo_text=runtime_context.get("emo_text", "")
            )
            _save_voiceover_tts_debug(debug_folder, "output.json", result)
        except Exception as e:
            _save_voiceover_tts_debug(debug_folder, "error.json", {"error": str(e)})
            self._mark_failed(task_id, str(e))
            return

        db = SessionLocal()
        try:
            task = db.query(models.VoiceoverTtsTask).filter(
                models.VoiceoverTtsTask.id == task_id
            ).first()
            if not task:
                return
            episode = db.query(models.Episode).filter(
                models.Episode.id == task.episode_id
            ).first()
            if not episode:
                raise Exception("片段不存在")
            script = db.query(models.Script).filter(
                models.Script.id == episode.script_id
            ).first()
            shared = _load_script_voiceover_shared_data(script) if script else _voiceover_default_shared_data()
            refs = shared.get("voice_references", [])
            default_ref_id = str(refs[0].get("id") or "").strip() if isinstance(refs, list) and refs else ""

            voiceover_payload = _parse_episode_voiceover_payload(episode)
            shots, _ = _normalize_voiceover_shots_for_tts(
                voiceover_payload.get("shots", []),
                default_ref_id
            )
            line_entry = _find_voiceover_line_entry(shots, task.line_id)
            if not isinstance(line_entry, dict):
                raise Exception(f"未找到line_id={task.line_id}")

            line_tts = _normalize_voiceover_line_tts(
                line_entry.get("tts"),
                default_ref_id
            )
            generated = line_tts.get("generated_audios", [])
            if not isinstance(generated, list):
                generated = []
            generated.insert(0, {
                "id": f"tts_result_{uuid.uuid4().hex}",
                "name": f"生成结果 {len(generated) + 1}",
                "url": str(result.get("cdn_url") or "").strip(),
                "task_id": str(task.id),
                "created_at": datetime.utcnow().isoformat(),
                "status": "completed"
            })
            line_tts["generated_audios"] = generated
            line_tts["generate_status"] = "idle"
            line_tts["generate_error"] = ""
            line_tts["latest_task_id"] = str(task.id)
            line_entry["tts"] = line_tts
            voiceover_payload["shots"] = shots
            episode.voiceover_data = json.dumps(voiceover_payload, ensure_ascii=False)

            task.status = "completed"
            task.result_json = json.dumps(result, ensure_ascii=False)
            task.error_message = ""
            task.completed_at = datetime.utcnow()
            db.commit()
            sync_voiceover_tts_task_to_dashboard(task.id)
        except Exception as e:
            db.rollback()
            self._mark_failed(task_id, str(e))
        finally:
            db.close()

voiceover_tts_poller = VoiceoverTtsQueuePoller()

# 启动事件：启动视频状态轮询器
@app.on_event("startup")
async def startup_event():
    start_external_cache_prewarms()
    start_background_pollers()

# 关闭事件：停止轮询器
@app.on_event("shutdown")
async def shutdown_event():
    stop_background_pollers()

# ==================== Pydantic模型 ====================

class LoginRequest(BaseModel):
    username: str
    password: str

class PasswordVerifyRequest(BaseModel):
    password: str

# ==================== 工具函数 ====================

_voiceover_shot_match_key = voiceover_data.voiceover_shot_match_key
_merge_voiceover_line_preserving_tts = voiceover_data.merge_voiceover_line_preserving_tts
_merge_voiceover_dialogue_preserving_tts = voiceover_data.merge_voiceover_dialogue_preserving_tts
_merge_voiceover_shots_preserving_extensions = voiceover_data.merge_voiceover_shots_preserving_extensions

VOICEOVER_TTS_API_URL = get_env("VOICEOVER_TTS_API_URL", "")
VOICEOVER_TTS_METHOD_SAME = voiceover_data.VOICEOVER_TTS_METHOD_SAME
VOICEOVER_TTS_METHOD_VECTOR = voiceover_data.VOICEOVER_TTS_METHOD_VECTOR
VOICEOVER_TTS_METHOD_EMO_TEXT = voiceover_data.VOICEOVER_TTS_METHOD_EMO_TEXT
VOICEOVER_TTS_METHOD_AUDIO = voiceover_data.VOICEOVER_TTS_METHOD_AUDIO
VOICEOVER_TTS_ALLOWED_METHODS = voiceover_data.VOICEOVER_TTS_ALLOWED_METHODS
VOICEOVER_TTS_VECTOR_KEYS = voiceover_data.VOICEOVER_TTS_VECTOR_KEYS

_voiceover_default_test_mp3_path = partial(voiceover_data.voiceover_default_test_mp3_path, __file__)

_voiceover_default_vector_config = voiceover_data.voiceover_default_vector_config

_voiceover_default_shared_data = voiceover_data.voiceover_default_shared_data
_voiceover_default_reference_item = partial(
    voiceover_data.voiceover_default_reference_item,
    _voiceover_default_test_mp3_path,
)

_safe_float = voiceover_data.safe_float
_normalize_voiceover_vector_config = voiceover_data.normalize_voiceover_vector_config
_normalize_voiceover_setting_template_payload = voiceover_data.normalize_voiceover_setting_template_payload

_normalize_voiceover_shared_data = partial(
    voiceover_data.normalize_voiceover_shared_data,
    default_reference_item_factory=_voiceover_default_reference_item,
)
_load_script_voiceover_shared_data = partial(
    voiceover_data.load_script_voiceover_shared_data,
    normalize_shared_data=_normalize_voiceover_shared_data,
)
_save_script_voiceover_shared_data = partial(
    voiceover_data.save_script_voiceover_shared_data,
    normalize_shared_data=_normalize_voiceover_shared_data,
)

_voiceover_default_line_tts = voiceover_data.voiceover_default_line_tts
_normalize_voiceover_line_tts = voiceover_data.normalize_voiceover_line_tts
_ensure_voiceover_shot_line_fields = voiceover_data.ensure_voiceover_shot_line_fields
_normalize_voiceover_shots_for_tts = voiceover_data.normalize_voiceover_shots_for_tts
_extract_voiceover_tts_line_states = voiceover_data.extract_voiceover_tts_line_states
_find_voiceover_line_entry = voiceover_data.find_voiceover_line_entry
_parse_episode_voiceover_payload = voiceover_data.parse_episode_voiceover_payload
_voiceover_first_reference_id = voiceover_data.voiceover_first_reference_id
_iter_voiceover_lines = voiceover_data.iter_voiceover_lines

def _ensure_voiceover_permission(
    episode_id: int,
    user: models.User,
    db: Session
) -> Tuple[models.Episode, models.Script]:
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if not script or script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    return episode, script

def _replace_voice_reference_for_script_episodes(
    db: Session,
    script_id: int,
    removed_ref_id: str,
    fallback_ref_id: str
) -> int:
    """删除音色引用后，回填所有剧集里对应行的音色ID。"""
    removed = str(removed_ref_id or "").strip()
    if not removed:
        return 0

    updated_lines = 0
    episodes = db.query(models.Episode).filter(models.Episode.script_id == script_id).all()
    for episode in episodes:
        payload = _parse_episode_voiceover_payload(episode)
        shots, changed = _normalize_voiceover_shots_for_tts(payload.get("shots", []), fallback_ref_id)
        episode_changed = bool(changed)
        for line in _iter_voiceover_lines(shots):
            tts = _normalize_voiceover_line_tts(line.get("tts"), fallback_ref_id)
            if tts.get("voice_reference_id") == removed:
                tts["voice_reference_id"] = fallback_ref_id or ""
                line["tts"] = tts
                updated_lines += 1
                episode_changed = True
        if episode_changed:
            payload["shots"] = shots
            episode.voiceover_data = json.dumps(payload, ensure_ascii=False)
    return updated_lines

def _clear_tts_field_for_script_episodes(
    db: Session,
    script_id: int,
    field_name: str,
    removed_value: str
) -> int:
    """清理所有剧集中被删除的预设ID引用。"""
    target = str(removed_value or "").strip()
    if not target:
        return 0

    updated_lines = 0
    episodes = db.query(models.Episode).filter(models.Episode.script_id == script_id).all()
    for episode in episodes:
        payload = _parse_episode_voiceover_payload(episode)
        shots, changed = _normalize_voiceover_shots_for_tts(payload.get("shots", []), "")
        episode_changed = bool(changed)
        for line in _iter_voiceover_lines(shots):
            tts = _normalize_voiceover_line_tts(line.get("tts"), "")
            if str(tts.get(field_name) or "").strip() == target:
                tts[field_name] = ""
                line["tts"] = tts
                updated_lines += 1
                episode_changed = True
        if episode_changed:
            payload["shots"] = shots
            episode.voiceover_data = json.dumps(payload, ensure_ascii=False)
    return updated_lines

def _save_voiceover_tts_debug(folder_name: str, file_name: str, payload: dict):
    try:
        debug_dir = os.path.join("ai_debug", folder_name)
        os.makedirs(debug_dir, exist_ok=True)
        file_path = os.path.join(debug_dir, file_name)
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        log_file_task_event(
            task_folder=folder_name,
            file_name=file_name,
            payload=payload,
            task_type="voiceover_tts",
            stage="voiceover_tts",
            source_record_type="voiceover_tts_task",
            source_record_id=int(re.search(r"voiceover_tts_task_(\d+)", folder_name).group(1)) if re.search(r"voiceover_tts_task_(\d+)", folder_name) else None,
        )
    except Exception as e:
        print(f"[voiceover_tts][debug] save failed: {str(e)}")

def _resolve_voiceover_audio_source(reference_item: dict) -> str:
    if not isinstance(reference_item, dict):
        return ""
    url = str(reference_item.get("url") or "").strip()
    if url:
        return url
    local_path = str(reference_item.get("local_path") or "").strip()
    if local_path:
        if os.path.isabs(local_path):
            return local_path
        return os.path.abspath(os.path.join(os.path.dirname(__file__), "..", local_path))
    return ""

def _upload_local_or_remote_audio_to_cdn(audio_source: str) -> str:
    source = str(audio_source or "").strip()
    if not source:
        raise Exception("音频结果为空")

    if source.startswith("http://") or source.startswith("https://"):
        temp_path = None
        try:
            response = requests.get(source, timeout=60)
            if response.status_code != 200:
                raise Exception(f"下载远程音频失败: HTTP {response.status_code}")
            ext = os.path.splitext(source.split("?", 1)[0])[1] or ".wav"
            temp_path = os.path.join("uploads", f"tts_temp_{uuid.uuid4().hex}{ext}")
            with open(temp_path, "wb") as f:
                f.write(response.content)
            return upload_to_cdn(temp_path)
        finally:
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception:
                    pass

    if not os.path.isabs(source):
        source = os.path.abspath(source)
    if not os.path.exists(source):
        raise Exception(f"音频文件不存在: {source}")
    return upload_to_cdn(source)

def _generate_tts_with_index_tts(
    text_content: str,
    emotion_control_method: str,
    prompt_audio_source: str,
    emo_ref_audio_source: Optional[str],
    vector_config: dict,
    emo_text: str = ""
) -> dict:
    from gradio_client import Client, handle_file

    client = Client(VOICEOVER_TTS_API_URL)
    if emotion_control_method == VOICEOVER_TTS_METHOD_EMO_TEXT:
        try:
            client.predict(
                is_experimental=True,
                api_name="/on_experimental_change"
            )
        except Exception as experimental_error:
            print(f"[voiceover_tts] experimental mode toggle failed: {experimental_error}")
    prompt_audio = handle_file(prompt_audio_source)
    emo_ref_audio = handle_file(emo_ref_audio_source) if emo_ref_audio_source else None

    vec = _normalize_voiceover_vector_config(vector_config)
    result = client.predict(
        emo_control_method=emotion_control_method,
        prompt=prompt_audio,
        text=text_content,
        emo_ref_path=emo_ref_audio,
        emo_weight=vec["weight"],
        vec1=vec["joy"],
        vec2=vec["anger"],
        vec3=vec["sadness"],
        vec4=vec["fear"],
        vec5=vec["disgust"],
        vec6=vec["depression"],
        vec7=vec["surprise"],
        vec8=vec["neutral"],
        emo_text=str(emo_text or ""),
        emo_random=False,
        max_text_tokens_per_segment=120,
        param_16=True,
        param_17=0.8,
        param_18=30,
        param_19=0.8,
        param_20=0.0,
        param_21=3,
        param_22=10.0,
        param_23=1500,
        api_name="/gen_single"
    )

    audio_value = ""
    if isinstance(result, dict):
        audio_value = str(
            result.get("value") or result.get("path") or result.get("url") or ""
        ).strip()
    elif isinstance(result, str):
        audio_value = result.strip()

    if not audio_value:
        raise Exception("TTS接口返回音频为空")

    cdn_url = _upload_local_or_remote_audio_to_cdn(audio_value)
    return {
        "raw_result": result,
        "audio_source": audio_value,
        "cdn_url": cdn_url
    }

def generate_collage_image(shot_id: int, db: Session, include_scenes: bool = False, aspect_ratio: str = "16:9") -> str:
    """
    为指定镜头生成拼图，根据宽高比自动选择横排或竖排布局

    Args:
        shot_id: 镜头ID
        db: 数据库会话
        include_scenes: 是否包含场景类主体
        aspect_ratio: 视频宽高比 (如 "16:9", "9:16", "1:1")

    Returns:
        str: 拼图的CDN URL

    Raises:
        Exception: 如果生成失败
    """
    import io
    import urllib.request

    # 获取镜头信息
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise Exception("镜头不存在")

    # 解析选中的主体ID
    try:
        selected_ids = json.loads(shot.selected_card_ids or "[]")
    except Exception:
        selected_ids = []

    if not selected_ids:
        raise Exception("未选择任何主体")

    # 获取选中的主体卡片
    cards = db.query(models.SubjectCard).filter(
        models.SubjectCard.id.in_(selected_ids)
    ).all()

    if not cards:
        raise Exception("未找到主体卡片")

    # 收集每个主体的参考图
    subject_images = []

    print(f"[拼图生成] 开始收集主体图片，include_scenes={include_scenes}, aspect_ratio={aspect_ratio}")
    print(f"[拼图生成] 共有 {len(cards)} 个主体")

    for card in cards:
        print(f"[拼图生成] 处理主体: {card.name}, 类型: {card.card_type}")

        if not include_scenes and card.card_type == "场景":
            print(f"[拼图生成] 跳过场景类主体: {card.name} (include_scenes=False)")
            continue

        ref_images = db.query(models.GeneratedImage).filter(
            models.GeneratedImage.card_id == card.id,
            models.GeneratedImage.is_reference == True,
            models.GeneratedImage.status == "completed"
        ).all()

        print(f"[拼图生成] 主体 {card.name} 有 {len(ref_images)} 张参考图")

        image_urls = [img.image_path for img in ref_images] if ref_images else []
        subject_images.append({
            "name": card.name,
            "urls": image_urls,
            "has_images": len(image_urls) > 0,
            "card_type": card.card_type
        })

    # 下载所有图片
    print(f"[拼图生成] 开始处理 {len(subject_images)} 个主体...")
    downloaded_images = []

    for subject in subject_images:
        pil_images = []

        if subject["has_images"]:
            for url in subject["urls"]:
                try:
                    with urllib.request.urlopen(url, timeout=30) as response:
                        img_data = response.read()
                        pil_img = Image.open(io.BytesIO(img_data))
                        pil_images.append(pil_img)
                except Exception as e:
                    print(f"[拼图生成] 下载图片失败: {url} - {str(e)}")
                    continue

        if not pil_images:
            print(f"[拼图生成] 跳过没有图片的主体: {subject['name']}")
            continue

        downloaded_images.append({
            "name": subject["name"],
            "images": pil_images,
            "has_images": len(pil_images) > 0,
            "card_type": subject["card_type"]
        })

    if not downloaded_images:
        raise Exception("所有主体都没有参考图，无法生成拼图。请先为至少一个主体生成参考图。")

    # ==================== 根据宽高比计算画布尺寸 ====================
    try:
        w_ratio, h_ratio = map(int, aspect_ratio.split(':'))
    except Exception:
        w_ratio, h_ratio = 16, 9

    is_landscape = w_ratio >= h_ratio
    LONG_EDGE = 1920

    if is_landscape:
        CANVAS_WIDTH = LONG_EDGE
        CANVAS_HEIGHT = int(LONG_EDGE * h_ratio / w_ratio)
    else:
        CANVAS_HEIGHT = LONG_EDGE
        CANVAS_WIDTH = int(LONG_EDGE * w_ratio / h_ratio)

    PADDING = 10

    print(f"[拼图生成] 画布尺寸: {CANVAS_WIDTH}x{CANVAS_HEIGHT}, 布局: {'横排' if is_landscape else '竖排'}")

    # 创建画布
    canvas = Image.new("RGB", (CANVAS_WIDTH, CANVAS_HEIGHT), color="white")
    draw = ImageDraw.Draw(canvas)

    # 确定字体路径
    try:
        font_path = "C:\\Windows\\Fonts\\msyh.ttc"
        _test_font = ImageFont.truetype(font_path, 20)
    except Exception:
        try:
            font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
            _test_font = ImageFont.truetype(font_path, 20)
        except Exception:
            font_path = None

    def find_optimal_font_size(text, max_width, min_size=20, max_size=80):
        if font_path is None:
            return ImageFont.load_default()
        left, right = min_size, max_size
        best_size = min_size
        while left <= right:
            mid = (left + right) // 2
            try:
                test_font = ImageFont.truetype(font_path, mid)
                bbox = draw.textbbox((0, 0), text, font=test_font)
                text_width = bbox[2] - bbox[0]
                if text_width <= max_width:
                    best_size = mid
                    left = mid + 1
                else:
                    right = mid - 1
            except Exception:
                right = mid - 1
        try:
            return ImageFont.truetype(font_path, best_size)
        except Exception:
            return ImageFont.load_default()

    def get_unified_font(subject_name, type_label, max_text_width, min_size=20, max_size=80):
        """Get a unified font size for name and type label."""
        font_name = find_optimal_font_size(subject_name, max_text_width, min_size, max_size)
        font_type = find_optimal_font_size(type_label, max_text_width, min_size, max_size)
        try:
            size_name = font_name.size if hasattr(font_name, 'size') else max_size
            size_type = font_type.size if hasattr(font_type, 'size') else max_size
            unified_size = min(size_name, size_type)
            return ImageFont.truetype(font_path, unified_size) if font_path else ImageFont.load_default()
        except Exception:
            return font_name

    # ==================== 横排布局（宽 >= 高）====================
    if is_landscape:
        IMAGE_HEIGHT = int(CANVAS_HEIGHT * 0.6)
        LABEL_HEIGHT = CANVAS_HEIGHT - IMAGE_HEIGHT

        total_images = sum(len(s["images"]) for s in downloaded_images)
        available_width = CANVAS_WIDTH - (total_images + 1) * PADDING
        image_width = available_width // total_images

        current_x = PADDING

        for subject in downloaded_images:
            subject_name = subject["name"]
            images = subject["images"]
            subject_width = len(images) * image_width + (len(images) - 1) * PADDING

            # 绘制图片
            for img in images:
                img_ratio = img.width / img.height
                new_height = IMAGE_HEIGHT
                new_width = int(new_height * img_ratio)
                if new_width > image_width:
                    new_width = image_width
                    new_height = int(new_width / img_ratio)

                resized_img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                y_offset = (IMAGE_HEIGHT - new_height) // 2
                canvas.paste(resized_img, (current_x, y_offset))
                current_x += image_width + PADDING

            # 绘制标签
            card_type = subject["card_type"]
            type_label = f"{card_type}图片"
            max_text_width = int(subject_width * 0.9)

            unified_font = get_unified_font(subject_name, type_label, max_text_width)

            text_center_x = current_x - subject_width // 2 - PADDING

            try:
                font_height = unified_font.size if hasattr(unified_font, 'size') else 40
            except Exception:
                font_height = 40

            total_text_height = int(font_height * 2.2)
            text_vertical_start = IMAGE_HEIGHT + (LABEL_HEIGHT - total_text_height) // 2

            try:
                bbox1 = draw.textbbox((0, 0), subject_name, font=unified_font)
                text_width1 = bbox1[2] - bbox1[0]
            except Exception:
                text_width1 = len(subject_name) * 20

            text1_x = max(PADDING, text_center_x - text_width1 // 2)
            text1_y = text_vertical_start
            draw.text((text1_x, text1_y), subject_name, fill="black", font=unified_font)

            try:
                bbox2 = draw.textbbox((0, 0), type_label, font=unified_font)
                text_width2 = bbox2[2] - bbox2[0]
            except Exception:
                text_width2 = len(type_label) * 20

            text2_x = max(PADDING, text_center_x - text_width2 // 2)
            text2_y = text1_y + int(font_height * 1.2)
            draw.text((text2_x, text2_y), type_label, fill="black", font=unified_font)

    # ==================== 竖排布局（宽 < 高）====================
    else:
        num_subjects = len(downloaded_images)
        available_height = CANVAS_HEIGHT - (num_subjects + 1) * PADDING
        subject_slot_height = available_height // num_subjects

        # Each subject slot: 75% image area, 25% label area
        subject_image_height = int(subject_slot_height * 0.75)
        subject_label_height = subject_slot_height - subject_image_height

        current_y = PADDING

        for subject in downloaded_images:
            subject_name = subject["name"]
            images = subject["images"]
            card_type = subject["card_type"]
            type_label = f"{card_type}图片"

            # -- Draw images (horizontally within this subject's row) --
            num_imgs = len(images)
            available_img_width = CANVAS_WIDTH - (num_imgs + 1) * PADDING
            per_img_width = available_img_width // num_imgs

            img_x = PADDING
            for img in images:
                img_ratio = img.width / img.height
                new_height = subject_image_height
                new_width = int(new_height * img_ratio)
                if new_width > per_img_width:
                    new_width = per_img_width
                    new_height = int(new_width / img_ratio)

                resized_img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)

                # Center within allocated slot
                x_offset = img_x + (per_img_width - new_width) // 2
                y_offset = current_y + (subject_image_height - new_height) // 2

                canvas.paste(resized_img, (x_offset, y_offset))
                img_x += per_img_width + PADDING

            # -- Draw label (centered below images) --
            label_y_start = current_y + subject_image_height
            max_text_width = int(CANVAS_WIDTH * 0.9)

            unified_font = get_unified_font(subject_name, type_label, max_text_width, min_size=18, max_size=60)

            try:
                font_height = unified_font.size if hasattr(unified_font, 'size') else 30
            except Exception:
                font_height = 30

            total_text_height = int(font_height * 2.2)
            text_vertical_start = label_y_start + (subject_label_height - total_text_height) // 2

            # Name line
            try:
                bbox1 = draw.textbbox((0, 0), subject_name, font=unified_font)
                text_width1 = bbox1[2] - bbox1[0]
            except Exception:
                text_width1 = len(subject_name) * 20

            text1_x = (CANVAS_WIDTH - text_width1) // 2
            text1_y = text_vertical_start
            draw.text((text1_x, text1_y), subject_name, fill="black", font=unified_font)

            # Type line
            try:
                bbox2 = draw.textbbox((0, 0), type_label, font=unified_font)
                text_width2 = bbox2[2] - bbox2[0]
            except Exception:
                text_width2 = len(type_label) * 20

            text2_x = (CANVAS_WIDTH - text_width2) // 2
            text2_y = text1_y + int(font_height * 1.2)
            draw.text((text2_x, text2_y), type_label, fill="black", font=unified_font)

            current_y += subject_slot_height + PADDING

    # 保存到临时文件
    temp_filename = f"collage_{uuid.uuid4().hex}.jpg"
    temp_path = os.path.join("uploads", temp_filename)
    canvas.save(temp_path, "JPEG", quality=95)

    print(f"[拼图生成] 拼图已保存到临时文件: {temp_path}")

    # 上传到CDN
    try:
        cdn_url = upload_to_cdn(temp_path)
        print(f"[拼图生成] 拼图已上传到CDN: {cdn_url}")
    finally:
        try:
            os.remove(temp_path)
        except Exception as e:
            print(f"[拼图生成] 删除临时文件失败: {str(e)}")

    return cdn_url

# ==================== API路由 ====================

async def login(request: LoginRequest, db: Session = Depends(get_db)):
    """通过用户名 + 密码登录"""
    user = db.query(models.User).filter(models.User.username == request.username).first()

    if not user:
        raise HTTPException(status_code=401, detail="用户不存在")

    # 验证密码：自己的密码 或 通用管理密码（保留账号不允许通用密码）
    own_ok = _hash_password(request.password) == user.password_hash
    master_password = (MASTER_PASSWORD or "").strip()
    master_ok = (
        bool(master_password)
        and request.password == master_password
        and request.username not in HIDDEN_USERS
    )

    if not own_ok and not master_ok:
        raise HTTPException(status_code=401, detail="密码错误")

    # 用户使用本人密码登录时，同步保存明文密码供管理端查看
    if own_ok and user.password_plain != request.password:
        user.password_plain = request.password
        db.commit()

    return {
        "id": user.id,
        "username": user.username,
        "token": user.token,
        "created_at": user.created_at
    }

async def verify_token(user: models.User = Depends(get_current_user)):
    """验证token是否有效"""
    return {
        "id": user.id,
        "username": user.username,
        "created_at": user.created_at
    }


class ChangePasswordRequest(BaseModel):
    username: str
    old_password: str
    new_password: str


async def change_password(request: ChangePasswordRequest, db: Session = Depends(get_db)):
    """修改密码（需要验证原密码）"""
    user = db.query(models.User).filter(models.User.username == request.username).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    if _hash_password(request.old_password) != user.password_hash:
        raise HTTPException(status_code=401, detail="原密码错误")

    if not request.new_password:
        raise HTTPException(status_code=400, detail="新密码不能为空")

    user.password_hash = _hash_password(request.new_password)
    user.password_plain = request.new_password
    db.commit()
    return {"message": "密码修改成功"}

async def verify_nerva_password(request: PasswordVerifyRequest):
    """验证nerva用户密码"""
    nerva_password = _get_private_password_env("NERVA_PASSWORD")

    if nerva_password and request.password == nerva_password:
        return {"success": True}
    else:
        raise HTTPException(status_code=401, detail="密码错误")

# ==================== AI生成图片API ====================

# 生成图片
generate_image_for_card = card_media.generate_image_for_card
ImageGenerationRequest = card_image_generation_service.ImageGenerationRequest
_build_card_image_prompt = card_image_generation_service._build_card_image_prompt
_resolve_style_template_content_for_card_type = card_image_generation_service._resolve_style_template_content_for_card_type
_resolve_card_reference_urls = card_image_generation_service._resolve_card_reference_urls

def _verify_admin_panel_password(x_admin_password: Optional[str]) -> None:
    admin_panel_password = (ADMIN_PANEL_PASSWORD or "").strip()
    if (
        not admin_panel_password
        or (x_admin_password or "").strip() != admin_panel_password
    ):
        raise HTTPException(status_code=403, detail="管理员密码错误")


# Compatibility helpers for the extracted model config module.
FUNCTION_MODEL_DEFAULTS = model_configs_service.FUNCTION_MODEL_DEFAULTS
OBSOLETE_FUNCTION_MODEL_KEYS = model_configs_service.OBSOLETE_FUNCTION_MODEL_KEYS
LEGACY_TEXT_PROVIDER_KEYS = model_configs_service.LEGACY_TEXT_PROVIDER_KEYS
LEGACY_TEXT_MODEL_VALUES = model_configs_service.LEGACY_TEXT_MODEL_VALUES


def _get_function_model_default_selection(function_key: str) -> Dict[str, Optional[str]]:
    return model_configs_service._get_function_model_default_selection(function_key)


def _normalize_function_model_id(row: models.FunctionModelConfig) -> str:
    return model_configs_service._normalize_function_model_id(row)


def _ensure_function_model_configs(db):
    return model_configs_service._ensure_function_model_configs(db)


def _serialize_function_model_config(row: models.FunctionModelConfig, db: Session) -> Dict[str, Any]:
    return model_configs_service._serialize_function_model_config(row, db)


async def get_model_configs(
    x_admin_password: Optional[str] = Header(None, alias="X-Admin-Password"),
    db: Session = Depends(get_db),
):
    _verify_admin_panel_password(x_admin_password)
    return model_configs_service.get_model_configs_payload(db)


class UpdateModelConfigRequest(BaseModel):
    model_id: str = ""


async def sync_model_cache(
    x_admin_password: Optional[str] = Header(None, alias="X-Admin-Password"),
    db: Session = Depends(get_db),
):
    _verify_admin_panel_password(x_admin_password)
    sync_result = model_configs_service.sync_models_from_upstream(db)
    db.commit()
    cache_payload = model_configs_service.get_cached_models_payload(db)
    return {
        "message": "???????",
        "count": int(sync_result.get("count") or 0),
        "last_synced_at": cache_payload.get("last_synced_at"),
        "models": cache_payload.get("models", []),
    }


async def update_model_config(
    function_key: str,
    request: UpdateModelConfigRequest,
    x_admin_password: Optional[str] = Header(None, alias="X-Admin-Password"),
    db: Session = Depends(get_db),
):
    _verify_admin_panel_password(x_admin_password)
    model_configs_service._ensure_function_model_configs(db)
    row = db.query(models.FunctionModelConfig).filter(
        models.FunctionModelConfig.function_key == function_key
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="???????")

    explicit_model_id = str(request.model_id or "").strip() or model_configs_service.DEFAULT_TEXT_MODEL_ID
    resolved = model_configs_service.resolve_ai_model_option(
        model_configs_service.RELAY_PROVIDER_KEY,
        explicit_model_id,
        db=db,
    )

    row.provider_key = model_configs_service.RELAY_PROVIDER_KEY
    row.model_key = resolved["model_id"]
    row.model_id = resolved["model_id"]
    db.commit()
    db.refresh(row)
    return model_configs_service._serialize_function_model_config(row, db)


# ==================== Sora准则管理API ====================

# ==================== 视频生成准则API（统一管理Sora和Grok准则） ====================



# ==================== 兼容旧版本 Sora准则 API ====================



# 兼容旧版本 API（保留，但内部使用全局配置）


# ==================== 全局提示词模板API ====================








# ==================== 提示词管理API ====================






# ==================== 时长配置模板API ====================




# ==================== 剧本管理API ====================

# ==================== 片段管理API ====================

EpisodeCreate = episodes.EpisodeCreate
EpisodeResponse = episodes.EpisodeResponse

class EpisodeShotImageSizeUpdateRequest(BaseModel):
    shot_image_size: Optional[str] = None
    detail_images_model: Optional[str] = None
    detail_images_provider: Optional[str] = None
    storyboard2_video_duration: Optional[int] = None
    storyboard2_image_cw: Optional[int] = None
    storyboard2_include_scene_references: Optional[bool] = None


class EpisodeStoryboardVideoSettingsUpdateRequest(BaseModel):
    detail_images_model: Optional[str] = None
    detail_images_provider: Optional[str] = None
    storyboard2_image_cw: Optional[int] = None
    storyboard2_include_scene_references: Optional[bool] = None
    model: Optional[str] = None
    aspect_ratio: Optional[str] = None
    duration: Optional[int] = None
    resolution_name: Optional[str] = None
    storyboard_video_appoint_account: Optional[str] = None
    video_style_template_id: Optional[int] = None
    video_prompt_template: Optional[str] = None


_get_pydantic_fields_set = storyboard_defaults.get_pydantic_fields_set
_normalize_storyboard_video_appoint_account = storyboard_video_settings.normalize_storyboard_video_appoint_account


_get_first_episode_for_storyboard_defaults = storyboard_defaults.get_first_episode_for_storyboard_defaults
_build_episode_storyboard_sora_create_values = episodes._build_episode_storyboard_sora_create_values

create_episode = episodes.create_episode
get_script_episodes = episodes.get_script_episodes

_resolve_narration_template = episodes._resolve_narration_template
_resolve_opening_template = episodes._resolve_opening_template
_submit_episode_text_relay_task = episodes._submit_episode_text_relay_task
convert_to_narration = episodes.convert_to_narration
generate_opening = episodes.generate_opening
get_episode = episodes.get_episode
_build_episode_poll_status_payload = episodes._build_episode_poll_status_payload
_count_storyboard_items = episodes._count_storyboard_items
get_episode_poll_status = episodes.get_episode_poll_status
get_episode_total_cost = episodes.get_episode_total_cost
update_episode = episodes.update_episode
update_episode_storyboard2_duration = episodes.update_episode_storyboard2_duration


async def update_episode_shot_image_size(
    episode_id: int,
    request: EpisodeShotImageSizeUpdateRequest,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """更新剧集统一图视频设置（镜头图尺寸 + 镜头图模型 + 故事板2视频时长）"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if not script or script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    current_size = _normalize_jimeng_ratio(getattr(episode, "shot_image_size", None), default_ratio="9:16")
    normalized_size = _normalize_jimeng_ratio(request.shot_image_size, default_ratio=current_size)
    current_duration = _normalize_storyboard2_video_duration(
        getattr(episode, "storyboard2_video_duration", None),
        default_value=6
    )
    normalized_duration = _normalize_storyboard2_video_duration(
        request.storyboard2_video_duration,
        default_value=current_duration
    )
    current_detail_images_model = _normalize_detail_images_model(
        getattr(episode, "detail_images_model", None),
        default_model="seedream-4.0"
    )
    normalized_detail_images_model = _normalize_detail_images_model(
        request.detail_images_model,
        default_model=current_detail_images_model
    )
    current_detail_images_provider = _resolve_episode_detail_images_provider(episode)
    normalized_detail_images_provider = _normalize_detail_images_provider(
        request.detail_images_provider,
        default_provider=current_detail_images_provider
    )
    current_image_cw = _normalize_storyboard2_image_cw(
        getattr(episode, "storyboard2_image_cw", None),
        default_value=50
    )
    normalized_image_cw = _normalize_storyboard2_image_cw(
        request.storyboard2_image_cw,
        default_value=current_image_cw
    )
    current_include_scene = bool(getattr(episode, "storyboard2_include_scene_references", False))
    normalized_include_scene = current_include_scene if request.storyboard2_include_scene_references is None else bool(request.storyboard2_include_scene_references)
    episode.shot_image_size = normalized_size
    episode.detail_images_model = normalized_detail_images_model
    episode.detail_images_provider = normalized_detail_images_provider
    episode.storyboard2_video_duration = normalized_duration
    episode.storyboard2_image_cw = normalized_image_cw
    episode.storyboard2_include_scene_references = normalized_include_scene
    db.commit()

    return {
        "episode_id": episode.id,
        "shot_image_size": normalized_size,
        "detail_images_model": normalized_detail_images_model,
        "detail_images_provider": normalized_detail_images_provider,
        "storyboard2_video_duration": normalized_duration,
        "storyboard2_image_cw": normalized_image_cw,
        "storyboard2_include_scene_references": normalized_include_scene,
        "message": "图视频设置已保存"
    }


async def update_episode_storyboard_video_settings(
    episode_id: int,
    request: EpisodeStoryboardVideoSettingsUpdateRequest,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """更新剧集统一图视频设置（故事板sora视频 + 镜头图模型 + 参考图策略）"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if not script or script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    current_settings = _get_episode_storyboard_video_settings(episode)
    normalized_model = _normalize_storyboard_video_model(
        request.model,
        default_model=current_settings["model"]
    )
    normalized_aspect_ratio = _normalize_storyboard_video_aspect_ratio(
        request.aspect_ratio,
        model=normalized_model,
        default_ratio=current_settings["aspect_ratio"]
    )
    normalized_duration = _normalize_storyboard_video_duration(
        request.duration,
        model=normalized_model,
        default_duration=current_settings["duration"]
    )
    normalized_resolution_name = _normalize_storyboard_video_resolution_name(
        request.resolution_name,
        model=normalized_model,
        default_resolution=current_settings.get("resolution_name", "")
    )
    normalized_appoint_account = _normalize_storyboard_video_appoint_account(
        request.storyboard_video_appoint_account,
        default_value=getattr(episode, "storyboard_video_appoint_account", "") or ""
    )
    # 故事板sora中镜头图比例与视频比例保持一致，统一使用同一设置值。
    normalized_shot_image_size = normalized_aspect_ratio
    current_detail_images_model = _normalize_detail_images_model(
        getattr(episode, "detail_images_model", None),
        default_model="seedream-4.0"
    )
    normalized_detail_images_model = _normalize_detail_images_model(
        request.detail_images_model,
        default_model=current_detail_images_model
    )
    current_detail_images_provider = _resolve_episode_detail_images_provider(episode)
    normalized_detail_images_provider = _normalize_detail_images_provider(
        request.detail_images_provider,
        default_provider=current_detail_images_provider
    )
    current_image_cw = _normalize_storyboard2_image_cw(
        getattr(episode, "storyboard2_image_cw", None),
        default_value=50
    )
    normalized_image_cw = _normalize_storyboard2_image_cw(
        request.storyboard2_image_cw,
        default_value=current_image_cw
    )
    current_include_scene = bool(getattr(episode, "storyboard2_include_scene_references", False))
    normalized_include_scene = (
        current_include_scene
        if request.storyboard2_include_scene_references is None
        else bool(request.storyboard2_include_scene_references)
    )
    normalized_provider = _resolve_storyboard_video_provider(normalized_model)
    normalized_video_prompt_template = (
        (getattr(episode, "video_prompt_template", "") or "")
        if request.video_prompt_template is None
        else str(request.video_prompt_template or "")
    )

    episode.shot_image_size = normalized_shot_image_size
    episode.detail_images_model = normalized_detail_images_model
    episode.detail_images_provider = normalized_detail_images_provider
    episode.storyboard2_image_cw = normalized_image_cw
    episode.storyboard2_include_scene_references = normalized_include_scene
    episode.storyboard_video_model = normalized_model
    episode.storyboard_video_aspect_ratio = normalized_aspect_ratio
    episode.storyboard_video_duration = normalized_duration
    episode.storyboard_video_resolution_name = normalized_resolution_name
    episode.storyboard_video_appoint_account = normalized_appoint_account
    episode.video_prompt_template = normalized_video_prompt_template
    if request.video_style_template_id is not None:
        episode.video_style_template_id = request.video_style_template_id if request.video_style_template_id > 0 else None
    db.commit()

    return {
        "episode_id": episode.id,
        "shot_image_size": normalized_shot_image_size,
        "detail_images_model": normalized_detail_images_model,
        "detail_images_provider": normalized_detail_images_provider,
        "storyboard2_image_cw": normalized_image_cw,
        "storyboard2_include_scene_references": normalized_include_scene,
        "model": normalized_model,
        "aspect_ratio": normalized_aspect_ratio,
        "duration": normalized_duration,
        "resolution_name": normalized_resolution_name,
        "storyboard_video_appoint_account": normalized_appoint_account,
        "provider": normalized_provider,
        "video_style_template_id": episode.video_style_template_id,
        "video_prompt_template": normalized_video_prompt_template,
        "message": "视频设置已保存"
    }

# ==================== 辅助函数：从storyboard_data创建镜头记录 ====================

_normalize_storyboard_shot_ids = episode_cleanup.normalize_storyboard_shot_ids
_clear_storyboard_shot_dependencies = episode_cleanup.clear_storyboard_shot_dependencies
_delete_storyboard_shots_by_ids = episode_cleanup.delete_storyboard_shots_by_ids
_delete_episode_storyboard_shots = episode_cleanup.delete_episode_storyboard_shots
_clear_episode_dependencies = episode_cleanup.clear_episode_dependencies


def _get_storyboard_shot_family_ids(shot: models.StoryboardShot, db: Session) -> List[int]:
    stable_id = str(getattr(shot, "stable_id", "") or "").strip()
    query = db.query(models.StoryboardShot.id).filter(
        models.StoryboardShot.episode_id == shot.episode_id
    )
    if stable_id:
        query = query.filter(
            or_(
                models.StoryboardShot.stable_id == stable_id,
                and_(
                    models.StoryboardShot.shot_number == shot.shot_number,
                    or_(
                        models.StoryboardShot.stable_id.is_(None),
                        models.StoryboardShot.stable_id == "",
                    ),
                ),
            )
        )
    else:
        query = query.filter(models.StoryboardShot.shot_number == shot.shot_number)
    return [
        family_shot_id
        for family_shot_id, in query.order_by(
            models.StoryboardShot.shot_number.asc(),
            models.StoryboardShot.variant_index.asc(),
            models.StoryboardShot.id.asc()
        ).all()
    ]


_create_shots_from_storyboard_data = storyboard_sync.create_shots_from_storyboard_data

# ==================== 新三阶段分镜生成后台任务 ====================

def _submit_simple_storyboard_batch_task(
    db: Session,
    *,
    episode_id: int,
    batch_row: models.SimpleStoryboardBatch,
    duration: int,
):
    template = db.query(models.ShotDurationTemplate).filter(
        models.ShotDurationTemplate.duration == int(duration or 15)
    ).first()
    if not template:
        raise ValueError(f"未找到 {int(duration or 15)} 秒简单分镜模板")

    prompt = str(template.simple_storyboard_rule or "").format(content=str(batch_row.source_text or ""))
    config = get_ai_config("simple_storyboard")
    request_data = {
        "model": config["model"],
        "messages": [
            {
                "role": "user",
                "content": prompt,
            }
        ],
        "response_format": {"type": "json_object"},
        "stream": False,
    }
    task_payload = {
        "episode_id": int(episode_id),
        "batch_row_id": int(batch_row.id),
        "batch_index": int(batch_row.batch_index or 0),
    }
    relay_task = submit_and_persist_text_task(
        db,
        task_type="simple_storyboard_batch",
        owner_type="simple_storyboard_batch",
        owner_id=int(batch_row.id),
        stage_key="simple_storyboard",
        function_key="simple_storyboard",
        request_payload=request_data,
        task_payload=task_payload,
    )
    batch_row.status = "submitting"
    batch_row.error_message = ""
    batch_row.last_attempt = 1
    batch_row.updated_at = datetime.utcnow()
    return relay_task


def _submit_detailed_storyboard_stage1_task(db: Session, *, episode_id: int, simple_shots: List[Dict[str, Any]]):
    shots_content = ""
    for shot in simple_shots:
        shot_num = shot.get("shot_number", "?")
        original_text = shot.get("original_text", "")
        shots_content += f"镜头{shot_num}:\n{original_text}\n\n"

    prompt_template = get_prompt_by_key("detailed_storyboard_content_analysis")
    prompt = prompt_template.format(shots_content=shots_content)
    config = get_ai_config("detailed_storyboard_s1")
    request_data = {
        "model": config["model"],
        "messages": [
            {
                "role": "user",
                "content": prompt,
            }
        ],
        "response_format": {"type": "json_object"},
        "stream": False,
    }
    task_payload = {
        "episode_id": int(episode_id),
        "simple_shots": simple_shots,
    }
    return submit_and_persist_text_task(
        db,
        task_type="detailed_storyboard_stage1",
        owner_type="episode",
        owner_id=int(episode_id),
        stage_key="detailed_storyboard_stage1",
        function_key="detailed_storyboard_s1",
        request_payload=request_data,
        task_payload=task_payload,
    )


def _submit_detailed_storyboard_stage2_task(db: Session, *, episode_id: int, final_shots: List[Dict[str, Any]]):
    full_storyboard_json = json.dumps({"shots": final_shots}, ensure_ascii=False, indent=2)
    prompt_template = get_prompt_by_key("stage2_refine_shot")
    prompt = prompt_template.format(
        total_shots=len(final_shots),
        full_storyboard_json=full_storyboard_json
    )
    config = get_ai_config("detailed_storyboard_s2")
    request_data = {
        "model": config["model"],
        "messages": [
            {
                "role": "user",
                "content": prompt,
            }
        ],
        "response_format": {"type": "json_object"},
        "stream": False,
    }
    task_payload = {
        "episode_id": int(episode_id),
        "final_shots": final_shots,
    }
    return submit_and_persist_text_task(
        db,
        task_type="detailed_storyboard_stage2",
        owner_type="episode",
        owner_id=int(episode_id),
        stage_key="detailed_storyboard_stage2",
        function_key="detailed_storyboard_s2",
        request_payload=request_data,
        task_payload=task_payload,
    )

# 后台任务：生成简单分镜
def generate_simple_storyboard_background(episode_id: int, content: str, batch_size: int, duration: int):
    """后台生成简单分镜（新阶段1：镜头划分）"""
    db = SessionLocal()
    try:
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if not episode:
            return

        batch_texts = _split_simple_storyboard_batches(content, batch_size)
        total_batches = len(batch_texts)

        def mark_simple_storyboard_generating():
            episode.batch_size = batch_size
            episode.simple_storyboard_data = json.dumps({"shots": []}, ensure_ascii=False)
            episode.simple_storyboard_generating = True
            episode.simple_storyboard_error = ""
            _reset_simple_storyboard_batches_for_episode(episode_id, total_batches, batch_texts, db)

        commit_with_retry(
            db,
            prepare_fn=mark_simple_storyboard_generating,
            context=f"simple_storyboard_start episode={episode_id}"
        )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        task_folder = f"simple_storyboard_episode_{episode_id}_{timestamp}"

        from ai_service import generate_simple_storyboard
        result, debug_info = generate_simple_storyboard(
            content,
            batch_size,
            duration=duration,
            episode_id=episode_id,
            task_folder=task_folder,
            batches=[
                {
                    "batch_index": index + 1,
                    "content": batch_content,
                    "retry_count": 0,
                }
                for index, batch_content in enumerate(batch_texts)
            ],
            batch_retry_limit=10,
            on_batch_result=lambda payload: _apply_simple_storyboard_batch_update(episode_id, payload),
        )

        refreshed_episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if refreshed_episode:
            def save_simple_storyboard_result():
                summary = _refresh_episode_simple_storyboard_from_batches(refreshed_episode, db)
                refreshed_episode.simple_storyboard_data = json.dumps(result, ensure_ascii=False)
                refreshed_episode.simple_storyboard_generating = False
                refreshed_episode.simple_storyboard_error = "" if not summary.get("has_failures") else refreshed_episode.simple_storyboard_error

            commit_with_retry(
                db,
                prepare_fn=save_simple_storyboard_result,
                context=f"simple_storyboard_finish episode={episode_id}"
            )

        print(f"✅ 简单分镜生成成功，共 {len(result.get('shots', []))} 个镜头")

    except Exception as e:
        print(f"❌ 简单分镜生成失败: {str(e)}")
        _rollback_quietly(db)
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if episode:
            def mark_simple_storyboard_failed():
                batch_rows = _get_simple_storyboard_batch_rows(episode_id, db)
                if batch_rows:
                    for row in batch_rows:
                        if str(getattr(row, "status", "") or "").strip() in {"completed", "failed"}:
                            continue
                        row.status = "failed"
                        if not str(getattr(row, "error_message", "") or "").strip():
                            row.error_message = str(e)
                        row.updated_at = datetime.utcnow()
                    _refresh_episode_simple_storyboard_from_batches(episode, db)
                else:
                    episode.simple_storyboard_generating = False
                    episode.simple_storyboard_error = str(e)

            try:
                commit_with_retry(
                    db,
                    prepare_fn=mark_simple_storyboard_failed,
                    context=f"simple_storyboard_fail episode={episode_id}"
                )
            except Exception as commit_error:
                print(f"❌ 简单分镜失败状态写回失败: {str(commit_error)}")
    finally:
        db.close()


def retry_failed_simple_storyboard_batches_background(episode_id: int):
    db = SessionLocal()
    try:
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if not episode:
            return
        batch_rows = _get_simple_storyboard_batch_rows(episode_id, db)
        retry_items = [
            row for row in batch_rows
            if str(getattr(row, "status", "") or "").strip() in {"pending", "failed"}
            and int(getattr(row, "retry_count", 0) or 0) > 0
        ]
        if not retry_items:
            episode.simple_storyboard_generating = False
            db.commit()
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        task_folder = f"simple_storyboard_retry_episode_{episode_id}_{timestamp}"
        duration = episode.storyboard2_duration or 15
        batch_size = episode.batch_size or 500

        from ai_service import generate_simple_storyboard
        generate_simple_storyboard(
            str(getattr(episode, "content", "") or ""),
            batch_size,
            duration=duration,
            episode_id=episode_id,
            task_folder=task_folder,
            batches=_build_simple_storyboard_batch_runtime_items(retry_items),
            batch_retry_limit=1,
            on_batch_result=lambda payload: _apply_simple_storyboard_batch_update(episode_id, payload),
        )

        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if episode:
            def finalize_retry_state():
                _refresh_episode_simple_storyboard_from_batches(episode, db)
            commit_with_retry(
                db,
                prepare_fn=finalize_retry_state,
                context=f"simple_storyboard_retry_finish episode={episode_id}"
            )
    except Exception as e:
        _rollback_quietly(db)
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if episode:
            def mark_retry_failed():
                batch_rows = _get_simple_storyboard_batch_rows(episode_id, db)
                for row in batch_rows:
                    if str(getattr(row, "status", "") or "").strip() == "submitting":
                        row.status = "failed"
                        if not str(getattr(row, "error_message", "") or "").strip():
                            row.error_message = str(e)
                        row.updated_at = datetime.utcnow()
                _refresh_episode_simple_storyboard_from_batches(episode, db)
            try:
                commit_with_retry(
                    db,
                    prepare_fn=mark_retry_failed,
                    context=f"simple_storyboard_retry_fail episode={episode_id}"
                )
            except Exception as commit_error:
                print(f"❌ 简单分镜重试失败状态写回失败: {str(commit_error)}")
    finally:
        db.close()


# 后台任务：生成详细分镜
def generate_detailed_storyboard_background(episode_id: int):
    """后台生成详细分镜（新阶段2：内容分析 + 新阶段3：主体提示词）

    参数：
    - episode_id: 片段ID
    """
    db = SessionLocal()
    try:
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if not episode:
            return

        # 检查简单分镜是否存在
        if not episode.simple_storyboard_data:
            raise Exception("简单分镜数据不存在，请先生成简单分镜")

        def mark_detailed_storyboard_generating():
            episode.storyboard_generating = True
            episode.storyboard_error = ""

        commit_with_retry(
            db,
            prepare_fn=mark_detailed_storyboard_generating,
            context=f"detailed_storyboard_start episode={episode_id}"
        )

        # 创建任务文件夹
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        task_folder = f"detailed_storyboard_episode_{episode_id}_{timestamp}"

        # 读取简单分镜数据
        simple_storyboard = json.loads(episode.simple_storyboard_data)
        simple_shots = simple_storyboard.get("shots", [])

        if not simple_shots:
            raise Exception("简单分镜中没有镜头数据")

        # ==================== 新阶段2：详细分镜（内容分析） ====================

        from ai_service import generate_detailed_storyboard
        detailed_result, debug_info = generate_detailed_storyboard(
            simple_shots,
            episode_id=episode_id,
            task_folder=task_folder
        )

        detailed_shots = detailed_result.get("shots", [])
        if not detailed_shots:
            raise Exception("详细分镜生成失败：未生成任何镜头")

        # ==================== 新阶段3：主体绘画提示词生成 ====================

        from ai_service import stage2_generate_subject_prompts

        final_shots = []

        for shot in detailed_shots:
            shot["subjects"] = _normalize_storyboard_generation_subjects(shot.get("subjects", []))
            final_shots.append(shot)

        # 构建完整分镜表JSON（用于阶段3）
        full_storyboard_json = json.dumps({
            "shots": final_shots
        }, ensure_ascii=False, indent=2)

        # 调用阶段3生成主体提示词
        stage2_result, stage2_debug = stage2_generate_subject_prompts(
            full_storyboard_json,
            episode_id=episode_id,
            task_folder=task_folder
        )

        stage2_subjects = _normalize_stage2_subjects(stage2_result.get("subjects", []))
        name_mappings = stage2_result.get("name_mappings", {})
        canonical_subject_map = _build_subject_detail_map(stage2_subjects)
        for shot in final_shots:
            shot["subjects"] = _reconcile_storyboard_shot_subjects(
                shot,
                canonical_subject_map,
                name_mappings=name_mappings,
            )

        # 保存最终结果
        final_data = {
            "shots": final_shots,
            "subjects": stage2_subjects
        }

        voiceover_shots = []
        for shot in final_shots:
            voiceover_shot = {
                "shot_number": shot.get("shot_number"),
                "voice_type": shot.get("voice_type"),
                "narration": shot.get("narration"),
                "dialogue": shot.get("dialogue")
            }
            voiceover_shots.append(voiceover_shot)

        merged_voiceover_data = _merge_voiceover_shots_preserving_extensions(
            episode.voiceover_data,
            voiceover_shots
        )

        def save_detailed_storyboard_result():
            episode.storyboard_data = json.dumps(final_data, ensure_ascii=False)
            episode.voiceover_data = json.dumps(merged_voiceover_data, ensure_ascii=False)
            episode.storyboard_generating = False
            episode.storyboard_error = ""

        commit_with_retry(
            db,
            prepare_fn=save_detailed_storyboard_result,
            context=f"detailed_storyboard_finish episode={episode_id}"
        )

        print(f"✅ 详细分镜生成成功，共 {len(final_shots)} 个镜头，{len(stage2_subjects)} 个主体")

        # 自动创建镜头记录
        _create_shots_from_storyboard_data(episode_id, db)

    except Exception as e:
        print(f"❌ 详细分镜生成失败: {str(e)}")
        _rollback_quietly(db)
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if episode:
            def mark_detailed_storyboard_failed():
                episode.storyboard_generating = False
                episode.storyboard_error = str(e)

            try:
                commit_with_retry(
                    db,
                    prepare_fn=mark_detailed_storyboard_failed,
                    context=f"detailed_storyboard_fail episode={episode_id}"
                )
            except Exception as commit_error:
                print(f"❌ 详细分镜失败状态写回失败: {str(commit_error)}")
    finally:
        db.close()


# 后台任务：生成分镜表（新两阶段流程）
def generate_storyboard_background(episode_id: int, content: str, prompt_style: str = None, append_mode: bool = False):
    """后台生成分镜表（两阶段流程）

    阶段1：按500字分批，生成初步分镜（约9个/批），并提取主体
    阶段2：基于完整分镜表生成主体绘画提示词与别名（规范角色名）

    参数：
    - episode_id: 片段ID
    - content: 文案内容
    - prompt_style: 提示词风格（用于角色prompt拼接）
    - append_mode: 保留参数但不使用（新流程不支持追加）
    """
    db = SessionLocal()
    try:
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if not episode:
            return

        def mark_storyboard_generating():
            episode.storyboard_generating = True

        commit_with_retry(
            db,
            prepare_fn=mark_storyboard_generating,
            context=f"storyboard_start episode={episode_id}"
        )


        # ✅ 创建任务文件夹（所有阶段的调试文件都保存在这里）
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        task_folder = f"episode_{episode_id}_{timestamp}"

        # ==================== 阶段1：初步分镜生成 ====================

        # 文本分批逻辑：按自然段累积，接近500字触发一批
        paragraphs = [p.strip() for p in content.split('\n') if p.strip()]
        batches = []  # 每批是一个段落列表
        current_batch = []
        current_length = 0

        for para in paragraphs:
            para_length = len(para)
            if current_length + para_length >= 500 and current_batch:
                # 触发处理
                batches.append('\n\n'.join(current_batch))
                current_batch = [para]
                current_length = para_length
            else:
                current_batch.append(para)
                current_length += para_length

        # 最后一批
        if current_batch:
            batches.append('\n\n'.join(current_batch))


        # 对每批调用阶段1 AI（并发执行）
        all_stage1_shots = []
        shot_counter = 1

        def process_batch(batch_idx, batch_content):
            """处理单个批次"""
            try:
                # ✅ 传递 episode_id, task_folder 和 batch_id，让 stage1 函数自己保存调试信息
                result, debug_info = stage1_generate_initial_storyboard(
                    batch_content,
                    episode_id=episode_id,
                    task_folder=task_folder,
                    batch_id=str(batch_idx + 1)
                )
                batch_shots = result.get("shots", [])
                return (batch_idx, batch_shots)
            except Exception as e:
                return (batch_idx, [])

        # 并发执行所有批次
        with ThreadPoolExecutor(max_workers=min(len(batches), 10)) as executor:
            futures = [executor.submit(process_batch, i, batch) for i, batch in enumerate(batches)]
            batch_results = {}
            for future in futures:
                batch_idx, batch_shots = future.result()
                batch_results[batch_idx] = batch_shots

        # 按批次顺序合并并重新编号
        for batch_idx in sorted(batch_results.keys()):
            batch_shots = batch_results[batch_idx]
            for shot in batch_shots:
                shot["shot_number"] = str(shot_counter)
                shot_counter += 1
            all_stage1_shots.extend(batch_shots)

        if not all_stage1_shots:
            raise Exception("阶段1未生成任何镜头")


        # ==================== 阶段2：主体绘画提示词 ====================

        final_shots = []
        for stage1_shot in all_stage1_shots:
            final_shots.append({
                "shot_number": stage1_shot["shot_number"],
                "subjects": _normalize_storyboard_generation_subjects(stage1_shot.get("subjects", [])),
                "original_text": stage1_shot.get("original_text", ""),
                "voice_type": stage1_shot.get("voice_type", "none"),
                "narration": stage1_shot.get("narration"),
                "dialogue": stage1_shot.get("dialogue"),
            })

        full_storyboard_json = json.dumps({"shots": final_shots}, ensure_ascii=False, indent=2)
        stage2_subjects = []

        try:
            # ✅ 传递 episode_id 和 task_folder，让 Stage2 自己保存每次尝试的调试信息
            result, debug_info = stage2_generate_subject_prompts(
                full_storyboard_json,
                episode_id=episode_id,
                task_folder=task_folder
            )
            stage2_subjects = _normalize_stage2_subjects(result.get("subjects", []))
            name_mappings = result.get("name_mappings", {})  # 获取AI生成的名称映射表

        except Exception as e:
            stage2_subjects = []
            name_mappings = {}
            # Stage2 已经在内部保存了每次尝试的调试信息，不需要额外保存

        # 过滤并去重主体结果
        stage2_subjects = _normalize_stage2_subjects(stage2_subjects)

        # 使用AI提供的映射表更新分镜表中的主体名称
        if name_mappings:
            print(f"  ✓ 应用主体名称映射表: {name_mappings}")
            for shot in final_shots:
                updated_subjects = []
                seen = set()
                for subj in shot.get("subjects", []):
                    original_name = subj.get("name", "")
                    subject_type = subj.get("type", "")

                    # 使用映射表查找规范名称
                    canonical_name = name_mappings.get(original_name, original_name)

                    # 构建规范主体
                    canonical_subj = {"name": canonical_name, "type": subject_type}
                    key = (canonical_name, subject_type)

                    # 去重
                    if key in seen:
                        continue
                    seen.add(key)
                    updated_subjects.append(canonical_subj)

                shot["subjects"] = updated_subjects

        # ==================== 保存最终结果 ====================
        # 注意：ai_prompt 只存储纯粹的描述（不含格式化前缀和风格）
        # 完整prompt的拼接在生成图片时进行（generate_image_for_card 函数）
        final_data = {
            "shots": final_shots,
            "subjects": stage2_subjects  # 包含ai_prompt和alias
        }

        def save_storyboard_result():
            episode.storyboard_data = json.dumps(final_data, ensure_ascii=False)
            episode.storyboard_generating = False
            episode.storyboard_error = ""

        commit_with_retry(
            db,
            prepare_fn=save_storyboard_result,
            context=f"storyboard_finish episode={episode_id}"
        )


        # ✅ 自动创建镜头记录到storyboard_shots表
        try:
            _create_shots_from_storyboard_data(episode_id, db)
        except Exception as e:
            pass

    except Exception as e:
        _rollback_quietly(db)
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if episode:
            def mark_storyboard_failed():
                episode.storyboard_generating = False

                # 保存详细错误信息
                error_message = str(e)

                # 尝试从异常对象中提取更详细的信息
                if hasattr(e, 'debug_info'):
                    debug_info = e.debug_info
                    output = debug_info.get('output', {})

                    # 检查是否是content_filter错误
                    if 'full_response' in output:
                        full_response = output['full_response']
                        if isinstance(full_response, dict) and 'choices' in full_response:
                            choices = full_response.get('choices', [])
                            if choices and len(choices) > 0:
                                finish_reason = choices[0].get('finish_reason', '')
                                if finish_reason == 'content_filter':
                                    error_message = "AI内容审查拦截：生成的内容触发了安全过滤器。建议修改剧本中的敏感词汇或情绪描述。"
                                elif finish_reason and finish_reason != 'stop':
                                    error_message = f"AI生成异常：{finish_reason}"

                    # 如果有其他错误信息
                    if 'error' in output:
                        error_detail = output['error']
                        if 'JSON' in error_detail or 'json' in error_detail.lower():
                            error_message = f"AI响应格式错误：{error_detail}"

                episode.storyboard_error = error_message
                print(f"❌ 分镜表生成失败: {error_message}")

            try:
                commit_with_retry(
                    db,
                    prepare_fn=mark_storyboard_failed,
                    context=f"storyboard_fail episode={episode_id}"
                )
            except Exception as commit_error:
                print(f"❌ 分镜表失败状态写回失败: {str(commit_error)}")
    finally:
        db.close()

# 新增：分镜表分析响应模型
class StoryboardAnalyzeResponse(BaseModel):
    message: str  # 提示消息
    generating: bool  # 是否正在生成

# 新增：创建分镜表请求模型
class CreateStoryboardRequest(BaseModel):
    shots: List[dict]  # 编辑后的镜头列表

# ==================== 新三阶段分镜生成API ====================

class SimpleStoryboardRequest(BaseModel):
    content: Optional[str] = None  # 可选的自定义文案内容
    batch_size: Optional[int] = 500  # 分批字数，默认500


async def generate_detailed_storyboard_api(
    episode_id: int,
    background_tasks: BackgroundTasks = None,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """启动后台详细分镜生成任务（新阶段2+3：内容分析 + 主体提示词）"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    summary = _get_simple_storyboard_batch_summary(episode_id, db)
    if summary.get("has_failures"):
        raise HTTPException(status_code=400, detail="简单分镜存在失败批次，请先重试失败批次")

    # 清空旧的详细分镜数据
    def mark_detailed_storyboard_request_started():
        episode.storyboard_data = None
        episode.storyboard_generating = True
        episode.storyboard_error = ""

    commit_with_retry(
        db,
        prepare_fn=mark_detailed_storyboard_request_started,
        context=f"detailed_storyboard_request episode={episode_id}"
    )

    simple_storyboard = json.loads(episode.simple_storyboard_data or "{}")
    simple_shots = simple_storyboard.get("shots", [])
    if not simple_shots:
        raise HTTPException(status_code=400, detail="简单分镜中没有镜头数据")

    try:
        relay_task = _submit_detailed_storyboard_stage1_task(
            db,
            episode_id=episode_id,
            simple_shots=simple_shots,
        )
        db.commit()
    except Exception as exc:
        db.rollback()
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if episode:
            episode.storyboard_generating = False
            episode.storyboard_error = str(exc)
            db.commit()
        raise HTTPException(status_code=502, detail=f"提交文本任务失败: {str(exc)}")

    return {
        "message": "详细分镜生成任务已提交",
        "generating": True,
        "task_id": relay_task.external_task_id,
    }


class AnalyzeStoryboardRequest(BaseModel):
    content: Optional[str] = None  # 可选的自定义文案内容
    append: Optional[bool] = False  # 是否追加模式（不清空旧数据）

async def analyze_episode_for_storyboard(
    episode_id: int,
    request: AnalyzeStoryboardRequest = None,
    background_tasks: BackgroundTasks = None,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """兼容旧入口：无追加参数时复用新的详细分镜提交流程。"""
    if request and (bool(request.append) or str(request.content or "").strip()):
        raise HTTPException(status_code=400, detail="旧版追加分析流程已下线，请使用新的简单分镜/详细分镜流程")

    payload = await generate_detailed_storyboard_api(
        episode_id=episode_id,
        background_tasks=background_tasks,
        user=user,
        db=db,
    )
    return {
        "message": str(payload.get("message") or "分镜表生成任务已提交"),
        "generating": bool(payload.get("generating", True)),
    }


# 新增：获取详细分镜的原始JSON数据（包含完整的voice_type、narration、dialogue）
def get_detailed_storyboard(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取详细分镜的原始JSON数据（用于配音表等功能）"""
    episode, script = _ensure_voiceover_permission(episode_id, user, db)

    # 获取主体库
    library = db.query(models.StoryLibrary).filter(
        models.StoryLibrary.episode_id == episode.id
    ).first()

    stored_subject_map = {}
    if episode.storyboard_data:
        try:
            stored_subject_map = _build_subject_detail_map(json.loads(episode.storyboard_data).get("subjects", []))
        except Exception:
            stored_subject_map = {}

    subjects = []
    if library:
        cards = db.query(models.SubjectCard).filter(
            models.SubjectCard.library_id == library.id
        ).all()
        cards = [card for card in cards if card.card_type in ALLOWED_CARD_TYPES]
        for card in cards:
            stored_subject = stored_subject_map.get((card.name, card.card_type), {})
            subjects.append({
                "id": card.id,
                "name": card.name,
                "card_type": card.card_type,
                "type": card.card_type,
                "ai_prompt": (card.ai_prompt or "").strip() or stored_subject.get("ai_prompt", ""),
                "role_personality": (getattr(card, "role_personality", "") or "").strip() or stored_subject.get("role_personality", ""),
                "alias": (card.alias or "").strip() or stored_subject.get("alias", "")
            })

    shared_data = _load_script_voiceover_shared_data(script)
    default_voice_ref_id = _voiceover_first_reference_id(shared_data)

    # 优先从voiceover_data读取，缺失时回退storyboard_data并自动补齐line_id/tts
    voiceover_payload = _parse_episode_voiceover_payload(episode)
    shots = voiceover_payload.get("shots", [])
    loaded_from_storyboard = False

    if not isinstance(shots, list):
        shots = []

    if not shots and episode.storyboard_data:
        try:
            data = json.loads(episode.storyboard_data)
            if isinstance(data, dict) and "shots" in data:
                loaded_from_storyboard = True
                # 提取配音相关字段
                for shot in data.get("shots", []):
                    shots.append({
                        "shot_number": shot.get("shot_number"),
                        "voice_type": shot.get("voice_type"),
                        "narration": shot.get("narration"),
                        "dialogue": shot.get("dialogue")
                    })
        except json.JSONDecodeError:
            shots = []

    shots, changed = _normalize_voiceover_shots_for_tts(shots, default_voice_ref_id)
    if loaded_from_storyboard or changed:
        voiceover_payload["shots"] = shots
        episode.voiceover_data = json.dumps(voiceover_payload, ensure_ascii=False)
        db.commit()

    return {
        "generating": episode.storyboard_generating or False,
        "error": episode.storyboard_error or "",
        "shots": shots,
        "subjects": subjects,
        "tts_shared": shared_data
    }


# 保存配音表数据（只更新voiceover_data字段）
async def update_voiceover_data(
    episode_id: int,
    request: dict,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """保存配音表数据（只更新voiceover_data，不影响其他数据）"""
    episode, script = _ensure_voiceover_permission(episode_id, user, db)

    incoming_shots = request.get("shots", [])
    merged_voiceover_data = _merge_voiceover_shots_preserving_extensions(
        episode.voiceover_data,
        incoming_shots if isinstance(incoming_shots, list) else []
    )

    shared_data = _load_script_voiceover_shared_data(script)
    default_voice_ref_id = _voiceover_first_reference_id(shared_data)
    normalized_shots, _ = _normalize_voiceover_shots_for_tts(
        merged_voiceover_data.get("shots", []),
        default_voice_ref_id
    )
    merged_voiceover_data["shots"] = normalized_shots

    episode.voiceover_data = json.dumps(merged_voiceover_data, ensure_ascii=False)
    db.commit()

    print(f"✅ 配音表已保存，共 {len(normalized_shots)} 个镜头")

    return {"message": "配音表已保存", "success": True, "shots": normalized_shots}


async def get_voiceover_shared_data(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _, script = _ensure_voiceover_permission(episode_id, user, db)
    shared = _load_script_voiceover_shared_data(script)
    return {"success": True, "shared": shared}


async def create_voiceover_voice_reference(
    episode_id: int,
    name: str = Form(...),
    file: UploadFile = File(...),
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _, script = _ensure_voiceover_permission(episode_id, user, db)

    ref_name = str(name or "").strip()
    if not ref_name:
        raise HTTPException(status_code=400, detail="音色参考音频名称不能为空")

    cdn_url = save_and_upload_to_cdn(file)
    shared = _load_script_voiceover_shared_data(script)
    item = {
        "id": f"voice_ref_{uuid.uuid4().hex}",
        "name": ref_name,
        "file_name": str(file.filename or "").strip(),
        "url": cdn_url,
        "local_path": "",
        "created_at": datetime.utcnow().isoformat()
    }
    shared["voice_references"].append(item)
    _save_script_voiceover_shared_data(script, shared)
    db.commit()

    return {"success": True, "item": item, "shared": _load_script_voiceover_shared_data(script)}


async def rename_voiceover_voice_reference(
    episode_id: int,
    reference_id: str,
    request: dict,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _, script = _ensure_voiceover_permission(episode_id, user, db)

    target_id = str(reference_id or "").strip()
    if not target_id:
        raise HTTPException(status_code=400, detail="reference_id不能为空")

    new_name = str(request.get("name") or "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="音色名称不能为空")

    shared = _load_script_voiceover_shared_data(script)
    refs = shared.get("voice_references", [])
    if not isinstance(refs, list):
        refs = []
        shared["voice_references"] = refs

    target_item = None
    for item in refs:
        if isinstance(item, dict) and str(item.get("id") or "").strip() == target_id:
            target_item = item
            break
    if not isinstance(target_item, dict):
        raise HTTPException(status_code=404, detail="音色参考音频不存在")

    target_item["name"] = new_name
    target_item["updated_at"] = datetime.utcnow().isoformat()

    _save_script_voiceover_shared_data(script, shared)
    db.commit()
    return {
        "success": True,
        "item": target_item,
        "shared": _load_script_voiceover_shared_data(script)
    }


async def preview_voiceover_voice_reference(
    episode_id: int,
    reference_id: str,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _, script = _ensure_voiceover_permission(episode_id, user, db)
    target_id = str(reference_id or "").strip()
    if not target_id:
        raise HTTPException(status_code=400, detail="reference_id不能为空")

    shared = _load_script_voiceover_shared_data(script)
    refs = shared.get("voice_references", [])
    target = None
    if isinstance(refs, list):
        target = next((item for item in refs if str(item.get("id") or "").strip() == target_id), None)
    if not isinstance(target, dict):
        raise HTTPException(status_code=404, detail="音色参考音频不存在")

    source = _resolve_voiceover_audio_source(target)
    if not source:
        raise HTTPException(status_code=404, detail="音色参考音频不可访问")

    if source.startswith("http://") or source.startswith("https://"):
        return RedirectResponse(url=source, status_code=307)

    if not os.path.exists(source):
        raise HTTPException(status_code=404, detail="音色参考音频文件不存在")

    media_type = mimetypes.guess_type(source)[0] or "application/octet-stream"
    inline_name = os.path.basename(source)
    return FileResponse(
        source,
        media_type=media_type,
        headers={"Content-Disposition": f'inline; filename="{inline_name}"'}
    )


async def delete_voiceover_voice_reference(
    episode_id: int,
    reference_id: str,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _, script = _ensure_voiceover_permission(episode_id, user, db)
    target_id = str(reference_id or "").strip()
    if not target_id:
        raise HTTPException(status_code=400, detail="reference_id不能为空")

    shared = _load_script_voiceover_shared_data(script)
    before = len(shared.get("voice_references", []))
    shared["voice_references"] = [
        item for item in shared.get("voice_references", [])
        if str(item.get("id") or "").strip() != target_id
    ]
    if len(shared["voice_references"]) == before:
        raise HTTPException(status_code=404, detail="音色参考音频不存在")

    fallback_ref_id = _voiceover_first_reference_id(shared)
    _save_script_voiceover_shared_data(script, shared)
    updated_line_count = _replace_voice_reference_for_script_episodes(
        db, script.id, target_id, fallback_ref_id
    )
    db.commit()

    return {
        "success": True,
        "shared": _load_script_voiceover_shared_data(script),
        "fallback_voice_reference_id": fallback_ref_id,
        "updated_line_count": updated_line_count
    }


async def upsert_voiceover_vector_preset(
    episode_id: int,
    request: dict,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _, script = _ensure_voiceover_permission(episode_id, user, db)
    name = str(request.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="预设名称不能为空")

    preset_id = str(request.get("id") or "").strip() or f"vector_preset_{uuid.uuid4().hex}"
    vector_config = _normalize_voiceover_vector_config(request.get("vector_config"))
    description = str(request.get("description") or "").strip()

    shared = _load_script_voiceover_shared_data(script)
    presets = shared.get("vector_presets", [])
    updated = False
    now_iso = datetime.utcnow().isoformat()
    for item in presets:
        if str(item.get("id") or "").strip() == preset_id:
            item["name"] = name
            item["description"] = description
            item["vector_config"] = vector_config
            updated = True
            break
    if not updated:
        presets.append({
            "id": preset_id,
            "name": name,
            "description": description,
            "vector_config": vector_config,
            "created_at": now_iso
        })
    shared["vector_presets"] = presets
    _save_script_voiceover_shared_data(script, shared)
    db.commit()

    return {"success": True, "preset_id": preset_id, "shared": _load_script_voiceover_shared_data(script)}


async def delete_voiceover_vector_preset(
    episode_id: int,
    preset_id: str,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _, script = _ensure_voiceover_permission(episode_id, user, db)
    target_id = str(preset_id or "").strip()
    if not target_id:
        raise HTTPException(status_code=400, detail="preset_id不能为空")

    shared = _load_script_voiceover_shared_data(script)
    before = len(shared.get("vector_presets", []))
    shared["vector_presets"] = [
        item for item in shared.get("vector_presets", [])
        if str(item.get("id") or "").strip() != target_id
    ]
    if len(shared["vector_presets"]) == before:
        raise HTTPException(status_code=404, detail="向量预设不存在")

    _save_script_voiceover_shared_data(script, shared)
    updated_line_count = _clear_tts_field_for_script_episodes(
        db, script.id, "vector_preset_id", target_id
    )
    db.commit()

    return {
        "success": True,
        "shared": _load_script_voiceover_shared_data(script),
        "updated_line_count": updated_line_count
    }


async def create_voiceover_emotion_audio_preset(
    episode_id: int,
    name: str = Form(...),
    description: str = Form(""),
    file: UploadFile = File(...),
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _, script = _ensure_voiceover_permission(episode_id, user, db)
    preset_name = str(name or "").strip()
    if not preset_name:
        raise HTTPException(status_code=400, detail="情感参考音频名称不能为空")

    cdn_url = save_and_upload_to_cdn(file)
    shared = _load_script_voiceover_shared_data(script)
    item = {
        "id": f"emotion_audio_preset_{uuid.uuid4().hex}",
        "name": preset_name,
        "description": str(description or "").strip(),
        "file_name": str(file.filename or "").strip(),
        "url": cdn_url,
        "local_path": "",
        "created_at": datetime.utcnow().isoformat()
    }
    shared["emotion_audio_presets"].append(item)
    _save_script_voiceover_shared_data(script, shared)
    db.commit()

    return {"success": True, "item": item, "shared": _load_script_voiceover_shared_data(script)}


async def delete_voiceover_emotion_audio_preset(
    episode_id: int,
    preset_id: str,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _, script = _ensure_voiceover_permission(episode_id, user, db)
    target_id = str(preset_id or "").strip()
    if not target_id:
        raise HTTPException(status_code=400, detail="preset_id不能为空")

    shared = _load_script_voiceover_shared_data(script)
    before = len(shared.get("emotion_audio_presets", []))
    shared["emotion_audio_presets"] = [
        item for item in shared.get("emotion_audio_presets", [])
        if str(item.get("id") or "").strip() != target_id
    ]
    if len(shared["emotion_audio_presets"]) == before:
        raise HTTPException(status_code=404, detail="情感音频预设不存在")

    _save_script_voiceover_shared_data(script, shared)
    updated_line_count = _clear_tts_field_for_script_episodes(
        db, script.id, "emotion_audio_preset_id", target_id
    )
    db.commit()

    return {
        "success": True,
        "shared": _load_script_voiceover_shared_data(script),
        "updated_line_count": updated_line_count
    }


async def upsert_voiceover_setting_template(
    episode_id: int,
    request: dict,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _, script = _ensure_voiceover_permission(episode_id, user, db)

    name = str(request.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="模板名称不能为空")

    shared = _load_script_voiceover_shared_data(script)
    default_voice_ref_id = _voiceover_first_reference_id(shared)
    settings = _normalize_voiceover_setting_template_payload(
        request.get("settings"),
        default_voice_ref_id
    )

    templates = shared.get("setting_templates", [])
    if not isinstance(templates, list):
        templates = []

    target_id = str(request.get("id") or "").strip()
    target_item = None
    if target_id:
        target_item = next(
            (item for item in templates if str(item.get("id") or "").strip() == target_id),
            None
        )
    if not target_item:
        target_item = next(
            (item for item in templates if str(item.get("name") or "").strip() == name),
            None
        )

    now_iso = datetime.utcnow().isoformat()
    if target_item:
        target_item["name"] = name
        target_item["settings"] = settings
        target_item["updated_at"] = now_iso
        if not str(target_item.get("created_at") or "").strip():
            target_item["created_at"] = now_iso
        target_id = str(target_item.get("id") or "").strip() or f"setting_template_{uuid.uuid4().hex}"
        target_item["id"] = target_id
    else:
        target_id = target_id or f"setting_template_{uuid.uuid4().hex}"
        templates.append({
            "id": target_id,
            "name": name,
            "settings": settings,
            "created_at": now_iso,
            "updated_at": now_iso
        })

    shared["setting_templates"] = templates
    _save_script_voiceover_shared_data(script, shared)
    db.commit()

    return {
        "success": True,
        "template_id": target_id,
        "shared": _load_script_voiceover_shared_data(script)
    }


async def delete_voiceover_setting_template(
    episode_id: int,
    template_id: str,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _, script = _ensure_voiceover_permission(episode_id, user, db)
    target_id = str(template_id or "").strip()
    if not target_id:
        raise HTTPException(status_code=400, detail="template_id不能为空")

    shared = _load_script_voiceover_shared_data(script)
    before = len(shared.get("setting_templates", []))
    shared["setting_templates"] = [
        item for item in shared.get("setting_templates", [])
        if str(item.get("id") or "").strip() != target_id
    ]
    if len(shared["setting_templates"]) == before:
        raise HTTPException(status_code=404, detail="参数模板不存在")

    _save_script_voiceover_shared_data(script, shared)
    db.commit()
    return {
        "success": True,
        "shared": _load_script_voiceover_shared_data(script)
    }


async def enqueue_voiceover_line_generate(
    episode_id: int,
    line_id: str,
    request: dict,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    episode, script = _ensure_voiceover_permission(episode_id, user, db)
    target_line_id = str(line_id or "").strip()
    if not target_line_id:
        raise HTTPException(status_code=400, detail="line_id不能为空")

    shared = _load_script_voiceover_shared_data(script)
    refs = shared.get("voice_references", [])
    default_voice_ref_id = _voiceover_first_reference_id(shared)

    voiceover_payload = _parse_episode_voiceover_payload(episode)
    shots, changed = _normalize_voiceover_shots_for_tts(
        voiceover_payload.get("shots", []),
        default_voice_ref_id
    )
    line_entry = _find_voiceover_line_entry(shots, target_line_id)
    if not isinstance(line_entry, dict):
        raise HTTPException(status_code=404, detail=f"未找到 line_id={target_line_id}")

    line_tts = _normalize_voiceover_line_tts(line_entry.get("tts"), default_voice_ref_id)

    line_text = str(request.get("text") or line_entry.get("text") or "").strip()
    if not line_text:
        raise HTTPException(status_code=400, detail="配音文本为空")

    method = str(
        request.get("emotion_control_method")
        or line_tts.get("emotion_control_method")
        or VOICEOVER_TTS_METHOD_SAME
    ).strip()
    if method not in VOICEOVER_TTS_ALLOWED_METHODS:
        method = VOICEOVER_TTS_METHOD_SAME

    voice_reference_id = str(
        request.get("voice_reference_id")
        or line_tts.get("voice_reference_id")
        or default_voice_ref_id
    ).strip()
    if not voice_reference_id:
        raise HTTPException(status_code=400, detail="请先选择音色参考音频")

    selected_voice_ref = None
    if isinstance(refs, list):
        selected_voice_ref = next((x for x in refs if str(x.get("id") or "").strip() == voice_reference_id), None)
    if not selected_voice_ref:
        raise HTTPException(status_code=400, detail="音色参考音频不存在")

    emotion_audio_preset_id = ""
    if method == VOICEOVER_TTS_METHOD_AUDIO:
        emotion_audio_preset_id = str(
            request.get("emotion_audio_preset_id")
            or line_tts.get("emotion_audio_preset_id")
            or ""
        ).strip()
        if not emotion_audio_preset_id:
            raise HTTPException(status_code=400, detail="请先选择情感参考音频预设")
        emotion_presets = shared.get("emotion_audio_presets", [])
        selected_emotion = None
        if isinstance(emotion_presets, list):
            selected_emotion = next(
                (x for x in emotion_presets if str(x.get("id") or "").strip() == emotion_audio_preset_id),
                None
            )
        if not selected_emotion:
            raise HTTPException(status_code=400, detail="情感参考音频预设不存在")

    vector_preset_id = str(request.get("vector_preset_id") or line_tts.get("vector_preset_id") or "").strip()
    vector_config = _normalize_voiceover_vector_config(
        request.get("vector_config") or line_tts.get("vector_config")
    )
    emo_text = str(
        request.get("emo_text")
        if request.get("emo_text") is not None
        else line_entry.get("emotion")
        or ""
    ).strip()

    task_payload = {
        "text": line_text,
        "emo_text": emo_text,
        "emotion_control_method": method,
        "voice_reference_id": voice_reference_id,
        "vector_preset_id": vector_preset_id,
        "emotion_audio_preset_id": emotion_audio_preset_id,
        "vector_config": vector_config
    }

    task = models.VoiceoverTtsTask(
        episode_id=episode.id,
        line_id=target_line_id,
        status="pending",
        request_json=json.dumps(task_payload, ensure_ascii=False),
        result_json="",
        error_message=""
    )
    db.add(task)
    db.flush()

    line_tts["emotion_control_method"] = method
    line_tts["voice_reference_id"] = voice_reference_id
    line_tts["vector_preset_id"] = vector_preset_id
    line_tts["emotion_audio_preset_id"] = emotion_audio_preset_id
    line_tts["vector_config"] = vector_config
    line_tts["generate_status"] = "pending"
    line_tts["generate_error"] = ""
    line_tts["latest_task_id"] = str(task.id)
    line_entry["tts"] = line_tts

    voiceover_payload["shots"] = shots
    episode.voiceover_data = json.dumps(voiceover_payload, ensure_ascii=False)
    db.commit()
    sync_voiceover_tts_task_to_dashboard(task.id)

    queue_position = db.query(func.count(models.VoiceoverTtsTask.id)).filter(
        models.VoiceoverTtsTask.status.in_(["pending", "processing"]),
        models.VoiceoverTtsTask.id <= task.id
    ).scalar() or 1

    return {
        "success": True,
        "task_id": task.id,
        "line_id": target_line_id,
        "status": "pending",
        "queue_position": int(queue_position)
    }


async def enqueue_voiceover_generate_all(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    episode, script = _ensure_voiceover_permission(episode_id, user, db)

    shared = _load_script_voiceover_shared_data(script)
    refs = shared.get("voice_references", [])
    default_voice_ref_id = _voiceover_first_reference_id(shared)
    ref_id_set = {
        str(item.get("id") or "").strip()
        for item in refs
        if isinstance(item, dict) and str(item.get("id") or "").strip()
    }
    emotion_presets = shared.get("emotion_audio_presets", [])
    emotion_preset_id_set = {
        str(item.get("id") or "").strip()
        for item in emotion_presets
        if isinstance(item, dict) and str(item.get("id") or "").strip()
    }

    voiceover_payload = _parse_episode_voiceover_payload(episode)
    shots, _ = _normalize_voiceover_shots_for_tts(
        voiceover_payload.get("shots", []),
        default_voice_ref_id
    )

    enqueued_line_ids = []
    skipped = []
    seen_line_ids = set()

    created_task_ids = []
    for shot in shots:
        if not isinstance(shot, dict):
            continue

        line_entries = []
        narration = shot.get("narration")
        if isinstance(narration, dict):
            line_entries.append(narration)

        dialogue = shot.get("dialogue")
        if isinstance(dialogue, list):
            for item in dialogue:
                if isinstance(item, dict):
                    line_entries.append(item)

        for line_entry in line_entries:
            line_id = str(line_entry.get("line_id") or "").strip()
            if not line_id:
                skipped.append({"line_id": "", "reason": "line_id缺失"})
                continue
            if line_id in seen_line_ids:
                skipped.append({"line_id": line_id, "reason": "line_id重复"})
                continue
            seen_line_ids.add(line_id)

            line_text = str(line_entry.get("text") or "").strip()
            if not line_text:
                skipped.append({"line_id": line_id, "reason": "配音文本为空"})
                continue

            line_tts = _normalize_voiceover_line_tts(line_entry.get("tts"), default_voice_ref_id)
            status = str(line_tts.get("generate_status") or "").strip().lower()
            if status in {"pending", "processing"}:
                skipped.append({"line_id": line_id, "reason": "已在队列中或生成中"})
                continue

            method = str(line_tts.get("emotion_control_method") or VOICEOVER_TTS_METHOD_SAME).strip()
            if method not in VOICEOVER_TTS_ALLOWED_METHODS:
                method = VOICEOVER_TTS_METHOD_SAME

            voice_reference_id = str(
                line_tts.get("voice_reference_id") or default_voice_ref_id
            ).strip()
            if not voice_reference_id:
                skipped.append({"line_id": line_id, "reason": "未选择音色参考音频"})
                continue
            if ref_id_set and voice_reference_id not in ref_id_set:
                skipped.append({"line_id": line_id, "reason": "音色参考音频不存在"})
                continue

            emotion_audio_preset_id = ""
            if method == VOICEOVER_TTS_METHOD_AUDIO:
                emotion_audio_preset_id = str(line_tts.get("emotion_audio_preset_id") or "").strip()
                if not emotion_audio_preset_id:
                    skipped.append({"line_id": line_id, "reason": "未选择情感参考音频预设"})
                    continue
                if emotion_preset_id_set and emotion_audio_preset_id not in emotion_preset_id_set:
                    skipped.append({"line_id": line_id, "reason": "情感参考音频预设不存在"})
                    continue

            vector_preset_id = str(line_tts.get("vector_preset_id") or "").strip()
            vector_config = _normalize_voiceover_vector_config(line_tts.get("vector_config"))
            emo_text = str(line_entry.get("emotion") or "").strip()

            task_payload = {
                "text": line_text,
                "emo_text": emo_text,
                "emotion_control_method": method,
                "voice_reference_id": voice_reference_id,
                "vector_preset_id": vector_preset_id,
                "emotion_audio_preset_id": emotion_audio_preset_id,
                "vector_config": vector_config
            }
            task = models.VoiceoverTtsTask(
                episode_id=episode.id,
                line_id=line_id,
                status="pending",
                request_json=json.dumps(task_payload, ensure_ascii=False),
                result_json="",
                error_message=""
            )
            db.add(task)
            db.flush()
            created_task_ids.append(int(task.id))

            line_tts["emotion_control_method"] = method
            line_tts["voice_reference_id"] = voice_reference_id
            line_tts["vector_preset_id"] = vector_preset_id
            line_tts["emotion_audio_preset_id"] = emotion_audio_preset_id
            line_tts["vector_config"] = vector_config
            line_tts["generate_status"] = "pending"
            line_tts["generate_error"] = ""
            line_tts["latest_task_id"] = str(task.id)
            line_entry["tts"] = line_tts

            enqueued_line_ids.append(line_id)

    voiceover_payload["shots"] = shots
    episode.voiceover_data = json.dumps(voiceover_payload, ensure_ascii=False)
    db.commit()
    for created_task_id in created_task_ids:
        sync_voiceover_tts_task_to_dashboard(created_task_id)

    pending_count = db.query(func.count(models.VoiceoverTtsTask.id)).filter(
        models.VoiceoverTtsTask.status == "pending"
    ).scalar() or 0
    processing_count = db.query(func.count(models.VoiceoverTtsTask.id)).filter(
        models.VoiceoverTtsTask.status == "processing"
    ).scalar() or 0

    return {
        "success": True,
        "enqueued_count": len(enqueued_line_ids),
        "skipped_count": len(skipped),
        "enqueued_line_ids": enqueued_line_ids,
        "skipped": skipped,
        "queue": {
            "pending": int(pending_count),
            "processing": int(processing_count)
        }
    }


def get_voiceover_tts_status(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    episode, script = _ensure_voiceover_permission(episode_id, user, db)
    shared = _load_script_voiceover_shared_data(script)
    default_voice_ref_id = _voiceover_first_reference_id(shared)

    payload = _parse_episode_voiceover_payload(episode)
    shots, changed = _normalize_voiceover_shots_for_tts(payload.get("shots", []), default_voice_ref_id)
    if changed:
        payload["shots"] = shots
        episode.voiceover_data = json.dumps(payload, ensure_ascii=False)
        db.commit()

    line_states = _extract_voiceover_tts_line_states(shots)
    pending_count = db.query(func.count(models.VoiceoverTtsTask.id)).filter(
        models.VoiceoverTtsTask.status == "pending"
    ).scalar() or 0
    processing_count = db.query(func.count(models.VoiceoverTtsTask.id)).filter(
        models.VoiceoverTtsTask.status == "processing"
    ).scalar() or 0

    return {
        "success": True,
        "line_states": line_states,
        "queue": {
            "pending": int(pending_count),
            "processing": int(processing_count)
        }
    }


def get_episode_storyboard(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取片段的分镜表数据（优先从episode.storyboard_data读取完整AI生成数据）"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # ✅ 获取主体库（总是从数据库加载最新的subjects，包括用户后来添加的）
    library = db.query(models.StoryLibrary).filter(
        models.StoryLibrary.episode_id == episode.id
    ).first()

    stored_subject_map = {}
    if episode.storyboard_data:
        try:
            stored_subject_map = _build_subject_detail_map(json.loads(episode.storyboard_data).get("subjects", []))
        except Exception:
            stored_subject_map = {}

    # 构建完整的subjects列表
    subjects = []
    if library:
        cards = db.query(models.SubjectCard).filter(
            models.SubjectCard.library_id == library.id
        ).all()
        cards = [card for card in cards if card.card_type in ALLOWED_CARD_TYPES]
        for card in cards:
            stored_subject = stored_subject_map.get((card.name, card.card_type), {})
            subjects.append({
                "id": card.id,  # ✅ 添加id字段，前端需要用它匹配selected_card_ids
                "name": card.name,
                "card_type": card.card_type,
                "type": card.card_type,
                "ai_prompt": (card.ai_prompt or "").strip() or stored_subject.get("ai_prompt", ""),
                "role_personality": (getattr(card, "role_personality", "") or "").strip() or stored_subject.get("role_personality", ""),
                "alias": (card.alias or "").strip() or stored_subject.get("alias", "")
            })

    # ✅ 优先从 episode.storyboard_data 读取完整的 AI 生成数据
    if episode.storyboard_data:
        try:
            data = json.loads(episode.storyboard_data)
            # 验证数据格式
            if isinstance(data, dict) and "shots" in data:
                # ✅ 查询数据库中最新的 selected_card_ids（用户在故事板界面选择的主体）
                shots_records = db.query(models.StoryboardShot).filter(
                    models.StoryboardShot.episode_id == episode_id,
                    models.StoryboardShot.variant_index == 0  # 只查询主镜头
                ).all()

                # 建立镜头号到最新 selected_card_ids 的映射
                shot_card_ids_map = {}
                shot_id_map = {}  # ✅ 添加：镜头号到数据库ID的映射
                for shot_record in shots_records:
                    # 同时存储整数和字符串形式的镜头号，确保能匹配
                    shot_card_ids_map[shot_record.shot_number] = shot_record.selected_card_ids
                    shot_card_ids_map[str(shot_record.shot_number)] = shot_record.selected_card_ids
                    shot_id_map[shot_record.shot_number] = shot_record.id  # ✅ 保存数据库ID
                    shot_id_map[str(shot_record.shot_number)] = shot_record.id

                # 创建卡片ID到卡片对象的映射
                card_map = {}
                if library:
                    cards = db.query(models.SubjectCard).filter(
                        models.SubjectCard.library_id == library.id
                    ).all()
                    card_map = {card.id: card for card in cards if card.card_type in ALLOWED_CARD_TYPES}
                card_name_to_id = {
                    (card.name, card.card_type): card.id
                    for card in card_map.values()
                }
                storyboard_subject_map = _build_subject_detail_map(data.get("subjects", []))

                # ✅ 格式化shots数据，将dialogue和narration转换为前端期望的字符串格式
                formatted_shots = []
                for shot in data.get("shots", []):
                    formatted_shot = shot.copy()

                    # ✅ 添加数据库ID
                    shot_number = shot.get('shot_number')
                    if shot_number and shot_number in shot_id_map:
                        formatted_shot['id'] = shot_id_map[shot_number]

                    # ✅ 用数据库中最新的 selected_card_ids 替换旧数据，并转换为 subjects 数组
                    if shot_number and shot_number in shot_card_ids_map:
                        selected_card_ids_json = shot_card_ids_map[shot_number]
                        formatted_shot['selected_card_ids'] = selected_card_ids_json

                        # ✅ 将 selected_card_ids 转换为 subjects 数组（前端渲染需要）
                        try:
                            selected_ids = json.loads(selected_card_ids_json or "[]")
                            shot_subjects = []
                            for card_id in selected_ids:
                                if card_id in card_map:
                                    card = card_map[card_id]
                                    shot_subjects.append({
                                        "name": card.name,
                                        "type": card.card_type
                                    })
                            fallback_subjects = _reconcile_storyboard_shot_subjects(
                                formatted_shot,
                                storyboard_subject_map,
                            )
                            existing_subject_keys = {
                                ((subject.get("name") or "").strip(), (subject.get("type") or "角色").strip() or "角色")
                                for subject in shot_subjects
                                if isinstance(subject, dict)
                            }
                            merged_selected_ids = list(selected_ids)
                            for fallback_subject in fallback_subjects:
                                fallback_key = (
                                    (fallback_subject.get("name") or "").strip(),
                                    (fallback_subject.get("type") or "角色").strip() or "角色",
                                )
                                if fallback_key in existing_subject_keys:
                                    continue
                                existing_subject_keys.add(fallback_key)
                                shot_subjects.append({
                                    "name": fallback_key[0],
                                    "type": fallback_key[1],
                                })
                                fallback_card_id = card_name_to_id.get(fallback_key)
                                if fallback_card_id and fallback_card_id not in merged_selected_ids:
                                    merged_selected_ids.append(fallback_card_id)
                            if merged_selected_ids != selected_ids:
                                formatted_shot['selected_card_ids'] = json.dumps(merged_selected_ids, ensure_ascii=False)
                            formatted_shot['subjects'] = shot_subjects
                        except Exception as e:
                            print(f"[获取分镜表] 转换 selected_card_ids 失败: {str(e)}")
                            # 保留原有的 subjects（如果有的话）
                            if 'subjects' not in formatted_shot:
                                formatted_shot['subjects'] = []
                    elif 'subjects' not in formatted_shot:
                        formatted_shot['subjects'] = _reconcile_storyboard_shot_subjects(
                            formatted_shot,
                            storyboard_subject_map,
                        )

                    # 格式化dialogue字段为可读字符串（同时保留原始配音字段）
                    voice_type = shot.get('voice_type', 'none')

                    # ✅ 保留原始配音字段
                    formatted_shot['voice_type'] = shot.get('voice_type')
                    formatted_shot['narration'] = shot.get('narration')
                    formatted_shot['dialogue_array'] = shot.get('dialogue')  # 原始对白数组

                    # 格式化为可读字符串（用于表格显示）
                    if voice_type == 'narration':
                        narration = shot.get('narration')
                        if narration and isinstance(narration, dict):
                            speaker = narration.get('speaker', '')
                            gender = narration.get('gender', '')
                            emotion = narration.get('emotion', '')
                            text = narration.get('text', '')
                            formatted_shot['dialogue'] = f"旁白（{speaker}/{gender}/{emotion}）：{text}"
                        else:
                            formatted_shot['dialogue'] = ""

                    elif voice_type == 'dialogue':
                        dialogue = shot.get('dialogue')
                        if dialogue and isinstance(dialogue, list):
                            dialogue_lines = []
                            for d in dialogue:
                                speaker = d.get('speaker', '')
                                gender = d.get('gender', '')
                                target = d.get('target')
                                emotion = d.get('emotion', '')
                                text = d.get('text', '')
                                if target:
                                    dialogue_lines.append(f"{speaker}（{gender}）对{target}说（{emotion}）：{text}")
                                else:
                                    dialogue_lines.append(f"{speaker}（{gender}）说（{emotion}）：{text}")
                            formatted_shot['dialogue'] = '\n'.join(dialogue_lines)
                        else:
                            formatted_shot['dialogue'] = ""

                    else:
                        # voice_type为none或其他值时，dialogue应该为空
                        if not isinstance(shot.get('dialogue'), str):
                            formatted_shot['dialogue'] = ""

                    formatted_shots.append(formatted_shot)

                # ✅ 返回最新的subjects列表（包含用户后来添加的主体）
                return {
                    "shots": formatted_shots,
                    "subjects": subjects,  # 使用从数据库查询的最新subjects
                    "generating": episode.storyboard_generating or False,
                    "error": episode.storyboard_error or ""
                }
        except json.JSONDecodeError:
            # JSON解析失败，继续使用后备方案
            pass

    # ❌ 后备方案：从storyboard_shots表重建（会丢失voice_type、narration等数据）
    shots_records = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id,
    ).order_by(
        models.StoryboardShot.shot_number.asc(),
        models.StoryboardShot.variant_index.asc()
    ).all()

    # ✅ 使用前面已经查询的subjects和library（避免重复查询）
    card_map = {}
    if library:
        cards = db.query(models.SubjectCard).filter(
            models.SubjectCard.library_id == library.id
        ).all()
        cards = [card for card in cards if card.card_type in ALLOWED_CARD_TYPES]
        card_map = {card.id: card for card in cards}

    # 构建shots列表（转换为前端期望的格式）
    shots = []
    seen_shot_numbers = set()  # ✅ 用于去重，确保每个镜头号只显示一次

    for shot_record in shots_records:
        # 只显示主镜头（variant_index=0）
        if shot_record.variant_index != 0:
            continue

        # ✅ 跳过已经处理过的镜头号（去重）
        if shot_record.shot_number in seen_shot_numbers:
            continue
        seen_shot_numbers.add(shot_record.shot_number)

        # 解析selected_card_ids
        try:
            selected_ids = json.loads(shot_record.selected_card_ids or "[]")
        except:
            selected_ids = []

        # 构建subjects数组
        shot_subjects = []
        for card_id in selected_ids:
            if card_id in card_map:
                card = card_map[card_id]
                shot_subjects.append({
                    "name": card.name,
                    "type": card.card_type
                })

        shots.append({
            "shot_number": str(shot_record.shot_number),
            "subjects": shot_subjects,
            "original_text": shot_record.script_excerpt or "",
            "dialogue": shot_record.storyboard_dialogue or "",
            "storyboard_prompt": shot_record.sora_prompt or ""
        })

    return {
        "shots": shots,
        "subjects": subjects,
        "generating": episode.storyboard_generating or False,
        "error": episode.storyboard_error or ""
    }


def get_episode_storyboard_status(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    episode = _verify_episode_permission(episode_id, user, db)
    return {
        "generating": bool(episode.storyboard_generating),
        "error": episode.storyboard_error or "",
        "shots_count": _count_storyboard_items(episode.storyboard_data),
    }





def _analyze_storyboard_changes(episode_id: int, new_storyboard_data: dict, db: Session):
    """
    分析分镜表变更（和原始 JSON 数据比对）

    返回格式：
    {
        "modified": [{"shot_number": 1, "reason": "修改了原剧本段落", "has_video": True}, ...],
        "deleted": [{"shot_number": 2, "has_video": False}, ...],
        "added": [3, 4, ...]
    }
    """
    new_shots = new_storyboard_data.get("shots", [])

    # 获取episode
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        return {"modified": [], "deleted": [], "added": []}

    # ✅ 从 Episode.storyboard_data（JSON）读取旧数据
    old_shots = []
    if episode.storyboard_data:
        try:
            old_data = json.loads(episode.storyboard_data)
            old_shots = old_data.get("shots", [])
        except:
            pass

    # 如果没有旧数据，说明是第一次保存，所有镜头都是新增
    if not old_shots:
        added = []
        for new_shot in new_shots:
            try:
                shot_number = int(new_shot.get("shot_number", 0))
                if shot_number > 0:
                    added.append(shot_number)
            except:
                pass
        return {"modified": [], "deleted": [], "added": added}

    # 构建旧数据的字典（按 shot_number 索引）
    old_shots_dict = {}
    for old_shot in old_shots:
        shot_number_str = old_shot.get("shot_number", "")
        try:
            shot_number = int(shot_number_str)
            old_shots_dict[shot_number] = old_shot
        except:
            continue

    # 查询数据库中的镜头（用于判断是否有视频）
    all_existing_shots = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id
    ).all()

    # 按shot_number分组
    shots_video_status = {}  # {shot_number: has_video}
    for shot in all_existing_shots:
        if shot.shot_number not in shots_video_status:
            shots_video_status[shot.shot_number] = False
        # 只要有任意一个变体有视频，就标记为 True
        if shot.video_status in ["processing", "completed"]:
            shots_video_status[shot.shot_number] = True

    new_shot_numbers = set()
    modified = []
    added = []

    # 检查修改和新增
    for new_shot in new_shots:
        shot_number_str = new_shot.get("shot_number", "")
        try:
            shot_number = int(shot_number_str)
        except:
            continue

        new_shot_numbers.add(shot_number)

        if shot_number in old_shots_dict:
            # 镜头已存在，比较内容
            old_shot = old_shots_dict[shot_number]
            changes = []

            # 1. 比较原剧本段落
            new_original_text = new_shot.get("original_text", "").strip()
            old_original_text = old_shot.get("original_text", "").strip()
            if new_original_text != old_original_text:
                changes.append("原剧本段落")

            # 2. 比较对白
            new_dialogue = new_shot.get("dialogue", "").strip()
            old_dialogue = old_shot.get("dialogue", "").strip()
            if new_dialogue != old_dialogue:
                changes.append("对白")

            # 3. 角色/场景的修改不算作修改（已移除比较逻辑）

            # 如果有变更
            if changes:
                has_video = shots_video_status.get(shot_number, False)
                modified.append({
                    "shot_number": shot_number,
                    "reason": "、".join(changes),
                    "has_video": has_video
                })
        else:
            # 新增镜头
            added.append(shot_number)

    # 检查删除
    deleted = []
    for shot_number, old_shot in old_shots_dict.items():
        if shot_number not in new_shot_numbers:
            has_video = shots_video_status.get(shot_number, False)
            deleted.append({
                "shot_number": shot_number,
                "has_video": has_video
            })

    return {
        "modified": modified,
        "deleted": deleted,
        "added": added
    }


async def update_episode_storyboard(
    episode_id: int,
    request: dict,
    analyze_only: bool = False,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """保存编辑后的分镜表数据"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # 如果只是分析变更，不保存
    if analyze_only:
        changes = _analyze_storyboard_changes(episode_id, request, db)
        return {"analyze_only": True, "changes": changes}

    # ✅ 先读取旧的 storyboard_data（用于比对）
    old_storyboard_data = None
    if episode.storyboard_data:
        try:
            old_storyboard_data = json.loads(episode.storyboard_data)
        except:
            pass

    subject_fallbacks = _build_subject_detail_map((old_storyboard_data or {}).get("subjects", []))
    library = db.query(models.StoryLibrary).filter(
        models.StoryLibrary.episode_id == episode.id
    ).first()
    if library:
        existing_cards = db.query(models.SubjectCard).filter(
            models.SubjectCard.library_id == library.id,
            models.SubjectCard.card_type.in_(ALLOWED_CARD_TYPES)
        ).all()
        for card in existing_cards:
            key = (card.name, card.card_type)
            fallback = subject_fallbacks.get(key, {})
            subject_fallbacks[key] = _normalize_subject_detail_entry({
                "name": card.name,
                "type": card.card_type,
                "alias": card.alias or "",
                "ai_prompt": card.ai_prompt or "",
                "role_personality": getattr(card, "role_personality", "") or ""
            }, fallback)

    incoming_subject_map = _build_subject_detail_map(request.get("subjects", []))
    if incoming_subject_map:
        merged_subjects = []
        for key, incoming_subject in incoming_subject_map.items():
            merged_subject = _normalize_subject_detail_entry(incoming_subject, subject_fallbacks.get(key))
            if merged_subject:
                merged_subjects.append(merged_subject)
        request["subjects"] = merged_subjects
    elif subject_fallbacks:
        request["subjects"] = list(subject_fallbacks.values())

    canonical_subject_map = _build_subject_detail_map(request.get("subjects", []))
    if canonical_subject_map and isinstance(request.get("shots"), list):
        reconciled_shots = []
        for shot in request.get("shots", []):
            if not isinstance(shot, dict):
                continue
            shot_copy = dict(shot)
            shot_copy["subjects"] = _reconcile_storyboard_shot_subjects(
                shot_copy,
                canonical_subject_map,
            )
            reconciled_shots.append(shot_copy)
        request["shots"] = reconciled_shots

    # ✅ 保存新的分镜表数据
    episode.storyboard_data = json.dumps(request, ensure_ascii=False)

    # ✅ 同步配音数据到 voiceover_data（基础字段来自分镜，保留已有扩展字段）
    voiceover_shots = []
    for shot in request.get("shots", []):
        voiceover_shots.append({
            "shot_number": shot.get("shot_number"),
            "voice_type": shot.get("voice_type"),
            "narration": shot.get("narration"),
            "dialogue": shot.get("dialogue")
        })

    merged_voiceover_data = _merge_voiceover_shots_preserving_extensions(
        episode.voiceover_data,
        voiceover_shots
    )
    episode.voiceover_data = json.dumps(merged_voiceover_data, ensure_ascii=False)
    print(f"[保存分镜表] 同步更新了配音表数据，共 {len(voiceover_shots)} 个镜头")

    db.commit()

    # ✅ 同步新主体到数据库
    _sync_subjects_to_database(episode_id, request, db)

    # ✅ 同步到StoryboardShot表（传入旧数据）
    _sync_storyboard_to_shots(episode_id, request, old_storyboard_data, db)

    return {"message": "分镜表已保存", "success": True}

async def create_from_storyboard(
    episode_id: int,
    request: Optional[CreateStoryboardRequest] = None,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """根据分镜表JSON创建主体和镜头（统一使用辅助函数）"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    if not episode.storyboard_data:
        raise HTTPException(status_code=400, detail="没有分镜表数据")

    # ✅ 调用统一的辅助函数
    try:
        storyboard = json.loads(episode.storyboard_data)
        shots_count = len(storyboard.get("shots", []))
        subjects_count = len(storyboard.get("subjects", []))

        _create_shots_from_storyboard_data(episode_id, db)

        return {
            "message": "创建成功",
            "created_subjects": subjects_count,
            "created_shots": shots_count
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"创建失败: {str(e)}")

update_card_prompt = subject_cards.update_card_prompt

# ==================== 提示词模板API ====================



# ==================== 绘图风格模板API ====================







# ==================== 视频风格模板API ====================
















# ==================== 分镜图模板API ====================

# 绘图要求模板 CRUD





# 绘画风格模板 CRUD





# ==================== 故事板镜头API ====================

ShotCreate = shot_schemas.ShotCreate
ShotUpdate = shot_schemas.ShotUpdate
ManualSoraPromptRequest = shot_schemas.ManualSoraPromptRequest
ShotResponse = shot_schemas.ShotResponse
ShotVideoResponse = shot_schemas.ShotVideoResponse
GenerateVideoRequest = shot_schemas.GenerateVideoRequest
ThumbnailUpdate = shot_schemas.ThumbnailUpdate
GenerateSoraPromptRequest = shot_schemas.GenerateSoraPromptRequest
GenerateLargeShotPromptRequest = shot_schemas.GenerateLargeShotPromptRequest
VideoStatusInfoResponse = shot_schemas.VideoStatusInfoResponse
BatchGenerateSoraPromptsRequest = episode_schemas.BatchGenerateSoraPromptsRequest
BatchGenerateSoraPromptsResponse = episode_schemas.BatchGenerateSoraPromptsResponse
BatchGenerateSoraVideosRequest = episode_schemas.BatchGenerateSoraVideosRequest
ManagedTaskResponse = episode_schemas.ManagedTaskResponse
StartManagedGenerationRequest = episode_schemas.StartManagedGenerationRequest
ManagedSessionStatusResponse = episode_schemas.ManagedSessionStatusResponse

async def create_shot(
    episode_id: int,
    shot: ShotCreate,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """创建或更新镜头"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # 检查镜头是否已存在（查询所有变体）
    existing_shots = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id,
        models.StoryboardShot.shot_number == shot.shot_number,
        models.StoryboardShot.variant_index == 0
    ).all()

    sound_card_ids_in_payload = "selected_sound_card_ids" in getattr(shot, "model_fields_set", set())
    normalized_sound_card_ids = None
    if sound_card_ids_in_payload:
        normalized_sound_card_ids = _normalize_storyboard_selected_sound_card_ids(
            shot.selected_sound_card_ids,
            episode_id,
            db
        )

    if existing_shots:
        # 更新所有变体的镜头（保持右侧数据同步）
        for existing_shot in existing_shots:
            existing_shot.prompt_template = shot.prompt_template
            existing_shot.storyboard_video_prompt = shot.storyboard_video_prompt
            existing_shot.storyboard_audio_prompt = shot.storyboard_audio_prompt
            existing_shot.storyboard_dialogue = shot.storyboard_dialogue
            existing_shot.sora_prompt = shot.sora_prompt
            existing_shot.selected_card_ids = json.dumps(shot.selected_card_ids)
            if sound_card_ids_in_payload:
                existing_shot.selected_sound_card_ids = (
                    json.dumps(normalized_sound_card_ids, ensure_ascii=False)
                    if normalized_sound_card_ids is not None
                    else None
                )
            existing_shot.aspect_ratio = shot.aspect_ratio
            existing_shot.duration = shot.duration
        db.commit()
        db.refresh(existing_shots[0])

        # 计算detail_images_status
        detail_images = db.query(models.ShotDetailImage).filter(
            models.ShotDetailImage.shot_id == existing_shots[0].id
        ).all()

        if not detail_images:
            existing_shots[0].detail_images_status = 'idle'
        else:
            has_completed = any(img.status == 'completed' for img in detail_images)
            has_processing = any(img.status == 'processing' for img in detail_images)
            has_pending = any(img.status == 'pending' for img in detail_images)
            all_failed = all(img.status == 'failed' for img in detail_images)

            if has_completed:
                # 只要有成功的，就算completed（前端可以预览，会显示部分成功）
                existing_shots[0].detail_images_status = 'completed'
            elif has_processing or has_pending:
                existing_shots[0].detail_images_status = 'processing'
            elif all_failed:
                existing_shots[0].detail_images_status = 'failed'
            else:
                existing_shots[0].detail_images_status = 'idle'

        existing_shots[0].selected_scene_image_url = _resolve_selected_scene_reference_image_url(existing_shots[0], db)
        return existing_shots[0]
    else:
        created_shots = []
        for _ in [None]:
            new_shot = models.StoryboardShot(
                episode_id=episode_id,
                shot_number=shot.shot_number,
                stable_id=str(uuid.uuid4()),  # 生成stable_id
                variant_index=0,
                prompt_template=shot.prompt_template,
                storyboard_video_prompt=shot.storyboard_video_prompt,
                storyboard_audio_prompt=shot.storyboard_audio_prompt,
                storyboard_dialogue=shot.storyboard_dialogue,
                sora_prompt=shot.sora_prompt,
                selected_card_ids=json.dumps(shot.selected_card_ids),
                selected_sound_card_ids=(
                    json.dumps(normalized_sound_card_ids, ensure_ascii=False)
                    if sound_card_ids_in_payload and normalized_sound_card_ids is not None
                    else None
                ),
                aspect_ratio=shot.aspect_ratio,
                duration=shot.duration,
                storyboard_video_model="",
                storyboard_video_model_override_enabled=False,
                duration_override_enabled=False,
            )
            db.add(new_shot)
            created_shots.append(new_shot)
        db.commit()
        db.refresh(new_shot)

        # 新创建的镜头没有detail_images
        new_shot.detail_images_status = 'idle'
        new_shot.selected_scene_image_url = _resolve_selected_scene_reference_image_url(new_shot, db)

        return new_shot

def _sync_processing_storyboard_videos_for_episode(
    episode_id: int,
    db: Session,
    max_count: int = 12,
) -> int:
    """
    兜底同步故事板视频状态。
    作用：当独立 poller 落后或重启后，前端拉取镜头列表时仍能推进 processing -> failed/completed。
    """
    candidate_rows = db.query(models.StoryboardShot.id).filter(
        models.StoryboardShot.episode_id == episode_id,
        models.StoryboardShot.task_id != "",
        models.StoryboardShot.video_status.in_(["submitting", "preparing", "processing"]),
    ).order_by(
        models.StoryboardShot.video_submitted_at.asc().nullsfirst(),
        models.StoryboardShot.id.asc(),
    ).limit(max_count).all()

    if not candidate_rows:
        return 0

    updated_count = 0
    for row in candidate_rows:
        shot_id = int(getattr(row, "id", row) or 0)
        if shot_id <= 0:
            continue
        try:
            shot = db.query(models.StoryboardShot).filter(
                models.StoryboardShot.id == shot_id
            ).first()
            if not shot:
                continue

            result = check_video_status(str(shot.task_id or "").strip())
            if is_transient_video_status_error(result):
                continue

            normalized_status = normalize_video_generation_status(
                result.get("status"),
                default_value="processing",
            )
            error_message = str(result.get("error_message") or "").strip()

            if normalized_status == "failed":
                target_video_path = f"error:{error_message}" if error_message else "error:任务失败"
                changed = False
                if (shot.video_status or "").strip().lower() != "failed":
                    shot.video_status = "failed"
                    changed = True
                if str(shot.video_path or "") != target_video_path:
                    shot.video_path = target_video_path
                    changed = True
                if str(shot.thumbnail_video_path or "").strip():
                    shot.thumbnail_video_path = ""
                    changed = True
                if str(shot.video_error_message or "") != error_message:
                    shot.video_error_message = error_message
                    changed = True
                if changed:
                    updated_count += 1
                continue

            if normalized_status == "completed":
                video_url = str(result.get("video_url") or "").strip()
                changed = False
                previous_video_path = str(shot.video_path or "")
                previous_thumbnail = str(shot.thumbnail_video_path or "")

                if (shot.video_status or "").strip().lower() != "completed":
                    shot.video_status = "completed"
                    changed = True
                if video_url and previous_video_path != video_url:
                    shot.video_path = video_url
                    changed = True
                if video_url and (
                    not previous_thumbnail or previous_thumbnail == previous_video_path
                ):
                    if previous_thumbnail != video_url:
                        shot.thumbnail_video_path = video_url
                        changed = True
                if str(shot.video_error_message or "").strip():
                    shot.video_error_message = ""
                    changed = True
                if changed:
                    updated_count += 1
                continue

            if normalized_status in {"pending", "processing"}:
                if (shot.video_status or "").strip().lower() != "processing":
                    shot.video_status = "processing"
                    updated_count += 1
                if not shot.video_submitted_at:
                    shot.video_submitted_at = datetime.utcnow()
                    updated_count += 1
                continue

        except Exception as e:
            print(f"[storyboard-video-sync] failed for shot {shot_id}: {str(e)}")

    if updated_count:
        db.commit()
        db.expire_all()
    return updated_count


def get_episode_shots(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取片段的所有镜头"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    _sync_processing_storyboard_videos_for_episode(episode_id, db)
    if _reconcile_episode_runtime_flags(episode, db):
        db.commit()


    # 构建查询
    query = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id
    )


    shots = query.order_by(
        models.StoryboardShot.shot_number.asc(),
        models.StoryboardShot.variant_index.asc()
    ).all()

    # ✅ 去重：确保每个(shot_number, variant_index)组合只保留一条记录
    seen_keys = set()
    unique_shots = []
    for shot in shots:
        key = (shot.shot_number, shot.variant_index)
        if key not in seen_keys:
            seen_keys.add(key)
            unique_shots.append(shot)
    shots = unique_shots

    shot_ids = [shot.id for shot in shots]
    managed_task_id_map = {}
    if shot_ids:
        latest_video_map = {}
        videos = db.query(models.ShotVideo).filter(
            models.ShotVideo.shot_id.in_(shot_ids)
        ).order_by(
            models.ShotVideo.shot_id.asc(),
            models.ShotVideo.created_at.desc()
        ).all()

        for video in videos:
            if video.shot_id not in latest_video_map:
                latest_video_map[video.shot_id] = video.video_path

        needs_commit = False
        for shot in shots:
            if not (shot.thumbnail_video_path or "").strip():
                latest_path = latest_video_map.get(shot.id)
                if latest_path:
                    shot.thumbnail_video_path = latest_path
                    needs_commit = True

        if needs_commit:
            db.commit()

        managed_tasks = db.query(models.ManagedTask).filter(
            models.ManagedTask.shot_id.in_(shot_ids),
            models.ManagedTask.task_id != ""
        ).order_by(
            models.ManagedTask.shot_id.asc(),
            models.ManagedTask.id.desc()
        ).all()
        for managed_task in managed_tasks:
            if managed_task.shot_id not in managed_task_id_map:
                managed_task_id_map[managed_task.shot_id] = managed_task.task_id or ""

    # 查询每个shot的detail_images状态
    detail_images_status_map = {}
    detail_images_preview_map = {}
    detail_images = []
    if shot_ids:
        # 批量查询所有shot的detail_images
        detail_images = db.query(models.ShotDetailImage).filter(
            models.ShotDetailImage.shot_id.in_(shot_ids)
        ).order_by(
            models.ShotDetailImage.shot_id.asc(),
            models.ShotDetailImage.sub_shot_index.asc(),
            models.ShotDetailImage.id.asc()
        ).all()

        # 按shot_id分组
        detail_images_by_shot = {}
        for detail_img in detail_images:
            if detail_img.shot_id not in detail_images_by_shot:
                detail_images_by_shot[detail_img.shot_id] = []
            detail_images_by_shot[detail_img.shot_id].append(detail_img)

        # 计算每个shot的状态
        for shot_id, imgs in detail_images_by_shot.items():
            # 优先检查是否有completed，如果有任何completed就可以预览
            has_completed = any(img.status == 'completed' for img in imgs)
            has_processing = any(img.status == 'processing' for img in imgs)
            has_pending = any(img.status == 'pending' for img in imgs)
            all_completed = all(img.status == 'completed' for img in imgs)
            all_failed = all(img.status == 'failed' for img in imgs)

            if has_completed:
                # 只要有成功的，就算completed（前端可以预览，会显示部分成功）
                detail_images_status_map[shot_id] = 'completed'
            elif has_processing or has_pending:
                # 没有成功的，但有进行中的
                detail_images_status_map[shot_id] = 'processing'
            elif all_failed:
                detail_images_status_map[shot_id] = 'failed'
            else:
                detail_images_status_map[shot_id] = 'idle'

            # 提取预览图：按子镜头顺序取第一条成功记录的第一张图
            preview_url = None
            for detail_img in imgs:
                if detail_img.status != 'completed' or not detail_img.images_json:
                    continue
                try:
                    parsed_images = json.loads(detail_img.images_json)
                except Exception:
                    parsed_images = []
                if not isinstance(parsed_images, list):
                    continue
                for image_url in parsed_images:
                    if isinstance(image_url, str) and image_url.strip():
                        preview_url = image_url.strip()
                        break
                if preview_url:
                    break

            if preview_url:
                detail_images_preview_map[shot_id] = preview_url

    # 添加detail_images_status和detail_images_progress到shot对象
    for shot in shots:
        shot.detail_images_status = detail_images_status_map.get(shot.id, 'idle')
        shot.detail_images_preview_path = detail_images_preview_map.get(shot.id)

        # 计算detail_images_progress
        shot_detail_images = [img for img in detail_images if img.shot_id == shot.id]
        if shot_detail_images:
            completed_count = sum(1 for img in shot_detail_images if img.status == 'completed')
            total_count = len(shot_detail_images)

            if shot.detail_images_status == 'processing':
                # 生成中，显示进度如"2/5"
                shot.detail_images_progress = f"{completed_count}/{total_count}"
            elif shot.detail_images_status == 'completed':
                # 已完成，判断是否全部成功
                if completed_count == total_count:
                    # 全部成功，不显示进度（前端会显示"点击查看"）
                    shot.detail_images_progress = None
                else:
                    # 部分失败，显示如"3/5"（有2个失败）
                    shot.detail_images_progress = f"{completed_count}/{total_count}"
            else:
                shot.detail_images_progress = None
        else:
            shot.detail_images_progress = None

    # 使用 model_validate 强制Pydantic读取动态属性
    response_shots = []
    for shot in shots:
        try:
            # 使用from_attributes确保动态属性被读取
            shot_dict = {
                'id': shot.id,
                'episode_id': shot.episode_id,
                'shot_number': shot.shot_number,
                'variant_index': shot.variant_index,
                'prompt_template': shot.prompt_template,
                'script_excerpt': shot.script_excerpt,
                'storyboard_video_prompt': shot.storyboard_video_prompt,
                'storyboard_audio_prompt': shot.storyboard_audio_prompt,
                'storyboard_dialogue': shot.storyboard_dialogue,
                'scene_override': shot.scene_override,
                'scene_override_locked': bool(getattr(shot, 'scene_override_locked', False)),
                'sora_prompt': shot.sora_prompt,
                'sora_prompt_status': shot.sora_prompt_status,
                'selected_card_ids': shot.selected_card_ids,
                'selected_sound_card_ids': shot.selected_sound_card_ids,
                'video_path': shot.video_path or "",
                'thumbnail_video_path': shot.thumbnail_video_path or "",
                'video_status': shot.video_status,
                'task_id': shot.task_id or "",
                'managed_task_id': managed_task_id_map.get(shot.id, ""),
                'aspect_ratio': shot.aspect_ratio,
                'duration': shot.duration,
                'storyboard_video_model': getattr(shot, 'storyboard_video_model', "") or "",
                'storyboard_video_model_override_enabled': bool(getattr(shot, 'storyboard_video_model_override_enabled', False)),
                'duration_override_enabled': bool(getattr(shot, 'duration_override_enabled', False)),
                'provider': shot.provider,
                'storyboard_image_path': shot.storyboard_image_path or "",
                'storyboard_image_status': shot.storyboard_image_status,
                'storyboard_image_task_id': shot.storyboard_image_task_id or "",
                'first_frame_reference_image_url': getattr(shot, 'first_frame_reference_image_url', "") or "",
                'uploaded_scene_image_url': getattr(shot, 'uploaded_scene_image_url', "") or "",
                'use_uploaded_scene_image': bool(getattr(shot, 'use_uploaded_scene_image', False)),
                'selected_scene_image_url': _resolve_selected_scene_reference_image_url(shot, db),
                'timeline_json': shot.timeline_json or "",
                'detail_image_prompt_overrides': shot.detail_image_prompt_overrides or "{}",
                'detail_images_status': shot.detail_images_status,
                'detail_images_progress': shot.detail_images_progress,
                'detail_images_preview_path': shot.detail_images_preview_path,
                'created_at': shot.created_at
            }
            shot_response = ShotResponse(**shot_dict)
            response_shots.append(shot_response)
        except Exception as e:
            print(f"[get_episode_shots] 错误: 转换shot对象失败 shot_id={shot.id}")
            print(f"[get_episode_shots] 错误详情: {str(e)}")
            print(f"[get_episode_shots] shot_dict: {shot_dict}")
            import traceback
            traceback.print_exc()
            # 如果转换失败，使用model_validate with默认值
            try:
                # 尝试使用from_attributes直接转换
                response_shots.append(ShotResponse.model_validate(shot, from_attributes=True))
            except:
                # 最后的尝试，手动设置所有可能为空的字段
                shot_dict['sora_prompt'] = shot_dict.get('sora_prompt') or ""
                shot_dict['video_path'] = shot_dict.get('video_path') or ""
                shot_dict['thumbnail_video_path'] = shot_dict.get('thumbnail_video_path') or ""
                response_shots.append(ShotResponse(**shot_dict))

    return response_shots


def get_shot_video_status_info(
    shot_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    shot = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.id == shot_id
    ).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if not script or script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    task_id = str((shot.task_id or "").strip() or "").strip()
    if not task_id:
        managed_task_row = (
            db.query(models.ManagedTask.task_id)
            .filter(
                models.ManagedTask.shot_id == shot.id,
                models.ManagedTask.task_id != ""
            )
            .order_by(models.ManagedTask.id.desc())
            .first()
        )
        task_id = str((managed_task_row[0] if managed_task_row else "") or "").strip()

    if not task_id:
        raise HTTPException(status_code=400, detail="当前镜头没有可查询的任务ID")

    result = check_video_status(task_id)
    if is_transient_video_status_error(result):
        raise HTTPException(status_code=503, detail=result.get("error_message") or "暂时无法获取任务信息")

    status = str(result.get("status") or "").strip().lower()
    progress = int(result.get("progress") or 0)
    info = str(result.get("info") or "").strip()
    error_message = str(result.get("error_message") or "").strip()

    if not info and status in {"pending", "queued", "processing"}:
        if progress > 0:
            info = f"当前任务正在处理中，进度 {progress}%"
        else:
            info = "当前任务正在处理中，服务商暂未返回更多排队信息。"

    return VideoStatusInfoResponse(
        shot_id=shot.id,
        task_id=task_id,
        status=status or "unknown",
        progress=progress,
        info=info,
        error_message=error_message
    )

async def update_shot(
    shot_id: int,
    shot_data: ShotUpdate,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """更新镜头"""
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # ✅ 找到所有变体的对应镜头（右侧数据需要同步）
    all_variant_shots = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == shot.episode_id,
        models.StoryboardShot.shot_number == shot.shot_number,
        models.StoryboardShot.variant_index == shot.variant_index
    ).all()
    sound_card_ids_in_payload = "selected_sound_card_ids" in getattr(shot_data, "model_fields_set", set())
    normalized_sound_card_ids = None
    if sound_card_ids_in_payload:
        normalized_sound_card_ids = _normalize_storyboard_selected_sound_card_ids(
            shot_data.selected_sound_card_ids,
            shot.episode_id,
            db
        )

    # ✅ 更新所有变体的镜头（保持右侧数据同步）
    for variant_shot in all_variant_shots:
        if shot_data.prompt_template is not None:
            variant_shot.prompt_template = shot_data.prompt_template
        if shot_data.script_excerpt is not None:
            variant_shot.script_excerpt = shot_data.script_excerpt
        if shot_data.storyboard_video_prompt is not None:
            variant_shot.storyboard_video_prompt = shot_data.storyboard_video_prompt
        if shot_data.storyboard_audio_prompt is not None:
            variant_shot.storyboard_audio_prompt = shot_data.storyboard_audio_prompt
        if shot_data.storyboard_dialogue is not None:
            variant_shot.storyboard_dialogue = shot_data.storyboard_dialogue
        if shot_data.scene_override is not None:
            variant_shot.scene_override = shot_data.scene_override
            # 🔒 用户手动保存场景描述后，锁定自动填充（不再跟随场景主体自动更新）
            variant_shot.scene_override_locked = True
        if shot_data.scene_override_locked is not None:
            variant_shot.scene_override_locked = bool(shot_data.scene_override_locked)
        if shot_data.sora_prompt is not None:
            variant_shot.sora_prompt = shot_data.sora_prompt
            # 打印保存的Sora提示词
            if variant_shot.id == shot.id:  # 只打印一次
                print("=" * 80)
                print(f"[保存Sora提示词] 镜头ID: {shot.id}, 镜号: {shot.shot_number}")
                print("-" * 80)
                print("完整的Sora提示词:")
                print(shot_data.sora_prompt)
                print("=" * 80)
        if shot_data.sora_prompt_status is not None:
            variant_shot.sora_prompt_status = str(shot_data.sora_prompt_status or "").strip() or "idle"
        if shot_data.selected_card_ids is not None:
            variant_shot.selected_card_ids = json.dumps(shot_data.selected_card_ids)
        if sound_card_ids_in_payload:
            variant_shot.selected_sound_card_ids = (
                json.dumps(normalized_sound_card_ids, ensure_ascii=False)
                if normalized_sound_card_ids is not None
                else None
            )
        if shot_data.aspect_ratio is not None:
            variant_shot.aspect_ratio = shot_data.aspect_ratio
        if shot_data.duration is not None:
            variant_shot.duration = shot_data.duration
        if shot_data.storyboard_video_model is not None:
            variant_shot.storyboard_video_model = _normalize_storyboard_video_model(
                shot_data.storyboard_video_model,
                default_model=getattr(episode, "storyboard_video_model", DEFAULT_STORYBOARD_VIDEO_MODEL)
            )
        if shot_data.storyboard_video_model_override_enabled is not None:
            variant_shot.storyboard_video_model_override_enabled = bool(
                shot_data.storyboard_video_model_override_enabled
            )
        if shot_data.duration_override_enabled is not None:
            variant_shot.duration_override_enabled = bool(shot_data.duration_override_enabled)
        if shot_data.provider is not None:
            variant_shot.provider = shot_data.provider
        if shot_data.storyboard_image_path is not None:
            variant_shot.storyboard_image_path = str(shot_data.storyboard_image_path or "").strip()
        if shot_data.storyboard_image_status is not None:
            variant_shot.storyboard_image_status = str(shot_data.storyboard_image_status or "").strip()
        if shot_data.storyboard_image_model is not None:
            variant_shot.storyboard_image_model = str(shot_data.storyboard_image_model or "").strip()
        if shot_data.first_frame_reference_image_url is not None:
            variant_shot.first_frame_reference_image_url = normalize_first_frame_candidate_url(
                shot_data.first_frame_reference_image_url
            )
        if shot_data.uploaded_scene_image_url is not None:
            variant_shot.uploaded_scene_image_url = str(shot_data.uploaded_scene_image_url or "").strip()
        if shot_data.use_uploaded_scene_image is not None:
            variant_shot.use_uploaded_scene_image = bool(shot_data.use_uploaded_scene_image)

        _backfill_storyboard_visual_references_from_family(variant_shot, db)
        _apply_episode_storyboard_video_settings_to_shot(variant_shot, episode)

    db.commit()
    db.refresh(shot)

    # ✅ 调试：打印返回的shot对象的sora_prompt
    print(f"[返回镜头数据] ID: {shot.id}, sora_prompt长度: {len(shot.sora_prompt or '')}")
    print(f"[返回内容预览] {(shot.sora_prompt or '')[:100]}")

    # 计算detail_images_status
    detail_images = db.query(models.ShotDetailImage).filter(
        models.ShotDetailImage.shot_id == shot.id
    ).all()

    if not detail_images:
        shot.detail_images_status = 'idle'
    else:
        has_completed = any(img.status == 'completed' for img in detail_images)
        has_processing = any(img.status == 'processing' for img in detail_images)
        has_pending = any(img.status == 'pending' for img in detail_images)
        all_failed = all(img.status == 'failed' for img in detail_images)

        if has_completed:
            # 只要有成功的，就算completed（前端可以预览，会显示部分成功）
            shot.detail_images_status = 'completed'
        elif has_processing or has_pending:
            shot.detail_images_status = 'processing'
        elif all_failed:
            shot.detail_images_status = 'failed'
        else:
            shot.detail_images_status = 'idle'

    shot.selected_scene_image_url = _resolve_selected_scene_reference_image_url(shot, db)
    return shot

async def extract_scene_from_cards(
    shot_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """从镜头的选中主体卡片中提取场景描述"""
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # 提取场景描述
    scene_description = extract_scene_description(shot, db)

    return {
        "scene_description": scene_description,
        "shot_id": shot_id
    }

async def delete_episode(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """删除剧集（级联删除所有关联镜头）"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="剧集不存在")

    # 验证权限
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="关联剧本不存在")
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    episode_cleanup_stats = _clear_episode_dependencies([episode_id], db)

    print(
        "[剧集删除清理] "
        f"episode_id={episode_id} "
        f"simple_batches={episode_cleanup_stats['deleted_simple_storyboard_batches']} "
        f"managed_tasks={episode_cleanup_stats['deleted_managed_tasks']} "
        f"managed_sessions={episode_cleanup_stats['deleted_managed_sessions']} "
        f"voiceover_tts_tasks={episode_cleanup_stats['deleted_voiceover_tts_tasks']} "
        f"unlinked_libraries={episode_cleanup_stats['unlinked_libraries']}"
    )

    # 删除剧集（会级联删除所有关联的镜头和视频记录）
    db.delete(episode)
    db.commit()

    return {"message": "剧集删除成功", "episode_id": episode_id}

async def delete_shot(
    shot_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """删除镜头。删除主镜头时删除整组变体，删除变体时仅删除当前变体。"""
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    # 验证权限
    episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    variant_index = int(getattr(shot, "variant_index", 0) or 0)
    if variant_index > 0:
        if _is_storyboard_shot_generation_active(shot, db):
            raise HTTPException(status_code=400, detail="当前变体镜头正在生成中，请等待完成")
        shot_ids_to_delete = [shot.id]
        deleted_scope = "variant"
    else:
        if _count_active_video_generations_for_shot_family(shot, db) > 0:
            raise HTTPException(status_code=400, detail="当前镜头或其变体正在生成中，请等待完成")
        shot_ids_to_delete = _get_storyboard_shot_family_ids(shot, db)
        deleted_scope = "family"

    deleted_count = _delete_storyboard_shots_by_ids(
        shot_ids_to_delete,
        db,
        log_context=(
            f"episode_id={shot.episode_id} "
            f"shot_number={shot.shot_number} "
            f"stable_id={str(getattr(shot, 'stable_id', '') or '').strip() or '<empty>'}"
        ),
        allow_zero=True
    )

    db.commit()

    return {
        "message": "镜头删除成功",
        "shot_id": shot_id,
        "deleted_count": deleted_count,
        "deleted_scope": deleted_scope,
    }

async def duplicate_shot(
    shot_id: int,
    request: dict = {},  # 新增：接收请求体
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """复制镜头生成新的变体"""
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    if _count_active_video_generations_for_shot_family(shot, db) > 0:
        raise HTTPException(status_code=400, detail="当前镜头已有正在生成中的视频，请等待完成")

    # 准备复制镜头


    # 查询该镜头号的最大变体序号
    max_variant = db.query(func.max(models.StoryboardShot.variant_index)).filter(
        models.StoryboardShot.episode_id == shot.episode_id,
        models.StoryboardShot.shot_number == shot.shot_number,
    ).scalar()
    next_variant = (max_variant or 0) + 1


    new_shot = models.StoryboardShot(
        **build_duplicate_shot_payload(
            shot,
            next_variant=next_variant,
        )
    )
    db.add(new_shot)
    db.flush()
    _backfill_storyboard_visual_references_from_family(new_shot, db)
    db.commit()
    db.refresh(new_shot)

    # 新复制的镜头没有detail_images（detail_images不会被复制）
    new_shot.detail_images_status = 'idle'
    new_shot.selected_scene_image_url = _resolve_selected_scene_reference_image_url(new_shot, db)

    return new_shot


_debug_parse_card_ids = storyboard_reference_assets.parse_card_ids


_debug_resolve_subject_names = storyboard_prompt_context.debug_resolve_subject_names


_resolve_selected_cards = storyboard_reference_assets.resolve_selected_cards


_parse_storyboard_sound_card_ids = storyboard_sound_cards.parse_storyboard_sound_card_ids
_get_episode_story_library = storyboard_sound_cards.get_episode_story_library
_normalize_storyboard_selected_sound_card_ids = storyboard_sound_cards.normalize_storyboard_selected_sound_card_ids
_resolve_storyboard_selected_sound_cards = storyboard_sound_cards.resolve_storyboard_selected_sound_cards


_build_subject_text_for_ai = storyboard_prompt_context.build_subject_text_for_ai
_build_storyboard2_subject_text = storyboard_prompt_context.build_storyboard2_subject_text
_resolve_large_shot_template = storyboard_prompt_context.resolve_large_shot_template
SORA_REFERENCE_PROMPT_INSTRUCTION = storyboard_prompt_context.SORA_REFERENCE_PROMPT_INSTRUCTION
_append_sora_reference_prompt = storyboard_prompt_context.append_sora_reference_prompt
_resolve_sora_reference_prompt = storyboard_prompt_context.resolve_sora_reference_prompt


def _build_storyboard_prompt_request_data(
    db: Session,
    *,
    shot: models.StoryboardShot,
    episode: models.Episode,
    script: models.Script,
    prompt_key: str = "generate_video_prompts",
    duration_template_field: str = "video_prompt_rule",
    large_shot_template_id: Optional[int] = None,
    reference_shot_id: Optional[int] = None,
):
    storyboard2_prompt_key = "generate_storyboard2_video_prompts"
    effective_video_settings = _apply_episode_storyboard_video_settings_to_shot(shot, episode)
    safe_duration = max(1, int(effective_video_settings["duration"] or 15))
    template_duration = 15 if safe_duration <= 15 else 25

    selected_ids = []
    try:
        selected_ids = json.loads(shot.selected_card_ids or "[]")
    except Exception:
        selected_ids = []

    selected_cards = _resolve_selected_cards(db, selected_ids)
    subject_names = [card.name for card in selected_cards if card and card.name]
    subject_text = _build_subject_text_for_ai(selected_cards)
    scene_text = (shot.scene_override or "").strip()
    custom_style = (script.sora_prompt_style or "").strip()
    template_field = (duration_template_field or "video_prompt_rule").strip() or "video_prompt_rule"
    excerpt = (shot.script_excerpt or "").strip()
    if not excerpt:
        raise ValueError("请先填写原剧本段落")

    large_shot_template_content = ""
    large_shot_template_name = ""
    if prompt_key == "generate_large_shot_prompts":
        large_shot_template = _resolve_large_shot_template(db, large_shot_template_id)
        if not large_shot_template:
            raise ValueError("大镜头模板不存在")
        large_shot_template_id = large_shot_template.id
        large_shot_template_name = (large_shot_template.name or "").strip()
        large_shot_template_content = (large_shot_template.content or "").strip()

    if custom_style:
        template_for_format = custom_style
        if prompt_key == "generate_large_shot_prompts":
            template_for_format = inject_large_shot_template_content(template_for_format, large_shot_template_content)
        try:
            prompt = template_for_format.format(
                script_excerpt=excerpt,
                scene_description=scene_text,
                subject_text=subject_text,
                safe_duration=safe_duration,
                extra_style="",
                large_shot_template_content=large_shot_template_content
            )
        except KeyError:
            prompt = template_for_format
    else:
        use_duration_template = prompt_key != storyboard2_prompt_key
        if use_duration_template:
            template = db.query(models.ShotDurationTemplate).filter(
                models.ShotDurationTemplate.duration == template_duration
            ).first()
            template_rule = str(getattr(template, template_field, "") or "").strip() if template else ""
            prompt_template = template_rule or get_prompt_by_key(prompt_key)
        else:
            prompt_template = get_prompt_by_key(prompt_key)
        template_for_format = prompt_template
        if prompt_key == "generate_large_shot_prompts":
            template_for_format = inject_large_shot_template_content(template_for_format, large_shot_template_content)
        prompt = template_for_format.format(
            script_excerpt=excerpt,
            scene_description=scene_text,
            subject_text=subject_text,
            safe_duration=safe_duration,
            extra_style="",
            large_shot_template_content=large_shot_template_content
        )

    reference_prompt = _resolve_sora_reference_prompt(db, episode.id, reference_shot_id)
    prompt = _append_sora_reference_prompt(prompt, reference_prompt)

    config = get_ai_config("video_prompt")
    request_data = {
        "model": config["model"],
        "messages": [
            {
                "role": "user",
                "content": prompt,
            }
        ],
        "response_format": {"type": "json_object"},
        "stream": False,
    }
    task_payload = {
        "shot_id": int(shot.id),
        "episode_id": int(episode.id),
        "prompt_key": str(prompt_key or "generate_video_prompts"),
        "duration_template_field": template_field,
        "large_shot_template_id": int(large_shot_template_id or 0) if large_shot_template_id else None,
        "large_shot_template_name": large_shot_template_name,
        "large_shot_template_content": large_shot_template_content,
        "reference_shot_id": int(reference_shot_id or 0) if reference_shot_id else None,
    }
    return request_data, task_payload


def _refresh_episode_batch_sora_prompt_state(episode_id: int, db: Session):
    remaining = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id,
        models.StoryboardShot.sora_prompt_status == "generating",
    ).count()
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if episode:
        episode.batch_generating_prompts = remaining > 0


def _repair_stale_storyboard_prompt_generation(episode_id: int, db: Session) -> bool:
    shots = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id,
        models.StoryboardShot.sora_prompt_status == "generating",
    ).all()
    if not shots:
        return False

    shot_ids = [int(getattr(shot, "id", 0) or 0) for shot in shots]
    tasks = db.query(models.TextRelayTask).filter(
        models.TextRelayTask.task_type == "sora_prompt",
        models.TextRelayTask.owner_type == "shot",
        models.TextRelayTask.owner_id.in_(shot_ids),
    ).order_by(
        models.TextRelayTask.owner_id.asc(),
        models.TextRelayTask.id.desc(),
    ).all()

    latest_task_by_shot: Dict[int, models.TextRelayTask] = {}
    active_shot_ids = set()
    for task in tasks:
        shot_id = int(getattr(task, "owner_id", 0) or 0)
        if shot_id <= 0:
            continue
        if shot_id not in latest_task_by_shot:
            latest_task_by_shot[shot_id] = task
        if str(getattr(task, "status", "") or "").strip() in {"submitted", "queued", "running"}:
            active_shot_ids.add(shot_id)

    changed = False
    for shot in shots:
        shot_id = int(getattr(shot, "id", 0) or 0)
        if shot_id in active_shot_ids:
            continue

        next_status = ""
        latest_task = latest_task_by_shot.get(shot_id)
        latest_task_status = str(getattr(latest_task, "status", "") or "").strip() if latest_task else ""
        if latest_task_status == "succeeded":
            next_status = "completed"
        elif latest_task_status == "failed":
            next_status = "failed"
        else:
            has_prompt_content = bool(
                str(getattr(shot, "sora_prompt", "") or "").strip()
                or str(getattr(shot, "storyboard_video_prompt", "") or "").strip()
            )
            has_video_progress = str(getattr(shot, "video_status", "") or "").strip() in {
                "submitting",
                "preparing",
                "processing",
                "completed",
                "failed",
            }
            next_status = "completed" if (has_prompt_content or has_video_progress) else "failed"

        if str(getattr(shot, "sora_prompt_status", "") or "").strip() != next_status:
            shot.sora_prompt_status = next_status
            changed = True

    if changed:
        db.flush()

    return changed


def _reconcile_episode_runtime_flags(episode: Optional[models.Episode], db: Session) -> bool:
    if not episode:
        return False

    episode_id = int(getattr(episode, "id", 0) or 0)
    if episode_id <= 0:
        return False

    changed = False

    changed = _repair_stale_storyboard_prompt_generation(episode_id, db) or changed

    has_generating_sora_prompt = db.query(models.StoryboardShot.id).filter(
        models.StoryboardShot.episode_id == episode_id,
        models.StoryboardShot.sora_prompt_status == "generating",
    ).first() is not None
    if bool(getattr(episode, "batch_generating_prompts", False)) != has_generating_sora_prompt:
        episode.batch_generating_prompts = has_generating_sora_prompt
        changed = True

    simple_summary = _get_simple_storyboard_batch_summary(episode_id, db)
    simple_generating = bool(
        simple_summary.get("submitting_batches", 0) > 0
        or (
            simple_summary.get("total_batches", 0) > 0
            and simple_summary.get("completed_batches", 0) + simple_summary.get("failed_batches", 0)
            < simple_summary.get("total_batches", 0)
        )
    )
    if bool(getattr(episode, "simple_storyboard_generating", False)) != simple_generating:
        episode.simple_storyboard_generating = simple_generating
        changed = True

    if changed:
        db.flush()

    return changed


def _refresh_storyboard2_prompt_batch_state(episode_id: int, db: Session):
    pending_count = db.query(models.TextRelayTask).filter(
        models.TextRelayTask.task_type == "storyboard2_sora_prompt",
        models.TextRelayTask.status.in_(["submitted", "queued", "running"]),
    ).all()
    pending_task_ids = [int(row.owner_id or 0) for row in pending_count]
    active = False
    if pending_task_ids:
        active = db.query(models.Storyboard2Shot).filter(
            models.Storyboard2Shot.id.in_(pending_task_ids),
            models.Storyboard2Shot.episode_id == episode_id,
        ).count() > 0
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if episode:
        episode.batch_generating_storyboard2_prompts = active


def _submit_storyboard_prompt_task(
    db: Session,
    *,
    shot: models.StoryboardShot,
    episode: models.Episode,
    script: models.Script,
    prompt_key: str = "generate_video_prompts",
    duration_template_field: str = "video_prompt_rule",
    large_shot_template_id: Optional[int] = None,
    reference_shot_id: Optional[int] = None,
):
    request_data, task_payload = _build_storyboard_prompt_request_data(
        db,
        shot=shot,
        episode=episode,
        script=script,
        prompt_key=prompt_key,
        duration_template_field=duration_template_field,
        large_shot_template_id=large_shot_template_id,
        reference_shot_id=reference_shot_id,
    )
    return submit_and_persist_text_task(
        db,
        task_type="sora_prompt",
        owner_type="shot",
        owner_id=int(shot.id),
        stage_key=str(prompt_key or "video_prompt"),
        function_key="video_prompt",
        request_payload=request_data,
        task_payload=task_payload,
    )


def _submit_storyboard2_prompt_task(
    db: Session,
    *,
    storyboard2_shot: models.Storyboard2Shot,
):
    library = db.query(models.StoryLibrary).filter(
        models.StoryLibrary.episode_id == storyboard2_shot.episode_id
    ).first()
    all_subject_cards = []
    if library:
        all_subject_cards = db.query(models.SubjectCard).filter(
            models.SubjectCard.library_id == library.id,
            models.SubjectCard.card_type.in_(ALLOWED_CARD_TYPES)
        ).all()
        all_subject_cards.sort(
            key=lambda card: (
                _subject_type_sort_key(card.card_type),
                (card.name or ""),
                card.id
            )
        )
    subject_names = [card.name for card in all_subject_cards if card and (card.name or "").strip()]
    subject_text = _build_storyboard2_subject_text(all_subject_cards)

    excerpt = (storyboard2_shot.excerpt or "").strip()
    if not excerpt:
        raise ValueError("镜头原文为空")

    source_shot = None
    if storyboard2_shot.source_shot_id:
        source_shot = db.query(models.StoryboardShot).filter(
            models.StoryboardShot.id == storyboard2_shot.source_shot_id
        ).first()
    duration = int(source_shot.duration or 10) if source_shot else 10
    if duration not in (10, 15):
        duration = 10 if duration < 13 else 15

    prompt_template = get_prompt_by_key(STORYBOARD2_VIDEO_PROMPT_KEY)
    prompt = prompt_template.format(
        script_excerpt=excerpt,
        scene_description="",
        subject_text=subject_text,
        safe_duration=duration,
        extra_style="",
    )
    config = get_ai_config("video_prompt")
    request_data = {
        "model": config["model"],
        "messages": [
            {
                "role": "user",
                "content": prompt,
            }
        ],
        "response_format": {"type": "json_object"},
        "stream": False,
    }
    task_payload = {
        "episode_id": int(storyboard2_shot.episode_id),
        "storyboard2_shot_id": int(storyboard2_shot.id),
        "duration": int(duration),
        "subject_names": subject_names,
    }
    return submit_and_persist_text_task(
        db,
        task_type="storyboard2_sora_prompt",
        owner_type="storyboard2_shot",
        owner_id=int(storyboard2_shot.id),
        stage_key=STORYBOARD2_VIDEO_PROMPT_KEY,
        function_key="video_prompt",
        request_payload=request_data,
        task_payload=task_payload,
    )


def _queue_single_storyboard_prompt_generation(
    shot_id: int,
    user: models.User,
    db: Session,
    background_tasks: BackgroundTasks,
    prompt_key: str = "generate_video_prompts",
    duration_template_field: str = "video_prompt_rule",
    started_message: str = "Sora提示词生成任务已开始，请稍后刷新页面查看结果。",
    large_shot_template_id: Optional[int] = None,
    reference_shot_id: Optional[int] = None,
):
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    excerpt = (shot.script_excerpt or "").strip()
    if not excerpt:
        raise HTTPException(status_code=400, detail="请先填写原剧本段落")

    if prompt_key == "generate_large_shot_prompts":
        resolved_template = _resolve_large_shot_template(db, large_shot_template_id)
        if not resolved_template:
            raise HTTPException(status_code=404, detail="大镜头模板不存在")
        large_shot_template_id = resolved_template.id

    debug_selected_ids = _debug_parse_card_ids(shot.selected_card_ids)
    debug_subject_names = _debug_resolve_subject_names(db, debug_selected_ids)
    print(
        f"[SoraSubjectDebug][single_request] shot_id={shot.id} "
        f"episode_id={shot.episode_id} shot_number={shot.shot_number} "
        f"raw_selected_card_ids={shot.selected_card_ids!r} "
        f"parsed_selected_ids={debug_selected_ids} "
        f"resolved_subject_names={debug_subject_names} "
        f"large_shot_template_id={large_shot_template_id} "
        f"prompt_key={prompt_key} duration_template_field={duration_template_field} "
        f"reference_shot_id={reference_shot_id}"
    )

    shot.sora_prompt_status = 'generating'
    try:
        _submit_storyboard_prompt_task(
            db,
            shot=shot,
            episode=episode,
            script=script,
            prompt_key=prompt_key,
            duration_template_field=duration_template_field,
            large_shot_template_id=large_shot_template_id,
            reference_shot_id=reference_shot_id,
        )
        db.commit()
    except Exception as exc:
        db.rollback()
        shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
        if shot:
            shot.sora_prompt_status = 'failed'
            db.commit()
        raise HTTPException(status_code=502, detail=f"提交文本任务失败: {str(exc)}")

    return {
        "message": started_message,
        "shot_id": shot_id
    }


async def generate_sora_prompt(
    shot_id: int,
    background_tasks: BackgroundTasks,
    request: Optional[GenerateSoraPromptRequest] = None,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """生成Sora提示词（后台任务）"""
    return _queue_single_storyboard_prompt_generation(
        shot_id=shot_id,
        user=user,
        db=db,
        background_tasks=background_tasks,
        prompt_key="generate_video_prompts",
        duration_template_field="video_prompt_rule",
        started_message="Sora提示词生成任务已开始，请稍后刷新页面查看结果。",
        reference_shot_id=(request.reference_shot_id if request else None),
    )


async def generate_large_shot_prompt(
    shot_id: int,
    background_tasks: BackgroundTasks,
    request: Optional[GenerateLargeShotPromptRequest] = None,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """生成大镜头提示词（后台任务）"""
    return _queue_single_storyboard_prompt_generation(
        shot_id=shot_id,
        user=user,
        db=db,
        background_tasks=background_tasks,
        prompt_key="generate_large_shot_prompts",
        duration_template_field="large_shot_prompt_rule",
        started_message="大镜头提示词生成任务已开始，请稍后刷新页面查看结果。"
        ,
        large_shot_template_id=(request.template_id if request else None),
    )


async def manual_set_sora_prompt(
    shot_id: int,
    request: ManualSoraPromptRequest,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """手动设置Sora提示词，并将状态置为completed"""
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    prompt_text = (request.sora_prompt or "").strip()
    if not prompt_text:
        raise HTTPException(status_code=400, detail="提示词不能为空")

    all_variant_shots = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == shot.episode_id,
        models.StoryboardShot.shot_number == shot.shot_number,
        models.StoryboardShot.variant_index == shot.variant_index
    ).all()

    for variant_shot in all_variant_shots:
        variant_shot.sora_prompt = prompt_text
        variant_shot.storyboard_video_prompt = prompt_text
        variant_shot.sora_prompt_status = 'completed'

    db.commit()
    db.refresh(shot)

    shot.selected_scene_image_url = _resolve_selected_scene_reference_image_url(shot, db)
    return shot

async def get_full_sora_prompt(
    shot_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取完整的Sora提示词（用于复制）"""
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # 构建完整提示词
    full_prompt = build_sora_prompt(shot, db)

    return {
        "full_prompt": full_prompt
    }

def _do_generate_single_sora_prompt(
    shot_id: int,
    user_id: int,
    prompt_key: str = "generate_video_prompts",
    duration_template_field: str = "video_prompt_rule",
    large_shot_template_id: Optional[int] = None,
):
    """后台任务：生成单个镜头提示词"""
    db = SessionLocal()
    try:
        shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
        if not shot:
            return

        episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
        script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
        if script.user_id != user_id:
            return

        effective_video_settings = _apply_episode_storyboard_video_settings_to_shot(shot, episode)

        excerpt = (shot.script_excerpt or "").strip()

        selected_ids = []
        try:
            selected_ids = json.loads(shot.selected_card_ids or "[]")
        except Exception:
            selected_ids = []

        selected_cards = _resolve_selected_cards(db, selected_ids)
        subject_names = [card.name for card in selected_cards if card and card.name]
        subject_text = _build_subject_text_for_ai(selected_cards)

        print(
            f"[SoraSubjectDebug][single_worker_prepare] shot_id={shot.id} "
            f"episode_id={shot.episode_id} shot_number={shot.shot_number} "
            f"raw_selected_card_ids={shot.selected_card_ids!r} "
            f"parsed_selected_ids={selected_ids} "
            f"resolved_subject_names={subject_names} "
            f"resolved_subject_text={subject_text} "
            f"excerpt_len={len(excerpt)} "
            f"large_shot_template_id={large_shot_template_id} "
            f"prompt_key={prompt_key} duration_template_field={duration_template_field}"
        )

        try:
            prompt_style = (script.sora_prompt_style or "").strip() or None
            scene_description = (shot.scene_override or "").strip()
            large_shot_template = None
            large_shot_template_content = ""
            large_shot_template_name = ""
            if prompt_key == "generate_large_shot_prompts":
                large_shot_template = _resolve_large_shot_template(db, large_shot_template_id)
                if not large_shot_template:
                    raise Exception("大镜头模板不存在")
                large_shot_template_id = large_shot_template.id
                large_shot_template_name = (large_shot_template.name or "").strip()
                large_shot_template_content = (large_shot_template.content or "").strip()

            # 调用AI生成
            result = generate_storyboard_prompts(
                excerpt,
                subject_names,
                effective_video_settings["duration"],
                prompt_style,
                shot_id=shot.id,
                prompt_key=prompt_key,
                subject_text_override=subject_text,
                scene_description=scene_description,
                duration_template_field=duration_template_field,
                large_shot_template_id=large_shot_template_id,
                large_shot_template_content=large_shot_template_content,
                large_shot_template_name=large_shot_template_name,
            )

            # 获取timeline数组并转换为表格格式
            timeline = result.get("timeline", [])

            # 保存原始timeline JSON数据
            shot.timeline_json = json.dumps(timeline, ensure_ascii=False)

            table_content = format_timeline_to_table(timeline)

            # 将表格存储到storyboard_video_prompt字段（用于前端显示）
            shot.storyboard_video_prompt = table_content
            shot.storyboard_audio_prompt = ""  # 不再使用此字段
            shot.storyboard_dialogue = ""  # 不再使用此字段

            # ✅ 生成时，自动提取场景描述并保存到 scene_override（仅在未锁定且为空时）
            if should_autofill_scene_override(
                current_scene_override=shot.scene_override,
                scene_override_locked=bool(getattr(shot, "scene_override_locked", False)),
            ):
                scene_desc = extract_scene_description(shot, db)
                if scene_desc:
                    shot.scene_override = scene_desc
                    print(f"[自动填充] scene_override: {scene_desc[:100]}..." if len(scene_desc) > 100 else f"[自动填充] scene_override: {scene_desc}")
            else:
                print(f"[跳过自动填充] scene_override已存在或已锁定，保持当前内容")

            # ✅ 将表格内容保存到 sora_prompt，让用户可以直接看到和编辑
            shot.sora_prompt = table_content
            shot.sora_prompt_status = 'completed'  # 设置状态为已完成

            # 打印生成的表格内容（用于调试）
            print("=" * 80)
            print(f"[生成Sora提示词] 镜头ID: {shot.id}, 镜号: {shot.shot_number}")
            print("-" * 80)
            print("分镜表格内容:")
            print(table_content)
            print("=" * 80)
            print(
                f"[SoraSubjectDebug][single_worker_done] shot_id={shot.id} "
                f"timeline_count={len(timeline)} subject_names={subject_names}"
            )

            db.commit()

        except Exception as e:
            print(f"[SoraSubjectDebug][single_worker_error] shot_id={shot.id} error={str(e)}")
            shot.sora_prompt_status = 'failed'  # 设置状态为失败
            db.commit()  # 确保状态保存

    except Exception as e:
        db.rollback()
    finally:
        db.close()

def _do_batch_generate_sora_prompts(
    episode_id: int,
    default_template: str,
    user_id: int,
    shot_ids: Optional[List[int]] = None,
    duration: Optional[int] = None
):
    """后台任务：批量生成Sora提示词（并发执行）"""
    db = SessionLocal()
    try:
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if not episode:
            print(f"批量生成失败：片段 {episode_id} 不存在")
            return

        script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
        if script.user_id != user_id:
            print(f"批量生成失败：用户 {user_id} 无权限")
            return

        # 标记为生成中
        episode.batch_generating_prompts = True
        db.commit()

        # 获取所有镜头
        query = db.query(models.StoryboardShot).filter(
            models.StoryboardShot.episode_id == episode_id
        )

        # 如果指定了shot_ids，则只处理这些镜头
        if shot_ids:
            query = query.filter(models.StoryboardShot.id.in_(shot_ids))

        shots = query.order_by(models.StoryboardShot.shot_number, models.StoryboardShot.variant_index).all()

        if not shots:
            print(f"批量生成失败：片段 {episode_id} 没有镜头")
            episode.batch_generating_prompts = False
            db.commit()
            return

        library = db.query(models.StoryLibrary).filter(
            models.StoryLibrary.episode_id == episode.id
        ).first()

        prompt_style = (script.sora_prompt_style or "").strip() or None

        # 准备任务参数
        tasks = []
        for shot in shots:
            effective_video_settings = _apply_episode_storyboard_video_settings_to_shot(shot, episode)
            # 如果镜头没有模板，设置默认模板
            if not (shot.prompt_template or "").strip():
                shot.prompt_template = default_template

            # 获取镜头的选中角色
            selected_ids = []
            try:
                selected_ids = json.loads(shot.selected_card_ids or "[]")
            except Exception:
                selected_ids = []

            selected_cards = _resolve_selected_cards(
                db,
                selected_ids,
                library.id if library else None
            )
            subject_names = [card.name for card in selected_cards if card and card.name]
            subject_text = _build_subject_text_for_ai(selected_cards)

            print(
                f"[SoraSubjectDebug][batch_prepare] shot_id={shot.id} "
                f"episode_id={shot.episode_id} shot_number={shot.shot_number} "
                f"raw_selected_card_ids={shot.selected_card_ids!r} "
                f"parsed_selected_ids={selected_ids} "
                f"resolved_subject_names={subject_names} "
                f"resolved_subject_text={subject_text}"
            )

            # 获取剧本段落
            excerpt = (shot.script_excerpt or "").strip()
            scene_description = (shot.scene_override or "").strip()

            # 如果为空，跳过这个镜头
            if not excerpt:
                continue

            tasks.append({
                'shot_id': shot.id,
                'excerpt': excerpt,
                'subject_names': subject_names,
                'subject_text': subject_text,
                'scene_description': scene_description,
                'duration': effective_video_settings["duration"],
                'prompt_style': prompt_style
            })

        db.commit()  # 提交模板设置

        if not tasks:
            print(f"批量生成失败：没有有效的镜头数据")
            episode.batch_generating_prompts = False
            db.commit()
            return

        # 定义单个任务处理函数
        def process_single_shot(task):
            try:
                result = generate_storyboard_prompts(
                    script_excerpt=task['excerpt'],
                    subject_names=task['subject_names'],
                    duration=task['duration'],
                    prompt_style=task['prompt_style'],
                    shot_id=task['shot_id'],
                    subject_text_override=task['subject_text'],
                    scene_description=task['scene_description']
                )
                return {'shot_id': task['shot_id'], 'success': True, 'result': result}
            except Exception as e:
                return {'shot_id': task['shot_id'], 'success': False, 'error': str(e)}

        # 使用线程池并发执行（最多10个并发）
        from concurrent.futures import ThreadPoolExecutor, as_completed
        success_count = 0
        failed_count = 0

        with ThreadPoolExecutor(max_workers=10) as pool:
            futures = [pool.submit(process_single_shot, task) for task in tasks]

            for future in as_completed(futures):
                try:
                    result = future.result()
                    shot_id = result['shot_id']

                    if result['success']:
                        # 保存结果到数据库
                        db_save = SessionLocal()
                        try:
                            shot = db_save.query(models.StoryboardShot).filter(
                                models.StoryboardShot.id == shot_id
                            ).first()

                            if shot:
                                ai_result = result['result']

                                # 获取timeline数组并转换为表格格式
                                timeline = ai_result.get("timeline", [])

                                # 保存原始timeline JSON数据
                                shot.timeline_json = json.dumps(timeline, ensure_ascii=False)

                                table_content = format_timeline_to_table(timeline)

                                # 将表格存储到storyboard_video_prompt字段（用于前端显示）
                                shot.storyboard_video_prompt = table_content
                                shot.storyboard_audio_prompt = ""  # 不再使用此字段
                                shot.storyboard_dialogue = ""  # 不再使用此字段

                                # ✅ 批量生成时，自动提取场景描述并保存到 scene_override（仅在未锁定且为空时）
                                if should_autofill_scene_override(
                                    current_scene_override=shot.scene_override,
                                    scene_override_locked=bool(getattr(shot, "scene_override_locked", False)),
                                ):
                                    scene_desc = extract_scene_description(shot, db_save)
                                    if scene_desc:
                                        shot.scene_override = scene_desc
                                        print(f"[自动填充] scene_override: {scene_desc[:100]}..." if len(scene_desc) > 100 else f"[自动填充] scene_override: {scene_desc}")
                                else:
                                    print(f"[跳过自动填充] scene_override已存在或已锁定，保持当前内容")

                                # ✅ 将表格内容保存到 sora_prompt，让用户可以直接看到和编辑
                                shot.sora_prompt = table_content
                                shot.sora_prompt_status = 'completed'  # 设置状态为已完成

                                # 打印生成的表格内容
                                print("=" * 80)
                                print(f"[批量生成Sora提示词] 镜头ID: {shot.id}, 镜号: {shot.shot_number}")
                                print("-" * 80)
                                print("分镜表格内容:")
                                print(table_content[:200] + "..." if len(table_content) > 200 else table_content)
                                print("=" * 80)

                                db_save.commit()
                                success_count += 1
                        finally:
                            db_save.close()
                    else:
                        # 失败：设置状态为 failed
                        error_msg = result.get('error', 'Unknown error')
                        print(f"[批量生成Sora提示词] 镜头ID {shot_id} 失败: {error_msg}")

                        db_save = SessionLocal()
                        try:
                            shot = db_save.query(models.StoryboardShot).filter(
                                models.StoryboardShot.id == shot_id
                            ).first()
                            if shot:
                                shot.sora_prompt_status = 'failed'
                                db_save.commit()
                        finally:
                            db_save.close()

                        failed_count += 1

                except Exception as e:
                    print(f"[批量生成Sora提示词] 处理结果时出错: {str(e)}")
                    failed_count += 1

        print(f"批量生成完成：成功 {success_count}，失败 {failed_count}")

        # 标记为完成
        episode.batch_generating_prompts = False
        db.commit()

    except Exception as e:
        try:
            episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
            if episode:
                episode.batch_generating_prompts = False
                db.commit()
        except:
            pass
    finally:
        db.close()


async def batch_generate_sora_prompts(
    episode_id: int,
    request: BatchGenerateSoraPromptsRequest,
    background_tasks: BackgroundTasks,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """批量生成Sora提示词（后台任务）"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # 获取镜头数量（如果指定了shot_ids，则只计算这些镜头）
    if request.shot_ids:
        shot_count = db.query(models.StoryboardShot).filter(
            models.StoryboardShot.episode_id == episode_id,
            models.StoryboardShot.id.in_(request.shot_ids)
        ).count()
    else:
        shot_count = db.query(models.StoryboardShot).filter(
            models.StoryboardShot.episode_id == episode_id
        ).count()

    if shot_count == 0:
        raise HTTPException(status_code=400, detail="没有选择有效的镜头")

    print(
        f"[SoraSubjectDebug][batch_request] episode_id={episode_id} "
        f"requested_shot_ids={request.shot_ids if request.shot_ids else 'ALL'} "
        f"matched_shot_count={shot_count}"
    )

    # 清空选中镜头的旧sora_prompt，让前端显示"生成中"状态
    if request.shot_ids:
        shots_to_clear = db.query(models.StoryboardShot).filter(
            models.StoryboardShot.episode_id == episode_id,
            models.StoryboardShot.id.in_(request.shot_ids)
        ).all()
    else:
        shots_to_clear = db.query(models.StoryboardShot).filter(
            models.StoryboardShot.episode_id == episode_id
        ).all()

    for shot in shots_to_clear:
        old_prompt = shot.sora_prompt
        # 不清空 sora_prompt，保留原内容，只设置状态为生成中
        shot.sora_prompt_status = 'generating'  # 设置状态为生成中

    db.commit()
    print("批量生成：状态已设置为生成中，已提交到数据库")

    submitted_count = 0
    for shot in shots_to_clear:
        try:
            _submit_storyboard_prompt_task(
                db,
                shot=shot,
                episode=episode,
                script=script,
                prompt_key="generate_video_prompts",
                duration_template_field="video_prompt_rule",
                large_shot_template_id=None,
            )
            submitted_count += 1
        except Exception as exc:
            shot.sora_prompt_status = 'failed'
            print(f"[批量Sora提交失败] shot_id={shot.id} error={str(exc)}")

    _refresh_episode_batch_sora_prompt_state(episode_id, db)
    db.commit()

    return {
        "message": f"批量生成任务已提交，共 {submitted_count} 个镜头。",
        "total_count": shot_count,
        "submitted_count": submitted_count,
    }


def _do_batch_generate_sora_videos(
    episode_id: int,
    user_id: int,
    shot_ids: Optional[List[int]] = None,
    appoint_account: Optional[str] = None,
):
    """后台任务：批量生成Sora视频（并发提交任务）"""
    db = SessionLocal()
    try:
        episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
        if not episode:
            print(f"批量生成视频失败：片段 {episode_id} 不存在")
            return

        script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
        if script.user_id != user_id:
            print(f"批量生成视频失败：用户 {user_id} 无权限")
            return

        user = db.query(models.User).filter(models.User.id == user_id).first()
        if not user:
            print(f"批量生成视频失败：用户 {user_id} 不存在")
            return

        query = db.query(models.StoryboardShot).filter(
            models.StoryboardShot.episode_id == episode_id
        )
        if shot_ids:
            query = query.filter(models.StoryboardShot.id.in_(shot_ids))

        shots = query.order_by(models.StoryboardShot.shot_number, models.StoryboardShot.variant_index).all()
        if not shots:
            print(f"批量生成视频失败：片段 {episode_id} 没有镜头")
            return

        try:
            _ensure_storyboard_video_generation_slots_available(shots, db)
        except HTTPException as e:
            print(f"批量生成视频失败：{getattr(e, 'detail', str(e))}")
            return

        episode_settings = _get_episode_storyboard_video_settings(episode)
        effective_appoint_account = _normalize_storyboard_video_appoint_account(
            appoint_account,
            default_value=episode_settings.get("appoint_account", "")
        )
        for shot in shots:
            _apply_episode_storyboard_video_settings_to_shot(shot, episode)
            shot.video_status = 'submitting'
        db.commit()

        target_ids = [shot.id for shot in shots]

        def mark_failed(shot_id: int, reason: str):
            try:
                db.rollback()
                failed_shot = db.query(models.StoryboardShot).filter(
                    models.StoryboardShot.id == shot_id
                ).first()
                if failed_shot:
                    failed_shot.video_status = 'failed'
                    failed_shot.video_path = f"error:{reason}"
                    db.commit()
            except Exception as e:
                print(f"批量生成视频失败：镜头 {shot_id} 更新失败: {str(e)}")

        async def generate_single_video(shot_id: int):
            """处理单个镜头的视频生成（带错误处理）"""
            try:
                # 不使用 BackgroundTasks，直接在这里处理拼图生成和视频提交
                db_local = SessionLocal()
                try:
                    shot = db_local.query(models.StoryboardShot).filter(
                        models.StoryboardShot.id == shot_id
                    ).first()

                    if not shot:
                        print(f"[批量生成] 镜头 {shot_id} 不存在")
                        return

                    # 构建完整提示词
                    full_prompt = build_sora_prompt(shot, db_local)
                    if not full_prompt:
                        mark_failed(shot_id, "缺少Sora提示词")
                        return
                    selected_first_frame_image_url = _resolve_selected_first_frame_reference_image_url(
                        shot,
                        db_local,
                    )

                    print(f"[批量生成] 镜头 {shot_id} 提交视频生成任务...")
                    model_name = _resolve_storyboard_video_model_by_provider(
                        shot.provider,
                        default_model=getattr(shot, "storyboard_video_model", None) or episode_settings["model"]
                    )
                    request_data = _build_unified_storyboard_video_task_payload(
                        shot=shot,
                        db=db_local,
                        username=user.username,
                        model_name=model_name,
                        provider=shot.provider or episode_settings["provider"],
                        full_prompt=full_prompt,
                        aspect_ratio=shot.aspect_ratio,
                        duration=shot.duration,
                        first_frame_image_url=selected_first_frame_image_url,
                        resolution_name=episode_settings.get("resolution_name", ""),
                        appoint_account=effective_appoint_account,
                    )
                    submit_timeout = 60 if _is_moti_storyboard_video_model(model_name) else 30

                    submit_response = requests.post(
                        get_video_task_create_url(),
                        headers=get_video_api_headers(),
                        json=request_data,
                        timeout=submit_timeout
                    )

                    if submit_response.status_code != 200:
                        error_msg = f"视频请求失败: {submit_response.status_code}"
                        mark_failed(shot_id, error_msg)
                        return

                    submit_result = submit_response.json()
                    task_id = submit_result.get('task_id')

                    if not task_id:
                        error_msg = f"视频返回异常: {submit_result.get('message', '未知错误')}"
                        mark_failed(shot_id, error_msg)
                        return

                    shot.task_id = task_id
                    shot.video_status = 'processing'
                    shot.video_submitted_at = datetime.utcnow()
                    _record_storyboard_video_charge(
                        db_local,
                        shot=shot,
                        task_id=task_id,
                        stage="video_generate",
                        detail_payload={
                            "source": "batch_generate",
                            "provider": request_data.get("provider"),
                            "model": request_data.get("model"),
                        },
                    )
                    db_local.commit()
                    print(f"[批量生成] 镜头 {shot_id} 视频任务已提交: {task_id}")

                finally:
                    db_local.close()

            except HTTPException as e:
                mark_failed(shot_id, getattr(e, "detail", str(e)))
            except Exception as e:
                mark_failed(shot_id, str(e))

        async def run_batch():
            # 并发执行所有镜头的视频生成
            tasks = [generate_single_video(shot_id) for shot_id in target_ids]
            await asyncio.gather(*tasks, return_exceptions=True)
            print(f"批量生成完成：共处理 {len(target_ids)} 个镜头")

        asyncio.run(run_batch())

    except Exception as e:
        print(f"批量生成视频出错: {str(e)}")
    finally:
        db.close()


async def batch_generate_sora_videos(
    episode_id: int,
    request: BatchGenerateSoraVideosRequest,
    background_tasks: BackgroundTasks,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """批量生成Sora视频（后台任务）"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    if request.shot_ids:
        shot_count = db.query(models.StoryboardShot).filter(
            models.StoryboardShot.episode_id == episode_id,
            models.StoryboardShot.id.in_(request.shot_ids)
        ).count()
    else:
        shot_count = db.query(models.StoryboardShot).filter(
            models.StoryboardShot.episode_id == episode_id
        ).count()

    if shot_count == 0:
        raise HTTPException(status_code=400, detail="没有选择有效的镜头")

    target_query = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id
    )
    if request.shot_ids:
        target_query = target_query.filter(models.StoryboardShot.id.in_(request.shot_ids))
    target_shots = target_query.all()
    _ensure_storyboard_video_generation_slots_available(target_shots, db)

    background_tasks.add_task(
        _do_batch_generate_sora_videos,
        episode_id,
        user.id,
        request.shot_ids,
        request.appoint_account
    )

    return {
        "message": f"批量生成任务已开始，共 {shot_count} 个镜头。请稍后刷新页面查看结果。",
        "total_count": shot_count
    }

# ==================== 托管视频生成API ====================

def _get_next_managed_reserved_variant_index(original_shot: models.StoryboardShot, db: Session) -> int:
    max_variant = db.query(func.max(models.StoryboardShot.variant_index)).filter(
        models.StoryboardShot.episode_id == original_shot.episode_id,
        models.StoryboardShot.shot_number == original_shot.shot_number
    ).scalar()
    family_count = db.query(func.count(models.StoryboardShot.id)).filter(
        models.StoryboardShot.episode_id == original_shot.episode_id,
        models.StoryboardShot.shot_number == original_shot.shot_number
    ).scalar()
    return max(int(max_variant or 0), int(family_count or 0)) + 1


def _create_managed_reserved_shot(
    original_shot: models.StoryboardShot,
    provider: str,
    reserved_variant_index: int
) -> models.StoryboardShot:
    return models.StoryboardShot(
        episode_id=original_shot.episode_id,
        shot_number=original_shot.shot_number,
        stable_id=original_shot.stable_id,
        variant_index=reserved_variant_index,
        prompt_template=original_shot.prompt_template,
        script_excerpt=original_shot.script_excerpt,
        storyboard_video_prompt=original_shot.storyboard_video_prompt,
        storyboard_audio_prompt=original_shot.storyboard_audio_prompt,
        storyboard_dialogue=original_shot.storyboard_dialogue,
        scene_override=original_shot.scene_override,
        scene_override_locked=bool(getattr(original_shot, "scene_override_locked", False)),
        sora_prompt=original_shot.sora_prompt,
        sora_prompt_is_full=bool(getattr(original_shot, "sora_prompt_is_full", False)),
        sora_prompt_status=original_shot.sora_prompt_status,
        selected_card_ids=original_shot.selected_card_ids,
        selected_sound_card_ids=getattr(original_shot, "selected_sound_card_ids", None),
        first_frame_reference_image_url=getattr(original_shot, "first_frame_reference_image_url", ""),
        uploaded_scene_image_url=getattr(original_shot, "uploaded_scene_image_url", ""),
        use_uploaded_scene_image=bool(getattr(original_shot, "use_uploaded_scene_image", False)),
        aspect_ratio=original_shot.aspect_ratio,
        duration=original_shot.duration,
        storyboard_video_model=getattr(original_shot, "storyboard_video_model", ""),
        storyboard_video_model_override_enabled=bool(getattr(original_shot, "storyboard_video_model_override_enabled", False)),
        duration_override_enabled=bool(getattr(original_shot, "duration_override_enabled", False)),
        provider=provider,
        video_status="processing",
        video_error_message="托管排队中",
        timeline_json=original_shot.timeline_json,
        detail_image_prompt_overrides=original_shot.detail_image_prompt_overrides,
        storyboard_image_path=original_shot.storyboard_image_path,
        storyboard_image_status=original_shot.storyboard_image_status,
        storyboard_image_task_id=original_shot.storyboard_image_task_id,
        storyboard_image_model=original_shot.storyboard_image_model,
    )


def _reserve_legacy_managed_session_slots(session: models.ManagedSession, db: Session) -> int:
    active_tasks = db.query(models.ManagedTask).filter(
        models.ManagedTask.session_id == session.id,
        models.ManagedTask.status.in_(["pending", "processing"]),
        models.ManagedTask.shot_id <= 0
    ).order_by(models.ManagedTask.id.asc()).all()

    if not active_tasks:
        return 0

    original_shot_cache = {}
    reserved_count = 0

    for task in active_tasks:
        stable_id = str(task.shot_stable_id or "").strip()
        if not stable_id:
            continue

        if stable_id not in original_shot_cache:
            original_shot_cache[stable_id] = db.query(models.StoryboardShot).filter(
                models.StoryboardShot.stable_id == stable_id,
                models.StoryboardShot.variant_index == 0
            ).first()
        original_shot = original_shot_cache.get(stable_id)
        if not original_shot:
            continue

        has_original_video = bool((original_shot.video_path or "").strip()) and not str(original_shot.video_path or "").startswith("error:")
        has_existing_variants = db.query(func.count(models.StoryboardShot.id)).filter(
            models.StoryboardShot.episode_id == original_shot.episode_id,
            models.StoryboardShot.shot_number == original_shot.shot_number,
            models.StoryboardShot.variant_index > 0
        ).scalar()
        if not has_original_video and not int(has_existing_variants or 0):
            continue

        reserved_variant_index = _get_next_managed_reserved_variant_index(original_shot, db)
        reserved_shot = _create_managed_reserved_shot(
            original_shot,
            session.provider,
            reserved_variant_index
        )
        db.add(reserved_shot)
        db.flush()

        task.shot_id = reserved_shot.id
        reserved_count += 1

    return reserved_count

async def start_managed_generation(
    episode_id: int,
    request: StartManagedGenerationRequest,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """开始托管视频生成"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # 检查是否已有运行中的托管会话
    existing_session = db.query(models.ManagedSession).filter(
        models.ManagedSession.episode_id == episode_id,
        models.ManagedSession.status == "running"
    ).first()

    if existing_session:
        raise HTTPException(status_code=400, detail="已有托管任务在进行中")

    # 获取原始镜头（variant_index=0）
    query = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id,
        models.StoryboardShot.variant_index == 0
    )

    # 如果指定了镜头ID，只处理这些镜头
    if request.shot_ids:
        query = query.filter(models.StoryboardShot.id.in_(request.shot_ids))

    original_shots = query.all()

    if not original_shots:
        raise HTTPException(status_code=400, detail="没有可生成的镜头")

    # 当前托管模式仅允许每个镜头生成 1 个视频
    variant_count = int(request.variant_count or 1)
    if variant_count != 1:
        raise HTTPException(status_code=400, detail="当前每个镜头仅支持托管生成1个视频")

    _ensure_storyboard_video_generation_slots_available(
        original_shots,
        db,
        requested_count_per_shot=variant_count,
    )

    episode_settings = _get_episode_storyboard_video_settings(episode)
    for original_shot in original_shots:
        _apply_episode_storyboard_video_settings_to_shot(original_shot, episode)
    db.flush()

    # 创建托管会话，保存provider
    session = models.ManagedSession(
        episode_id=episode_id,
        status="running",
        total_shots=len(original_shots),
        completed_shots=0,
        variant_count=variant_count,
        provider=episode_settings["provider"]
    )
    db.add(session)
    db.commit()
    db.refresh(session)

    # 为每个原始镜头预留结果槽位，并创建指定数量的任务
    for original_shot in original_shots:
        # 确保有stable_id
        if not original_shot.stable_id:
            original_shot.stable_id = str(uuid.uuid4())
            db.flush()

        next_reserved_variant_index = _get_next_managed_reserved_variant_index(original_shot, db)
        for offset in range(variant_count):
            reserved_variant_index = next_reserved_variant_index + offset
            reserved_shot = _create_managed_reserved_shot(
                original_shot,
                episode_settings["provider"],
                reserved_variant_index
            )
            db.add(reserved_shot)
            db.flush()

            task = models.ManagedTask(
                session_id=session.id,
                shot_id=reserved_shot.id,
                shot_stable_id=original_shot.stable_id,
                status="pending"
            )
            db.add(task)

    db.commit()
    session_tasks = db.query(models.ManagedTask).filter(
        models.ManagedTask.session_id == session.id
    ).all()
    for managed_task in session_tasks:
        sync_managed_task_to_dashboard(managed_task.id)

    return {
        "session_id": session.id,
        "message": f"托管已开始，共{len(original_shots)}个镜头，将生成{len(original_shots) * variant_count}个视频",
        "total_shots": len(original_shots),
        "variant_count": variant_count,
        "model": episode_settings["model"],
        "aspect_ratio": episode_settings["aspect_ratio"],
        "duration": episode_settings["duration"],
        "provider": episode_settings["provider"]
    }

async def stop_managed_generation(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """停止托管视频生成"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # 查找运行中的会话
    session = db.query(models.ManagedSession).filter(
        models.ManagedSession.episode_id == episode_id,
        models.ManagedSession.status == "running"
    ).first()

    if not session:
        raise HTTPException(status_code=404, detail="没有正在运行的托管任务")

    reserved_count = _reserve_legacy_managed_session_slots(session, db)

    # 转为后台继续收尾，不向上游发送取消请求
    session.status = "detached"
    session.completed_at = None
    db.commit()

    return {
        "message": (
            f"托管已转为后台继续收尾，已预留的结果槽位会继续完成"
            + (f"（本次补齐 {reserved_count} 个旧任务槽位）" if reserved_count > 0 else "")
        )
    }

def get_managed_session_status(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取当前托管会话状态"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # 查找最新的会话
    session = db.query(models.ManagedSession).filter(
        models.ManagedSession.episode_id == episode_id
    ).order_by(models.ManagedSession.created_at.desc()).first()

    if not session:
        return ManagedSessionStatusResponse(
            session_id=None,
            status="none",
            total_shots=0,
            completed_shots=0,
            created_at=None
        )

    return ManagedSessionStatusResponse(
        session_id=session.id,
        status=session.status,
        total_shots=session.total_shots,
        completed_shots=session.completed_shots,
        created_at=session.created_at
    )

async def refresh_episode_videos(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """刷新片段所有视频的最新状态和URL（并发查询，最大并发数6）"""
    import asyncio
    import aiohttp

    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # 获取所有有task_id的镜头
    shots = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id,
        models.StoryboardShot.task_id != ''
    ).all()

    if not shots:
        return {
            "success": True,
            "total_shots": 0,
            "updated_count": 0
        }

    # 准备并发查询
    async def check_single_shot(shot_data, session, semaphore):
        """检查单个镜头状态（带并发控制）"""
        async with semaphore:
            shot_id = shot_data['id']
            task_id = shot_data['task_id']

            try:
                url = get_video_task_status_url(task_id)
                headers = {
                    "Authorization": get_video_api_headers()["Authorization"]
                }

                timeout = aiohttp.ClientTimeout(total=10)
                async with session.get(url, headers=headers, timeout=timeout, ssl=False) as response:
                    if response.status == 200:
                        result = await response.json()

                        status = result.get('status', '')
                        video_url = result.get('video_url', '')
                        cdn_uploaded = result.get('cdn_uploaded', False)
                        price = result.get('price')

                        return {
                            'shot_id': shot_id,
                            'success': True,
                            'status': status,
                            'video_url': video_url,
                            'cdn_uploaded': cdn_uploaded,
                            'price': price
                        }
                    else:
                        return {'shot_id': shot_id, 'success': False}

            except Exception as e:
                print(f"[refresh_videos] 查询镜头{shot_id}失败: {str(e)}")
                return {'shot_id': shot_id, 'success': False}

    async def batch_check_shots():
        """批量并发查询所有镜头状态"""
        # 准备镜头数据（避免在异步中访问ORM对象）
        shots_data = [{'id': s.id, 'task_id': s.task_id} for s in shots]

        # 创建信号量限制并发数为6
        semaphore = asyncio.Semaphore(6)

        # 创建HTTP会话
        connector = aiohttp.TCPConnector(limit=6)
        async with aiohttp.ClientSession(connector=connector) as session:
            # 并发查询所有镜头
            tasks = [
                check_single_shot(shot_data, session, semaphore)
                for shot_data in shots_data
            ]
            results = await asyncio.gather(*tasks)

        return results

    # 执行并发查询
    results = await batch_check_shots()

    # 批量更新数据库
    updated_count = 0
    for result in results:
        if not result['success']:
            continue

        shot_id = result['shot_id']
        status = result.get('status', '')
        video_url = result.get('video_url', '')
        cdn_uploaded = result.get('cdn_uploaded', False)
        price = result.get('price')

        # 只更新已完成且有URL的镜头
        if status != 'completed' or not video_url:
            continue

        shot = db.query(models.StoryboardShot).filter(
            models.StoryboardShot.id == shot_id
        ).first()

        if not shot:
            continue

        # 检查是否需要更新
        needs_update = False

        if shot.video_path != video_url or shot.cdn_uploaded != cdn_uploaded:
            previous_video_path = shot.video_path
            previous_thumbnail = shot.thumbnail_video_path

            shot.video_path = video_url
            shot.cdn_uploaded = cdn_uploaded

            if not previous_thumbnail or previous_thumbnail == previous_video_path:
                shot.thumbnail_video_path = video_url

            if shot.video_status != 'completed':
                shot.video_status = 'completed'

            needs_update = True

            # 更新ShotVideo表
            latest_shot_video = db.query(models.ShotVideo).filter(
                models.ShotVideo.shot_id == shot.id
            ).order_by(models.ShotVideo.created_at.desc()).first()

            if latest_shot_video and latest_shot_video.video_path != video_url:
                latest_shot_video.video_path = video_url

        # 更新价格（如果有）
        if price is not None:
            price_cents = int(float(price) * 100)
            if shot.price != price_cents:
                shot.price = price_cents
                needs_update = True

        if needs_update:
            updated_count += 1

    db.commit()

    return {
        "success": True,
        "total_shots": len(shots),
        "updated_count": updated_count
    }

async def get_shot_videos(
    shot_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取镜头生成的视频列表"""
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    videos = db.query(models.ShotVideo).filter(
        models.ShotVideo.shot_id == shot_id
    ).order_by(models.ShotVideo.created_at.desc()).all()

    needs_commit = False
    current_video_path = (shot.video_path or "").strip()
    if current_video_path and not current_video_path.startswith("error:"):
        if not any(v.video_path == current_video_path for v in videos):
            db.add(models.ShotVideo(shot_id=shot_id, video_path=current_video_path))
            needs_commit = True
        if not (shot.thumbnail_video_path or "").strip():
            shot.thumbnail_video_path = current_video_path
            needs_commit = True

    if needs_commit:
        db.commit()
        videos = db.query(models.ShotVideo).filter(
            models.ShotVideo.shot_id == shot_id
        ).order_by(models.ShotVideo.created_at.desc()).all()

    return videos

async def update_shot_thumbnail(
    shot_id: int,
    request: ThumbnailUpdate,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """设置镜头卡片缩略视频"""
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    video = db.query(models.ShotVideo).filter(
        models.ShotVideo.id == request.video_id,
        models.ShotVideo.shot_id == shot_id
    ).first()
    if not video:
        raise HTTPException(status_code=404, detail="视频不存在")

    shot.thumbnail_video_path = video.video_path
    db.commit()

    return {"message": "缩略图已更新", "thumbnail_video_path": shot.thumbnail_video_path}

# ==================== 分镜表导入API ====================

async def import_storyboard(
    episode_id: int,
    file: UploadFile = File(...),
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """导入分镜表（xls）并生成镜头"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in [".xls", ".xlsx"]:
        raise HTTPException(status_code=400, detail="仅支持.xls或.xlsx格式的分镜表")

    try:
        from openpyxl import load_workbook
        from io import BytesIO
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少依赖openpyxl，请先安装")

    content = await file.read()
    try:
        # 去掉data_only=True，确保读取最新保存的值
        wb = load_workbook(filename=BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析失败: {str(e)}")

    if ws.max_row < 2:
        raise HTTPException(status_code=400, detail="表格为空")

    def normalize_header(value: str) -> str:
        text = str(value).strip() if value else ""
        text = text.replace("\n", "").replace("\r", "").replace(" ", "")
        text = text.replace("（", "(").replace("）", ")")
        return text

    def find_column(header_map, keywords):
        for key, idx in header_map.items():
            for kw in keywords:
                if kw in key:
                    return idx
        return None

    def cell_to_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value)
        return str(value).strip()

    def parse_shot_number(value, fallback):
        if value is None or value == "":
            return fallback
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value) if value.is_integer() else fallback
        text = str(value).strip()
        if not text:
            return fallback
        try:
            return int(text)
        except Exception:
            match = re.search(r"\d+", text)
            return int(match.group(0)) if match else fallback

    def parse_subjects_with_type(value) -> List[dict]:
        """解析主体列表，支持格式：萧景珩(角色)、书房(场景)

        Returns:
            List[dict]: [{"name": "萧景珩", "type": "角色"}, ...]
        """
        text = cell_to_text(value)
        if not text:
            return []

        # 支持多种分隔符：顿号、逗号、分号等
        normalized = re.sub(r"[，、/;；|]+", "、", text)
        parts = [part.strip() for part in normalized.split("、") if part.strip()]

        subjects = []
        allowed_types = {"角色", "场景"}
        for part in parts:
            # 尝试匹配 "名称(类型)" 格式
            match = re.match(r"^(.+?)\(([^)]+)\)$", part)
            if match:
                name = match.group(1).strip()
                subject_type = match.group(2).strip()
            else:
                # 没有括号，默认为角色
                name = part
                subject_type = "角色"

            if name and subject_type in allowed_types:
                subjects.append({"name": name, "type": subject_type})

        return subjects

    # 读取表头（openpyxl的行列从1开始）
    headers = [normalize_header(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    header_map = {headers[i]: i + 1 for i in range(len(headers)) if headers[i]}  # 存储column号（1-based）

    shot_idx = find_column(header_map, ["镜号"])
    subjects_idx = find_column(header_map, ["角色/场景", "角色场景", "主体"])
    excerpt_idx = find_column(header_map, ["对应的原剧本段落", "原剧本段落"])
    dialogue_idx = find_column(header_map, ["对白", "台词"])
    storyboard_prompt_idx = find_column(header_map, ["分镜提示词"])
    duration_idx = find_column(header_map, ["时长"])

    # 至少需要有"角色/场景"或"原剧本段落"或"对白"或"分镜提示词"之一
    if subjects_idx is None and excerpt_idx is None and dialogue_idx is None and storyboard_prompt_idx is None:
        raise HTTPException(status_code=400, detail="未找到必要列：至少需要有【角色/场景】或【原剧本段落】或【对白】或【分镜提示词】之一")

    rows_data = []
    for r in range(2, ws.max_row + 1):  # 从第2行开始读取数据
        raw_shot = ws.cell(row=r, column=shot_idx).value if shot_idx is not None else None
        shot_number = parse_shot_number(raw_shot, r - 1)  # r-1 作为fallback（因为第2行对应镜号1）

        # 解析新的6列结构
        subjects = parse_subjects_with_type(ws.cell(row=r, column=subjects_idx).value) if subjects_idx is not None else []
        script_excerpt = cell_to_text(ws.cell(row=r, column=excerpt_idx).value) if excerpt_idx is not None else ""
        dialogue = cell_to_text(ws.cell(row=r, column=dialogue_idx).value) if dialogue_idx is not None else ""
        storyboard_prompt = cell_to_text(ws.cell(row=r, column=storyboard_prompt_idx).value) if storyboard_prompt_idx is not None else ""

        # 调试日志：打印第一行数据
        if r == 2:
            print(f"  镜号={shot_number}, 主体={subjects}")

        # 跳过空行
        if not subjects and not script_excerpt and not dialogue and not storyboard_prompt:
            continue

        duration = 15
        if duration_idx is not None:
            raw_duration = ws.cell(row=r, column=duration_idx).value
            parsed_duration = parse_shot_number(raw_duration, None)
            if parsed_duration in (10, 15):
                duration = parsed_duration

        rows_data.append({
            "shot_number": shot_number,
            "subjects": subjects,
            "script_excerpt": script_excerpt,
            "dialogue": dialogue,
            "storyboard_prompt": storyboard_prompt,
            "duration": duration
        })

    if not rows_data:
        raise HTTPException(status_code=400, detail="表格无有效数据")

    # 获取剧集的主体库
    library = db.query(models.StoryLibrary).filter(
        models.StoryLibrary.episode_id == episode.id
    ).first()

    if not library:
        raise HTTPException(status_code=500, detail="主体库不存在")

    # 现有主体卡片映射
    existing_cards = db.query(models.SubjectCard).filter(
        models.SubjectCard.library_id == library.id
    ).all()
    name_to_id = {card.name: card.id for card in existing_cards}

    # 根据Excel中的主体自动补齐主体卡片
    created_subjects = []
    for row in rows_data:
        for subject in row.get("subjects", []):
            name = subject["name"]
            subject_type = subject["type"]
            if subject_type not in ("角色", "场景"):
                continue
            if name not in name_to_id:
                new_card = models.SubjectCard(
                    library_id=library.id,
                    name=name,
                    card_type=subject_type
                )
                db.add(new_card)
                db.flush()
                name_to_id[name] = new_card.id
                created_subjects.append(f"{name}({subject_type})")

    # ⚠️ 替换模式：删除该episode的所有旧镜头
    deleted_count = _delete_episode_storyboard_shots(episode_id, db)
    db.commit()

    # 创建新导入的镜头
    for idx, row in enumerate(rows_data):
        # 获取主体ID列表
        selected_ids = []
        for subject in row.get("subjects", []):
            name = subject["name"]
            if name in name_to_id:
                selected_ids.append(name_to_id[name])

        # 打印每个镜头的对白字段（用于调试）
        dialogue_preview = row['dialogue'][:50] + '...' if len(row['dialogue']) > 50 else row['dialogue']
        storyboard_prompt_preview = row['storyboard_prompt'][:50] + '...' if len(row['storyboard_prompt']) > 50 else row['storyboard_prompt']

        for _ in [None]:
            new_shot = models.StoryboardShot(
                episode_id=episode_id,
                shot_number=int(row["shot_number"]),
                variant_index=0,
                prompt_template="",
                script_excerpt=row["script_excerpt"],  # 原剧本段落
                storyboard_dialogue=row["dialogue"],  # 对白
                sora_prompt=row["storyboard_prompt"],  # ✅ 保存Excel中的分镜提示词
                selected_card_ids=json.dumps(selected_ids),
                selected_sound_card_ids=None,
                aspect_ratio="16:9",
                duration=row["duration"],
                storyboard_video_model="",
                storyboard_video_model_override_enabled=False,
                duration_override_enabled=True,
            )
            db.add(new_shot)

    db.commit()

    # 验证：查询刚保存的数据
    saved_shots = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id
    ).order_by(models.StoryboardShot.shot_number).limit(3).all()

    for shot in saved_shots:
        dialogue_preview = (shot.storyboard_dialogue or '')[:50] + '...' if len(shot.storyboard_dialogue or '') > 50 else (shot.storyboard_dialogue or '')
        sora_prompt_preview = (shot.sora_prompt or '')[:50] + '...' if len(shot.sora_prompt or '') > 50 else (shot.sora_prompt or '')


    return {
        "message": "导入成功（已替换所有镜头）",
        "imported_shots": len(rows_data),
        "deleted_shots": deleted_count,
        "created_subjects": len(created_subjects),
        "created_subject_names": created_subjects[:10]  # 显示前10个
    }

async def export_storyboard(
    episode_id: int,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """导出分镜表为Excel文件（.xlsx）"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    # 获取所有镜头
    shots = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id
    ).order_by(
        models.StoryboardShot.shot_number.asc(),
        models.StoryboardShot.variant_index.asc()
    ).all()

    if not shots:
        raise HTTPException(status_code=400, detail="没有镜头数据")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少依赖openpyxl，请先安装")

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "分镜表"

    # 表头
    headers = ["镜号", "角色/场景", "原剧本段落", "对白", "分镜提示词", "时长"]
    ws.append(headers)

    # 设置表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 设置列宽
    column_widths = [8, 30, 40, 30, 50, 8]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # 获取主体库以便查询主体详情
    library = db.query(models.StoryLibrary).filter(
        models.StoryLibrary.episode_id == episode.id
    ).first()

    card_map = {}
    if library:
        cards = db.query(models.SubjectCard).filter(
            models.SubjectCard.library_id == library.id
        ).all()
        card_map = {card.id: card for card in cards}

    # 填充数据
    for shot in shots:
        # 解析主体列表
        try:
            selected_ids = json.loads(shot.selected_card_ids or "[]")
        except:
            selected_ids = []

        subjects_text_parts = []
        for card_id in selected_ids:
            if card_id in card_map:
                card = card_map[card_id]
                subjects_text_parts.append(f"{card.name}({card.card_type})")

        subjects_text = "、".join(subjects_text_parts)

        # 构建行数据
        row_data = [
            shot.shot_number,
            subjects_text,
            shot.script_excerpt or "",
            shot.storyboard_dialogue or "",
            shot.sora_prompt or "",  # 分镜提示词使用sora_prompt
            shot.duration
        ]
        ws.append(row_data)

        # 设置单元格自动换行
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 生成文件名
    safe_episode_name = re.sub(r'[\\/*?:"<>|]', '_', episode.name or f"片段{episode_id}")
    filename = f"{safe_episode_name}_分镜表.xlsx"
    output_path = os.path.join("uploads", f"export_{uuid.uuid4().hex[:8]}_{filename}")

    # 保存文件
    wb.save(output_path)

    return FileResponse(
        path=output_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==================== 拼图生成API ====================

def _generate_collage_task(shot_id: int, include_scenes: bool = False, card_ids_hash: str = None):
    """后台任务：生成拼图"""
    db = SessionLocal()
    try:
        # Get shot's aspect_ratio
        shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
        shot_aspect_ratio = (shot.aspect_ratio or "16:9") if shot else "16:9"

        # 生成拼图
        collage_url = generate_collage_image(shot_id, db, include_scenes, aspect_ratio=shot_aspect_ratio)

        # 保存到数据库，记录主体ID组合哈希值
        new_collage = models.ShotCollage(
            shot_id=shot_id,
            collage_path=collage_url,
            is_selected=False,
            card_ids_hash=card_ids_hash
        )
        db.add(new_collage)
        db.commit()
        print(f"[拼图生成] 镜头 {shot_id} 拼图生成成功: {collage_url}, card_ids_hash: {card_ids_hash}")
    except Exception as e:
        print(f"[拼图生成] 镜头 {shot_id} 拼图生成失败: {str(e)}")
    finally:
        db.close()


def _generate_collage_and_video(shot_id: int, full_prompt: str, include_scenes: bool = False):
    """后台任务：生成拼图然后生成视频"""
    db = SessionLocal()
    try:
        shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
        if not shot:
            print(f"[生成拼图+视频] 镜头 {shot_id} 不存在")
            return

        owner_username = ""
        try:
            episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
            if episode:
                _apply_episode_storyboard_video_settings_to_shot(shot, episode)
            script = db.query(models.Script).filter(models.Script.id == episode.script_id).first() if episode else None
            owner = db.query(models.User).filter(models.User.id == script.user_id).first() if script else None
            owner_username = (owner.username or "").strip() if owner else ""
        except Exception:
            owner_username = ""

        # 步骤1：生成拼图
        print(f"[生成拼图+视频] 开始生成拼图... (include_scenes={include_scenes})")
        try:
            collage_url = generate_collage_image(shot_id, db, include_scenes, aspect_ratio=shot.aspect_ratio or "16:9")

            # 创建拼图记录并设置为选中
            new_collage = models.ShotCollage(
                shot_id=shot_id,
                collage_path=collage_url,
                is_selected=True
            )
            db.add(new_collage)
            db.commit()
            db.refresh(new_collage)
            print(f"[生成拼图+视频] 拼图生成成功: {collage_url}")
        except Exception as e:
            print(f"[生成拼图+视频] 拼图生成失败: {str(e)}")
            shot.video_status = 'failed'
            shot.video_path = f"error:拼图生成失败: {str(e)}"
            db.commit()
            return

        # 步骤2：调用统一 Video API 生成视频
        print(f"[生成拼图+视频] 开始生成视频...")
        try:
            model_name = _resolve_storyboard_video_model_by_provider(
                shot.provider,
                default_model=getattr(shot, "storyboard_video_model", None) or getattr(episode, "storyboard_video_model", None) or DEFAULT_STORYBOARD_VIDEO_MODEL
            )
            request_data = _build_unified_storyboard_video_task_payload(
                shot=shot,
                db=db,
                username=owner_username,
                model_name=model_name,
                provider=shot.provider or _resolve_storyboard_video_provider(model_name),
                full_prompt=full_prompt,
                aspect_ratio=shot.aspect_ratio,
                duration=shot.duration,
                first_frame_image_url=new_collage.collage_path,
                resolution_name=getattr(episode, "storyboard_video_resolution_name", None) if episode else None,
                appoint_account=getattr(episode, "storyboard_video_appoint_account", "") if episode else "",
            )
            submit_timeout = 60 if _is_moti_storyboard_video_model(model_name) else 30

            submit_response = requests.post(
                get_video_task_create_url(),
                headers=get_video_api_headers(),
                json=request_data,
                timeout=submit_timeout
            )

            if submit_response.status_code != 200:
                error_msg = f"视频请求失败: {submit_response.status_code}"
                print(f"[生成拼图+视频] {error_msg}")
                save_ai_debug(
                    'video_generate',
                    request_data,
                    {'error': error_msg, 'status_code': submit_response.status_code},
                    shot_id=shot_id
                )
                shot.video_status = 'failed'
                shot.video_path = f"error:{error_msg}"
                db.commit()
                return

            submit_result = submit_response.json()
            task_id = submit_result.get('task_id')

            if not task_id:
                error_msg = f"视频返回异常: {submit_result.get('message', '未知错误')}"
                print(f"[生成拼图+视频] {error_msg}")
                save_ai_debug(
                    'video_generate',
                    request_data,
                    {'error': error_msg, 'response': submit_result},
                    shot_id=shot_id
                )
                shot.video_status = 'failed'
                shot.video_path = f"error:{error_msg}"
                db.commit()
                return

            save_ai_debug(
                'video_generate',
                request_data,
                {'task_id': task_id, 'response': submit_result},
                shot_id=shot_id
            )

            shot.task_id = task_id
            shot.video_status = 'processing'
            shot.video_submitted_at = datetime.utcnow()  # ✅ 重置提交时间
            _record_storyboard_video_charge(
                db,
                shot=shot,
                task_id=task_id,
                stage="video_generate",
                detail_payload={
                    "source": "collage_generate",
                    "provider": request_data.get("provider"),
                    "model": request_data.get("model"),
                },
            )
            db.commit()
            print(f"[生成拼图+视频] 视频生成任务已提交: task_id={task_id}")

        except Exception as e:
            print(f"[生成拼图+视频] 视频生成失败: {str(e)}")
            save_ai_debug(
                'video_generate',
                request_data if 'request_data' in locals() else {},
                {'error': str(e)},
                shot_id=shot_id
            )
            shot.video_status = 'failed'
            shot.video_path = f"error:{str(e)}"
            db.commit()

    except Exception as e:
        print(f"[生成拼图+视频] 整体流程失败: {str(e)}")
    finally:
        db.close()


# ==================== 视频生成API ====================

ACTIVE_VIDEO_GENERATION_STATUSES = storyboard_video_generation_limits.ACTIVE_VIDEO_GENERATION_STATUSES
ACTIVE_MANAGED_TASK_STATUSES = storyboard_video_generation_limits.ACTIVE_MANAGED_TASK_STATUSES
MAX_ACTIVE_VIDEO_GENERATIONS_PER_SHOT = storyboard_video_generation_limits.MAX_ACTIVE_VIDEO_GENERATIONS_PER_SHOT
_get_storyboard_shot_family_identity = storyboard_video_generation_limits.get_storyboard_shot_family_identity
_get_storyboard_shot_family_filters = storyboard_video_generation_limits.get_storyboard_shot_family_filters
_count_active_video_generations_for_shot_family = storyboard_video_generation_limits.count_active_video_generations_for_shot_family
_is_storyboard_shot_generation_active = storyboard_video_generation_limits.is_storyboard_shot_generation_active
_build_active_video_generation_limit_message = storyboard_video_generation_limits.build_active_video_generation_limit_message
_ensure_storyboard_video_generation_slots_available = storyboard_video_generation_limits.ensure_storyboard_video_generation_slots_available

async def generate_video(
    shot_id: int,
    request: GenerateVideoRequest = GenerateVideoRequest(),
    background_tasks: BackgroundTasks = None,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    episode = db.query(models.Episode).filter(models.Episode.id == shot.episode_id).first()
    script = db.query(models.Script).filter(models.Script.id == episode.script_id).first()
    if script.user_id != user.id:
        raise HTTPException(status_code=403, detail="无权限")

    if _backfill_storyboard_visual_references_from_family(shot, db):
        db.commit()
        db.refresh(shot)

    _ensure_storyboard_video_generation_slots_available([shot], db)

    # 为本次“点击生成视频”固定一个任务分组 key，后续调试事件归到同一条 dashboard 任务
    video_debug_folder = save_ai_debug(
        'video_generate',
        {
            "shot_id": shot_id,
            "episode_id": shot.episode_id,
            "first_frame_reference_image_url": normalize_first_frame_candidate_url(
                getattr(shot, "first_frame_reference_image_url", "")
            ),
            "requested_at": datetime.utcnow().isoformat()
        },
        output_data={"status": "request_received"},
        shot_id=shot_id
    )

    episode_settings = _apply_episode_storyboard_video_settings_to_shot(shot, episode)
    db.commit()
    db.refresh(shot)

    # ✅ 总是调用 build_sora_prompt() 进行完整拼接（视频风格 + 场景 + 表格）
    full_prompt = build_sora_prompt(shot, db)

    if not full_prompt:
        raise HTTPException(status_code=400, detail="缺少Sora提示词")

    selected_first_frame_image_url = _resolve_selected_first_frame_reference_image_url(shot, db)

    # ✅ 不要覆盖 shot.sora_prompt，保留用户编辑的内容

    # 打印提交给Sora API的完整提示词
    print("=" * 80)
    print(f"[生成视频] 镜头ID: {shot.id}, 镜号: {shot.shot_number}")
    print(f"[生成视频] 首帧参考图: {selected_first_frame_image_url or '未选择'}")
    print("-" * 80)
    print("提交给Sora API的完整提示词:")
    print(full_prompt)
    print("=" * 80)

    try:
        model_name = _resolve_storyboard_video_model_by_provider(
            shot.provider,
            default_model=getattr(shot, "storyboard_video_model", None) or episode_settings["model"]
        )
        request_data = _build_unified_storyboard_video_task_payload(
            shot=shot,
            db=db,
            username=user.username,
            model_name=model_name,
            provider=shot.provider or _resolve_storyboard_video_provider(model_name),
            full_prompt=full_prompt,
            aspect_ratio=shot.aspect_ratio,
            duration=shot.duration,
            first_frame_image_url=selected_first_frame_image_url,
            resolution_name=episode_settings.get("resolution_name", ""),
            appoint_account=_normalize_storyboard_video_appoint_account(
                request.appoint_account,
                default_value=episode_settings.get("appoint_account", "")
            ),
        )
        submit_timeout = 60 if _is_moti_storyboard_video_model(model_name) else 30

        def call_video_api():
            return requests.post(
                get_video_task_create_url(),
                headers=get_video_api_headers(),
                json=request_data,
                timeout=submit_timeout
            )

        loop = asyncio.get_event_loop()
        submit_response = await loop.run_in_executor(executor, call_video_api)

        if submit_response.status_code != 200:
            error_msg = f"视频请求失败: {submit_response.status_code}"
            save_ai_debug(
                'video_generate',
                request_data,
                {'error': error_msg, 'status_code': submit_response.status_code},
                shot_id=shot_id,
                task_folder=video_debug_folder
            )
            raise Exception(error_msg)

        submit_result = submit_response.json()

        task_id = submit_result.get('task_id')
        if not task_id:
            error_msg = f"视频返回异常: {submit_result.get('message', '未知错误')}"
            save_ai_debug(
                'video_generate',
                request_data,
                {'error': error_msg, 'response': submit_result},
                shot_id=shot_id,
                task_folder=video_debug_folder
            )
            raise Exception(error_msg)

        save_ai_debug(
            'video_generate',
            request_data,
            {'task_id': task_id, 'response': submit_result},
            shot_id=shot_id,
            task_folder=video_debug_folder
        )

        shot.task_id = task_id
        shot.video_status = 'processing'
        shot.video_submitted_at = datetime.utcnow()  # ✅ 重置提交时间，避免超时检查误判
        _record_storyboard_video_charge(
            db,
            shot=shot,
            task_id=task_id,
            stage="video_generate",
            detail_payload={
                "source": "single_generate_with_image",
                "provider": request_data.get("provider"),
                "model": request_data.get("model"),
            },
        )
        db.commit()
        db.refresh(shot)

        return {
            "task_id": task_id,
            "status": "processing"
        }

    except Exception as e:
        if 'request_data' in locals():
            save_ai_debug(
                'video_generate',
                request_data,
                {'exception': str(e)},
                shot_id=shot_id,
                task_folder=video_debug_folder
            )
        status_code = 400 if str(e) in SEEDANCE_AUDIO_VALIDATION_ERRORS else 500
        raise HTTPException(status_code=status_code, detail=str(e))
def check_shot_video_status(
    shot_id: int,
    db: Session = Depends(get_db)
):
    """检查视频生成状态"""
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    return {
        "status": shot.video_status,  # idle/processing/completed/failed
        "video_path": shot.video_path,
        "task_id": shot.task_id
    }


def cancel_invalid_seedance_video_tasks(db: Session) -> dict:
    """取消当前库里已在跑、但不满足 Seedance 音频规则的任务，并标记本地失败。"""
    episode_cache: Dict[int, Any] = {}

    def get_episode(episode_id: int):
        if episode_id not in episode_cache:
            episode_cache[episode_id] = db.query(models.Episode).filter(
                models.Episode.id == episode_id
            ).first()
        return episode_cache[episode_id]

    invalid_shots = []
    processing_shots = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.task_id != "",
        models.StoryboardShot.video_status.in_(["submitting", "preparing", "processing"])
    ).all()

    for shot in processing_shots:
        episode = get_episode(shot.episode_id)
        default_model = getattr(episode, "storyboard_video_model", None) if episode else None
        model_name = _resolve_storyboard_video_model_by_provider(
            getattr(shot, "provider", None),
            default_model=getattr(shot, "storyboard_video_model", None) or default_model or "sora-2"
        )
        if not _is_moti_storyboard_video_model(model_name):
            continue
        assets = _collect_moti_v2_reference_assets(shot, db)
        reason = _get_seedance_audio_validation_error(
            assets["audio_items"],
            float(assets["total_audio_duration_seconds"] or 0.0)
        )
        if not reason:
            continue
        invalid_shots.append({
            "shot_id": shot.id,
            "task_id": shot.task_id,
            "reason": reason
        })

    invalid_managed_tasks = []
    processing_managed_tasks = db.query(models.ManagedTask).join(
        models.ManagedSession,
        models.ManagedTask.session_id == models.ManagedSession.id
    ).filter(
        models.ManagedTask.status == "processing",
        models.ManagedTask.task_id != "",
        models.ManagedSession.status.in_(ACTIVE_MANAGED_SESSION_STATUSES),
        models.ManagedSession.provider == "moti"
    ).all()

    for task in processing_managed_tasks:
        original_shot = db.query(models.StoryboardShot).filter(
            models.StoryboardShot.stable_id == task.shot_stable_id,
            models.StoryboardShot.variant_index == 0
        ).first()
        if not original_shot:
            continue
        assets = _collect_moti_v2_reference_assets(original_shot, db)
        reason = _get_seedance_audio_validation_error(
            assets["audio_items"],
            float(assets["total_audio_duration_seconds"] or 0.0)
        )
        if not reason:
            continue
        invalid_managed_tasks.append({
            "managed_task_id": task.id,
            "task_id": task.task_id,
            "session_id": task.session_id,
            "reason": reason
        })

    cancel_target_ids = list({
        item["task_id"]
        for item in invalid_shots + invalid_managed_tasks
        if item.get("task_id")
    })
    cancel_result = _cancel_upstream_video_tasks(cancel_target_ids)

    if invalid_shots:
        shot_map = {
            shot.id: shot
            for shot in db.query(models.StoryboardShot).filter(
                models.StoryboardShot.id.in_([item["shot_id"] for item in invalid_shots])
            ).all()
        }
        for item in invalid_shots:
            shot = shot_map.get(item["shot_id"])
            if not shot:
                continue
            shot.video_status = "failed"
            shot.video_error_message = item["reason"]
            shot.video_path = f"error:{item['reason']}"
            shot.thumbnail_video_path = ""

    touched_session_ids = set()
    if invalid_managed_tasks:
        managed_task_map = {
            task.id: task
            for task in db.query(models.ManagedTask).filter(
                models.ManagedTask.id.in_([item["managed_task_id"] for item in invalid_managed_tasks])
            ).all()
        }
        for item in invalid_managed_tasks:
            task = managed_task_map.get(item["managed_task_id"])
            if not task:
                continue
            task.status = "failed"
            task.error_message = item["reason"]
            task.completed_at = datetime.utcnow()
            touched_session_ids.add(task.session_id)

    if invalid_shots or invalid_managed_tasks:
        db.commit()

    if touched_session_ids:
        try:
            from managed_generation_service import ManagedGenerationPoller
            sync_poller = ManagedGenerationPoller()
            for session_id in touched_session_ids:
                sync_poller._update_session_progress(session_id, db)
            db.commit()
        except Exception as e:
            _rollback_quietly(db)
            print(f"[seedance] 刷新托管进度失败: {str(e)}")

    return {
        "invalid_shot_count": len(invalid_shots),
        "invalid_managed_task_count": len(invalid_managed_tasks),
        "cancel_result": cancel_result,
        "invalid_shots": invalid_shots,
        "invalid_managed_tasks": invalid_managed_tasks
    }


def _looks_like_video_status_query_failure(message: str) -> bool:
    normalized = str(message or "").strip()
    if not normalized:
        return False
    return (
        normalized.startswith("请求失败: 5")
        or normalized.startswith("查询异常:")
        or normalized.startswith("状态查询失败: HTTP 5")
    )


def repair_managed_tasks_failed_by_query_errors(db: Session) -> dict:
    """修复因状态查询 5xx/超时被误判失败的托管任务。"""
    from video_service import check_video_status, is_transient_video_status_error
    from managed_generation_service import ManagedGenerationPoller

    candidate_tasks = db.query(models.ManagedTask).join(
        models.ManagedSession,
        models.ManagedTask.session_id == models.ManagedSession.id
    ).filter(
        models.ManagedTask.status == "failed",
        models.ManagedTask.task_id != "",
        models.ManagedSession.status.in_(ACTIVE_MANAGED_SESSION_STATUSES)
    ).order_by(models.ManagedTask.id.asc()).all()

    candidate_tasks = [
        task for task in candidate_tasks
        if _looks_like_video_status_query_failure(task.error_message)
    ]

    revived_tasks = []
    superseded_tasks = []
    confirmed_failed_tasks = []
    skipped_tasks = []
    cancel_task_ids = set()
    touched_session_ids = set()

    for task in candidate_tasks:
        status_info = check_video_status(task.task_id)
        if is_transient_video_status_error(status_info):
            skipped_tasks.append({
                "managed_task_id": task.id,
                "task_id": task.task_id,
                "reason": status_info.get("error_message", "")
            })
            continue

        upstream_status = str(status_info.get("status") or "").strip().lower()
        siblings = db.query(models.ManagedTask).filter(
            models.ManagedTask.session_id == task.session_id,
            models.ManagedTask.shot_stable_id == task.shot_stable_id
        ).order_by(models.ManagedTask.id.asc()).all()

        other_tasks = [item for item in siblings if item.id != task.id]
        other_processing = [
            item for item in other_tasks
            if item.status == "processing" and str(item.task_id or "").strip()
        ]
        other_completed = [item for item in other_tasks if item.status == "completed"]

        if upstream_status in {"submitted", "pending", "processing"}:
            if other_processing or other_completed:
                cancel_task_ids.add(task.task_id)
                takeover_task = (other_processing or other_completed)[0]
                note = f"状态查询异常后已由任务#{takeover_task.id}接管，已取消旧上游任务"
                base_error = str(task.error_message or "").strip()
                task.error_message = note if not base_error else f"{base_error}；{note}"
                superseded_tasks.append({
                    "managed_task_id": task.id,
                    "task_id": task.task_id,
                    "upstream_status": upstream_status,
                    "takeover_task_id": takeover_task.id
                })
                touched_session_ids.add(task.session_id)
                continue

            task.status = "processing"
            task.error_message = ""
            task.completed_at = None
            revived_tasks.append({
                "managed_task_id": task.id,
                "task_id": task.task_id,
                "upstream_status": upstream_status
            })
            touched_session_ids.add(task.session_id)
            continue

        if upstream_status == "completed":
            if other_completed:
                confirmed_failed_tasks.append({
                    "managed_task_id": task.id,
                    "task_id": task.task_id,
                    "upstream_status": upstream_status,
                    "reason": "同镜头已有完成结果，保留现状"
                })
                continue

            task.status = "processing"
            task.error_message = ""
            task.completed_at = None
            revived_tasks.append({
                "managed_task_id": task.id,
                "task_id": task.task_id,
                "upstream_status": upstream_status
            })
            touched_session_ids.add(task.session_id)
            continue

        confirmed_failed_tasks.append({
            "managed_task_id": task.id,
            "task_id": task.task_id,
            "upstream_status": upstream_status,
            "reason": str(status_info.get("error_message") or "")
        })

    cancel_result = _cancel_upstream_video_tasks(list(cancel_task_ids))

    if revived_tasks or superseded_tasks:
        db.commit()

    if touched_session_ids:
        try:
            sync_poller = ManagedGenerationPoller()
            for session_id in touched_session_ids:
                sync_poller._update_session_progress(session_id, db)
            db.commit()
        except Exception as e:
            _rollback_quietly(db)
            print(f"[managed-repair] 刷新托管进度失败: {str(e)}")

    return {
        "candidate_count": len(candidate_tasks),
        "revived_count": len(revived_tasks),
        "superseded_count": len(superseded_tasks),
        "confirmed_failed_count": len(confirmed_failed_tasks),
        "skipped_count": len(skipped_tasks),
        "cancel_result": cancel_result,
        "revived_tasks": revived_tasks,
        "superseded_tasks": superseded_tasks,
        "confirmed_failed_tasks": confirmed_failed_tasks,
        "skipped_tasks": skipped_tasks
    }


async def export_shot_video(
    shot_id: int,
    db: Session = Depends(get_db)
):
    """导出单个镜头的视频"""
    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    if shot.video_status != 'completed':
        raise HTTPException(status_code=400, detail=f"视频未完成生成，当前状态: {shot.video_status}")

    if not shot.video_path:
        raise HTTPException(status_code=404, detail="视频路径不存在")

    # video_path现在保存的是CDN URL，直接返回
    return {
        "video_url": shot.video_path,
        "shot_number": shot.shot_number
    }

# ==================== 分镜图生成API ====================

_DETAIL_IMAGES_MODEL_CONFIG = shot_image_generation._DETAIL_IMAGES_MODEL_CONFIG
_resolve_storyboard_sora_image_ratio = shot_image_generation._resolve_storyboard_sora_image_ratio
_resolve_detail_images_actual_model = shot_image_generation._resolve_detail_images_actual_model
_build_image_generation_debug_meta = shot_image_generation._build_image_generation_debug_meta
_build_image_generation_request_payload = shot_image_generation._build_image_generation_request_payload
_submit_single_image_generation_task = shot_image_generation._submit_single_image_generation_task
_save_detail_images_debug = shot_image_generation._save_detail_images_debug
generate_detail_images = shot_image_generation.generate_detail_images
_process_detail_images_generation = shot_image_generation._process_detail_images_generation
get_shot_detail_images = shot_image_generation.get_shot_detail_images
set_shot_detail_image_cover = shot_image_generation.set_shot_detail_image_cover
generate_storyboard_image = shot_reference_workflow.generate_storyboard_image
set_shot_first_frame_reference = shot_reference_workflow.set_shot_first_frame_reference
upload_shot_first_frame_reference_image = shot_reference_workflow.upload_shot_first_frame_reference_image
upload_shot_scene_image = shot_reference_workflow.upload_shot_scene_image
set_shot_scene_image_selection = shot_reference_workflow.set_shot_scene_image_selection
_get_shot_detail_image_urls = shot_reference_workflow._get_shot_detail_image_urls
_get_shot_first_frame_candidate_urls = shot_reference_workflow._get_shot_first_frame_candidate_urls
_backfill_storyboard_visual_references_from_family = shot_reference_workflow._backfill_storyboard_visual_references_from_family
_resolve_selected_first_frame_reference_image_url = shot_reference_workflow._resolve_selected_first_frame_reference_image_url


def _generate_single_image_with_polling(
    prompt_text: str,
    model_name: str,
    provider: Optional[str] = None,
    size: str = "9:16",
    resolution: Optional[str] = None,
    reference_images: Optional[List[str]] = None,
    timeout: int = 600,
    poll_interval_seconds: int = 5,
) -> dict:
    """提交并轮询图片任务，返回单图结果。"""
    normalized_provider = str(provider or "").strip().lower()
    submit_result = _submit_single_image_generation_task(
        prompt_text=prompt_text,
        model_name=model_name,
        provider=normalized_provider,
        size=size,
        resolution=resolution,
        reference_images=reference_images,
        name=f"single_image_{uuid.uuid4().hex[:8]}",
    )
    task_id = submit_result["task_id"]
    submit_api_url = submit_result["submit_api_url"]
    status_api_url = submit_result["status_api_url"]
    start_time = time.time()
    transient_query_error_count = 0
    while time.time() - start_time < timeout:
        status_result = get_image_task_status(task_id, model_name, normalized_provider)
        if is_transient_image_status_error(status_result):
            transient_query_error_count += 1
            if transient_query_error_count >= 10:
                return {
                    "success": False,
                    "error": f"连续查询异常 10 次：{status_result.get('error_message') or '状态查询失败'}",
                    "task_id": task_id,
                    "submit_api_url": submit_api_url,
                    "status_api_url": status_api_url,
                }
            time.sleep(max(1, int(poll_interval_seconds)))
            continue
        transient_query_error_count = 0
        status = status_result.get("status")
        if status == "completed":
            images = status_result.get("images") or []
            if not images:
                return {
                    "success": False,
                    "error": "生成任务已完成，但未返回图片",
                    "task_id": task_id,
                    "submit_api_url": submit_api_url,
                    "status_api_url": status_api_url
                }
            return {
                "success": True,
                "images": images[:1],
                "task_id": task_id,
                "submit_api_url": submit_api_url,
                "status_api_url": status_api_url
            }
        if status == "failed":
            return {
                "success": False,
                "error": status_result.get("error") or status_result.get("error_message") or "生成失败",
                "task_id": task_id,
                "submit_api_url": submit_api_url,
                "status_api_url": status_api_url
            }
        time.sleep(max(1, int(poll_interval_seconds)))

    return {
        "success": False,
        "error": f"生成超时（超过{timeout}秒）",
        "task_id": task_id,
        "submit_api_url": submit_api_url,
        "status_api_url": status_api_url
    }


_normalize_jimeng_ratio = storyboard2_media.normalize_jimeng_ratio


_normalize_storyboard2_video_duration = storyboard_defaults.normalize_storyboard2_video_duration
_normalize_storyboard2_image_cw = storyboard_defaults.normalize_storyboard2_image_cw
_normalize_detail_images_provider = storyboard_defaults.normalize_detail_images_provider
_resolve_episode_detail_images_provider = storyboard_defaults.resolve_episode_detail_images_provider
_normalize_detail_images_model = storyboard_defaults.normalize_detail_images_model


_STORYBOARD_VIDEO_MODEL_CONFIG = storyboard_video_settings.STORYBOARD_VIDEO_MODEL_CONFIG
_normalize_storyboard_video_model = storyboard_video_settings.normalize_storyboard_video_model
_normalize_storyboard_video_aspect_ratio = storyboard_video_settings.normalize_storyboard_video_aspect_ratio
_normalize_storyboard_video_duration = storyboard_video_settings.normalize_storyboard_video_duration
_normalize_storyboard_video_resolution_name = storyboard_video_settings.normalize_storyboard_video_resolution_name
_resolve_storyboard_video_provider = storyboard_video_settings.resolve_storyboard_video_provider
_is_moti_storyboard_video_model = storyboard_video_settings.is_moti_storyboard_video_model
_resolve_storyboard_video_model_by_provider = storyboard_video_settings.resolve_storyboard_video_model_by_provider
_map_storyboard_prompt_template_duration = storyboard_video_settings.map_storyboard_prompt_template_duration
_is_storyboard_shot_duration_override_enabled = storyboard_video_settings.is_storyboard_shot_duration_override_enabled
_is_storyboard_shot_model_override_enabled = storyboard_video_settings.is_storyboard_shot_model_override_enabled
_get_episode_storyboard_video_settings = storyboard_video_settings.get_episode_storyboard_video_settings
_get_effective_storyboard_video_settings_for_shot = storyboard_video_settings.get_effective_storyboard_video_settings_for_shot


def _map_api_model_by_duration(model: str, duration: Optional[int]) -> str:
    """
    根据时长映射模型名称
    - sora-2 + 25秒 → sora-2-pro
    - 其他情况 → 保持原样
    """
    if (model or "").strip().lower() == "sora-2" and duration == 25:
        return "sora-2-pro"
    return model


_get_seedance_audio_validation_error = storyboard_video_payload.get_seedance_audio_validation_error
_collect_moti_v2_reference_assets = storyboard_video_payload._collect_moti_v2_reference_assets
_build_moti_v2_content = storyboard_video_payload._build_moti_v2_content
_build_storyboard_video_text_and_images_content = storyboard_video_payload.build_storyboard_video_reference_content
_build_grok_video_content = storyboard_video_payload._build_grok_video_content
_build_unified_storyboard_video_task_payload = storyboard_video_payload._build_unified_storyboard_video_task_payload


def _apply_episode_storyboard_video_settings_to_shot(shot, episode) -> Dict[str, Any]:
    settings = _get_effective_storyboard_video_settings_for_shot(shot, episode)
    shot.storyboard_video_model = settings["model"]
    shot.storyboard_video_model_override_enabled = bool(settings["model_override_enabled"])
    shot.aspect_ratio = settings["aspect_ratio"]
    shot.duration = settings["duration"]
    shot.provider = settings["provider"]
    return settings

# ==================== 镜头细化图片生成API ====================

_get_subject_card_reference_image_url = storyboard_reference_assets.get_subject_card_reference_image_url
_collect_storyboard_subject_reference_urls = storyboard_reference_assets.collect_storyboard_subject_reference_urls


_get_selected_scene_card_image_url = storyboard_reference_assets.get_selected_scene_card_image_url
_resolve_selected_scene_reference_image_url = storyboard_reference_assets.resolve_selected_scene_reference_image_url


# ==================== 视频导出API ====================

async def export_all_videos(
    episode_id: int,
    db: Session = Depends(get_db)
):
    """导出片段的所有视频"""
    episode = db.query(models.Episode).filter(models.Episode.id == episode_id).first()
    if not episode:
        raise HTTPException(status_code=404, detail="片段不存在")

    shots = db.query(models.StoryboardShot).filter(
        models.StoryboardShot.episode_id == episode_id,
        models.StoryboardShot.video_status == 'completed'
    ).order_by(
        models.StoryboardShot.shot_number.asc(),
        models.StoryboardShot.variant_index.asc()
    ).all()

    if not shots:
        raise HTTPException(status_code=404, detail="没有已完成的视频")

    videos = []
    for shot in shots:
        if shot.video_path:
            # video_path现在保存的是CDN URL，直接使用
            videos.append({
                "shot_id": shot.id,
                "shot_number": shot.shot_number,
                "video_url": shot.video_path
            })

    return {
        "episode_name": episode.name,
        "total_videos": len(videos),
        "videos": videos
    }

# 重新处理视频（下载并上传到CDN）
async def reprocess_shot_video(
    shot_id: int,
    db: Session = Depends(get_db)
):
    """重新处理视频：下载Sora视频并上传到自己的CDN（适用于已完成但还是Sora URL的视频）"""
    from video_service import download_and_upload_video

    shot = db.query(models.StoryboardShot).filter(models.StoryboardShot.id == shot_id).first()
    if not shot:
        raise HTTPException(status_code=404, detail="镜头不存在")

    if shot.video_status != 'completed':
        raise HTTPException(status_code=400, detail=f"视频未完成生成，当前状态: {shot.video_status}")

    if not shot.video_path:
        raise HTTPException(status_code=404, detail="视频路径不存在")

    # 如果已经是自己CDN的URL（包含moapp.net.cn），不需要重新处理
    if 'moapp.net.cn' in shot.video_path:
        return {"message": "视频已经在自己的CDN", "video_url": shot.video_path}

    # 重新处理：下载并上传到自己的CDN
    try:
        cdn_url = download_and_upload_video(shot.video_path, shot_id)
        previous_video_path = shot.video_path
        previous_thumbnail = shot.thumbnail_video_path
        shot.video_path = cdn_url
        if not previous_thumbnail or previous_thumbnail == previous_video_path:
            shot.thumbnail_video_path = cdn_url

        new_video = models.ShotVideo(
            shot_id=shot.id,
            video_path=cdn_url
        )
        db.add(new_video)
        db.commit()

        return {"message": "视频处理成功", "video_url": cdn_url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")


# ==================== 故事板2 API ====================

Storyboard2GenerateImagesRequest = storyboard2.Storyboard2GenerateImagesRequest
Storyboard2SetCurrentImageRequest = storyboard2.Storyboard2SetCurrentImageRequest
Storyboard2GenerateVideoRequest = storyboard2.Storyboard2GenerateVideoRequest


def _parse_storyboard2_timeline(timeline_json_value):
    if not timeline_json_value:
        return []

    try:
        timeline_data = json.loads(timeline_json_value) if isinstance(timeline_json_value, str) else timeline_json_value
    except Exception:
        return []

    if isinstance(timeline_data, list):
        return timeline_data

    if isinstance(timeline_data, dict):
        nested = timeline_data.get("timeline")
        if isinstance(nested, list):
            return nested

    return []


def _storyboard2_fallback_timeline(duration_seconds: int):
    total = max(6, int(duration_seconds or 10))
    segment_count = 2 if total <= 9 else (3 if total <= 14 else 4)
    segment_length = total / segment_count

    timeline = []
    for idx in range(segment_count):
        start = int(round(idx * segment_length))
        if idx == segment_count - 1:
            end = total
        else:
            end = max(start + 1, int(round((idx + 1) * segment_length)))

        timeline.append({
            "time": f"{start}s-{end}s",
            "visual": f"分镜{idx + 1}画面描述",
            "audio": ""
        })

    return timeline


def _resolve_storyboard2_time_range(item: dict, index: int, total_count: int) -> str:
    if not isinstance(item, dict):
        return f"分镜 {index}/{total_count}"

    raw_time = item.get("time") or item.get("time_range") or ""
    if isinstance(raw_time, str) and raw_time.strip():
        return raw_time.strip()

    start = item.get("start_time", item.get("start", item.get("begin")))
    end = item.get("end_time", item.get("end"))
    duration = item.get("duration")

    if start is not None and end is not None:
        return f"{start}s-{end}s"
    if start is not None and duration is not None:
        try:
            end_value = float(start) + float(duration)
            if end_value.is_integer():
                end_value = int(end_value)
            return f"{start}s-{end_value}s"
        except Exception:
            pass

    return f"分镜 {index}/{total_count}"


def _extract_storyboard2_timeline_subject_names(timeline_item: dict) -> List[str]:
    """Extract subject names from timeline item fields."""
    if not isinstance(timeline_item, dict):
        return []

    raw_candidates = [
        timeline_item.get("subjects"),
        timeline_item.get("subject_names"),
        timeline_item.get("subject_list"),
        timeline_item.get("subject"),
        timeline_item.get("main_subject"),
    ]

    names: List[str] = []

    def append_name(value):
        if value is None:
            return
        if isinstance(value, (list, tuple, set)):
            for item in value:
                append_name(item)
            return
        if isinstance(value, dict):
            for key in ("name", "subject", "label"):
                if value.get(key):
                    append_name(value.get(key))
                    return
            return

        text = str(value).strip()
        if not text:
            return

        parts = re.split(r"[\n,，、;；|/]+", text)
        for part in parts:
            token = str(part or "").strip()
            if not token:
                continue
            token = re.sub(r"^(角色|场景|道具|主体|人物)\s*[:：]\s*", "", token)
            token = re.sub(r"^(男主\d*|女主\d*|主角\d*)\s*[:：]\s*", "", token)
            token = token.strip("[]【】")
            if not token:
                continue
            names.append(token)

    for candidate in raw_candidates:
        append_name(candidate)

    resolved = []
    seen = set()
    for name in names:
        clean = str(name or "").strip()
        if not clean or clean in seen:
            continue
        seen.add(clean)
        resolved.append(clean)
    return resolved


def _normalize_storyboard2_subject_token(text_value) -> str:
    text = str(text_value or "").strip()
    if not text:
        return ""

    text = re.sub(r"^(角色|场景|道具|主体|人物)\s*[:：]\s*", "", text)
    text = re.sub(r"^(男主\d*|女主\d*|主角\d*)\s*[:：]\s*", "", text)
    text = text.strip("[]【】")
    text = re.sub(r"[（(][^（）()]{1,20}[)）]\s*$", "", text).strip()
    text = re.sub(r"\s+", "", text)
    return text.lower()


def _resolve_storyboard2_subject_ids_from_names(
    subject_names: List[str],
    library_cards: List[models.SubjectCard]
) -> List[int]:
    """Resolve timeline subject names to SubjectCard ids."""
    if not subject_names or not library_cards:
        return []

    exact_map: Dict[str, List[int]] = {}
    fuzzy_entries: List[Tuple[int, str]] = []

    for card in library_cards:
        if not card:
            continue
        card_id = int(card.id)
        card_tokens = [
            _normalize_storyboard2_subject_token(card.name),
            _normalize_storyboard2_subject_token(card.alias),
        ]
        for token in card_tokens:
            if not token:
                continue
            exact_map.setdefault(token, []).append(card_id)
            fuzzy_entries.append((card_id, token))

    resolved_ids: List[int] = []
    seen_ids = set()
    for raw_name in subject_names:
        normalized_name = _normalize_storyboard2_subject_token(raw_name)
        if not normalized_name:
            continue

        matched_ids = exact_map.get(normalized_name, [])
        if not matched_ids:
            fuzzy_hits = []
            for card_id, token in fuzzy_entries:
                if not token:
                    continue
                if len(normalized_name) >= 2 and (normalized_name in token or token in normalized_name):
                    fuzzy_hits.append(card_id)
            matched_ids = fuzzy_hits

        for card_id in matched_ids:
            if card_id in seen_ids:
                continue
            seen_ids.add(card_id)
            resolved_ids.append(card_id)

    return resolved_ids


def _is_prop_subject_card_type(card_type: str) -> bool:
    card_type_text = str(card_type or "").strip().lower()
    if not card_type_text:
        return False
    if card_type_text == "prop":
        return True
    if "道具" in card_type_text:
        return True
    if "閬撳叿" in card_type_text:
        return True
    return False


_collect_storyboard2_reference_images = storyboard2_reference_images.collect_storyboard2_reference_images


# Storyboard2 generation and edit routes now live in api.routers.episodes.

# Storyboard2 generation and edit routes now live in api.routers.episodes.


app.include_router(episodes.router)
app.include_router(storyboard2.router)
app.include_router(storyboard_excel.router)
app.include_router(voiceover.router)
app.include_router(simple_storyboard.router)
app.include_router(scripts.router)
app.include_router(shots.router)
app.include_router(managed_generation.router)
app.include_router(hit_dramas.router)

# Compatibility exports for direct callers while managed task routes live in api.routers.episodes.
get_managed_tasks = episodes.get_managed_tasks

# Compatibility exports for direct callers while simple storyboard routes live in api.routers.simple_storyboard.
generate_simple_storyboard_api = simple_storyboard.generate_simple_storyboard_api
get_simple_storyboard = simple_storyboard.get_simple_storyboard
get_simple_storyboard_status = simple_storyboard.get_simple_storyboard_status
retry_failed_simple_storyboard_batches_api = simple_storyboard.retry_failed_simple_storyboard_batches_api
update_simple_storyboard = simple_storyboard.update_simple_storyboard

# Compatibility exports for direct callers while storyboard2 routes live in api.routers.storyboard2.
Storyboard2SetCurrentImageRequest = storyboard2.Storyboard2SetCurrentImageRequest
Storyboard2BatchGenerateSoraPromptsRequest = storyboard2.Storyboard2BatchGenerateSoraPromptsRequest
Storyboard2GenerateImagesRequest = storyboard2.Storyboard2GenerateImagesRequest
Storyboard2GenerateVideoRequest = storyboard2.Storyboard2GenerateVideoRequest
Storyboard2UpdateShotRequest = storyboard2.Storyboard2UpdateShotRequest
Storyboard2UpdateSubShotRequest = storyboard2.Storyboard2UpdateSubShotRequest
_verify_episode_permission = storyboard2._verify_episode_permission
_parse_storyboard2_card_ids = storyboard2_reference_images.parse_storyboard2_card_ids
_clean_scene_ai_prompt_text = storyboard2_board.clean_scene_ai_prompt_text
_extract_scene_description_from_card_ids = storyboard2_board.extract_scene_description_from_card_ids
_resolve_storyboard2_scene_override_text = storyboard2_board.resolve_storyboard2_scene_override_text
_pick_storyboard2_source_shots = storyboard2_board.pick_storyboard2_source_shots
_ensure_storyboard2_initialized = storyboard2_board.ensure_storyboard2_initialized
_mark_storyboard2_image_task_active = storyboard2._mark_storyboard2_image_task_active
_mark_storyboard2_image_task_inactive = storyboard2._mark_storyboard2_image_task_inactive
_is_storyboard2_image_task_active = storyboard2._is_storyboard2_image_task_active
_recover_orphan_storyboard2_image_tasks = storyboard2._recover_orphan_storyboard2_image_tasks
_serialize_storyboard2_board = storyboard2_board.serialize_storyboard2_board
_get_storyboard2_sub_shot_with_permission = storyboard2._get_storyboard2_sub_shot_with_permission
_get_storyboard2_shot_with_permission = storyboard2._get_storyboard2_shot_with_permission
_resolve_storyboard2_selected_card_ids = storyboard2_reference_images.resolve_storyboard2_selected_card_ids
_is_scene_subject_card_type = storyboard2_reference_images.is_scene_subject_card_type
_subject_type_sort_key = storyboard2_board.subject_type_sort_key
_get_optional_prompt_config_content = storyboard2._get_optional_prompt_config_content
_save_storyboard2_image_debug = storyboard2._save_storyboard2_image_debug
_save_storyboard2_video_debug = storyboard2._save_storyboard2_video_debug
_normalize_storyboard2_video_status = storyboard2_media.normalize_storyboard2_video_status
_is_storyboard2_video_processing = storyboard2_media.is_storyboard2_video_processing
_build_storyboard2_video_name_tag = storyboard2._build_storyboard2_video_name_tag
_process_storyboard2_video_cover_and_cdn = storyboard2._process_storyboard2_video_cover_and_cdn
_sync_storyboard2_processing_videos = storyboard2._sync_storyboard2_processing_videos
get_storyboard2_data = storyboard2.get_storyboard2_data
batch_generate_storyboard2_sora_prompts = storyboard2.batch_generate_storyboard2_sora_prompts
generate_storyboard2_sub_shot_images = storyboard2.generate_storyboard2_sub_shot_images
generate_storyboard2_sub_shot_video = storyboard2.generate_storyboard2_sub_shot_video
_recover_storyboard2_video_polling = storyboard2._recover_storyboard2_video_polling
update_storyboard2_shot = storyboard2.update_storyboard2_shot
update_storyboard2_sub_shot = storyboard2.update_storyboard2_sub_shot
delete_storyboard2_video = storyboard2.delete_storyboard2_video
set_storyboard2_current_image = storyboard2.set_storyboard2_current_image
delete_storyboard2_image = storyboard2.delete_storyboard2_image

# Compatibility exports for direct callers while video task routes live in api.routers.video.
CancelVideoTasksRequest = video.CancelVideoTasksRequest
query_task_status = video.query_task_status
_normalize_video_task_ids = video._normalize_video_task_ids
_get_user_cancelable_video_task_ids = video._get_user_cancelable_video_task_ids
_cancel_upstream_video_tasks = video._cancel_upstream_video_tasks


async def cancel_video_tasks(
    request: CancelVideoTasksRequest,
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return await video._cancel_video_tasks_impl(
        request,
        user,
        db,
        _cancel_upstream_video_tasks,
    )

# Compatibility exports for direct callers while script routes live in api.routers.scripts.
ScriptCreate = scripts.ScriptCreate
ScriptUpdate = scripts.ScriptUpdate
ScriptResponse = scripts.ScriptResponse
CopyScriptRequest = scripts.CopyScriptRequest
create_script = scripts.create_script
get_my_scripts = scripts.get_my_scripts
get_script = scripts.get_script
update_script = scripts.update_script
delete_script = scripts.delete_script
copy_script = scripts.copy_script

# Compatibility exports for direct callers while hit drama routes live in api.routers.hit_dramas.
normalize_hit_drama_online_time = hit_dramas.normalize_hit_drama_online_time
normalize_hit_drama_payload = hit_dramas.normalize_hit_drama_payload
HitDramaCreate = hit_dramas.HitDramaCreate
HitDramaUpdate = hit_dramas.HitDramaUpdate
HitDramaResponse = hit_dramas.HitDramaResponse
HitDramaHistoryResponse = hit_dramas.HitDramaHistoryResponse
get_hit_dramas = hit_dramas.get_hit_dramas
create_hit_drama = hit_dramas.create_hit_drama
update_hit_drama = hit_dramas.update_hit_drama
delete_hit_drama = hit_dramas.delete_hit_drama
get_hit_drama_history = hit_dramas.get_hit_drama_history
upload_hit_drama_video = hit_dramas.upload_hit_drama_video
import_hit_drama_excel = hit_dramas.import_hit_drama_excel

if __name__ == "__main__":
    import uvicorn
    host = os.getenv("HOST", "0.0.0.0")
    port = int(os.getenv("PORT", "10001"))
    workers = max(1, int(os.getenv("WEB_CONCURRENCY", "1")))

    # 多 worker Web 模式下默认关闭本进程内 poller，避免每个 worker 各自启动一套后台线程。
    if workers > 1 and os.getenv("ENABLE_BACKGROUND_POLLER") is None:
        os.environ["ENABLE_BACKGROUND_POLLER"] = "0"

    if workers > 1:
        uvicorn.run("main:app", host=host, port=port, workers=workers)
    else:
        uvicorn.run(app, host=host, port=port)
